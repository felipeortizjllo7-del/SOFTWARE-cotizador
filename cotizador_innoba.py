# -*- coding: utf-8 -*-
"""
Cotizador INNOBA Colombia DMC
- Itinerario combinado: hasta 5 destinos en una sola cotizacion.
- Tarifas 2026. Precios en USD. TRM en vivo (dolar-colombia.com, -100), NO se muestra.
- Descripciones de tours (TARIFARIO - TOURS 2026).
- Al generar el PDF, se envia por correo (Office 365) al email del cliente.
"""
import os
import sys
import re
import ssl
import json
import uuid
import shutil
import smtplib
import difflib
import tempfile
import datetime
import calendar
import threading
import unicodedata
import urllib.request
from email.message import EmailMessage
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk
from fpdf import FPDF
from PIL import Image


# ============================================================================
# Rutas
# ============================================================================
def app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def recurso(nombre):
    base = getattr(sys, "_MEIPASS", app_dir())
    return os.path.join(base, nombre)

def datos_dir():
    """Carpeta ESCRIBIBLE para datos del usuario (config, clientes).
       Necesaria cuando el programa se instala en Program Files."""
    base = os.path.join(os.environ.get("APPDATA") or app_dir(), "CotizadorInnoba")
    try:
        os.makedirs(base, exist_ok=True)
    except Exception:
        base = app_dir()
    return base

CONFIG_PATH = os.path.join(datos_dir(), "config_empresa.json")


# ============================================================================
# Version del software y actualizaciones automaticas
# ============================================================================
# IMPORTANTE: este numero se incrementa en cada ajuste (lo hace publicar_version.py).
# Esquema resumido de 2 digitos: 1.0 -> 1.1 -> ... -> 1.9 -> 2.0
VERSION = "10.3"
GITHUB_OWNER = "felipeortizjllo7-del"
GITHUB_REPO = "SOFTWARE-cotizador"
# Webhook (Google Apps Script /exec) por donde el HTML de los clientes envia sus
# cotizaciones; el .exe las importa aqui.
WEBHOOK_URL = "https://script.google.com/macros/s/AKfycbzzl9r500lei3AdM2LnIgtw6_n9wmvx1mDlnNPL-xccCX6Bsovb4PbLJBEX9bElYJIpBg/exec"
# Clave para LEER las cotizaciones (solo el .exe la tiene). El HTML NUNCA lee, solo
# envia; asi los clientes no pueden ver lo que cotizan los demas.
WEBHOOK_KEY = "inb_9f3Kx72Qp_seg2026"
# Archivo con la ultima version publicada (rama main del repositorio)
UPDATE_URL = (f"https://raw.githubusercontent.com/{GITHUB_OWNER}/{GITHUB_REPO}"
              f"/main/version.json")

def _ver_tuple(s):
    nums = re.findall(r"\d+", str(s or "0"))
    return tuple(int(x) for x in nums[:3]) if nums else (0,)

def obtener_version_remota():
    """Lee la ultima version publicada. Primero la API de contenidos (sin cache del
       CDN, deteccion instantanea); si falla, el archivo raw."""
    ctx = ssl.create_default_context()
    fuentes = [
        (f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}"
         f"/contents/version.json?ref=main", "application/vnd.github.raw"),
        (UPDATE_URL, "*/*"),
    ]
    for url, accept in fuentes:
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "CotizadorInnoba", "Accept": accept,
                "Cache-Control": "no-cache"})
            with urllib.request.urlopen(req, context=ctx, timeout=12) as r:
                return json.loads(r.read().decode("utf-8"))
        except Exception:
            continue
    return None

def hay_actualizacion():
    """Devuelve el dict de la version remota si es MAYOR a la instalada, si no None."""
    info = obtener_version_remota()
    if info and _ver_tuple(info.get("version", "0")) > _ver_tuple(VERSION):
        return info
    return None


# ============================================================================
# Paleta de marca
# ============================================================================
NAVY = "#013984"; NAVY2 = "#00285F"; BLUE = "#1466C7"; BLUE_H = "#0F4FA0"
CYAN = "#2E8BE6"; BG = "#EEF3FA"; CARD = "#FFFFFF"; CARD2 = "#F4F8FD"
TEXT = "#16233D"; MUTED = "#64748B"; GREEN = "#1E9E5A"; GREEN_H = "#178049"
LINE = "#D7E1EF"; RED = "#C0392B"

def _alto_util_pantalla(fallback=760):
    """Altura utilizable de la pantalla (sin la barra de tareas), en Windows."""
    try:
        import ctypes
        from ctypes import wintypes
        rect = wintypes.RECT()
        # SPI_GETWORKAREA = 0x0030
        ctypes.windll.user32.SystemParametersInfoW(0x0030, 0, ctypes.byref(rect), 0)
        alto = rect.bottom - rect.top
        if alto > 200:
            return alto
    except Exception:
        pass
    return fallback


def aclarar(hexc, f=0.86):
    """Mezcla un color hacia el blanco (f=0..1) para obtener un tinte suave."""
    try:
        h = hexc.lstrip("#")
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        r = int(r + (255 - r) * f); g = int(g + (255 - g) * f); b = int(b + (255 - b) * f)
        return f"#{r:02x}{g:02x}{b:02x}"
    except Exception:
        return "#EEF3FA"


# ============================================================================
# Tasa de cambio (TRM en vivo)
# ============================================================================
TRM_URL = "https://www.dolar-colombia.com/"
# Fuente OFICIAL de la TRM (Superfinanciera via datos.gov.co): estable y con CORS.
TRM_API = ("https://www.datos.gov.co/resource/32sa-8pi3.json"
           "?$order=vigenciadesde%20DESC&$limit=1")
DESCUENTO_DOLAR = 100.0   # descuento por defecto

def _trm_valida(v):
    """Acepta solo valores de TRM con sentido (evita precios negativos por un
       parseo malo). La TRM COP/USD ronda 3000-5000; damos margen amplio."""
    try:
        v = float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return None
    return v if 1000.0 <= v <= 20000.0 else None

# Quien realiza la cotizacion (nombre, cargo) -> firma del PDF
COTIZADORES = [
    ("Felipe Ortiz Jaramillo", "Gerente - Innoba DMC"),
    ("Carlos Ortiz Jaramillo", "Gerente Comercial - Innoba DMC"),
]

def periodo_por_fecha(fecha):
    """Segun la fecha de IDA -> (descuento_pesos, margen_hotel, margen_terrestre).
       margen None = usar el margen normal de cada hoja. Reglas:
       - 2027 (todo el ano): -300, margenes 0.82 / 0.69
       - sep-dic 2026: -200, margenes normales
       - resto (por defecto): -100, margenes normales."""
    if fecha:
        y, m = fecha.year, fecha.month
        if y >= 2027:
            return 300.0, 0.82, 0.69
        if y == 2026 and 9 <= m <= 12:
            return 200.0, None, None
    return 100.0, None, None

def obtener_trm(timeout=12):
    """Devuelve la TRM del dia (COP por USD) o None. Primero la fuente oficial;
       si falla, dolar-colombia.com. Siempre validando el rango."""
    # 1) Fuente oficial (datos.gov.co) - estable y diaria
    for ctx in (ssl.create_default_context(), ssl._create_unverified_context()):
        try:
            req = urllib.request.Request(TRM_API,
                                         headers={"User-Agent": "CotizadorInnoba"})
            with urllib.request.urlopen(req, timeout=timeout, context=ctx) as r:
                data = json.loads(r.read().decode("utf-8"))
            v = _trm_valida(data[0].get("valor")) if data else None
            if v:
                return v
        except Exception:
            continue
    # 2) Respaldo: dolar-colombia.com
    req = urllib.request.Request(
        TRM_URL, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})
    html = None
    for ctx in (ssl.create_default_context(), ssl._create_unverified_context()):
        try:
            with urllib.request.urlopen(req, timeout=timeout, context=ctx) as r:
                html = r.read().decode("utf-8", "replace")
            break
        except Exception:
            continue
    if not html:
        return None
    # tomar el primer numero con formato de miles (ej. 3,252.11) y validarlo
    for m in re.finditer(r"(\d{1,2}[.,]\d{3}[.,]\d{2})", html):
        v = _trm_valida(m.group(1).replace(",", ""))
        if v:
            return v
    return None


# ============================================================================
# Configuracion de la empresa
# ============================================================================
DEFAULT_CONFIG = {
    "empresa": "INNOBA Colombia DMC",
    "nit": "", "direccion": "", "telefono": "", "email": "", "web": "",
    "logo": "",
    "firma_nombre": "Felipe Ortiz",
    "firma_cargo": "Gerente - INNOBA Colombia DMC",
    "correo_remitente": "",
    "smtp_servidor": "smtp.office365.com",
    "smtp_puerto": "587",
    "smtp_password": "",
    "ultima_trm": "", "ultima_trm_fecha": "",
    "notas": ("Tarifas sujetas a disponibilidad al momento de la reserva. "
              "Precios en dolares americanos (USD) por el total indicado."),
}

def cargar_config():
    cfg = dict(DEFAULT_CONFIG)
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg.update(json.load(f))
        except Exception:
            pass
    return cfg

def guardar_config(cfg):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


# ---- Clientes / Empresas (con sus vendedores) ----
CLIENTES_PATH = os.path.join(datos_dir(), "clientes.json")

def _sembrar_clientes():
    """Primera apertura tras instalar: copia la base de clientes incluida
       en el paquete hacia la carpeta escribible del usuario."""
    if os.path.exists(CLIENTES_PATH):
        return
    for semilla in (recurso("clientes.json"),
                    os.path.join(app_dir(), "clientes.json")):
        if os.path.exists(semilla):
            try:
                shutil.copyfile(semilla, CLIENTES_PATH)
            except Exception:
                pass
            return

def cargar_clientes():
    _sembrar_clientes()
    if os.path.exists(CLIENTES_PATH):
        try:
            with open(CLIENTES_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return []

def guardar_clientes(lst):
    with open(CLIENTES_PATH, "w", encoding="utf-8") as f:
        json.dump(lst, f, ensure_ascii=False, indent=2)


# ---- Historial de cotizaciones (con consecutivo) ----
COTIZACIONES_PATH = os.path.join(datos_dir(), "cotizaciones.json")

def cargar_cotizaciones():
    if os.path.exists(COTIZACIONES_PATH):
        try:
            with open(COTIZACIONES_PATH, "r", encoding="utf-8") as f:
                d = json.load(f)
            if isinstance(d, dict) and "items" in d:
                return d
        except Exception:
            pass
    return {"seq": 0, "items": []}

def guardar_cotizaciones(data):
    with open(COTIZACIONES_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def respaldar_datos(cfg):
    """Guarda copias de seguridad de cotizaciones/clientes en %APPDATA%\\...\\respaldos.
       Siempre deja un respaldo '_ultimo'; y si la version cambio (hubo una
       actualizacion) deja una copia permanente por version, para que las
       cotizaciones NUNCA se pierdan al actualizar."""
    try:
        carpeta = os.path.join(datos_dir(), "respaldos")
        os.makedirs(carpeta, exist_ok=True)
        cambio_version = cfg.get("ultima_version_vista", "") != VERSION
        hoy = datetime.date.today().strftime("%Y%m%d")
        for nombre in ("cotizaciones.json", "clientes.json", "config_empresa.json",
                       "reservas.json", "tareas.json"):
            src = os.path.join(datos_dir(), nombre)
            if not os.path.exists(src):
                continue
            base = nombre.rsplit(".", 1)[0]
            shutil.copyfile(src, os.path.join(carpeta, base + "_ultimo.json"))
            if cambio_version:
                shutil.copyfile(
                    src, os.path.join(carpeta, f"{base}_v{VERSION}_{hoy}.json"))
        if cambio_version:
            cfg["ultima_version_vista"] = VERSION
            try:
                guardar_config(cfg)
            except Exception:
                pass
    except Exception:
        pass


def respaldo_antes_actualizar():
    """Copia PERMANENTE de los datos justo antes de instalar una actualizacion, con sello
       de fecha/hora, para que nunca se pierda nada. Devuelve la carpeta de respaldos."""
    carpeta = os.path.join(datos_dir(), "respaldos")
    try:
        os.makedirs(carpeta, exist_ok=True)
        sello = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        for nombre in ("cotizaciones.json", "reservas.json", "clientes.json",
                       "tareas.json", "config_empresa.json"):
            src = os.path.join(datos_dir(), nombre)
            if os.path.exists(src):
                base = nombre.rsplit(".", 1)[0]
                shutil.copyfile(
                    src, os.path.join(carpeta, f"{base}_ANTES_v{VERSION}_{sello}.json"))
    except Exception:
        pass
    return carpeta

def peek_numero_cotizacion():
    """Devuelve el proximo consecutivo (sin reservarlo aun), ej. COT-00001."""
    return f"COT-{cargar_cotizaciones().get('seq', 0) + 1:05d}"

def registrar_cotizacion(rec):
    """Asigna el consecutivo, guarda el registro y devuelve el numero asignado.
       Ademas sube la cotizacion a la nube para compartirla con los demas equipos."""
    data = cargar_cotizaciones()
    data["seq"] = int(data.get("seq", 0)) + 1
    numero = f"COT-{data['seq']:05d}"
    rec = dict(rec); rec["numero"] = numero
    if not rec.get("uid"):
        rec["uid"] = "COTX-" + uuid.uuid4().hex[:12]
    rec.setdefault("origen", "Escritorio")
    data["items"].append(rec)
    guardar_cotizaciones(data)
    try:
        if enviar_cotizacion_nube(rec):
            rec["nube_ok"] = True
            guardar_cotizaciones(data)
    except Exception:
        pass
    return numero

def importar_cotizaciones_html():
    """Trae las cotizaciones creadas por clientes en el HTML (via WEBHOOK_URL) y
       las agrega al historial local sin duplicar. Devuelve cuantas nuevas."""
    if not WEBHOOK_URL:
        return 0
    try:
        ctx = ssl.create_default_context()
        url = WEBHOOK_URL + ("&" if "?" in WEBHOOK_URL else "?") + "key=" + WEBHOOK_KEY
        req = urllib.request.Request(url, headers={"User-Agent": "CotizadorInnoba"})
        with urllib.request.urlopen(req, context=ctx, timeout=20) as r:
            remotas = json.loads(r.read().decode("utf-8"))
    except Exception:
        return 0
    if isinstance(remotas, dict):
        remotas = remotas.get("items", [])
    if not isinstance(remotas, list):
        return 0
    data = cargar_cotizaciones()
    por_wid = {str(it.get("web_id")): it for it in data["items"] if it.get("web_id")}
    por_uid = {str(it.get("uid")): it for it in data["items"] if it.get("uid")}
    cambios = 0
    borrados = set()
    for rc in remotas:
        if not isinstance(rc, dict):
            continue
        uid = str(rc.get("uid") or "")
        wid = str(rc.get("id") or rc.get("web_id") or "")
        # marca de borrado (tombstone): quitar esa cotizacion en este equipo
        if rc.get("_accion") == "borrar" or rc.get("borrar"):
            borrados.add(uid); borrados.add(wid)
            continue
        es_exe = bool(uid) and not wid.startswith("WEB-") and rc.get("origen", "") != "HTML (cliente)"
        if es_exe:
            # --- Cotizacion hecha en el .exe (Felipe/Carlos) ---
            local = por_uid.get(uid)
            if local is not None:
                if _merge_cot(local, rc):
                    cambios += 1
                continue
            rc2 = {k: v for k, v in rc.items() if k != "tipo"}
            data["seq"] = int(data.get("seq", 0)) + 1
            rc2["numero"] = f"COT-{data['seq']:05d}"   # numero unico en este equipo
            rc2.setdefault("estado", "Pendiente")
            rc2["pdf"] = ""            # el PDF vive en el otro equipo; se regenera del snapshot
            rc2["nube_ok"] = True      # ya esta en la nube: este equipo no la re-sube
            data["items"].append(rc2); por_uid[uid] = rc2; cambios += 1
            continue
        # --- Cotizacion del HTML (cliente/agencia) ---
        if not wid:
            continue
        dests = rc.get("destinos", [])
        if isinstance(dests, str):
            dests = [d.strip() for d in dests.split(",") if d.strip()]
        try:
            total = float(rc.get("total", 0) or 0)
        except (TypeError, ValueError):
            total = 0.0
        try:
            total_cliente = float(rc.get("total_cliente", 0) or 0)
        except (TypeError, ValueError):
            total_cliente = 0.0
        try:
            ganancia = float(str(rc.get("ganancia", "") or "").replace(",", "") or 0)
        except (TypeError, ValueError):
            ganancia = 0.0
        local = por_wid.get(wid)
        if local is not None:
            # ya existe: sincronizar gestion + ACTUALIZAR contenido si el cliente la re-envio
            # (editada). Conserva numero/estado/tareas/seguimiento (gestion del .exe).
            cambio = _merge_cot(local, rc)
            for campo, val in (("cliente", rc.get("cliente", "")), ("asesor", rc.get("asesor", "")),
                               ("asesor_tel", rc.get("asesor_tel", "")), ("email", rc.get("email", "")),
                               ("fecha", rc.get("fecha", "")), ("fechas_viaje", rc.get("fechas_viaje", "")),
                               ("destinos", dests), ("total", total), ("total_cliente", total_cliente),
                               ("ganancia_agencia", ganancia), ("snapshot", rc.get("snapshot"))):
                if val not in (None, "", []) and local.get(campo) != val:
                    local[campo] = val; cambio = True
            if cambio:
                cambios += 1
            continue
        data["seq"] = int(data.get("seq", 0)) + 1
        nuevo = {
            "numero": f"COT-{data['seq']:05d}", "web_id": wid, "origen": "HTML (cliente)",
            "cliente": rc.get("cliente", ""), "asesor": rc.get("asesor", ""),
            "asesor_tel": rc.get("asesor_tel", ""), "email": rc.get("email", ""),
            "fecha": rc.get("fecha", ""), "fechas_viaje": rc.get("fechas_viaje", ""),
            "destinos": dests, "total": total, "total_cliente": total_cliente,
            "ganancia_agencia": ganancia, "estado": rc.get("estado", "") or "Pendiente",
            "pdf": "", "snapshot": rc.get("snapshot")}
        # traer tambien la gestion si ya venia en la nube (estado/seguimiento/reserva)
        _merge_cot(nuevo, rc)
        data["items"].append(nuevo); por_wid[wid] = nuevo; cambios += 1
    # aplicar borrados (tombstones)
    if borrados:
        borrados.discard("")
        antes = len(data["items"])
        data["items"] = [x for x in data["items"]
                         if str(x.get("uid")) not in borrados
                         and str(x.get("web_id")) not in borrados]
        cambios += (antes - len(data["items"]))
    if cambios:
        guardar_cotizaciones(data)
    return cambios


# Campos de GESTION que se sincronizan entre equipos (no tocan numero/pdf/uid locales)
_COT_SYNC_FIELDS = ("estado", "fecha_seg", "correo_seg", "notas", "tareas",
                    "reserva_creada", "auto_correo_seg", "correo_seg_enviado")


def _merge_cot(local, rc):
    """Copia a 'local' los campos de gestion que cambiaron en 'rc'. Devuelve True si hubo
       cambios. Conserva numero, pdf, uid, web_id y origen del registro local."""
    cambio = False
    for f in _COT_SYNC_FIELDS:
        if f in rc and local.get(f) != rc.get(f):
            local[f] = rc.get(f)
            cambio = True
    return cambio


def enviar_cotizacion_nube(rec):
    """Sube/actualiza una cotizacion en el webhook para que su gestion (estado, seguimiento,
       'enviada a reserva'...) se vea en los demas equipos. Sirve para las del .exe y las del
       HTML. Best-effort. Requiere 'uid' o 'web_id'."""
    if not WEBHOOK_URL:
        return False
    try:
        key = rec.get("uid") or rec.get("web_id")
        if not key:
            return False
        payload = {k: v for k, v in rec.items()
                   if not str(k).startswith("_") and k != "pdf"}
        payload["tipo"] = "cotizacion"          # bucket 'cotizaciones' del webhook
        payload["id"] = str(key)                # clave con la que el webhook deduplica
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            WEBHOOK_URL, data=body, method="POST",
            headers={"Content-Type": "text/plain;charset=utf-8",
                     "User-Agent": "CotizadorInnoba"})
        ctx = ssl.create_default_context()
        with urllib.request.urlopen(req, context=ctx, timeout=15) as r:
            r.read()
        return True
    except Exception:
        return False


def depurar_cotizaciones_vencidas():
    """Elimina las cotizaciones que pasado 1 MES desde su fecha de elaboracion NO se
       convirtieron en reserva. Respeta las 'Ganada' y las 'En seguimiento' (gestion
       activa). Tambien las borra de la nube. Devuelve la lista de las eliminadas."""
    data = cargar_cotizaciones()
    hoy = datetime.date.today()
    quitar = []
    for it in data["items"]:
        if it.get("reserva_creada"):
            continue                       # ya se convirtio en reserva
        if it.get("estado") in ("Ganada", "En seguimiento"):
            continue                       # gestion activa / ganada -> respetar
        f = parse_fecha(it.get("fecha", ""))
        if not f:
            continue
        if add_months(f, 1) < hoy:         # ya paso 1 mes desde la elaboracion
            quitar.append(it)
    if not quitar:
        return []
    claves = [(it.get("uid") or it.get("web_id")) for it in quitar]
    ids_quitar = {id(it) for it in quitar}
    data["items"] = [it for it in data["items"] if id(it) not in ids_quitar]
    guardar_cotizaciones(data)
    for k in claves:
        try:
            enviar_borrado_nube("cotizacion", k)
        except Exception:
            pass
    return quitar


def enviar_borrado_nube(tipo, clave):
    """Le pide al webhook que BORRE de la nube el registro (reserva/cotizacion) con esa
       clave (uid/web_id), para que el borrado se propague a todos. Best-effort."""
    if not WEBHOOK_URL or not clave:
        return False
    try:
        payload = {"tipo": tipo, "id": str(clave), "uid": str(clave), "_accion": "borrar"}
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            WEBHOOK_URL, data=body, method="POST",
            headers={"Content-Type": "text/plain;charset=utf-8",
                     "User-Agent": "CotizadorInnoba"})
        ctx = ssl.create_default_context()
        with urllib.request.urlopen(req, context=ctx, timeout=15) as r:
            r.read()
        return True
    except Exception:
        return False


def subir_cotizaciones_exe():
    """Sube a la nube todas las cotizaciones locales hechas en el .exe (para compartirlas
       con los demas equipos). Devuelve cuantas subio. Asigna 'uid' si falta."""
    data = cargar_cotizaciones()
    cambio = False
    subidas = 0
    for it in data["items"]:
        if it.get("origen", "") == "HTML (cliente)":
            continue
        if it.get("nube_ok"):
            continue   # ya subida (o importada de otro equipo): no repetir
        if not it.get("uid"):
            it["uid"] = "COTX-" + uuid.uuid4().hex[:12]
            cambio = True
        if enviar_cotizacion_nube(it):
            it["nube_ok"] = True
            cambio = True
            subidas += 1
    if cambio:
        guardar_cotizaciones(data)
    return subidas


# ---- Sincronizacion de RESERVAS en la nube (mismo webhook, tipo=reserva) ----
def enviar_reserva_nube(rec):
    """Sube (o actualiza) una reserva al webhook para que los demas equipos la vean.
       Best-effort: no lanza excepciones. Devuelve True/False."""
    if not WEBHOOK_URL:
        return False
    try:
        payload = {k: v for k, v in rec.items() if not str(k).startswith("_")}
        payload["tipo"] = "reserva"
        if not payload.get("uid"):
            return False
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            WEBHOOK_URL, data=body, method="POST",
            headers={"Content-Type": "text/plain;charset=utf-8",
                     "User-Agent": "CotizadorInnoba"})
        ctx = ssl.create_default_context()
        with urllib.request.urlopen(req, context=ctx, timeout=15) as r:
            r.read()
        return True
    except Exception:
        return False


def importar_reservas_nube():
    """Trae las reservas de la nube y las mezcla en el archivo local (por 'uid', sin
       duplicar; si ya existe y cambio, la actualiza). Devuelve cuantas nuevas/cambiadas."""
    if not WEBHOOK_URL:
        return 0
    try:
        ctx = ssl.create_default_context()
        url = (WEBHOOK_URL + ("&" if "?" in WEBHOOK_URL else "?")
               + "key=" + WEBHOOK_KEY + "&tipo=reserva")
        req = urllib.request.Request(url, headers={"User-Agent": "CotizadorInnoba"})
        with urllib.request.urlopen(req, context=ctx, timeout=20) as r:
            remotas = json.loads(r.read().decode("utf-8"))
    except Exception:
        return 0
    if isinstance(remotas, dict):
        remotas = remotas.get("items", [])
    if not isinstance(remotas, list):
        return 0
    data = cargar_reservas()
    por_uid = {str(it.get("uid")): it for it in data["items"] if it.get("uid")}
    nums_local = {str(it.get("numero", "")) for it in data["items"]}
    cambios = 0
    max_num = 0
    borrados = set()
    for rc in remotas:
        if not isinstance(rc, dict):
            continue
        uid = str(rc.get("uid") or "")
        if not uid:
            continue
        # marca de borrado (tombstone): eliminar esa reserva en este equipo
        if rc.get("_accion") == "borrar" or rc.get("borrar"):
            borrados.add(uid)
            continue
        rc = {k: v for k, v in rc.items() if k != "tipo"}
        try:
            max_num = max(max_num, int(re.sub(r"\D", "", str(rc.get("numero", "0"))) or 0))
        except Exception:
            pass
        if uid in por_uid:
            local = por_uid[uid]
            a = json.dumps(local, sort_keys=True, ensure_ascii=False)
            b = json.dumps(rc, sort_keys=True, ensure_ascii=False)
            if a != b:
                # adoptar el registro remoto (incluye el numero canonico de la nube);
                # asi una reorganizacion del consecutivo se propaga a todos los equipos
                local.clear(); local.update(rc); cambios += 1
        else:
            # si el numero choca con uno local de OTRA reserva, renumerar la entrante
            num = str(rc.get("numero", "") or "")
            if not num or num in nums_local:
                nuevo = max(int(data.get("seq", RES_SEQ_INICIAL) or RES_SEQ_INICIAL),
                            max_num) + 1
                data["seq"] = nuevo
                rc["numero"] = str(nuevo)
                num = str(nuevo)
            nums_local.add(num)
            por_uid[uid] = rc
            data["items"].append(rc); cambios += 1
    # aplicar borrados (tombstones): quitar localmente lo que se borro en otro equipo
    if borrados:
        antes = len(data["items"])
        data["items"] = [x for x in data["items"] if str(x.get("uid")) not in borrados]
        cambios += (antes - len(data["items"]))
    # mantener el consecutivo por encima del mayor numero visto
    if max_num:
        data["seq"] = max(int(data.get("seq", RES_SEQ_INICIAL) or RES_SEQ_INICIAL), max_num)
    if cambios:
        guardar_reservas(data)
    return cambios


def compactar_consecutivo_reservas():
    """Renumera las reservas de forma CONTINUA (2989, 2990, 2991, ...) en orden de
       creacion, para quitar saltos en el consecutivo, y las re-sube a la nube para
       que todos los equipos adopten el mismo numero. Devuelve el ultimo numero usado."""
    data = cargar_reservas()

    def clave(it):
        f = str(it.get("fecha_creacion", "") or "")
        try:
            n = int(re.sub(r"\D", "", str(it.get("numero", "0"))) or 0)
        except Exception:
            n = 0
        return (f or "9999", n)

    orden = sorted(data.get("items", []), key=clave)
    seq = RES_SEQ_INICIAL
    for it in orden:
        seq += 1
        it["numero"] = str(seq)
        if not it.get("uid"):
            it["uid"] = "RES-" + uuid.uuid4().hex[:12]
    data["seq"] = seq
    guardar_reservas(data)
    # re-subir para que los demas equipos adopten los numeros nuevos
    try:
        for it in orden:
            enviar_reserva_nube(it)
    except Exception:
        pass
    return seq


def sincronizar_reservas_nube():
    """Sube las reservas locales que aun no estan en la nube y luego importa las remotas.
       Devuelve (subidas, importadas)."""
    subidas = 0
    try:
        for r in cargar_reservas().get("items", []):
            if r.get("uid") and enviar_reserva_nube(r):
                subidas += 1
    except Exception:
        pass
    importadas = importar_reservas_nube()
    return subidas, importadas


def parse_fecha(s):
    """Convierte 'dd/mm/aaaa' (o similares) a date; None si no se puede."""
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def seguimientos_pendientes(data=None):
    """Cotizaciones (no cerradas) cuya fecha de seguimiento o alguna tarea ya
       vencio (<= hoy). Devuelve lista de (item, motivo)."""
    data = data or cargar_cotizaciones()
    hoy = datetime.date.today()
    res = []
    for it in data.get("items", []):
        if it.get("estado") in ("Ganada", "Perdida"):
            continue
        motivos = []
        fs = parse_fecha(it.get("fecha_seg", ""))
        if fs and fs <= hoy:
            motivos.append(f"seguimiento {it.get('fecha_seg')}")
        for t in it.get("tareas", []):
            if t.get("hecha"):
                continue
            ft = parse_fecha(t.get("fecha", ""))
            if ft and ft <= hoy:
                motivos.append(f"tarea: {t.get('texto','')[:30]} ({t.get('fecha')})")
        if motivos:
            res.append((it, "; ".join(motivos)))
    return res

def exportar_clientes_excel(lst, ruta):
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Clientes"
    ws.append(["Empresa", "NIT/Documento", "Telefono", "Email", "Sitio web", "Pais",
               "Vendedor", "Vendedor telefono", "Vendedor email", "Vendedor cargo"])
    for c in lst:
        vends = c.get("vendedores") or [{}]
        for v in vends:
            ws.append([c.get("empresa", ""), c.get("nit", ""), c.get("telefono", ""),
                       c.get("email", ""), c.get("web", ""), c.get("pais", ""),
                       v.get("nombre", ""), v.get("telefono", ""), v.get("email", ""),
                       v.get("cargo", "")])
    wb.save(ruta)

def _nz(s):
    s = "".join(ch for ch in unicodedata.normalize("NFD", str(s or ""))
                if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", s).lower().strip()


def _num(s, defecto=0):
    """Convierte a numero; devuelve int si es entero (100), float si tiene decimales."""
    try:
        v = float(str(s).replace(",", "").strip() or defecto)
        return int(v) if v == int(v) else v
    except Exception:
        return defecto

def _leer_filas(ruta):
    """Lee filas de un .xlsx o .csv (detecta codificacion y delimitador)."""
    ext = os.path.splitext(ruta)[1].lower()
    if ext in (".csv", ".txt"):
        import csv, io
        raw = open(ruta, "rb").read()
        texto = None
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                texto = raw.decode(enc); break
            except UnicodeDecodeError:
                continue
        muestra = texto[:2000]
        delim = ";" if muestra.count(";") >= muestra.count(",") else ","
        return list(csv.reader(io.StringIO(texto), delimiter=delim))
    import openpyxl
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    return [row for row in wb.active.iter_rows(values_only=True)]

def importar_clientes_excel(ruta):
    """Importa clientes de un Excel/CSV. Reconoce nuestro formato y exportaciones
       tipo Bitrix (Nombre + Apellido + Compania)."""
    filas = _leer_filas(ruta)
    if not filas:
        return []
    header = [_nz(x) for x in filas[0]]
    def col(keys, exclude=()):
        for i, h in enumerate(header):
            if any(k in h for k in keys) and not any(e in h for e in exclude):
                return i
        return None
    i_comp = col(["compania", "empresa", "company", "razon social"])
    i_cli = col(["cliente"], exclude=["tipo"])
    i_emp = i_comp if i_comp is not None else i_cli
    i_nom = col(["nombre"], exclude=["compania", "empresa", "segundo"])
    i_ape = col(["apellido"])
    i_nit = col(["nit", "documento", "ruc", "id fiscal", "identificacion"], exclude=["tipo"])
    i_tel = col(["telefono del trabajo", "telefono trabajo", "telefono", "phone", "celular", "movil"],
                exclude=["vendedor", "sms", "localizador", "fax", "casa"])
    i_email = col(["e-mail del trabajo", "email del trabajo", "email", "correo", "e-mail"],
                  exclude=["vendedor", "boletines", "casa"])
    i_web = col(["sitio web", "web", "url"], exclude=["facebook", "vk"])
    i_pais = col(["pais", "country"], exclude=["codigo"])
    i_cargo = col(["cargo", "rol", "puesto"])
    i_vnom = col(["vendedor", "asesor"], exclude=["telefono", "email", "correo", "cargo"])
    i_vtel = col(["vendedor telefono", "telefono vendedor"])
    i_vemail = col(["vendedor email", "vendedor correo", "email vendedor"])
    if i_emp is None and i_nom is None:
        i_emp = 0
    def val(row, i):
        return str(row[i]).strip() if (i is not None and i < len(row) and row[i] is not None) else ""

    def es_junk(t):
        t = t.lower()
        return any(k in t for k in ("no-reply", "noreply", "no_reply", "bitrix24",
                                    "google-noreply", "@google.com", "imol|facebook"))

    empresas = {}
    orden = []
    for row in filas[1:]:
        if not any(row):
            continue
        nombre = (val(row, i_nom) + " " + val(row, i_ape)).strip().lstrip("?").strip()
        email = val(row, i_email)
        tel = val(row, i_tel)
        if tel and "e+" in tel.lower():
            tel = ""
        cargo = val(row, i_cargo)
        emp = val(row, i_emp) or val(row, i_vnom)
        if not emp:
            emp = nombre or email
        if not emp or es_junk(nombre + " " + email + " " + emp):
            continue
        key = _nz(emp)
        if key not in empresas:
            empresas[key] = {"empresa": emp, "nit": val(row, i_nit), "telefono": tel,
                             "email": email, "web": val(row, i_web),
                             "pais": val(row, i_pais), "vendedores": []}
            orden.append(key)
        e = empresas[key]
        if not e["email"] and email:
            e["email"] = email
        if not e["telefono"] and tel:
            e["telefono"] = tel
        if not e["nit"]:
            e["nit"] = val(row, i_nit)
        vnom = val(row, i_vnom) or (nombre if _nz(nombre) != key else "")
        if vnom and not any(_nz(v["nombre"]) == _nz(vnom) for v in e["vendedores"]):
            e["vendedores"].append({"nombre": vnom, "telefono": val(row, i_vtel) or tel,
                                    "email": val(row, i_vemail) or email, "cargo": cargo})
    return [empresas[k] for k in orden]


def add_months(fecha, meses):
    m = fecha.month - 1 + meses
    y = fecha.year + m // 12
    m = m % 12 + 1
    d = min(fecha.day, calendar.monthrange(y, m)[1])
    return datetime.date(y, m, d)


# ============================================================================
# Correo (Office 365 / SMTP)
# ============================================================================
def enviar_correo(cfg, destinatario, asunto, cuerpo, adjunto):
    remit = cfg.get("correo_remitente", "").strip()
    servidor = (cfg.get("smtp_servidor", "") or "smtp.office365.com").strip()
    try:
        puerto = int(cfg.get("smtp_puerto", "587") or 587)
    except Exception:
        puerto = 587
    pw = cfg.get("smtp_password", "")
    if not remit or not pw:
        raise ValueError("Falta configurar el correo remitente y su contrasena "
                         "en 'Datos de mi empresa'.")
    if not destinatario:
        raise ValueError("El cliente no tiene email.")
    msg = EmailMessage()
    msg["From"] = remit
    msg["To"] = destinatario
    msg["Subject"] = asunto
    msg.set_content(cuerpo)
    with open(adjunto, "rb") as f:
        msg.add_attachment(f.read(), maintype="application", subtype="pdf",
                           filename=os.path.basename(adjunto))
    with smtplib.SMTP(servidor, puerto, timeout=30) as s:
        s.ehlo()
        s.starttls(context=ssl.create_default_context())
        s.ehlo()
        s.login(remit, pw)
        s.send_message(msg)


def enviar_correo_texto(cfg, destinatario, asunto, cuerpo):
    """Envia un correo solo de texto (sin adjunto), p.ej. el seguimiento comercial."""
    remit = cfg.get("correo_remitente", "").strip()
    servidor = (cfg.get("smtp_servidor", "") or "smtp.office365.com").strip()
    try:
        puerto = int(cfg.get("smtp_puerto", "587") or 587)
    except Exception:
        puerto = 587
    pw = cfg.get("smtp_password", "")
    if not remit or not pw:
        raise ValueError("Falta configurar el correo remitente y su contrasena "
                         "en 'Datos de mi empresa'.")
    if not destinatario:
        raise ValueError("No hay correo del cliente.")
    msg = EmailMessage()
    msg["From"] = remit
    msg["To"] = destinatario
    msg["Subject"] = asunto
    msg.set_content(cuerpo)
    with smtplib.SMTP(servidor, puerto, timeout=30) as s:
        s.ehlo()
        s.starttls(context=ssl.create_default_context())
        s.ehlo()
        s.login(remit, pw)
        s.send_message(msg)


def cuerpo_seguimiento_cotizacion(item, cfg):
    """Texto del correo de seguimiento a una cotizacion, firmado por quien la hizo."""
    quien = (item.get("cotizado_por") or item.get("asesor") or "").strip()
    if not quien and cfg:
        quien = (cfg.get("firma_nombre", "") or "").strip()
    empresa = (cfg.get("empresa", "") if cfg else "") or "INNOBA Colombia DMC"
    tel = (cfg.get("telefono", "") if cfg else "")
    cliente = (item.get("cliente", "") or "").strip()
    saludo = f"Hola{(' ' + cliente) if cliente else ''},\n\n"
    firma = "\n".join(x for x in [quien, empresa, ("Cel: " + tel) if tel else ""] if x)
    return (saludo +
            "Espero que te encuentres muy bien.\n\n"
            "Queria hacer un breve seguimiento a la cotizacion que te envie hace unos dias "
            "para conocer si tuviste la oportunidad de revisarla y saber si existe alguna "
            "inquietud o informacion adicional en la que podamos apoyarte.\n\n"
            "Para nosotros sera un gusto acompanarte en este proyecto y adaptar la propuesta, "
            "si es necesario, para que se ajuste a tus necesidades.\n\n"
            "Quedo muy atento(a) a tus comentarios y a cualquier consulta que tengas. Sera un "
            "placer ayudarte a hacer realidad esta experiencia.\n\n"
            "Muchas gracias por tu tiempo y quedo pendiente de tu respuesta.\n\n"
            "Cordialmente,\n" + firma)


def cfg_remitente_cotizador(cfg, item):
    """Copia de cfg con el correo remitente/clave del cotizador que hizo la cotizacion.
       Felipe -> correo_felipe/pass_felipe; Carlos -> correo_carlos/pass_carlos; si no,
       usa el remitente por defecto (correo_remitente/smtp_password)."""
    c = dict(cfg or {})
    quien = _quien_cerro(item)
    if quien == "Felipe" and c.get("correo_felipe"):
        c["correo_remitente"] = c.get("correo_felipe", "")
        c["smtp_password"] = c.get("pass_felipe", "")
    elif quien == "Carlos" and c.get("correo_carlos"):
        c["correo_remitente"] = c.get("correo_carlos", "")
        c["smtp_password"] = c.get("pass_carlos", "")
    return c


def procesar_correos_programados(cfg):
    """Envia los correos de seguimiento PROGRAMADOS cuya fecha ya llego (una vez).
       Cada uno sale del correo del cotizador que la hizo. Devuelve numeros enviados."""
    data = cargar_cotizaciones()
    hoy = datetime.date.today()
    enviados = []
    cambiado = False
    for it in data.get("items", []):
        if not it.get("auto_correo_seg"):
            continue
        if it.get("estado") in ("Ganada", "Perdida"):
            continue
        fseg = parse_fecha(it.get("fecha_seg", ""))
        if not fseg or fseg > hoy:
            continue
        ya = _parse_fecha_iso(it.get("correo_seg_enviado", ""))
        if ya and ya >= fseg:      # ya se envio para este ciclo de seguimiento
            continue
        dest = (it.get("email", "") or "").strip()
        if not dest:
            continue
        cfgr = cfg_remitente_cotizador(cfg, it)
        if not (cfgr.get("correo_remitente") and cfgr.get("smtp_password")):
            continue
        asunto = f"Seguimiento de su cotizacion {it.get('numero','')} - {cfgr.get('empresa','')}"
        try:
            enviar_correo_texto(cfgr, dest, asunto, cuerpo_seguimiento_cotizacion(it, cfgr))
            it["correo_seg_enviado"] = hoy.isoformat()
            enviados.append(it.get("numero", ""))
            cambiado = True
        except Exception:
            pass
    if cambiado:
        guardar_cotizaciones(data)
    return enviados


def _ics_recordatorio(item, fecha_dt):
    """Construye un evento de calendario (todo el dia) con recordatorio para el
       seguimiento de una cotizacion."""
    uid = (item.get("numero", "COT") + "-seg@innobadmc.com")
    ymd = fecha_dt.strftime("%Y%m%d")
    fin = (fecha_dt + datetime.timedelta(days=1)).strftime("%Y%m%d")
    stamp = datetime.datetime.now().strftime("%Y%m%dT%H%M%S")
    cli = (item.get("cliente", "") or "").replace(",", " ")
    ase = (item.get("asesor", "") or "")
    summ = f"Seguimiento cotizacion {item.get('numero','')} - {cli}"
    desc = (f"Dar seguimiento a la cotizacion {item.get('numero','')} de {cli}. "
            f"Asesor: {ase}. Destinos: {', '.join(item.get('destinos', []))}. "
            f"Total: {usd(item.get('total', 0))}.")
    return ("BEGIN:VCALENDAR\r\nVERSION:2.0\r\nPRODID:-//INNOBA//Cotizador//ES\r\n"
            "CALSCALE:GREGORIAN\r\nMETHOD:PUBLISH\r\nBEGIN:VEVENT\r\n"
            f"UID:{uid}\r\nDTSTAMP:{stamp}\r\n"
            f"DTSTART;VALUE=DATE:{ymd}\r\nDTEND;VALUE=DATE:{fin}\r\n"
            f"SUMMARY:{summ}\r\nDESCRIPTION:{desc}\r\n"
            "BEGIN:VALARM\r\nTRIGGER:-PT9H\r\nACTION:DISPLAY\r\n"
            "DESCRIPTION:Recordatorio de seguimiento\r\nEND:VALARM\r\n"
            "END:VEVENT\r\nEND:VCALENDAR\r\n")

def enviar_recordatorio_ics(cfg, destinatarios, item, fecha_dt):
    """Envia por correo una invitacion de calendario (.ics) con el recordatorio."""
    remit = cfg.get("correo_remitente", "").strip()
    pw = cfg.get("smtp_password", "")
    if not remit or not pw:
        raise ValueError("Falta configurar el correo remitente y su contrasena.")
    dest = [d for d in destinatarios if d]
    if not dest:
        raise ValueError("No hay destinatario para el recordatorio.")
    servidor = (cfg.get("smtp_servidor", "") or "smtp.office365.com").strip()
    try:
        puerto = int(cfg.get("smtp_puerto", "587") or 587)
    except Exception:
        puerto = 587
    ics = _ics_recordatorio(item, fecha_dt)
    msg = EmailMessage()
    msg["From"] = remit
    msg["To"] = ", ".join(dest)
    msg["Subject"] = (f"Recordatorio seguimiento {item.get('numero','')} - "
                      f"{item.get('cliente','')} ({fecha_dt.strftime('%d/%m/%Y')})")
    msg.set_content(
        f"Recordatorio automatico: dar seguimiento a la cotizacion "
        f"{item.get('numero','')} de {item.get('cliente','')} el "
        f"{fecha_dt.strftime('%d/%m/%Y')}.\n\nAsesor: {item.get('asesor','')}\n"
        f"Adjuntamos una invitacion de calendario con recordatorio.")
    msg.add_attachment(ics.encode("utf-8"), maintype="text", subtype="calendar",
                       filename="seguimiento.ics", params={"method": "PUBLISH"})
    with smtplib.SMTP(servidor, puerto, timeout=30) as s:
        s.ehlo(); s.starttls(context=ssl.create_default_context()); s.ehlo()
        s.login(remit, pw); s.send_message(msg)


# ============================================================================
# Base de precios y logica
# ============================================================================
def cargar_precios():
    for ruta in (os.path.join(app_dir(), "precios_2026.json"), recurso("precios_2026.json")):
        if os.path.exists(ruta):
            with open(ruta, "r", encoding="utf-8") as f:
                return json.load(f)
    raise FileNotFoundError("No se encontro 'precios_2026.json'.")

def cargar_precios_seguro():
    """Como cargar_precios pero devuelve {} si falla (para el modulo Reservas)."""
    try:
        return cargar_precios()
    except Exception:
        return {}

def hoteles_por_destino(precios, destino):
    """Nombres de hoteles que maneja INNOBA para ese destino (desde precios_2026)."""
    if not precios or not destino:
        return []
    clave = None
    for k in precios:
        if k.strip().lower() == destino.strip().lower():
            clave = k
            break
    if clave is None:
        return []
    nombres = []
    for h in ((precios.get(clave, {}) or {}).get("hoteles", {}) or {}).get("hoteles", []):
        n = (h.get("nombre", "") or "").strip()
        if n and n not in nombres:
            nombres.append(n)
    return sorted(nombres)


def _clave_destino(precios, destino):
    for k in (precios or {}):
        if k.strip().lower() == (destino or "").strip().lower():
            return k
    return None


def _tasa_tarifario(precios, destino, cfg):
    """TRM a usar: la del config si es valida, si no la embebida en el tarifario."""
    if cfg:
        t = _trm_valida(cfg.get("ultima_trm", ""))
        if t:
            return t
    dd = precios.get(destino, {}) if precios else {}
    for sec in ("hoteles", "terrestres"):
        try:
            v = float(dd.get(sec, {}).get("tasa", 0) or 0)
            if v > 100:
                return v
        except Exception:
            pass
    return 4000.0


def hoteles_detalle(precios, destino):
    """Lista de dicts de hoteles del destino (nombre, temporada, sencilla/doble/triple...)."""
    clave = _clave_destino(precios, destino)
    if clave is None:
        return []
    return ((precios.get(clave, {}) or {}).get("hoteles", {}) or {}).get("hoteles", [])


def servicios_terrestres(precios, destino):
    """Lista de servicios terrestres (tours/traslados) del destino."""
    clave = _clave_destino(precios, destino)
    if clave is None:
        return []
    return ((precios.get(clave, {}) or {}).get("terrestres", {}) or {}).get("servicios", [])


def precio_hotel_usd_pp(precios, destino, hotel, acomodacion="doble", noches=1, cfg=None):
    """Precio por persona en USD de un hotel (acomodacion) para 'noches' noches."""
    clave = _clave_destino(precios, destino)
    dd = (precios.get(clave, {}) or {}).get("hoteles", {}) if clave else {}
    margen = float(dd.get("margen", 0.88) or 0.88)
    tasa = _tasa_tarifario(precios, clave or destino, cfg)
    ocup = {"sencilla": 1, "doble": 2, "triple": 3}.get(acomodacion, 2)
    room = float(hotel.get(acomodacion, 0) or 0)
    if not room or not tasa or not margen:
        return 0.0
    pp_noche = (room / ocup) / margen / tasa
    return round(pp_noche * max(1, int(noches or 1)), 2)


def precio_servicio_usd_pp(precios, destino, serv, grupo=2, cfg=None):
    """Precio por persona en USD de un servicio terrestre para un grupo de N."""
    clave = _clave_destino(precios, destino)
    dd = (precios.get(clave, {}) or {}).get("terrestres", {}) if clave else {}
    margen = float(dd.get("margen", 0.75) or 0.75)
    tasa = _tasa_tarifario(precios, clave or destino, cfg)
    pr = serv.get("precios", {}) or {}
    if not pr:
        return 0.0
    key = str(int(grupo or 2))
    if key not in pr:
        nums = sorted(int(k) for k in pr.keys() if str(k).isdigit())
        menores = [n for n in nums if n <= (grupo or 2)]
        elegido = (menores[-1] if menores else (nums[0] if nums else None))
        key = str(elegido) if elegido is not None else None
    val = float(pr.get(key, 0) or 0) if key else 0.0
    if not val or not tasa or not margen:
        return 0.0
    return round(val / margen / tasa, 2)


def es_transporte(nombre):
    n = nombre.lower()
    return any(k in n for k in ("traslado", "asistencia", "asitencia", "transporte"))

def precio_terrestre_usd(servicio, pax, tasa, margen):
    precios = servicio["precios"]
    disp = sorted(int(k) for k in precios.keys())
    if not disp or not tasa:
        return 0.0
    col = pax if pax in disp else min(disp, key=lambda x: abs(x - pax))
    if pax > max(disp):
        col = max(disp)
    ppc = precios.get(str(col)) or precios.get(col)
    total_cop = ppc * pax
    venta_cop = total_cop / margen if margen else total_cop
    return venta_cop / tasa

def precio_hotel_usd_noche(valor_cop, tasa, margen):
    if not valor_cop or not tasa:
        return 0.0
    venta_cop = valor_cop / margen if margen else valor_cop
    return venta_cop / tasa


# ---- Reglas de ninos ----
# Edad en anos cumplidos. 0 = "0-11 meses" (menor de 1 ano).
EDAD_OPCIONES = ["0-11 meses", "1 ano", "2 anos", "3 anos", "4 anos", "5 anos",
                 "6 anos", "7 anos", "8 anos", "9 anos"]
CHILD_PRIVADO_USD = 10.0     # 12 meses a 2 anos, servicio PRIVADO: 10 USD
CHILD_HOTEL_COP = 70000.0    # 3 a 9 anos: 70.000 COP/noche

def es_privado(nombre):
    n = _norm_txt(nombre)
    return "privado" in n or "privada" in n   # si no dice -> compartido

def precio_servicio_grupo(serv, adultos, ninos_ages, tasa, margen, privado):
    """Total USD de un terrestre/tour para el grupo, aplicando reglas de ninos.
       0-11m: cortesia | 1-2 anos: privado 10USD / compartido 100% | 3-9: 50%."""
    N = adultos + sum(1 for a in ninos_ages if a >= 1)   # bebes 0-11m no ocupan cupo
    N = max(N, 1)
    total_N = precio_terrestre_usd(serv, N, tasa, margen)
    pp = total_N / N                                     # precio por persona (100%)
    total = adultos * pp
    for a in ninos_ages:
        if a == 0:
            continue                                     # cortesia
        elif a <= 2:
            total += CHILD_PRIVADO_USD if privado else pp
        else:
            total += 0.5 * pp
    return total

def precio_hotel_nino_noche(tasa, margen):
    """USD por noche por nino de 3-9 anos (70.000 / margen hotel / TRM)."""
    if not tasa:
        return 0.0
    return (CHILD_HOTEL_COP / margen) / tasa if margen else CHILD_HOTEL_COP / tasa


# ---- Descripciones de tours ----
_STOP = {"de", "del", "la", "el", "en", "y", "a", "por", "con", "los", "las",
         "tour", "tours", "visita", "dia", "full", "the", "compartido", "compartida",
         "privado", "privada", "especial", "sencilla", "sencillo", "doble", "grupo",
         "pax", "round", "trip", "in", "out", "ok", "o", "u", "para", "desde", "hacia"}

def _norm_txt(s):
    s = "".join(c for c in unicodedata.normalize("NFD", str(s))
                if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).lower().strip()

def _toks(s):
    return set(t for t in _norm_txt(s).split() if t not in _STOP and len(t) > 2)

def _match_score(a, b):
    ta, tb = _toks(a), _toks(b)
    if not ta or not tb:
        return 0.0, 0
    inter = len(ta & tb)
    if inter == 0:
        return 0.0, 0          # sin palabra clave en comun -> no forzar match
    cont = inter / len(ta)     # fraccion de palabras clave de A presentes en B
    jacc = inter / len(ta | tb)
    return max(cont, jacc), inter

def cargar_descripciones():
    for ruta in (os.path.join(app_dir(), "descripciones_tours.json"),
                 recurso("descripciones_tours.json")):
        if os.path.exists(ruta):
            try:
                with open(ruta, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
    return {}

def buscar_descripcion(nombre, destino, data, umbral=0.34):
    """Devuelve la MEJOR descripcion (misma ciudad) que comparta palabra clave.
       Prioriza no omitir descripciones correctas, sin poner una equivocada."""
    best = None; bs = 0.0
    for c in data.get(destino, []):
        sc, inter = _match_score(nombre, c["nombre"])
        if sc > bs:
            bs, best = sc, c
    if best and bs >= umbral:
        return best
    return None

def texto_descripcion(reg, maxlen=1200):
    d = (reg.get("descripcion", "") or "").strip()
    dur = (reg.get("duracion", "") or "").strip()
    inc = (reg.get("incluye", "") or "").strip()
    extra = []
    if dur and len(dur) < 40:
        extra.append("Duracion: " + dur)
    if inc:
        extra.append("Incluye: " + inc)
    txt = d
    if extra:
        txt = (d + "  " if d else "") + " | ".join(extra)
    txt = re.sub(r"\s+", " ", txt).strip()
    if len(txt) > maxlen:
        txt = txt[:maxlen - 1].rsplit(" ", 1)[0] + "..."
    return txt


# ============================================================================
# PDF
# ============================================================================
PDF_PRIM = (1, 57, 132); PDF_BLUE = (20, 102, 199)
PDF_CLARO = (233, 240, 250); PDF_TXT = (30, 40, 60)

def usd(v):
    try:
        return f"USD {v:,.2f}"
    except Exception:
        return f"USD {v}"


class CotizacionPDF(FPDF):
    def __init__(self, cfg):
        super().__init__(orientation="P", unit="mm", format="A4")
        self.cfg = cfg
        self.set_auto_page_break(auto=True, margin=18)
        self.set_margins(15, 15, 15)

    def _guard_ancho(self, w):
        # Evita el error de fpdf "Not enough horizontal space..." cuando una celda
        # de ancho automatico (w=0) queda con ~0 de ancho por tener el cursor
        # pegado al margen derecho: en ese caso se vuelve al margen izquierdo.
        if not w:
            disponible = (self.w - self.r_margin) - self.get_x()
            if disponible < 3:
                self.set_x(self.l_margin)

    def cell(self, w=0, *args, **kwargs):
        self._guard_ancho(w)
        return super().cell(w, *args, **kwargs)

    def multi_cell(self, w=0, *args, **kwargs):
        self._guard_ancho(w)
        return super().multi_cell(w, *args, **kwargs)

    def header(self):
        cfg = self.cfg; y0 = 12; logo = cfg.get("logo", ""); text_x = 15
        if logo and os.path.exists(logo):
            try:
                with Image.open(logo) as im:
                    w_px, h_px = im.size
                max_w, max_h = 54, 30
                ratio = min(max_w / w_px, max_h / h_px)
                w_mm = w_px * ratio * 0.2645833; h_mm = h_px * ratio * 0.2645833
                if h_mm > max_h:
                    s = max_h / h_mm; w_mm *= s; h_mm *= s
                self.image(logo, x=15, y=y0, h=h_mm)
                text_x = 15 + w_mm + 8
            except Exception:
                text_x = 15
        self.set_xy(text_x, y0)
        self.set_text_color(*PDF_PRIM); self.set_font("Helvetica", "B", 16)
        self.cell(0, 7, self._t(cfg.get("empresa", "")), ln=1)
        self.set_x(text_x); self.set_text_color(*PDF_TXT); self.set_font("Helvetica", "", 9)
        l2 = []
        if cfg.get("nit"): l2.append("NIT/RUC: " + cfg["nit"])
        if cfg.get("direccion"): l2.append(cfg["direccion"])
        if l2:
            self.set_x(text_x); self.cell(0, 5, self._t("  |  ".join(l2)), ln=1)
        l3 = []
        if cfg.get("telefono"): l3.append("Tel: " + cfg["telefono"])
        if cfg.get("email"): l3.append(cfg["email"])
        if cfg.get("web"): l3.append(cfg["web"])
        if l3:
            self.set_x(text_x); self.cell(0, 5, self._t("  |  ".join(l3)), ln=1)
        self.set_draw_color(*PDF_PRIM); self.set_line_width(0.6)
        self.line(15, 44, 195, 44); self.set_y(48)

    def footer(self):
        self.set_y(-14); self.set_font("Helvetica", "I", 8)
        self.set_text_color(130, 130, 130)
        self.cell(0, 10, self._t(f"{self.cfg.get('empresa','')}  -  Pagina {self.page_no()}"),
                  align="C")

    def _t(self, texto):
        if texto is None:
            return ""
        return str(texto).encode("latin-1", "replace").decode("latin-1")


def _seccion_tabla(pdf, titulo, filas, total_seccion):
    if not filas:
        return
    T = pdf._t
    pdf.ln(1)
    pdf.set_font("Helvetica", "B", 10); pdf.set_text_color(*PDF_BLUE)
    pdf.cell(0, 6, T(titulo), ln=1)
    w_desc, w_det, w_pp, w_val = 90, 30, 30, 30
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(*PDF_PRIM); pdf.set_text_color(255, 255, 255)
    pdf.cell(w_desc, 7, T("  Concepto"), fill=True)
    pdf.cell(w_det, 7, T("Detalle"), fill=True, align="C")
    pdf.cell(w_pp, 7, T("Por pasajero"), fill=True, align="C")
    pdf.cell(w_val, 7, T("Total (USD)"), fill=True, align="C", ln=1)
    pdf.set_text_color(*PDF_TXT)
    f = 0
    for fila in filas:
        desc, det, val = fila[0], fila[1], fila[2]
        pp = fila[3] if len(fila) > 3 else None
        descripcion = fila[4] if len(fila) > 4 else ""
        relleno = (f % 2 == 1)
        pdf.set_font("Helvetica", "B", 9)
        lin_c = pdf.multi_cell(w_desc - 4, 4.6, T("  " + desc), border=0,
                               align="L", split_only=True)
        alto_c = len(lin_c) * 4.6
        lin_d = []
        if descripcion:
            pdf.set_font("Helvetica", "I", 7.5)
            lin_d = pdf.multi_cell(w_desc - 6, 3.6, T("   " + descripcion), border=0,
                                   align="L", split_only=True)
        alto = max(7, alto_c + len(lin_d) * 3.6 + 3)
        x0 = pdf.get_x(); y0 = pdf.get_y()
        if y0 + alto > (297 - 18):
            pdf.add_page(); x0 = pdf.get_x(); y0 = pdf.get_y()
        if relleno:
            pdf.set_fill_color(*PDF_CLARO)
        pdf.multi_cell(w_desc, alto, "", border=0, fill=relleno)
        pdf.set_xy(x0, y0 + 1)
        pdf.set_text_color(*PDF_TXT); pdf.set_font("Helvetica", "B", 9)
        pdf.multi_cell(w_desc - 4, 4.6, T("  " + desc), border=0, align="L")
        if descripcion:
            pdf.set_xy(x0, y0 + 1 + alto_c)
            pdf.set_text_color(110, 120, 135); pdf.set_font("Helvetica", "I", 7.5)
            pdf.multi_cell(w_desc - 6, 3.6, T("   " + descripcion), border=0, align="L")
            pdf.set_text_color(*PDF_TXT)
        pdf.set_xy(x0 + w_desc, y0); pdf.set_font("Helvetica", "", 9)
        pdf.cell(w_det, alto, T(det), align="C", fill=relleno)
        pdf.cell(w_pp, alto, T(usd(pp) if pp else "-"), align="R", fill=relleno)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(w_val, alto, T(usd(val)), align="R", fill=relleno, ln=1)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_y(y0 + alto)
        f += 1


def _banda_destino(pdf, texto):
    T = pdf._t
    pdf.ln(3)
    if pdf.get_y() > 250:
        pdf.add_page()
    pdf.set_fill_color(*PDF_BLUE); pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, T("  " + texto), ln=1, fill=True)
    pdf.set_text_color(*PDF_TXT)


def _tabla_hoteles_combinada(pdf, con_op):
    """Multidestino: una sola tabla que empareja el hotel de cada destino (por
       orden) y SUMA los precios por persona (Sencilla/Doble/Triple)."""
    T = pdf._t
    dests = [b["destino"] for b in con_op]
    pdf.ln(2)
    if pdf.get_y() > 230:
        pdf.add_page()
    pdf.set_font("Helvetica", "B", 10); pdf.set_text_color(*PDF_BLUE)
    pdf.cell(0, 6, T("OPCIONES DE HOTEL - precio por persona SUMANDO los "
                     f"{len(dests)} destinos (el cliente elige)"), ln=1)
    n = len(con_op)
    w_price, w_cat = 22.0, 24.0
    w_h = max(24.0, (180.0 - w_cat - 3 * w_price) / n)
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(*PDF_PRIM); pdf.set_text_color(255, 255, 255)
    for b in con_op:
        pdf.cell(w_h, 7, T("  Hotel " + b["destino"][:13]), fill=True)
    pdf.cell(w_cat, 7, T("Categoria"), fill=True, align="C")
    pdf.cell(w_price, 7, T("Sencilla"), fill=True, align="C")
    pdf.cell(w_price, 7, T("Doble"), fill=True, align="C")
    pdf.cell(w_price, 7, T("Triple"), fill=True, align="C", ln=1)
    pdf.set_text_color(*PDF_TXT)

    def money(v):
        return f"{v:,.2f}" if v else "-"

    maxn = max(len(b["opciones"]) for b in con_op)
    for i in range(maxn):
        fila = [b["opciones"][min(i, len(b["opciones"]) - 1)] for b in con_op]
        relleno = (i % 2 == 1)
        if relleno:
            pdf.set_fill_color(*PDF_CLARO)
        if pdf.get_y() > 275:
            pdf.add_page()
        pdf.set_font("Helvetica", "B", 8)
        for o in fila:
            pdf.cell(w_h, 8, T("  " + o["nombre"][:24]), fill=relleno)
        cats = []
        for o in fila:
            c = (o.get("categoria") or "").strip()
            if c and c not in cats:
                cats.append(c)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(w_cat, 8, T(("/".join(cats))[:14] or "-"), align="C", fill=relleno)

        def suma(acc):
            vals = [o.get(acc) for o in fila]
            return sum(vals) if all(v for v in vals) else None
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(w_price, 8, T(money(suma("sencilla"))), align="R", fill=relleno)
        pdf.cell(w_price, 8, T(money(suma("doble"))), align="R", fill=relleno)
        pdf.cell(w_price, 8, T(money(suma("triple"))), align="R", fill=relleno, ln=1)
    pdf.set_font("Helvetica", "I", 8); pdf.set_text_color(110, 120, 135)
    pdf.cell(0, 5, T("  Valores en USD POR PERSONA (adulto), SUMANDO "
                     + " + ".join(dests) + ", por todo el viaje. Incluye traslados "
                     "y actividades."), ln=1)
    pdf.set_text_color(*PDF_TXT)


def generar_pdf(cfg, datos, bloques, total, ruta_salida):
    """bloques: lista de dicts {destino, subtitulo, secciones:[(t,filas,sub)], subtotal}."""
    pdf = CotizacionPDF(cfg)
    pdf.add_page()
    T = pdf._t
    multi = len(bloques) > 1
    titulo = "  COTIZACION" + (" - ITINERARIO" if multi
                               else (" - " + bloques[0]["destino"] if bloques else ""))
    pdf.set_fill_color(*PDF_PRIM); pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 9, T(titulo), ln=1, fill=True)
    pdf.ln(2)

    destinos_txt = datos.get("destinos_txt") or ", ".join(
        b.get("destino", "") for b in bloques if b.get("destino"))

    pdf.set_text_color(*PDF_TXT)
    y_b = pdf.get_y()
    pdf.set_font("Helvetica", "B", 10); pdf.set_text_color(*PDF_PRIM)
    pdf.cell(90, 6, T("AGENCIA / CLIENTE"), ln=1)
    pdf.set_text_color(*PDF_TXT)
    for etq, clave in [("Nombre agencia", "cliente"), ("Asesor", "asesor"),
                       ("Email", "cli_email"), ("Telefono", "cli_tel"),
                       ("Tel. asesor", "asesor_tel")]:
        val = datos.get(clave, "")
        if val:
            pdf.set_font("Helvetica", "B", 9); pdf.cell(28, 5, T(etq + ":"))
            pdf.set_font("Helvetica", "", 9); pdf.cell(62, 5, T(val), ln=1)
    y_izq = pdf.get_y()
    pdf.set_xy(110, y_b)
    pdf.set_font("Helvetica", "B", 10); pdf.set_text_color(*PDF_PRIM)
    pdf.cell(85, 6, T("DETALLES DE LA COTIZACION"), ln=1)
    pdf.set_text_color(*PDF_TXT)
    for etq, val in [("No. Cotizacion", datos.get("numero", "")),
                     ("Fecha de cotizacion", datos.get("fecha", "")),
                     ("Destino", destinos_txt),
                     ("Valida hasta", datos.get("valida_hasta", "")),
                     ("Fechas de viaje", datos.get("fechas_viaje", "")),
                     ("Pasajeros", datos.get("pax_txt", ""))]:
        if val:
            pdf.set_x(110)
            pdf.set_font("Helvetica", "B", 9); pdf.cell(34, 5, T(etq + ":"))
            pdf.set_font("Helvetica", "", 9); pdf.cell(51, 5, T(str(val)), ln=1)
    pdf.set_y(max(y_izq, pdf.get_y()) + 2)

    def edad_txt(a):
        return "Bebe 0-11 meses" if a == 0 else f"{a} " + ("ano" if a == 1 else "anos")

    def cel(v):
        return usd(v) if v else "-"

    con_op = [b for b in bloques if b["opciones"]]
    combinar = len(con_op) > 1   # multidestino: tabla de hoteles combinada al final
    for b in bloques:
        _banda_destino(pdf, b["subtitulo"])
        for titulo_s, filas, sub in b["base_secciones"]:
            _seccion_tabla(pdf, titulo_s, filas, sub)
        ops = b["opciones"]
        # Opciones de hotel POR DESTINO (solo si NO es multidestino)
        if ops and not combinar:
            pdf.ln(2)
            if pdf.get_y() > 240:
                pdf.add_page()
            pdf.set_font("Helvetica", "B", 10); pdf.set_text_color(*PDF_BLUE)
            titulo_op = "OPCIONES DE HOTEL - precio por persona (el cliente elige)" \
                if len(ops) > 1 else "ALOJAMIENTO - precio por persona"
            pdf.cell(0, 6, T(titulo_op), ln=1)
            w_h, w_c, w_p = 74, 30, 25
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_fill_color(*PDF_PRIM); pdf.set_text_color(255, 255, 255)
            pdf.cell(w_h, 7, T("  Hotel"), fill=True)
            pdf.cell(w_c, 7, T("Categoria"), fill=True, align="C")
            pdf.cell(w_p, 7, T("Sencilla"), fill=True, align="C")
            pdf.cell(w_p, 7, T("Doble"), fill=True, align="C")
            pdf.cell(w_p, 7, T("Triple"), fill=True, align="C", ln=1)
            pdf.set_text_color(*PDF_TXT)
            for j, op in enumerate(ops):
                relleno = (j % 2 == 1)
                if relleno:
                    pdf.set_fill_color(*PDF_CLARO)
                pdf.set_font("Helvetica", "B", 9)
                pdf.cell(w_h, 8, T("  " + op["nombre"][:38]), fill=relleno)
                pdf.set_font("Helvetica", "", 9)
                pdf.cell(w_c, 8, T(op["categoria"] or "-"), align="C", fill=relleno)
                pdf.set_font("Helvetica", "B", 9)
                pdf.cell(w_p, 8, T(cel(op["sencilla"])), align="R", fill=relleno)
                pdf.cell(w_p, 8, T(cel(op["doble"])), align="R", fill=relleno)
                pdf.cell(w_p, 8, T(cel(op["triple"])), align="R", fill=relleno, ln=1)
            pdf.set_font("Helvetica", "I", 8); pdf.set_text_color(110, 120, 135)
            pdf.cell(0, 5, T("  Valores POR PERSONA (adulto) segun acomodacion, "
                             "por todo el viaje. Incluye traslados y actividades."), ln=1)
            pdf.set_text_color(*PDF_TXT)
        # Precio por nino (fijo, no depende del hotel ni de la acomodacion)
        if b["ninos"]:
            pdf.set_font("Helvetica", "", 8); pdf.set_text_color(*PDF_TXT)
            partes = [f"{edad_txt(a)} (x{c}): {usd(pr)}" for a, c, pr in b["ninos"]]
            pdf.multi_cell(0, 4.5, T("  Precio por nino: " + "   |   ".join(partes)))

    # ---- Multidestino: tabla de hoteles combinada (suma de destinos) ----
    if combinar:
        _tabla_hoteles_combinada(pdf, con_op)

    # ---- Costo total de la reserva: 1a opcion de hotel + habitaciones indicadas ----
    hab = datos.get("habitaciones") or {}
    OCCP = {"sencilla": 1, "doble": 2, "triple": 3}
    ocup = sum(hab.get(k, 0) * OCCP[k] for k in OCCP)
    if con_op and ocup > 0:
        n_ni = sum(c for b in con_op for a, c, pr in b["ninos"])
        total_res = 0.0; detalle_hab = []
        for acc in ("sencilla", "doble", "triple"):
            n = hab.get(acc, 0)
            if not n:
                continue
            detalle_hab.append(f"{n} {acc}")
            for b in con_op:
                pp = b["opciones"][0].get(acc)
                if pp:
                    total_res += n * OCCP[acc] * pp
        total_res += sum(c * pr for b in con_op for a, c, pr in b["ninos"])
        hoteles_op1 = " + ".join(b["opciones"][0]["nombre"] for b in con_op)
        pdf.ln(4)
        if pdf.get_y() > 245:
            pdf.add_page()
        pax_txt = (f"{con_op[0]['n_adultos']} adulto(s)"
                   + (f" + {n_ni} nino(s)" if n_ni else ""))
        pdf.set_font("Helvetica", "B", 10); pdf.set_fill_color(*PDF_PRIM)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(0, 8, T(f"  COSTO TOTAL DE LA RESERVA - 1a opcion ({pax_txt})"),
                 ln=1, fill=True)
        pdf.set_text_color(*PDF_TXT); pdf.set_font("Helvetica", "", 9)
        pdf.set_fill_color(*PDF_CLARO)
        pdf.cell(0, 7, T("  Habitaciones solicitadas: " + ", ".join(detalle_hab)),
                 fill=True, ln=1)
        pdf.set_font("Helvetica", "B", 12); pdf.set_fill_color(*PDF_PRIM)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(120, 9, T("  TOTAL DE LA RESERVA (USD)"), fill=True)
        pdf.cell(0, 9, T(usd(total_res)), align="R", fill=True, ln=1)
        pdf.set_text_color(110, 120, 135); pdf.set_font("Helvetica", "I", 8)
        pdf.cell(0, 5, T(f"  Calculado con la 1a opcion: {hoteles_op1[:110]}"), ln=1)
        pdf.set_text_color(*PDF_TXT)

    # ---- Itinerario dia por dia (opcional) ----
    itin = (datos.get("itinerario") or "").strip()
    if itin:
        pdf.add_page()
        pdf.set_fill_color(*PDF_PRIM); pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 9, T("  ITINERARIO DIA POR DIA"), ln=1, fill=True); pdf.ln(2)
        pdf.set_text_color(*PDF_TXT)
        for par in itin.split("\n"):
            par = par.strip()
            if not par:
                pdf.ln(1); continue
            if pdf.get_y() > 270:
                pdf.add_page()
            if re.match(r"(?i)^d[ií]a\s*\d", par):
                pdf.ln(1); pdf.set_fill_color(*PDF_BLUE); pdf.set_text_color(255, 255, 255)
                pdf.set_font("Helvetica", "B", 10)
                pdf.cell(0, 7, T("  " + par), ln=1, fill=True)
                pdf.set_text_color(*PDF_TXT)
            else:
                pdf.set_font("Helvetica", "", 9)
                pdf.multi_cell(0, 4.8, T(par))

    pdf.ln(6); pdf.set_text_color(*PDF_PRIM); pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, T("Notas y condiciones"), ln=1)
    pdf.set_text_color(*PDF_TXT); pdf.set_font("Helvetica", "", 9)
    notas = datos.get("notas", "") or cfg.get("notas", "")
    if notas:
        pdf.multi_cell(0, 5, T(notas))

    firma_nom = (datos.get("firma_nombre") or cfg.get("firma_nombre", "")).strip()
    firma_cargo = (datos.get("firma_cargo") or cfg.get("firma_cargo", "")).strip()
    if firma_nom:
        if pdf.get_y() > 240:
            pdf.add_page()
        pdf.ln(16); pdf.set_text_color(*PDF_TXT); pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 5, T("Cordialmente,"), ln=1)
        pdf.ln(10)
        pdf.set_draw_color(*PDF_PRIM); pdf.set_line_width(0.4)
        y = pdf.get_y(); pdf.line(15, y, 85, y); pdf.ln(1)
        pdf.set_font("Helvetica", "B", 10); pdf.set_text_color(*PDF_PRIM)
        pdf.cell(0, 5, T(firma_nom), ln=1)
        if firma_cargo:
            pdf.set_font("Helvetica", "", 9); pdf.set_text_color(*PDF_TXT)
            pdf.cell(0, 5, T(firma_cargo), ln=1)

    pdf.output(ruta_salida)


# ============================================================================
# MODULO RESERVAS: datos, consecutivo, rotacion de asesores y vouchers PDF
# ============================================================================
RESERVAS_PATH = os.path.join(datos_dir(), "reservas.json")
RES_SEQ_INICIAL = 2988   # el proximo consecutivo asignado sera 2989

ESTADOS_RES = ["Confirmada", "Confirmada con pago", "Aplazada", "Anulada"]
# Semaforo: aprobada (con pago) = verde | en seguimiento (confirmada/aplazada) = amarillo |
#           cancelada (anulada) = rojo
ESTADO_RES_COLOR = {"Confirmada": "#D9A400", "Confirmada con pago": GREEN,
                    "Aplazada": "#D9A400", "Anulada": RED}
ESTADO_RES_FILA = {"Confirmada": "#FFF3C4", "Confirmada con pago": "#E3F5EA",
                   "Aplazada": "#FFF3C4", "Anulada": "#FBE6E6"}

# Estado de la reserva con cada proveedor (seguimiento)
ESTADOS_PROV = ["Pendiente", "Reservado sin pago", "Reservado con pago"]
ESTADO_PROV_COLOR = {"Pendiente": MUTED, "Reservado sin pago": "#D9A400",
                     "Reservado con pago": GREEN}


def cargar_reservas():
    if os.path.exists(RESERVAS_PATH):
        try:
            with open(RESERVAS_PATH, "r", encoding="utf-8") as f:
                d = json.load(f)
            if isinstance(d, dict) and "items" in d:
                d.setdefault("seq", RES_SEQ_INICIAL)
                d.setdefault("rot", 0)
                # Piso del consecutivo: nunca por debajo de RES_SEQ_INICIAL (no baja si
                # ya se paso ese numero). Asi el proximo consecutivo arranca en 2989.
                try:
                    d["seq"] = max(int(d.get("seq", RES_SEQ_INICIAL) or RES_SEQ_INICIAL),
                                   RES_SEQ_INICIAL)
                except Exception:
                    d["seq"] = RES_SEQ_INICIAL
                return d
        except Exception:
            pass
    return {"seq": RES_SEQ_INICIAL, "rot": 0, "items": []}


def guardar_reservas(data):
    with open(RESERVAS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def peek_numero_reserva():
    return str(cargar_reservas().get("seq", RES_SEQ_INICIAL) + 1)


def asesores_reservas(cfg):
    """Lista de asesores de reservas configurados (hasta 3, con nombre)."""
    lst = cfg.get("asesores_reservas") or []
    return [a for a in lst if isinstance(a, dict) and (a.get("nombre") or "").strip()]


def registrar_reserva(rec, cfg):
    """Asigna consecutivo + asesor por rotacion equitativa y guarda la reserva.
       Devuelve (numero, asesor_asignado)."""
    data = cargar_reservas()
    data["seq"] = int(data.get("seq", RES_SEQ_INICIAL)) + 1
    numero = str(data["seq"])
    ases = asesores_reservas(cfg)
    asesor = {}
    if ases:
        idx = int(data.get("rot", 0)) % len(ases)
        asesor = ases[idx]
        data["rot"] = idx + 1
    rec = dict(rec)
    rec["numero"] = numero
    if not rec.get("uid"):
        rec["uid"] = "RES-" + uuid.uuid4().hex[:12]
    if not rec.get("asesor"):
        rec["asesor"] = asesor
    data["items"].append(rec)
    guardar_reservas(data)
    # Notificar por correo a la asesora asignada (para que le llegue para su gestion).
    # Solo si la reserva ya trae datos (las 'en blanco' se avisan luego al enviar).
    # No bloquea el registro: el resultado se adjunta al rec devuelto (no se persiste).
    if rec.get("cliente") or rec.get("destinos") or rec.get("destinos_detalle"):
        try:
            ok, info = notificar_reserva_asesora(rec, cfg)
        except Exception as e:
            ok, info = False, str(e)
    else:
        ok, info = None, "reserva en blanco"
    rec["_notif_ok"] = ok
    rec["_notif_info"] = info
    # Subir a la nube para que la asesora la vea en su equipo (best-effort).
    try:
        enviar_reserva_nube(rec)
    except Exception:
        pass
    return numero, rec


def _servicios_reserva_texto(rec):
    """Resumen de servicios de la reserva (por destino) para el correo a la asesora."""
    lineas = []
    try:
        for d in destinos_detalle_de(rec):
            nom = d.get("nombre", "") or "(destino)"
            servs = []
            for k in ("hotel", "transporte", "guia", "actividad"):
                for s in (d.get(k, []) or []):
                    sv = s.get("servicio", "") or s.get("proveedor", "")
                    if sv:
                        servs.append(sv)
            lineas.append(f"  - {nom}: " + (", ".join(servs) if servs else "(sin servicios)"))
    except Exception:
        pass
    return "\n".join(lineas)


def notificar_reserva_asesora(rec, cfg):
    """Envia a la asesora asignada un correo con los datos de la reserva (y el voucher del
       cliente si se puede) para su gestion. Devuelve (ok, info). No lanza excepciones."""
    ase = rec.get("asesor", {}) or {}
    if isinstance(ase, dict):
        email = (ase.get("email", "") or "").strip()
        nombre_ase = (ase.get("nombre", "") or "").strip() or "asesora"
    else:
        email, nombre_ase = "", str(ase)
    if not email:
        return False, "la asesora asignada no tiene correo configurado (Reservas > Asesores)"
    numero = rec.get("numero", "")
    creador = (rec.get("cotizado_por") or cfg.get("firma_nombre", "") or "").strip()
    dests = ", ".join(rec.get("destinos", []) or []) or "-"
    try:
        monto = float(rec.get("monto", 0) or 0)
    except Exception:
        monto = 0.0
    cuerpo = (
        f"Hola {nombre_ase},\n\n"
        "Se te asigno una nueva reserva para gestion:\n\n"
        f"  Reserva N.: {numero}\n"
        f"  Cliente / agencia: {rec.get('cliente','') or '-'}\n"
        f"  Contacto: {rec.get('contacto','') or '-'}\n"
        f"  Destinos: {dests}\n"
        f"  Fechas de viaje: {rec.get('fechas_viaje','') or '-'}\n"
        f"  Pasajeros: {rec.get('pax_txt','') or '-'}\n"
        f"  Monto: USD {monto:,.2f}\n\n"
        "Servicios:\n" + (_servicios_reserva_texto(rec) or "  (ver detalle en el sistema)") + "\n\n"
        "Por favor abre el modulo Reservas para confirmar proveedores y enviar los vouchers.\n"
        + (f"\nCreada por: {creador}\n" if creador else "")
        + f"\n{cfg.get('empresa','INNOBA Colombia DMC')}"
    )
    asunto = f"Nueva reserva {numero} asignada - {rec.get('cliente','')}"
    # Intentar adjuntar el voucher del cliente; si falla, enviar solo texto.
    try:
        ruta = os.path.join(tempfile.gettempdir(), f"Reserva_{numero}.pdf")
        generar_voucher_cliente(cfg, rec, ruta)
        enviar_correo(cfg, email, asunto, cuerpo, ruta)
        return True, email
    except Exception:
        try:
            enviar_correo_texto(cfg, email, asunto, cuerpo)
            return True, email
        except Exception as e:
            return False, str(e)


def _texto_notif_reserva(guardado):
    """Linea para el messagebox que informa si se le aviso a la asesora por correo."""
    ok = guardado.get("_notif_ok")
    if ok is True:
        return f"\n\n✉ Se notifico a la asesora por correo ({guardado.get('_notif_info','')})."
    if ok is None:
        return ("\n\nℹ Cuando completes los datos, usa el boton '✉ Enviar a la asesora' "
                "en la reserva para avisarle.")
    return ("\n\n⚠ No se pudo avisar a la asesora por correo: "
            f"{guardado.get('_notif_info','') or 'sin correo'}.\n"
            "Verifica el correo de la asesora (Reservas > Asesores) y el correo remitente "
            "con su contrasena en 'Datos de mi empresa'.")


def actualizar_reserva(numero, cambios):
    """Aplica cambios a la reserva con ese numero, guarda y sincroniza a la nube."""
    data = cargar_reservas()
    actualizada = None
    for it in data["items"]:
        if it.get("numero") == numero:
            it.update(cambios)
            actualizada = it
            break
    guardar_reservas(data)
    if actualizada is not None:
        try:
            if not actualizada.get("uid"):
                actualizada["uid"] = "RES-" + uuid.uuid4().hex[:12]
                guardar_reservas(data)
            enviar_reserva_nube(actualizada)
        except Exception:
            pass


def _pax_desde_snapshot(snap):
    ad = int(snap.get("adultos", 0) or 0)
    ninos = len(snap.get("ages", []) or [])
    partes = []
    if ad:
        partes.append(f"{ad} adulto" + ("s" if ad != 1 else ""))
    if ninos:
        partes.append(f"{ninos} nino" + ("s" if ninos != 1 else ""))
    return ", ".join(partes) or "-"


def _acom_de_hab(hab):
    """Acomodacion dominante (sencilla/doble/triple) a partir de la habitacion."""
    if isinstance(hab, dict):
        best = max(("sencilla", "doble", "triple"),
                   key=lambda k: float(hab.get(k, 0) or 0))
        if float(hab.get(best, 0) or 0) > 0:
            return best
    elif isinstance(hab, str):
        for k in ("sencilla", "doble", "triple"):
            if k in hab.lower():
                return k
    return "doble"


def items_cobro_desde_snapshot(snap, cfg=None):
    """Reconstruye los items cotizados (hoteles + tours/traslados) con su precio POR
       PERSONA del tarifario, para precargarlos en la reserva. cant = # de pasajeros."""
    snap = snap or {}
    try:
        precios = cargar_precios_seguro()
    except Exception:
        precios = {}
    ad = int(snap.get("adultos", 2) or 2)
    ninos = len(snap.get("ages", []) or [])
    pax = max(1, ad + ninos)
    acom = _acom_de_hab(snap.get("hab", ""))
    items = []
    for tr in snap.get("tramos", []):
        dest = tr.get("destino", "")
        noches = int(tr.get("noches", 1) or 1)
        hoteles = hoteles_detalle(precios, dest)
        for hname in tr.get("hoteles", []):
            hd = next((h for h in hoteles if h.get("nombre") == hname), None)
            val = precio_hotel_usd_pp(precios, dest, hd, acom, noches, cfg) if hd else 0.0
            items.append({"desc": f"{dest} - {hname} ({acom}, {noches}N, p.p.)",
                          "valor": round(float(val or 0), 2), "cant": pax})
        servs = servicios_terrestres(precios, dest)
        for sname in list(tr.get("trans", [])) + list(tr.get("act", [])):
            sd = next((s for s in servs if s.get("nombre") == sname), None)
            val = precio_servicio_usd_pp(precios, dest, sd, pax, cfg) if sd else 0.0
            items.append({"desc": f"{dest} - {sname} (p.p., grupo {pax})",
                          "valor": round(float(val or 0), 2), "cant": pax})
    return items


def reserva_desde_cotizacion(cot):
    """Construye el borrador de reserva a partir de una cotizacion del historial.
       Extrae hoteles y servicios (traslados/tours) como renglones asignables a
       proveedor."""
    snap = cot.get("snapshot") or {}
    detalle = []
    hotel0 = ""
    for tr in snap.get("tramos", []):
        dest = tr.get("destino", "")
        item = {"nombre": dest, "hotel": [], "transporte": [], "guia": [], "actividad": []}
        for h in tr.get("hoteles", []):
            item["hotel"].append({"servicio": h, "proveedor": h, "correo": "",
                                  "enviado": False, "fecha_envio": ""})
            hotel0 = hotel0 or h
        for t in tr.get("trans", []):
            item["transporte"].append({"servicio": t, "proveedor": "", "correo": "",
                                       "enviado": False, "fecha_envio": ""})
        for a in tr.get("act", []):
            item["actividad"].append({"servicio": a, "proveedor": "", "correo": "",
                                      "enviado": False, "fecha_envio": ""})
        detalle.append(item)
    ini, fin = _fechas_in_out(cot.get("fechas_viaje", ""))
    ciudad = (cot.get("destinos", []) or [""])[0]
    # Precargar los items cobrados con el detalle cotizado (servicios + precio p.p.)
    items_cobro = items_cobro_desde_snapshot(snap)
    monto = (round(sum(it["valor"] * it.get("cant", 1) for it in items_cobro), 2)
             if items_cobro else float(cot.get("total", 0) or 0))
    rec = {
        "cot_origen": cot.get("numero", ""),
        "cot_uid": cot.get("uid") or cot.get("web_id") or cot.get("numero", ""),
        "cliente": cot.get("cliente", ""),
        "contacto": cot.get("asesor", ""),
        "email": cot.get("email", "") or snap.get("email", ""),
        "destinos": cot.get("destinos", []),
        "fechas_viaje": cot.get("fechas_viaje", ""),
        "pax_txt": _pax_desde_snapshot(snap),
        "hab": snap.get("hab", ""),
        "estado": "Confirmada",
        "monto": monto,
        "moneda": "USD",
        "items_cobro": items_cobro,
        "destinos_detalle": detalle,
        "itinerario": snap.get("itinerario", ""),
        "notas": "",
        "snapshot": snap,
        "voucher_cliente": "",
        "fecha_creacion": datetime.date.today().isoformat(),
    }
    rec.update(_voucher_defaults(ciudad, hotel0, ini, fin, snap.get("hab", "")))
    return rec


# Categorias de servicio por destino (clave interna, etiqueta, tipo para el voucher)
CATEGORIAS_SERV = [("hotel", "Hotel", "Hotel"),
                   ("transporte", "Transporte / Traslados", "Transporte"),
                   ("guia", "Guia", "Guia"),
                   ("actividad", "Actividades / Tours", "Actividad")]
CAT_KEYS = [c[0] for c in CATEGORIAS_SERV]
MAX_DESTINOS_RES = 5
# Migracion de claves antiguas -> nuevas
_CAT_MIGRA = {"hoteles": "hotel", "traslados": "transporte", "tours": "actividad"}


def _servicio_vacio():
    return {"servicio": "", "proveedor": "", "correo": "", "enviado": False,
            "fecha_envio": "", "estado_prov": "Pendiente", "fecha": "", "hora": "",
            "origen": "", "vehiculo": "", "observacion": ""}


def _norm_serv(s):
    """Asegura que un servicio tenga todas las claves (seguimiento + voucher)."""
    for k, v in (("servicio", ""), ("proveedor", ""), ("correo", ""),
                 ("enviado", False), ("fecha_envio", ""), ("estado_prov", "Pendiente"),
                 ("fecha", ""), ("hora", ""), ("origen", ""), ("vehiculo", ""),
                 ("observacion", "")):
        s.setdefault(k, v)
    return s


def _destino_vacio(nombre=""):
    d = {"nombre": nombre}
    for k in CAT_KEYS:
        d[k] = []
    return d


def destinos_detalle_de(res):
    """Devuelve la estructura de servicios por destino. Migra reservas antiguas que
       guardaban 'renglones' plano a la nueva estructura anidada por destino."""
    dd = res.get("destinos_detalle")
    if isinstance(dd, list) and dd:
        for d in dd:
            d.setdefault("nombre", "")
            # migrar claves antiguas (hoteles/traslados/tours) a las nuevas
            for viejo, nuevo in _CAT_MIGRA.items():
                if viejo in d:
                    d.setdefault(nuevo, [])
                    d[nuevo] = d[nuevo] + d.pop(viejo)
            for k in CAT_KEYS:
                d.setdefault(k, [])
    else:
        # Migrar desde renglones + destinos
        nombres = list(res.get("destinos", []) or [])
        for r in res.get("renglones", []):
            d = r.get("destino", "")
            if d and d not in nombres:
                nombres.append(d)
        dd = [_destino_vacio(n) for n in nombres]
        idx = {n: i for i, n in enumerate(nombres)}
        tipomap = {"Hotel": "hotel", "Traslado": "transporte", "Tour": "actividad"}
        for r in res.get("renglones", []):
            n = r.get("destino", "")
            if n not in idx:
                continue
            s = {"servicio": r.get("servicio", ""), "proveedor": r.get("proveedor", ""),
                 "correo": r.get("correo", ""), "enviado": r.get("enviado", False),
                 "fecha_envio": r.get("fecha_envio", "")}
            clave = tipomap.get(r.get("tipo", ""), "actividad")
            dd[idx[n]][clave].append(s)
        res["destinos_detalle"] = dd
    # normalizar todos los servicios (agrega estado_prov, hora, etc.)
    for d in dd:
        for k in CAT_KEYS:
            for s in d.get(k, []):
                _norm_serv(s)
    return dd


def resumen_seguimiento(res):
    """Cuenta la gestion de proveedores de una reserva: total, con voucher enviado
       y por estado (Reservado con/sin pago, Pendiente)."""
    total = enviados = 0
    por_estado = {e: 0 for e in ESTADOS_PROV}
    for d in destinos_detalle_de(res):
        for k in CAT_KEYS:
            for s in d.get(k, []):
                if not (s.get("servicio") or s.get("proveedor")):
                    continue
                total += 1
                if s.get("enviado"):
                    enviados += 1
                por_estado[s.get("estado_prov", "Pendiente")] = \
                    por_estado.get(s.get("estado_prov", "Pendiente"), 0) + 1
    return total, enviados, por_estado


def renglon_de(res, di, cat, si):
    """Construye el dict de renglon (para el voucher) desde un servicio anidado."""
    tipo = next((t for k, _l, t in CATEGORIAS_SERV if k == cat), cat)
    dest = res["destinos_detalle"][di]
    s = dest[cat][si]
    reng = {"tipo": tipo, "destino": dest.get("nombre", ""),
            "servicio": s.get("servicio", ""), "proveedor": s.get("proveedor", ""),
            "correo": s.get("correo", ""), "fecha": s.get("fecha", ""), "hora": s.get("hora", ""),
            "origen": s.get("origen", ""), "vehiculo": s.get("vehiculo", ""),
            "observacion": s.get("observacion", "")}
    return reng, s


def _dir_vouchers():
    ruta = os.path.join(datos_dir(), "vouchers")
    os.makedirs(ruta, exist_ok=True)
    return ruta


def generar_voucher_prov_archivo(cfg, res, di, cat, si):
    reng, s = renglon_de(res, di, cat, si)
    fn = os.path.join(_dir_vouchers(),
                      f"Voucher_prov_{res.get('numero','')}_{di+1}_{cat}_{si+1}.pdf")
    generar_voucher_proveedor(cfg, res, reng, fn)
    return fn, reng, s


def enviar_voucher_prov(cfg, res, di, cat, si):
    """Genera y envia por correo el voucher del proveedor; marca 'enviado' y guarda."""
    fn, reng, s = generar_voucher_prov_archivo(cfg, res, di, cat, si)
    if not reng["correo"]:
        raise ValueError("El proveedor no tiene correo.")
    asunto = f"Reserva {res.get('numero','')} - {reng['servicio']} - {cfg.get('empresa','')}"
    cuerpo = (f"Estimado {reng['proveedor']}:\n\n"
              f"Adjuntamos el voucher de la reserva {res.get('numero','')} para "
              f"{res.get('cliente','')}.\nDestino: {reng['destino']}\n"
              f"Fechas de viaje: {res.get('fechas_viaje','')}\n"
              f"Pasajeros: {res.get('pax_txt','')}\n\n"
              f"Favor confirmar disponibilidad y remitir la facturacion a nombre de "
              f"{cfg.get('empresa','')}.\n\nCordialmente,\n{cfg.get('empresa','')}")
    enviar_correo(cfg, reng["correo"], asunto, cuerpo, fn)
    s["enviado"] = True
    s["fecha_envio"] = datetime.date.today().strftime("%d/%m/%Y")
    actualizar_reserva(res.get("numero", ""), {"destinos_detalle": res["destinos_detalle"]})
    return fn


# ---- Voucher CONSOLIDADO por proveedor (transporte / guia): un solo voucher con
#      todos los servicios de ese proveedor en el destino ----
def servicios_por_proveedor(res, di, cat):
    """Agrupa los servicios de una categoria en un destino por proveedor.
       Devuelve [(proveedor, correo, [(si, s), ...]), ...] en orden de aparicion.
       Si un servicio no tiene proveedor, va en su propio grupo."""
    try:
        lst = res["destinos_detalle"][di].get(cat, [])
    except Exception:
        return []
    grupos = {}
    orden = []
    for si, s in enumerate(lst):
        if not (s.get("servicio") or "").strip():
            continue
        prov = (s.get("proveedor", "") or "").strip()
        key = _nz(prov) if prov else f"__solo_{si}"
        if key not in grupos:
            grupos[key] = {"proveedor": prov, "correo": s.get("correo", ""), "items": []}
            orden.append(key)
        grupos[key]["items"].append((si, s))
        if not grupos[key]["correo"] and s.get("correo"):
            grupos[key]["correo"] = s.get("correo")
    return [(grupos[k]["proveedor"], grupos[k]["correo"], grupos[k]["items"]) for k in orden]


def generar_voucher_proveedor_multi(cfg, res, cat, dest, proveedor, servicios, ruta):
    """Voucher con VARIOS servicios del mismo proveedor (transporte o guia)."""
    pdf = VoucherPDF(cfg); pdf.add_page(); T = pdf._t
    n_pax = len(pasajeros_de(res)) or res.get("pax_txt", "")
    fdef = res.get("os_fecha_in", "") or res.get("fechas_viaje", "")
    if cat == "guia":
        _os_encab_prov(pdf, cfg, res, "VOUCHER RESERVA GUIA  -  ORDEN " + res.get("numero", ""))
        _os_band(pdf, "RESERVA GUIA")
        _os_row(pdf, "GUIA", proveedor)
        _os_row(pdf, "DESTINO", dest)
        _os_row(pdf, "# PASAJEROS", str(n_pax))
        _os_band(pdf, "SERVICIOS DEL GUIA")
        pdf.set_x(12); pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO)
        pdf.set_text_color(*PDF_PRIM)
        pdf.cell(30, 6.5, T(" FECHA"), border=1, fill=True)
        pdf.cell(126, 6.5, T(" SERVICIO"), border=1, fill=True)
        pdf.cell(30, 6.5, T(" HORA"), border=1, fill=True, ln=1)
        pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
        for s in servicios:
            f = s.get("fecha", "") or fdef
            y0 = pdf.get_y()
            pdf.set_x(12); pdf.multi_cell(30, 5.5, T(" " + str(f)), border=1)
            h1 = pdf.get_y() - y0
            pdf.set_xy(42, y0); pdf.multi_cell(126, 5.5, T(" " + str(s.get("servicio", ""))), border=1)
            h2 = pdf.get_y() - y0
            hh = max(h1, h2, 5.5)
            pdf.set_xy(168, y0); pdf.cell(30, hh, T(" " + str(s.get("hora", ""))), border=1, ln=1)
            pdf.set_y(y0 + hh)
    else:  # transporte
        _os_encab_prov(pdf, cfg, res, "VOUCHER TRANSPORTE  -  ORDEN " + res.get("numero", ""))
        _os_band(pdf, "SERVICIO DE TRANSPORTE")
        _os_row(pdf, "PROVEEDOR", proveedor)
        _os_row(pdf, "DESTINO", dest)
        _os_row(pdf, "# PASAJEROS", str(n_pax))
        _os_band(pdf, "TRAYECTOS / SERVICIOS")
        pdf.set_x(12); pdf.set_font("Helvetica", "B", 8); pdf.set_fill_color(*PDF_CLARO)
        pdf.set_text_color(*PDF_PRIM)
        anchos = [(26, "FECHA"), (38, "ORIGEN"), (48, "SERVICIO"), (18, "HORA"),
                  (18, "PAX"), (38, "VEHICULO")]
        for w, h in anchos:
            pdf.cell(w, 6.5, T(" " + h), border=1, fill=True, align="C")
        pdf.ln(6.5); pdf.set_font("Helvetica", "", 8); pdf.set_text_color(*PDF_TXT)
        for s in servicios:
            f = s.get("fecha", "") or fdef
            vals = [str(f), s.get("origen", ""), s.get("servicio", ""), s.get("hora", ""),
                    str(n_pax), s.get("vehiculo", "")]
            y0 = pdf.get_y(); x = 12
            for (w, _h), v in zip(anchos, vals):
                pdf.set_xy(x, y0); pdf.cell(w, 7, T(" " + str(v)[:40]), border=1); x += w
            pdf.ln(7)
    obs = "   ".join(s.get("observacion", "") for s in servicios if s.get("observacion"))
    if obs:
        _os_row(pdf, "OBSERVACIONES", obs, wl=40, wv=146)
    _os_pasajeros_tabla(pdf, res)
    pdf.output(ruta)
    return ruta


def generar_voucher_consolidado_archivo(cfg, res, di, cat, proveedor, items):
    dest = res["destinos_detalle"][di].get("nombre", "")
    safe = re.sub(r"\W+", "", (proveedor or "prov"))[:20] or "prov"
    fn = os.path.join(_dir_vouchers(),
                      f"Voucher_{cat}_{res.get('numero','')}_{di+1}_{safe}.pdf")
    generar_voucher_proveedor_multi(cfg, res, cat, dest, proveedor,
                                    [s for _si, s in items], fn)
    return fn


def enviar_voucher_consolidado(cfg, res, di, cat, proveedor, correo, items):
    fn = generar_voucher_consolidado_archivo(cfg, res, di, cat, proveedor, items)
    if not correo:
        raise ValueError("El proveedor no tiene correo.")
    dest = res["destinos_detalle"][di].get("nombre", "")
    asunto = f"Reserva {res.get('numero','')} - {cat.capitalize()} {dest} - {cfg.get('empresa','')}"
    cuerpo = (f"Estimado {proveedor}:\n\n"
              f"Adjuntamos el voucher de la reserva {res.get('numero','')} para "
              f"{res.get('cliente','')}, con TODOS los servicios de {dest}.\n"
              f"Fechas de viaje: {res.get('fechas_viaje','')}\n"
              f"Pasajeros: {res.get('pax_txt','')}\n\n"
              f"Favor confirmar disponibilidad y remitir la facturacion a nombre de "
              f"{cfg.get('empresa','')}.\n\nCordialmente,\n{cfg.get('empresa','')}")
    enviar_correo(cfg, correo, asunto, cuerpo, fn)
    hoy = datetime.date.today().strftime("%d/%m/%Y")
    for _si, s in items:
        s["enviado"] = True
        s["fecha_envio"] = hoy
    actualizar_reserva(res.get("numero", ""), {"destinos_detalle": res["destinos_detalle"]})
    return fn


def _mes_de_iso(s):
    s = (s or "").strip()
    return s[:7] if len(s) >= 7 else "?"


def _estilo_encabezado_xlsx(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="013984")
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")


def _autoancho_xlsx(ws, minimo=10, maximo=48):
    for col in ws.columns:
        largo = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(minimo, min(maximo, largo + 2))


def exportar_reporte_reservas(ruta, mes=None):
    """Reporte de reservas (Excel): detalle + resumen por mes."""
    import openpyxl
    data = cargar_reservas()
    items = list(data.get("items", []))
    if mes:
        items = [it for it in items if _mes_de_iso(it.get("fecha_creacion", "")) == mes]
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Reservas"
    ws.append(["N. Reserva", "Fecha creacion", "Mes", "Cliente", "Asesor", "Destinos",
               "Fechas viaje", "Estado", "Monto USD", "Servicios", "Vouchers enviados"])
    for it in items:
        ase = it.get("asesor", {}) or {}
        try:
            tot_serv, env, _pe = resumen_seguimiento(it)
        except Exception:
            tot_serv, env = 0, 0
        ws.append([it.get("numero", ""), it.get("fecha_creacion", ""),
                   _mes_de_iso(it.get("fecha_creacion", "")), it.get("cliente", ""),
                   ase.get("nombre", ""), ", ".join(it.get("destinos", [])),
                   it.get("fechas_viaje", ""), it.get("estado", ""),
                   round(float(it.get("monto", 0) or 0), 2), tot_serv, env])
    _estilo_encabezado_xlsx(ws); _autoancho_xlsx(ws); ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Resumen por mes")
    ws2.append(["Mes", "# Reservas", "Monto total USD", "Confirmadas", "Con pago",
                "Aplazadas", "Anuladas"])
    resumen = {}
    for it in items:
        m = _mes_de_iso(it.get("fecha_creacion", ""))
        r = resumen.setdefault(m, {"n": 0, "monto": 0.0, "Confirmada": 0,
                                   "Confirmada con pago": 0, "Aplazada": 0, "Anulada": 0})
        r["n"] += 1
        r["monto"] += float(it.get("monto", 0) or 0)
        e = it.get("estado", "Confirmada")
        if e in r:
            r[e] += 1
    for m in sorted(resumen):
        r = resumen[m]
        ws2.append([m, r["n"], round(r["monto"], 2), r["Confirmada"],
                    r["Confirmada con pago"], r["Aplazada"], r["Anulada"]])
    _estilo_encabezado_xlsx(ws2); _autoancho_xlsx(ws2)
    wb.save(ruta)
    return len(items)


def _quien_cerro(it):
    c = (it.get("cotizado_por", "") or it.get("asesor", "") or "").strip()
    cl = c.lower()
    if "felipe" in cl:
        return "Felipe"
    if "carlos" in cl:
        return "Carlos"
    return c or "(sin asignar)"


def exportar_reporte_ventas(ruta, mes=None):
    """Reporte de ventas cerradas (cotizaciones Ganadas) por mes, separadas por
       Felipe y Carlos (Excel)."""
    import openpyxl
    data = cargar_cotizaciones()
    items = [it for it in data.get("items", []) if it.get("estado") == "Ganada"]

    def mes_de(it):
        f = parse_fecha(it.get("fecha", ""))
        return f.strftime("%Y-%m") if f else "?"

    if mes:
        items = [it for it in items if mes_de(it) == mes]
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Ventas cerradas"
    ws.append(["Cotizacion", "Fecha", "Mes", "Cerrada por", "Cliente / Agencia", "Asesor",
               "Destinos", "Total USD"])
    for it in items:
        ws.append([it.get("numero", ""), it.get("fecha", ""), mes_de(it), _quien_cerro(it),
                   it.get("cliente", ""), it.get("asesor", ""),
                   ", ".join(it.get("destinos", [])), round(float(it.get("total", 0) or 0), 2)])
    _estilo_encabezado_xlsx(ws); _autoancho_xlsx(ws); ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Resumen por mes")
    ws2.append(["Mes", "Felipe #", "Felipe USD", "Carlos #", "Carlos USD",
                "Otros #", "Otros USD", "Total #", "Total USD"])
    resumen = {}
    for it in items:
        m = mes_de(it); q = _quien_cerro(it); t = float(it.get("total", 0) or 0)
        r = resumen.setdefault(m, {"Felipe": [0, 0.0], "Carlos": [0, 0.0], "Otros": [0, 0.0]})
        key = q if q in ("Felipe", "Carlos") else "Otros"
        r[key][0] += 1; r[key][1] += t
    for m in sorted(resumen):
        f, c, o = resumen[m]["Felipe"], resumen[m]["Carlos"], resumen[m]["Otros"]
        ws2.append([m, f[0], round(f[1], 2), c[0], round(c[1], 2), o[0], round(o[1], 2),
                    f[0] + c[0] + o[0], round(f[1] + c[1] + o[1], 2)])
    _estilo_encabezado_xlsx(ws2); _autoancho_xlsx(ws2)
    wb.save(ruta)
    return len(items)


# ============================================================================
# MODULO COMERCIAL: tareas de gestion + indicadores
# ============================================================================
TAREAS_PATH = os.path.join(datos_dir(), "tareas.json")
ESTADOS_TAREA = ["Pendiente", "En progreso", "Completada"]
ESTADOS_CLIENTE = ["Sin clasificar", "Cliente actual (vigente en compra)",
                   "En seguimiento (para ser cliente)", "Descartado"]
ESTADO_TAREA_COLOR = {"Pendiente": MUTED, "En progreso": BLUE, "Completada": GREEN, "Vencida": RED}
ESTADO_TAREA_FILA = {"Pendiente": "#F1F5FB", "En progreso": "#EAF2FD",
                     "Completada": "#E3F5EA", "Vencida": "#FBE6E6"}
PRIORIDADES_TAREA = ["Alta", "Media", "Baja"]
PRIORIDAD_COLOR = {"Alta": RED, "Media": "#D9A400", "Baja": MUTED}


def cargar_tareas():
    if os.path.exists(TAREAS_PATH):
        try:
            with open(TAREAS_PATH, "r", encoding="utf-8") as f:
                d = json.load(f)
            if isinstance(d, dict) and "items" in d:
                d.setdefault("seq", 0)
                return d
        except Exception:
            pass
    return {"seq": 0, "items": []}


def guardar_tareas(data):
    with open(TAREAS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def registrar_tarea(rec):
    data = cargar_tareas()
    data["seq"] = int(data.get("seq", 0)) + 1
    rec = dict(rec)
    rec["numero"] = f"TAR-{data['seq']:04d}"
    rec.setdefault("fecha_creacion", datetime.date.today().isoformat())
    data["items"].append(rec)
    guardar_tareas(data)
    return rec


def actualizar_tarea(numero, cambios):
    data = cargar_tareas()
    for it in data["items"]:
        if it.get("numero") == numero:
            it.update(cambios)
            break
    guardar_tareas(data)


def eliminar_tarea(numero):
    data = cargar_tareas()
    data["items"] = [it for it in data["items"] if it.get("numero") != numero]
    guardar_tareas(data)


def _parse_fecha_iso(s):
    try:
        return datetime.date.fromisoformat((s or "").strip())
    except Exception:
        return None


def estado_tarea_efectivo(t):
    """Estado real: 'Vencida' si no esta completada y su fecha limite ya paso."""
    est = t.get("estado", "Pendiente")
    if est != "Completada":
        f = _parse_fecha_iso(t.get("fecha_limite", ""))
        if f and f < datetime.date.today():
            return "Vencida"
    return est


def _progreso_checklist(t):
    ch = t.get("checklist", []) or []
    hechas = sum(1 for c in ch if c.get("hecha"))
    return hechas, len(ch)


def indicadores_comerciales():
    """KPIs de gestion comercial: tareas + ventas (cotizaciones) + reservas."""
    hoy = datetime.date.today()
    mes = hoy.strftime("%Y-%m")
    tareas = cargar_tareas().get("items", [])
    ind = {"tareas_total": len(tareas), "pendientes": 0, "en_progreso": 0,
           "completadas": 0, "vencidas": 0}
    for t in tareas:
        e = estado_tarea_efectivo(t)
        clave = {"Pendiente": "pendientes", "En progreso": "en_progreso",
                 "Completada": "completadas", "Vencida": "vencidas"}.get(e)
        if clave:
            ind[clave] += 1

    cots = cargar_cotizaciones().get("items", [])
    ganadas = [c for c in cots if c.get("estado") == "Ganada"]

    def mes_cot(c):
        f = parse_fecha(c.get("fecha", ""))
        return f.strftime("%Y-%m") if f else ""

    gan_mes = [c for c in ganadas if mes_cot(c) == mes]
    ind["cot_total"] = len(cots)
    ind["cot_ganadas"] = len(ganadas)
    ind["ventas_mes_n"] = len(gan_mes)
    ind["ventas_mes_usd"] = round(sum(float(c.get("total", 0) or 0) for c in gan_mes), 2)
    ind["conversion"] = round(100.0 * len(ganadas) / len(cots), 1) if cots else 0.0
    # ventas por cotizador (Felipe / Carlos) del mes
    ind["ventas_felipe"] = round(sum(float(c.get("total", 0) or 0) for c in gan_mes
                                     if _quien_cerro(c) == "Felipe"), 2)
    ind["ventas_carlos"] = round(sum(float(c.get("total", 0) or 0) for c in gan_mes
                                     if _quien_cerro(c) == "Carlos"), 2)

    resv = cargar_reservas().get("items", [])
    res_mes = [r for r in resv if _mes_de_iso(r.get("fecha_creacion", "")) == mes]
    ind["reservas_mes_n"] = len(res_mes)
    ind["reservas_mes_usd"] = round(sum(float(r.get("monto", 0) or 0) for r in res_mes
                                        if r.get("estado") != "Anulada"), 2)
    return ind


def indicadores_cotizaciones():
    """KPIs del modulo de cotizaciones."""
    data = cargar_cotizaciones().get("items", [])
    hoy = datetime.date.today(); mes = hoy.strftime("%Y-%m")
    ind = {"total": len(data)}
    for e in ESTADOS_COT:
        ind[e] = sum(1 for c in data if c.get("estado") == e)
    ind["monto_total"] = round(sum(float(c.get("total", 0) or 0) for c in data), 2)
    gan = [c for c in data if c.get("estado") == "Ganada"]
    ind["monto_ganado"] = round(sum(float(c.get("total", 0) or 0) for c in gan), 2)
    ind["conversion"] = round(100.0 * len(gan) / len(data), 1) if data else 0.0
    ind["ticket"] = round(ind["monto_ganado"] / len(gan), 2) if gan else 0.0

    def mesc(c):
        f = parse_fecha(c.get("fecha", ""))
        return f.strftime("%Y-%m") if f else ""

    delmes = [c for c in data if mesc(c) == mes]
    ind["mes_n"] = len(delmes)
    ind["mes_usd"] = round(sum(float(c.get("total", 0) or 0) for c in delmes), 2)
    try:
        ind["seg_vencidos"] = len(seguimientos_pendientes())
    except Exception:
        ind["seg_vencidos"] = 0
    return ind


def indicadores_reservas():
    """KPIs del modulo de reservas."""
    data = cargar_reservas().get("items", [])
    hoy = datetime.date.today(); mes = hoy.strftime("%Y-%m")
    ind = {"total": len(data)}
    for e in ESTADOS_RES:
        ind[e] = sum(1 for r in data if r.get("estado") == e)
    activos = [r for r in data if r.get("estado") != "Anulada"]
    ind["monto_total"] = round(sum(float(r.get("monto", 0) or 0) for r in activos), 2)
    conpago = [r for r in data if r.get("estado") == "Confirmada con pago"]
    ind["con_pago_usd"] = round(sum(float(r.get("monto", 0) or 0) for r in conpago), 2)
    delmes = [r for r in data if _mes_de_iso(r.get("fecha_creacion", "")) == mes]
    ind["mes_n"] = len(delmes)
    ind["mes_usd"] = round(sum(float(r.get("monto", 0) or 0) for r in delmes
                               if r.get("estado") != "Anulada"), 2)
    tot_serv = env = 0
    for r in data:
        try:
            t, e, _pe = resumen_seguimiento(r)
        except Exception:
            t, e = 0, 0
        tot_serv += t; env += e
    ind["serv_total"] = tot_serv; ind["serv_enviados"] = env
    ind["serv_pendientes"] = tot_serv - env
    return ind


# ---- Liquidacion de rentabilidad por orden de servicio (reserva) ----
# Tasas por defecto (sobre las VENTAS en pesos), tomadas del Excel de INNOBA.
TASAS_LIQ_DEF = {"comision": 0.0265, "iva_comision": 0.19, "retefuente": 0.015,
                 "reteica": 0.002, "cuatromil": 0.004, "bonificacion": 0.03}
GASTOS_OPER = [("hotel", "Hotel"), ("tld_in", "Traslado IN"), ("tld_out", "Traslado OUT"),
               ("transportes", "Transportes"), ("guia", "Guia"), ("compartido", "Compartido"),
               ("actividades", "Actividades"), ("poliza", "Poliza")]


def tasas_liq(cfg):
    t = dict(TASAS_LIQ_DEF)
    for k, v in (cfg.get("tasas_liq", {}) or {}).items():
        try:
            t[k] = float(v)
        except Exception:
            pass
    return t


def trm_del_mes(cfg, mes):
    """TRM promedio guardado para ese mes (o el ultimo/general/4000 por defecto)."""
    tm = cfg.get("trm_mensual", {}) or {}
    try:
        v = float(tm.get(mes, 0) or 0)
        if v > 0:
            return v
    except Exception:
        pass
    return float(cfg.get("trm_liq", 4000) or 4000)


def liquidar_reserva(res, cfg, trm_default=None):
    """Liquida una orden de servicio: ventas (USD*TRM) - gastos operativos - gastos
       financieros = utilidad; rentabilidad % y bonificacion (3% de la utilidad).
       El TRM es: el de la reserva (override) si lo tiene; si no, el promedio del mes
       (trm_default); si no, el general de la config."""
    t = tasas_liq(cfg)
    liq = res.get("liq", {}) or {}
    monto_usd = float(res.get("monto", 0) or 0)
    try:
        trm = float(liq.get("trm", 0) or 0)
    except Exception:
        trm = 0.0
    if not trm:
        try:
            trm = float(trm_default or 0)
        except Exception:
            trm = 0.0
    if not trm:
        trm = float(cfg.get("trm_liq", 4000) or 4000)
    ventas = monto_usd * trm
    operativos = 0.0
    for k, _ in GASTOS_OPER:
        try:
            operativos += float(liq.get(k, 0) or 0)
        except Exception:
            pass
    comision = ventas * t["comision"]
    iva_comision = comision * t["iva_comision"]
    retefuente = ventas * t["retefuente"]
    reteica = ventas * t["reteica"]
    cuatromil = ventas * t["cuatromil"]
    financieros = comision + iva_comision + retefuente + reteica + cuatromil
    total_gasto = operativos + financieros
    utilidad = ventas - total_gasto
    rent = (utilidad / ventas) if ventas else 0.0
    bonif = utilidad * t["bonificacion"] if utilidad > 0 else 0.0
    return {"trm": trm, "ventas": ventas, "operativos": operativos, "comision": comision,
            "iva_comision": iva_comision, "retefuente": retefuente, "reteica": reteica,
            "cuatromil": cuatromil, "financieros": financieros, "total_gasto": total_gasto,
            "utilidad": utilidad, "rent": rent, "bonif": bonif}


def _cop(v):
    """Formatea un valor en pesos colombianos: $ 1.426.746"""
    try:
        return "$ " + f"{float(v):,.0f}".replace(",", ".")
    except Exception:
        return "$ 0"


def _kpi_card(parent, titulo, valor, color, ancho=140, alto=66, on_click=None, active=False):
    """Tarjeta compacta de indicador (valor grande + titulo). Si on_click se pasa,
       la tarjeta es clicable (para filtrar) y 'active' la resalta."""
    card = ctk.CTkFrame(parent, fg_color=("#EAF2FD" if active else CARD), corner_radius=12,
                        border_width=(2 if active else 1),
                        border_color=(color if active else LINE), width=ancho, height=alto)
    card.pack(side="left", padx=4, fill="y"); card.pack_propagate(False)
    l1 = ctk.CTkLabel(card, text=str(valor), text_color=color, font=("Segoe UI", 18, "bold"))
    l1.pack(pady=(9, 0), padx=8)
    l2 = ctk.CTkLabel(card, text=titulo, text_color=MUTED, font=("Segoe UI", 10),
                      wraplength=ancho - 14)
    l2.pack(pady=(0, 8), padx=6)
    if on_click:
        for w in (card, l1, l2):
            w.configure(cursor="hand2")
            w.bind("<Button-1>", lambda e: on_click())
    return card


def exportar_reporte_tareas(ruta, mes=None):
    """Reporte de tareas comerciales (Excel)."""
    import openpyxl
    items = cargar_tareas().get("items", [])
    if mes:
        items = [it for it in items if _mes_de_iso(it.get("fecha_creacion", "")) == mes]
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Tareas"
    ws.append(["Tarea", "Titulo", "Cliente", "Responsable", "Prioridad", "Fecha limite",
               "Estado", "Checklist", "Creada"])
    for t in items:
        h, n = _progreso_checklist(t)
        ws.append([t.get("numero", ""), t.get("titulo", ""), t.get("cliente", ""),
                   t.get("responsable", ""), t.get("prioridad", ""), t.get("fecha_limite", ""),
                   estado_tarea_efectivo(t), f"{h}/{n}", t.get("fecha_creacion", "")])
    _estilo_encabezado_xlsx(ws); _autoancho_xlsx(ws); ws.freeze_panes = "A2"
    wb.save(ruta)
    return len(items)


def ranking_contactos():
    """Ranking de contactos (vendedores de agencia) por reservas hechas con nosotros:
       del mes, del ano y total. Para premiar a los mejores contactos."""
    hoy = datetime.date.today()
    mes = hoy.strftime("%Y-%m"); ano = str(hoy.year)
    r = {}
    for it in cargar_reservas().get("items", []):
        con = (it.get("contacto", "") or "").strip() or "(sin contacto)"
        d = r.setdefault(con, {"contacto": con, "empresas": set(), "mes_n": 0,
                               "ano_n": 0, "ano_usd": 0.0, "total_n": 0, "total_usd": 0.0})
        if it.get("cliente"):
            d["empresas"].add(it.get("cliente"))
        fc = it.get("fecha_creacion", "")
        anulada = it.get("estado") == "Anulada"
        monto = float(it.get("monto", 0) or 0)
        d["total_n"] += 1
        if not anulada:
            d["total_usd"] += monto
        if fc[:7] == mes:
            d["mes_n"] += 1
        if fc[:4] == ano:
            d["ano_n"] += 1
            if not anulada:
                d["ano_usd"] += monto
    filas = list(r.values())
    for f in filas:
        f["empresas"] = ", ".join(sorted(f["empresas"]))
        f["ano_usd"] = round(f["ano_usd"], 2); f["total_usd"] = round(f["total_usd"], 2)
    filas.sort(key=lambda x: (x["ano_n"], x["ano_usd"], x["total_n"]), reverse=True)
    return filas


def exportar_reporte_contactos(ruta, mes=None):
    """Reporte del ranking de contactos (Excel)."""
    import openpyxl
    filas = ranking_contactos()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Ranking contactos"
    ws.append(["#", "Contacto", "Agencia(s)", "Reservas del mes", "Reservas del ano",
               "Monto ano USD", "Reservas total", "Monto total USD"])
    for i, f in enumerate(filas, 1):
        ws.append([i, f["contacto"], f["empresas"], f["mes_n"], f["ano_n"], f["ano_usd"],
                   f["total_n"], f["total_usd"]])
    _estilo_encabezado_xlsx(ws); _autoancho_xlsx(ws); ws.freeze_panes = "A2"
    wb.save(ruta)
    return len(filas)


def _fechas_in_out(fechas_viaje):
    s = fechas_viaje or ""
    if " al " in s:
        a, b = s.split(" al ", 1)
        return a.strip(), b.strip()
    return s.strip(), ""


def _parse_ddmmyyyy(s):
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime((s or "").strip(), fmt).date()
        except Exception:
            continue
    return None


def _voucher_defaults(ciudad="", hotel="", fecha_in="", fecha_out="", habitaciones=""):
    """Campos de la Orden de Servicio (voucher al cliente), parametrizables."""
    return {
        "os_ciudad": ciudad, "os_hotel": hotel,
        "os_fecha_in": fecha_in, "os_fecha_out": fecha_out,
        "os_habitaciones": habitaciones, "os_acomodacion": "",
        "os_alimentacion": "", "os_origen": "",
        "os_contacto_principal": "", "os_contacto_secundario": "",
        "os_vuelo_llegada": "", "os_hora_llegada": "",
        "os_vuelo_salida": "", "os_hora_salida": "",
        "os_vuelo_interno1": "", "os_vuelo_interno2": "",
        "os_contacto_emergencia": "", "os_info_adicional": "",
        "os_pasajeros": "", "os_actividades": "",
    }


def eliminar_reserva(numero, uid=None):
    """Borra UNA sola reserva. Por 'uid' (identificador unico) si se conoce; si no,
       borra solo la PRIMERA que coincida por numero (nunca todas las del mismo numero).
       Tambien la borra de la nube para que desaparezca en los demas equipos."""
    data = cargar_reservas()
    quitado_uid = uid
    if uid:
        data["items"] = [it for it in data["items"] if str(it.get("uid")) != str(uid)]
    else:
        nuevo = []
        borrado = False
        for it in data["items"]:
            if not borrado and it.get("numero") == numero:
                borrado = True
                quitado_uid = it.get("uid")
                continue
            nuevo.append(it)
        data["items"] = nuevo
    guardar_reservas(data)
    try:
        if quitado_uid:
            enviar_borrado_nube("reserva", quitado_uid)
    except Exception:
        pass


def reparar_reservas():
    """Repara el archivo local: asigna 'uid' a las que no lo tengan y renumera las
       reservas que tengan un numero repetido (para que cada una sea unica y borrar
       una no afecte a las demas). Devuelve cuantas reparo."""
    data = cargar_reservas()
    vistos_num = set()
    seq = int(data.get("seq", RES_SEQ_INICIAL) or RES_SEQ_INICIAL)
    cambios = 0
    for it in data["items"]:
        if not it.get("uid"):
            it["uid"] = "RES-" + uuid.uuid4().hex[:12]
            cambios += 1
        num = str(it.get("numero", "") or "")
        if not num or num in vistos_num:
            seq += 1
            it["numero"] = str(seq)
            cambios += 1
        vistos_num.add(str(it.get("numero", "")))
    if cambios:
        data["seq"] = max(seq, int(data.get("seq", RES_SEQ_INICIAL) or RES_SEQ_INICIAL))
        guardar_reservas(data)
    return cambios


def _parse_pasajeros(txt):
    """Cada linea: 'Nombre, Documento' -> lista de (nombre, documento)."""
    out = []
    for ln in (txt or "").splitlines():
        ln = ln.strip()
        if not ln:
            continue
        if "," in ln:
            n, doc = ln.split(",", 1)
            out.append((n.strip(), doc.strip()))
        else:
            out.append((ln, ""))
    return out


def pasajeros_de(res):
    """Lista de (nombre, documento, telefono) de la reserva. Usa la lista
       estructurada 'pasajeros_list'; si no existe, migra del texto 'os_pasajeros'."""
    lst = res.get("pasajeros_list")
    if isinstance(lst, list) and lst:
        return [(p.get("nombre", ""), p.get("documento", ""), p.get("telefono", ""))
                for p in lst if (p.get("nombre", "") or p.get("documento", ""))]
    return [(n, d, "") for n, d in _parse_pasajeros(res.get("os_pasajeros", ""))]


_MESES_ES = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO",
             "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]


def fecha_larga_es(d):
    """Formatea una fecha como '30 DE AGOSTO DE 2026'."""
    try:
        return f"{d.day} DE {_MESES_ES[d.month]} DE {d.year}"
    except Exception:
        return ""


def _parse_actividades(txt, itinerario=""):
    """Cada linea: 'Fecha | Actividad | Observacion'. Si vacio, usa el itinerario
       (lineas 'DIA N - texto') como (Dia, Actividad, '')."""
    out = []
    for ln in (txt or "").splitlines():
        ln = ln.strip()
        if not ln:
            continue
        partes = [p.strip() for p in ln.split("|")]
        while len(partes) < 3:
            partes.append("")
        out.append((partes[0], partes[1], partes[2]))
    if out:
        return out
    for ln in (itinerario or "").splitlines():
        ln = ln.strip()
        if not ln:
            continue
        if " - " in ln:
            dia, resto = ln.split(" - ", 1)
            out.append((dia.strip(), resto.strip(), ""))
        else:
            out.append(("", ln, ""))
    return out


def _monto_fmt(v, moneda="USD"):
    try:
        return f"{moneda} {float(v):,.2f}"
    except Exception:
        return f"{moneda} {v}"


def _voucher_encabezado(pdf, titulo):
    T = pdf._t
    pdf.set_fill_color(*PDF_PRIM); pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 9, T("  " + titulo), ln=1, fill=True)
    pdf.ln(3)
    pdf.set_text_color(*PDF_TXT)


def _voucher_fila(pdf, etq, val):
    if not val:
        return
    T = pdf._t
    y = pdf.get_y()
    pdf.set_xy(15, y)
    pdf.set_font("Helvetica", "B", 10); pdf.set_text_color(*PDF_PRIM)
    pdf.cell(46, 7, T(etq + ":"))
    pdf.set_font("Helvetica", "", 10); pdf.set_text_color(*PDF_TXT)
    pdf.set_xy(61, y)
    pdf.multi_cell(134, 7, T(str(val)))


class VoucherPDF(FPDF):
    """PDF de la Orden de Servicio (voucher al cliente), estilo tabla."""
    def __init__(self, cfg):
        super().__init__(orientation="P", unit="mm", format="A4")
        self.cfg = cfg
        self.set_auto_page_break(auto=True, margin=12)
        self.set_margins(12, 12, 12)

    def _guard_ancho(self, w):
        if not w:
            disponible = (self.w - self.r_margin) - self.get_x()
            if disponible < 3:
                self.set_x(self.l_margin)

    def cell(self, w=0, *args, **kwargs):
        self._guard_ancho(w)
        return super().cell(w, *args, **kwargs)

    def multi_cell(self, w=0, *args, **kwargs):
        self._guard_ancho(w)
        return super().multi_cell(w, *args, **kwargs)

    def footer(self):
        self.set_y(-12); self.set_font("Helvetica", "I", 7)
        self.set_text_color(140, 140, 140)
        self.cell(0, 8, self._t(f"{self.cfg.get('empresa','')}  -  Pagina {self.page_no()}"),
                  align="C")

    def _t(self, texto):
        if texto is None:
            return ""
        return str(texto).encode("latin-1", "replace").decode("latin-1")


def _os_band(pdf, texto, w=186, h=7):
    pdf.set_x(12)
    pdf.set_font("Helvetica", "B", 9); pdf.set_fill_color(*PDF_PRIM)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(w, h, pdf._t(texto), border=1, ln=1, fill=True, align="C")


def _os_row(pdf, label, value, wl=50, wv=136, h=6.5, val_align="L"):
    pdf.set_x(12)
    pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO)
    pdf.set_text_color(*PDF_PRIM)
    pdf.cell(wl, h, pdf._t(" " + label), border=1, fill=True)
    pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
    pdf.cell(wv, h, pdf._t(" " + str(value)), border=1, ln=1, align=val_align)


def generar_voucher_cliente(cfg, res, ruta):
    """Orden de Servicio (voucher de confirmacion) para el cliente, parametrizable."""
    pdf = VoucherPDF(cfg); pdf.add_page(); T = pdf._t
    ase = res.get("asesor", {}) or {}
    numero = res.get("numero", "")

    # --- Encabezado: titulo + asesor + bienvenida (izq) y logo (der) ---
    logo = cfg.get("logo", "")
    top_y = 12
    if logo and os.path.exists(logo):
        try:
            pdf.image(logo, x=158, y=12, h=22)
        except Exception:
            pass
    wtxt = 144
    pdf.set_xy(12, 13)
    pdf.set_font("Helvetica", "B", 12); pdf.set_fill_color(*PDF_PRIM)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(wtxt, 9, T("ORDEN DE SERVICIO N. " + numero), border=1, ln=2, fill=True, align="C")
    pdf.set_x(12)
    pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO)
    pdf.set_text_color(*PDF_PRIM)
    pdf.cell(wtxt, 6, T(cfg.get("empresa", "") + "  -  Asesor: " + ase.get("nombre", "")),
             border=1, ln=2, fill=True)
    pdf.set_x(12)
    pdf.set_font("Helvetica", "", 8); pdf.set_text_color(*PDF_TXT)
    pdf.multi_cell(wtxt, 4.4, T("Para nosotros es un gusto poder contar con ustedes y su reserva. "
                                "Su reserva esta confirmada. Por favor conserve este voucher para "
                                "garantizar la estadia de los pasajeros en destino."), border=1)
    if pdf.get_y() < top_y + 24:
        pdf.set_y(top_y + 24)
    pdf.ln(3)

    # --- RESERVA HOTELERA (una fila Ciudad/Hotel por destino) ---
    _os_band(pdf, "RESERVA HOTELERA")
    dd = destinos_detalle_de(res)
    filas_hotel = []
    for d in dd:
        hoteles = [h.get("servicio", "") for h in d.get("hotel", []) if h.get("servicio")]
        if d.get("nombre") or hoteles:
            filas_hotel.append((d.get("nombre", ""), " / ".join(hoteles)))
    if not filas_hotel:
        filas_hotel = [(res.get("os_ciudad", "") or ", ".join(res.get("destinos", [])),
                        res.get("os_hotel", ""))]
    for ciudad_d, hotel_d in filas_hotel:
        _os_row(pdf, "CIUDAD", ciudad_d)
        _os_row(pdf, "HOTEL", hotel_d)
    ini = res.get("os_fecha_in", ""); fin = res.get("os_fecha_out", "")
    pdf.set_x(12)
    pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO); pdf.set_text_color(*PDF_PRIM)
    pdf.cell(40, 6.5, T(" FECHA IN"), border=1, fill=True)
    pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
    pdf.cell(53, 6.5, T(" " + ini), border=1)
    pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO); pdf.set_text_color(*PDF_PRIM)
    pdf.cell(40, 6.5, T(" FECHA OUT"), border=1, fill=True)
    pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
    pdf.cell(53, 6.5, T(" " + fin), border=1, ln=1)
    hab = res.get("os_habitaciones", "") or res.get("hab", "")
    acom = res.get("os_acomodacion", "")
    _os_row(pdf, "N. HABITACIONES", hab)
    _os_row(pdf, "ACOMODACION", acom)

    # --- PASAJEROS / IDENTIFICACION ---
    pdf.ln(1)
    pdf.set_x(12); pdf.set_font("Helvetica", "B", 9); pdf.set_fill_color(*PDF_PRIM)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(96, 7, T("PASAJEROS"), border=1, fill=True, align="C")
    pdf.cell(46, 7, T("IDENTIFICACION"), border=1, fill=True, align="C")
    pdf.cell(44, 7, T("TELEFONO"), border=1, ln=1, fill=True, align="C")
    pax = pasajeros_de(res)
    if not pax:
        pax = [(res.get("pax_txt", "") or "-", "", "")]
    pdf.set_text_color(*PDF_TXT)
    for nom, doc, tel in pax:
        pdf.set_x(12); pdf.set_font("Helvetica", "", 8.5)
        pdf.cell(96, 6, T(" " + nom), border=1)
        pdf.cell(46, 6, T(" " + doc), border=1, align="C")
        pdf.cell(44, 6, T(" " + tel), border=1, ln=1, align="C")

    # --- Datos generales ---
    pdf.ln(1)
    _os_row(pdf, "ALIMENTACION", res.get("os_alimentacion", ""))
    _os_row(pdf, "ORIGEN", res.get("os_origen", ""))
    _os_row(pdf, "CONTACTO PRINCIPAL", res.get("os_contacto_principal", ""))
    _os_row(pdf, "SEGUNDO CONTACTO", res.get("os_contacto_secundario", ""))

    # --- ITINERARIO DE VUELO (con hora de llegada y salida) ---
    pdf.ln(1)
    _os_band(pdf, "ITINERARIO DE VUELO")

    def _fila_vuelo(etq, vuelo, hora):
        pdf.set_x(12); pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO)
        pdf.set_text_color(*PDF_PRIM)
        pdf.cell(46, 6.5, T(" " + etq), border=1, fill=True)
        pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
        pdf.cell(80, 6.5, T(" " + vuelo), border=1)
        pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO); pdf.set_text_color(*PDF_PRIM)
        pdf.cell(20, 6.5, T(" HORA"), border=1, fill=True)
        pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
        pdf.cell(40, 6.5, T(" " + hora), border=1, ln=1)

    _fila_vuelo("VUELO DE LLEGADA", res.get("os_vuelo_llegada", ""), res.get("os_hora_llegada", ""))
    _fila_vuelo("VUELO DE SALIDA", res.get("os_vuelo_salida", ""), res.get("os_hora_salida", ""))
    _os_row(pdf, "VUELO INTERNO 1", res.get("os_vuelo_interno1", ""))
    _os_row(pdf, "VUELO INTERNO 2", res.get("os_vuelo_interno2", ""))

    # --- DESCRIPCION DE ACTIVIDADES ---
    pdf.ln(1)
    _os_band(pdf, "DESCRIPCION DE ACTIVIDADES")
    pdf.set_x(12); pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO)
    pdf.set_text_color(*PDF_PRIM)
    pdf.cell(40, 6.5, T(" FECHA"), border=1, fill=True, align="C")
    pdf.cell(96, 6.5, T(" ACTIVIDADES"), border=1, fill=True, align="C")
    pdf.cell(50, 6.5, T(" OBSERVACIONES"), border=1, ln=1, fill=True, align="C")
    acts = _parse_actividades(res.get("os_actividades", ""), res.get("itinerario", ""))
    pdf.set_text_color(*PDF_TXT)
    if not acts:
        acts = [("", "", "")]
    for fecha, act, obs in acts:
        pdf.set_x(12); pdf.set_font("Helvetica", "", 8.5)
        y0 = pdf.get_y()
        # altura dinamica segun la actividad
        pdf.multi_cell(40, 5.2, T(" " + fecha), border=1)
        h1 = pdf.get_y() - y0
        pdf.set_xy(52, y0)
        pdf.multi_cell(96, 5.2, T(" " + act), border=1)
        h2 = pdf.get_y() - y0
        pdf.set_xy(148, y0)
        pdf.multi_cell(50, 5.2, T(" " + obs), border=1)
        h3 = pdf.get_y() - y0
        pdf.set_y(y0 + max(h1, h2, h3))

    # --- INFORMACION ADICIONAL ---
    info = res.get("os_info_adicional", "")
    if info:
        pdf.ln(1)
        _os_band(pdf, "INFORMACION ADICIONAL")
        pdf.set_x(12); pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
        pdf.multi_cell(186, 5, T(info), border=1)

    # --- CONTACTO DE EMERGENCIA ---
    emerg = res.get("os_contacto_emergencia", "") or cfg.get("telefono", "")
    if emerg:
        pdf.ln(1)
        _os_band(pdf, "CONTACTO DE EMERGENCIA:  " + emerg)

    # --- CONDICIONES GENERALES Y DE PAGO (pagina propia) ---
    _condiciones_cliente(pdf, cfg)

    pdf.output(ruta)


def _condiciones_cliente(pdf, cfg):
    """Pagina de Condiciones Generales y de Pago en el voucher del cliente."""
    T = pdf._t
    pdf.add_page()
    _os_band(pdf, "CONDICIONES GENERALES Y DE PAGO")
    pdf.ln(2)

    def titulo(txt):
        pdf.set_x(12); pdf.set_font("Helvetica", "B", 9); pdf.set_text_color(*PDF_PRIM)
        pdf.multi_cell(186, 4.8, T(txt))

    def cuerpo(txt):
        pdf.set_x(12); pdf.set_font("Helvetica", "", 8.3); pdf.set_text_color(*PDF_TXT)
        pdf.multi_cell(186, 4.4, T(txt)); pdf.ln(1.2)

    def item(tit, cpo):
        titulo(tit); cuerpo(cpo)

    item("IVA:",
         "Los precios NO incluyen IVA por ser servicio a extranjeros. En caso de que la venta "
         "sea a un colombiano se debe aplicar la tarifa del 19% de IVA.")
    item("Forma de pago:",
         "1 mes antes de la llegada del grupo. Por link de pagos con sobrecargo del 3%, o "
         "transferencia bancaria a la cuenta de ahorros BANISTMO PANAMA 0120179743.")
    item("Validez de la oferta:", "30 dias.")
    item("Pago:",
         "El pago debe realizarse con 30 dias de antelacion a la llegada de los pasajeros. De no "
         "presentarse el pago no se garantiza la reserva y queda sujeta a cambio de tarifas y "
         "disponibilidad de los servicios.")
    titulo("Cancelaciones:")
    for b in ("Si la cancelacion se realiza con 30 dias de antelacion, se realiza la devolucion "
              "del 100% del valor pagado.",
              "Si la cancelacion se realiza con 20 dias a la llegada de los pasajeros, se devuelve "
              "el 50% del valor pagado.",
              "Si la cancelacion se realiza con 10 dias o menos a la llegada de los pasajeros, no "
              "se realiza devolucion del dinero pagado."):
        pdf.set_x(14); pdf.set_font("Helvetica", "", 8.3); pdf.set_text_color(*PDF_TXT)
        pdf.multi_cell(184, 4.4, T("-  " + b))
    pdf.ln(1.5)

    cuerpo("La presente cotizacion no implica reserva ni bloqueo de lugares. Todas las tarifas "
           "estan sujetas a disponibilidad al momento de realizar la reserva en firme. Precios "
           "indicados en Dolares Americanos, de caracter informativo, y deben ser confirmados "
           "para realizar la reservacion ya que estan sujetos a modificaciones sin previo aviso. "
           "Precios de contado en dolares americanos; para transferencia electronica o link de "
           "pagos aplica un sobrecargo del 3% adicional. Favor tener en cuenta que los pasaportes "
           "para todo viaje internacional deben tener una vigencia minima de 6 meses al momento "
           "de abordar; el pasajero es responsable de comunicarse con los consulados de los "
           "paises que visitara para reconfirmar Visas y Vacunas requeridas para el ingreso. En "
           "caso de no recibir copias de pasaportes en la fecha establecida, INNOBA DMC S.A.S, "
           "Operadora Mayorista, no se hace responsable por la informacion recibida; cualquier "
           "cambio o modificacion sera responsabilidad de la agencia y/o pasajero y estara sujeto "
           "a las condiciones y cargos de la aerolinea. Este documento es INDISPENSABLE para la "
           "emision de los boletos de avion y tren cuando corresponda.")
    cuerpo("Los productos y/o servicios que se venden a los clientes finales de cada destino "
           "estan sujetos a cambios por efectos tales como accidentes, huelgas, revueltas, "
           "terremotos y otros acontecimientos de fuerza mayor que puedan ocurrir durante el "
           "viaje, por lo que los organizadores se reservan el derecho de hacer los cambios "
           "necesarios para asegurar el exito del producto turistico o servicio adquirido por el "
           "cliente. Si por alguna razon de fuerza mayor el operador se ve obligado a cancelar los "
           "servicios programados, se solicitara a los organizadores la devolucion del importe de "
           "los servicios cancelados como parte de los terminos y condiciones de la reserva.")
    cuerpo("La agencia o proveedor seleccionado tienen la autoridad para retirar de la gira a "
           "cualquier persona que, por motivos graves de conducta moral o disciplinaria, pueda "
           "danar o socavar el exito del tour o actividad; caso en el que el usuario tiene derecho "
           "al reembolso del valor de los servicios turisticos no disfrutados dependiendo de las "
           "politicas de los proveedores seleccionados para el plan adquirido. La agencia no "
           "asume ninguna responsabilidad respecto a asuntos legales o cualquier otra cuestion "
           "que pueda resultar en que el usuario se vea obligado a retirarse de la excursion, ni "
           "sobre los costos personales que el pasajero pueda incurrir.")
    item("AVISO DE CONFIDENCIALIDAD:",
         "Con fundamento en la Ley de Proteccion de Datos Personales, INNOBA DMC S.A.S es "
         "responsable de recabar sus datos personales, del uso que se les de y de su proteccion. "
         "Usted tiene derecho al Acceso, Rectificacion y Cancelacion de sus datos personales, a "
         "Oponerse a su tratamiento o a revocar el consentimiento otorgado. Para ello envie la "
         "solicitud al correo felipe@innobadmc.com y comuniquese a nuestra oficina para confirmar "
         "su correcta recepcion. Para consultar el Aviso de Privacidad completo o sus "
         "modificaciones, visite www.innobadmc.com")
    pdf.ln(2)
    pdf.set_x(12); pdf.set_font("Helvetica", "", 8.3); pdf.set_text_color(*PDF_TXT)
    pdf.multi_cell(186, 4.4, T("Cordialmente,"))
    pdf.set_font("Helvetica", "B", 9); pdf.set_text_color(*PDF_PRIM)
    pdf.set_x(12); pdf.multi_cell(186, 4.6, T("Felipe Ortiz Jaramillo"))
    pdf.set_font("Helvetica", "", 8.3); pdf.set_text_color(*PDF_TXT)
    pdf.set_x(12); pdf.multi_cell(186, 4.4, T("GERENTE GENERAL  -  INNOBA DMC\n"
                                              "Cel: +57 313 595 2944   ·   Correo: felipe@innobadmc.com"))


def _voucher_itinerario(pdf, texto):
    """Renderiza el itinerario dia por dia en el voucher (lineas 'DIA ...' en azul)."""
    texto = (texto or "").strip()
    if not texto:
        return
    T = pdf._t
    if pdf.get_y() > 245:
        pdf.add_page()
    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 11); pdf.set_text_color(*PDF_BLUE)
    pdf.cell(0, 7, T("ITINERARIO DE VIAJE"), ln=1)
    for linea in texto.splitlines():
        ln = linea.rstrip()
        if not ln.strip():
            pdf.ln(1); continue
        if pdf.get_y() > 268:
            pdf.add_page()
        es_dia = ln.strip().upper().startswith(("DIA", "DÍA"))
        pdf.set_x(15)
        if es_dia:
            pdf.set_font("Helvetica", "B", 10); pdf.set_text_color(*PDF_PRIM)
        else:
            pdf.set_font("Helvetica", "", 9); pdf.set_text_color(*PDF_TXT)
        pdf.multi_cell(180, 5.5, T(ln))


def _os_encab_prov(pdf, cfg, res, titulo):
    """Encabezado comun de los vouchers a proveedor (bienvenida + orden + logo)."""
    T = pdf._t
    logo = cfg.get("logo", "")
    if logo and os.path.exists(logo):
        try:
            pdf.image(logo, x=158, y=12, h=20)
        except Exception:
            pass
    wtxt = 144
    pdf.set_xy(12, 13)
    pdf.set_font("Helvetica", "B", 12); pdf.set_fill_color(*PDF_PRIM)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(wtxt, 9, T(titulo), border=1, ln=2, fill=True, align="C")
    pdf.set_x(12)
    pdf.set_font("Helvetica", "", 8); pdf.set_text_color(*PDF_TXT)
    pdf.multi_cell(wtxt, 4.4, T("Para nosotros es un placer atenderlos, su reserva esta confirmada. "
                                "Por favor mantener este voucher para facilitar la coordinacion de "
                                "los servicios en destino."), border=1)
    if pdf.get_y() < 34:
        pdf.set_y(34)
    pdf.ln(2)
    n_pax = len([1 for p in pasajeros_de(res)]) or res.get("pax_txt", "")
    pdf.set_x(12); pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO)
    pdf.set_text_color(*PDF_PRIM)
    pdf.cell(50, 6.5, T(" ORDEN OPERADOR"), border=1, fill=True)
    pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
    pdf.cell(50, 6.5, T(" " + res.get("numero", "")), border=1)
    pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO); pdf.set_text_color(*PDF_PRIM)
    pdf.cell(40, 6.5, T(" # TURISTAS"), border=1, fill=True)
    pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
    pdf.cell(46, 6.5, T(" " + str(n_pax)), border=1, ln=1)


def _os_pasajeros_tabla(pdf, res):
    T = pdf._t
    pax = pasajeros_de(res)
    if not pax:
        return
    pdf.ln(1)
    pdf.set_x(12); pdf.set_font("Helvetica", "B", 9); pdf.set_fill_color(*PDF_PRIM)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(96, 7, T("NOMBRE PASAJEROS"), border=1, fill=True, align="C")
    pdf.cell(46, 7, T("IDENTIFICACION"), border=1, fill=True, align="C")
    pdf.cell(44, 7, T("TELEFONO"), border=1, ln=1, fill=True, align="C")
    pdf.set_text_color(*PDF_TXT)
    for nom, doc, tel in pax:
        pdf.set_x(12); pdf.set_font("Helvetica", "", 8.5)
        pdf.cell(96, 6, T(" " + nom), border=1)
        pdf.cell(46, 6, T(" " + doc), border=1, align="C")
        pdf.cell(44, 6, T(" " + tel), border=1, ln=1, align="C")


def _os_pie_prov(pdf, cfg, res):
    T = pdf._t
    ase = res.get("asesor", {}) or {}
    obs = res.get("os_info_adicional", "")
    emerg = res.get("os_contacto_emergencia", "") or cfg.get("telefono", "")
    if emerg:
        pdf.ln(1); _os_band(pdf, "CONTACTO DE EMERGENCIA:  " + emerg)
    pdf.ln(1)
    _os_row(pdf, "OBSERVACIONES", obs, wl=40, wv=146)
    pdf.ln(2)
    pdf.set_x(12); pdf.set_font("Helvetica", "", 8); pdf.set_text_color(110, 110, 110)
    nit = (" (NIT/RUC " + cfg.get("nit", "") + ")") if cfg.get("nit") else ""
    pdf.multi_cell(186, 4.5, T("Estimado proveedor: confirmamos la reserva de los servicios descritos "
                               "para los pasajeros indicados. Favor confirmar disponibilidad y remitir "
                               "la facturacion a nombre de " + cfg.get("empresa", "") + nit + ".  "
                               "Asesor: " + ase.get("nombre", "") + "  " + ase.get("email", "")))


def generar_voucher_proveedor(cfg, res, renglon, ruta):
    """Voucher a proveedor con un modelo distinto segun el tipo:
       Guia, Transporte, Hotel o Actividad."""
    tipo = renglon.get("tipo", "")
    pdf = VoucherPDF(cfg); pdf.add_page(); T = pdf._t
    prov = renglon.get("proveedor", "")
    dest = renglon.get("destino", "")
    serv = renglon.get("servicio", "")
    # la fecha propia del servicio manda; si no tiene, usa la de la reserva
    fecha = renglon.get("fecha", "") or res.get("os_fecha_in", "") or res.get("fechas_viaje", "")
    hora = renglon.get("hora", "")
    obs = renglon.get("observacion", "")
    n_pax = len(pasajeros_de(res)) or res.get("pax_txt", "")

    if tipo == "Guia":
        _os_encab_prov(pdf, cfg, res, "VOUCHER RESERVA GUIA  -  ORDEN " + res.get("numero", ""))
        _os_band(pdf, "RESERVA GUIA")
        _os_row(pdf, "GUIA", prov)
        _os_row(pdf, "DESTINO / SERVICIO", (dest + "  -  " + serv).strip(" -"))
        pdf.set_x(12); pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO)
        pdf.set_text_color(*PDF_PRIM)
        pdf.cell(40, 6.5, T(" FECHA"), border=1, fill=True)
        pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
        pdf.cell(53, 6.5, T(" " + fecha), border=1)
        pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO); pdf.set_text_color(*PDF_PRIM)
        pdf.cell(40, 6.5, T(" HORA"), border=1, fill=True)
        pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
        pdf.cell(53, 6.5, T(" " + hora), border=1, ln=1)
        _os_row(pdf, "# PASAJEROS", str(n_pax))
        _os_pasajeros_tabla(pdf, res)
    elif tipo == "Transporte":
        _os_encab_prov(pdf, cfg, res, "VOUCHER TRANSPORTE  -  ORDEN " + res.get("numero", ""))
        _os_band(pdf, "SERVICIO DE TRANSPORTE")
        _os_row(pdf, "PROVEEDOR", prov)
        # tabla estilo servicio de transporte
        pdf.set_x(12); pdf.set_font("Helvetica", "B", 8); pdf.set_fill_color(*PDF_CLARO)
        pdf.set_text_color(*PDF_PRIM)
        anchos = [(30, "FECHA"), (44, "ORIGEN"), (37, "DESTINO"), (20, "HORA"),
                  (25, "N. PAX"), (30, "VEHICULO")]
        for w, h in anchos:
            pdf.cell(w, 6.5, T(" " + h), border=1, fill=True, align="C")
        pdf.ln(6.5)
        pdf.set_x(12); pdf.set_font("Helvetica", "", 8); pdf.set_text_color(*PDF_TXT)
        vals = [fecha, renglon.get("origen", ""), dest or serv, hora, str(n_pax),
                renglon.get("vehiculo", "")]
        for (w, _h), v in zip(anchos, vals):
            pdf.cell(w, 6.5, T(" " + str(v)), border=1)
        pdf.ln(6.5)
        _os_pasajeros_tabla(pdf, res)
    elif tipo == "Hotel":
        _os_encab_prov(pdf, cfg, res, "VOUCHER RESERVA HOTELERA  -  ORDEN " + res.get("numero", ""))
        _os_band(pdf, "RESERVA HOTELERA")
        _os_row(pdf, "HOTEL / PROVEEDOR", prov or serv)
        _os_row(pdf, "CIUDAD", dest)
        pdf.set_x(12); pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO)
        pdf.set_text_color(*PDF_PRIM)
        pdf.cell(40, 6.5, T(" FECHA IN"), border=1, fill=True)
        pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
        pdf.cell(53, 6.5, T(" " + (res.get("os_fecha_in", "") or fecha)), border=1)
        pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO); pdf.set_text_color(*PDF_PRIM)
        pdf.cell(40, 6.5, T(" FECHA OUT"), border=1, fill=True)
        pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
        pdf.cell(53, 6.5, T(" " + res.get("os_fecha_out", "")), border=1, ln=1)
        _os_row(pdf, "N. HABITACIONES", res.get("os_habitaciones", "") or res.get("hab", ""))
        _os_row(pdf, "ACOMODACION", res.get("os_acomodacion", ""))
        _os_pasajeros_tabla(pdf, res)
    else:  # Actividad
        _os_encab_prov(pdf, cfg, res, "VOUCHER RESERVA ACTIVIDAD  -  ORDEN " + res.get("numero", ""))
        _os_band(pdf, "RESERVA DE ACTIVIDAD")
        _os_row(pdf, "ACTIVIDAD", serv)
        _os_row(pdf, "PROVEEDOR", prov)
        _os_row(pdf, "DESTINO", dest)
        pdf.set_x(12); pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO)
        pdf.set_text_color(*PDF_PRIM)
        pdf.cell(40, 6.5, T(" FECHA"), border=1, fill=True)
        pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
        pdf.cell(53, 6.5, T(" " + fecha), border=1)
        pdf.set_font("Helvetica", "B", 8.5); pdf.set_fill_color(*PDF_CLARO); pdf.set_text_color(*PDF_PRIM)
        pdf.cell(40, 6.5, T(" HORA"), border=1, fill=True)
        pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
        pdf.cell(53, 6.5, T(" " + hora), border=1, ln=1)
        _os_row(pdf, "# PASAJEROS", str(n_pax))
        _os_pasajeros_tabla(pdf, res)

    if obs:
        pdf.ln(1); _os_row(pdf, "DETALLE / INCLUYE", obs, wl=40, wv=146)
    _os_pie_prov(pdf, cfg, res)
    pdf.output(ruta)


# ============================================================================
# Widgets auxiliares
# ============================================================================
class Stepper(ctk.CTkFrame):
    def __init__(self, master, value=0, minimo=0, maximo=99, width=104,
                 command=None, **kw):
        super().__init__(master, fg_color=CARD2, corner_radius=8, **kw)
        self.minimo, self.maximo = minimo, maximo
        self.command = command
        self.var = tk.StringVar(value=str(value))
        ctk.CTkButton(self, text="-", width=28, height=28, corner_radius=8,
                      fg_color=NAVY, hover_color=BLUE, font=("Segoe UI", 15, "bold"),
                      command=self._menos).pack(side="left", padx=3, pady=3)
        ctk.CTkEntry(self, textvariable=self.var, width=width - 74, height=28,
                     justify="center", border_width=0, fg_color=CARD2, text_color=TEXT,
                     font=("Segoe UI", 13, "bold")).pack(side="left")
        ctk.CTkButton(self, text="+", width=28, height=28, corner_radius=8,
                      fg_color=NAVY, hover_color=BLUE, font=("Segoe UI", 15, "bold"),
                      command=self._mas).pack(side="left", padx=3, pady=3)
        self.var.trace_add("write", lambda *a: self.command() if self.command else None)

    def get(self):
        try:
            return int(float(self.var.get()))
        except Exception:
            return self.minimo

    def set(self, v):
        self.var.set(str(v))

    def _menos(self):
        self.set(max(self.minimo, self.get() - 1))

    def _mas(self):
        self.set(min(self.maximo, self.get() + 1))


# ============================================================================
# Ventana de busqueda rapida de cliente (lupa)
# ============================================================================
class BuscadorClientes(ctk.CTkToplevel):
    def __init__(self, master, clientes, on_pick, on_editar=None, on_eliminar=None):
        super().__init__(master)
        self.clientes = clientes; self.on_pick = on_pick
        self.on_editar = on_editar; self.on_eliminar = on_eliminar
        self.title("Buscar cliente")
        self.geometry("480x540"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        self.grid_columnconfigure(0, weight=1); self.grid_rowconfigure(2, weight=1)
        ctk.CTkLabel(self, text="Buscar cliente", text_color=NAVY,
                     font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w",
                                                         padx=16, pady=(14, 2))
        self.var_q = tk.StringVar()
        e = ctk.CTkEntry(self, textvariable=self.var_q, height=40, corner_radius=10,
                         border_color=BLUE, border_width=2, fg_color=CARD,
                         font=("Segoe UI", 13),
                         placeholder_text="Escribe la empresa o el asesor...")
        e.grid(row=1, column=0, sticky="ew", padx=16, pady=(2, 8))
        self.var_q.trace_add("write", lambda *a: self._pintar())
        self.lista = ctk.CTkScrollableFrame(self, fg_color=CARD, corner_radius=12)
        self.lista.grid(row=2, column=0, sticky="nsew", padx=16, pady=(0, 14))
        self._pintar()
        self.after(120, e.focus_set)
        self.bind("<Escape>", lambda ev: self.destroy())

    def _pintar(self):
        for w in self.lista.winfo_children():
            w.destroy()
        q = _nz(self.var_q.get())
        n = 0
        for c in self.clientes:
            emp = c.get("empresa", "")
            vends = c.get("vendedores", []) or []
            if q and not (q in _nz(emp) or any(q in _nz(v.get("nombre", "")) for v in vends)):
                continue
            n += 1
            if n > 250:
                break
            sub = ", ".join(v.get("nombre", "") for v in vends[:3])
            txt = emp + (("\n   " + sub) if sub else "")
            fila = ctk.CTkFrame(self.lista, fg_color=CARD2, corner_radius=8)
            fila.pack(fill="x", padx=4, pady=2)
            fila.grid_columnconfigure(0, weight=1)
            ctk.CTkButton(fila, text=txt, anchor="w", height=42, corner_radius=8,
                          fg_color=CARD2, text_color=NAVY, hover_color=LINE,
                          font=("Segoe UI", 12, "bold"),
                          command=lambda x=emp: self._pick(x)).grid(row=0, column=0, sticky="ew")
            if self.on_editar:
                ctk.CTkButton(fila, text="✏", width=34, height=42, corner_radius=8,
                              fg_color=CARD2, text_color=BLUE, hover_color=LINE,
                              font=("Segoe UI", 14),
                              command=lambda x=emp: self._editar(x)).grid(row=0, column=1)
            if self.on_eliminar:
                ctk.CTkButton(fila, text="🗑", width=34, height=42, corner_radius=8,
                              fg_color=CARD2, text_color=RED, hover_color=LINE,
                              font=("Segoe UI", 14),
                              command=lambda x=emp: self._eliminar(x)).grid(row=0, column=2)
        if not n:
            ctk.CTkLabel(self.lista, text="Sin resultados.", text_color=MUTED).pack(pady=24)

    def _pick(self, emp):
        self.on_pick(emp); self.destroy()

    def _editar(self, emp):
        self.destroy()
        if self.on_editar:
            self.on_editar(emp)

    def _eliminar(self, emp):
        if messagebox.askyesno("Eliminar cliente",
                               f"¿Eliminar definitivamente a:\n\n{emp}\n\nEsta accion no se puede deshacer.",
                               parent=self):
            if self.on_eliminar:
                self.on_eliminar(emp)
            self._pintar()


class SelectorContacto(ctk.CTkToplevel):
    """Busca una empresa y permite elegir el CONTACTO (vendedor) en un clic.
       Llama on_pick(empresa_dict, contacto_dict)."""
    def __init__(self, master, on_pick, titulo="Buscar cliente y contacto"):
        super().__init__(master)
        self.on_pick = on_pick
        self.clientes = cargar_clientes()
        self.title(titulo)
        self.geometry("580x620"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        ctk.CTkLabel(self, text=titulo, font=("Segoe UI", 15, "bold"),
                     text_color=NAVY).pack(anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(self, text="Elige el contacto (vendedor) de la agencia para vincularlo de una vez.",
                     text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=16)
        self.q = tk.StringVar()
        e = ctk.CTkEntry(self, textvariable=self.q, height=40, corner_radius=10, border_color=BLUE,
                         border_width=2, fg_color=CARD, font=("Segoe UI", 12),
                         placeholder_text="Buscar empresa o contacto...")
        e.pack(fill="x", padx=16, pady=(8, 8))
        self.q.trace_add("write", lambda *a: self._pintar())
        self.lista = ctk.CTkScrollableFrame(self, fg_color=CARD, corner_radius=12)
        self.lista.pack(fill="both", expand=True, padx=16, pady=(0, 14))
        self._pintar()
        self.after(120, e.focus_set)
        self.bind("<Escape>", lambda ev: self.destroy())

    def _pintar(self):
        for w in self.lista.winfo_children():
            w.destroy()
        q = _nz(self.q.get())
        n = 0
        for c in self.clientes:
            emp = c.get("empresa", "")
            vends = c.get("vendedores", []) or []
            if q and not (q in _nz(emp) or any(q in _nz(v.get("nombre", "")) for v in vends)):
                continue
            n += 1
            if n > 200:
                break
            card = ctk.CTkFrame(self.lista, fg_color=CARD2, corner_radius=8)
            card.pack(fill="x", padx=4, pady=3)
            ctk.CTkLabel(card, text=emp or "(sin nombre)", text_color=NAVY, anchor="w",
                         font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=8, pady=(6, 2))
            if vends:
                for v in vends:
                    nom = v.get("nombre", "") or "(sin nombre)"
                    sub = v.get("cargo", "") or v.get("telefono", "")
                    ctk.CTkButton(card, text="👤  " + nom + (("   ·   " + sub) if sub else ""),
                                  height=30, corner_radius=6, fg_color=CARD, text_color=NAVY,
                                  hover_color=LINE, anchor="w",
                                  command=lambda cc=c, vv=v: self._pick(cc, vv)).pack(
                        fill="x", padx=8, pady=2)
                ctk.CTkButton(card, text="Usar solo la empresa (sin contacto)", height=26,
                              corner_radius=6, fg_color="transparent", text_color=MUTED,
                              hover_color=LINE, command=lambda cc=c: self._pick(cc, {})).pack(
                    fill="x", padx=8, pady=(2, 6))
            else:
                ctk.CTkButton(card, text="Usar esta empresa", height=30, corner_radius=6,
                              fg_color=GREEN, hover_color=GREEN_H,
                              command=lambda cc=c: self._pick(cc, {})).pack(fill="x", padx=8, pady=(0, 6))
        if n == 0:
            ctk.CTkLabel(self.lista, text="Sin resultados.\nCrea el cliente en 'Clientes'.",
                         text_color=MUTED).pack(pady=20)

    def _pick(self, c, v):
        try:
            self.on_pick(c, v)
        finally:
            self.destroy()


# ============================================================================
# Ventana de Historial de Cotizaciones (consecutivo + busqueda + seguimiento)
# ============================================================================
ESTADOS_COT = ["Pendiente", "Enviada", "En seguimiento", "Ganada", "Perdida"]
ESTADO_COLOR = {"Pendiente": MUTED, "Enviada": BLUE, "En seguimiento": "#D9A400",
                "Ganada": GREEN, "Perdida": RED}
# color de fondo de la fila segun estado (Ganada verde, Perdida rojo, En segui. AMARILLO)
ESTADO_FILA = {"Pendiente": "#F1F5FB", "Enviada": "#EAF2FD", "En seguimiento": "#FFF3C4",
               "Ganada": "#E3F5EA", "Perdida": "#FBE6E6"}


class VentanaCotizacionDetalle(ctk.CTkToplevel):
    """Editar la cotizacion y gestionar tareas de seguimiento."""
    def __init__(self, master, item, on_save):
        super().__init__(master)
        self.item = item; self.on_save = on_save
        self.title("Cotizacion " + item.get("numero", ""))
        self.geometry("660x680+140+10"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        try:
            self.after(60, lambda: self.state("zoomed"))
        except Exception:
            pass
        # Barra FIJA ARRIBA (siempre visible, no depende del alto ni del DPI) con Guardar / Cancelar
        footer = ctk.CTkFrame(self, fg_color=CARD2, height=60, corner_radius=0)
        footer.pack(side="top", fill="x")
        footer.pack_propagate(False)
        ctk.CTkButton(footer, text="💾  Guardar", fg_color=GREEN, hover_color=GREEN_H,
                      height=40, width=170, font=("Segoe UI", 13, "bold"),
                      command=self._guardar).pack(side="right", padx=(6, 16), pady=10)
        ctk.CTkButton(footer, text="Cancelar", fg_color=CARD, text_color=NAVY, hover_color=LINE,
                      height=40, width=120, border_width=1, border_color=LINE,
                      font=("Segoe UI", 12, "bold"),
                      command=self.destroy).pack(side="right", padx=6, pady=10)
        cont = ctk.CTkScrollableFrame(self, fg_color=BG)
        cont.pack(fill="both", expand=True, padx=16, pady=16)
        ctk.CTkLabel(cont, text=f"{item.get('numero','')}   ·   {item.get('fecha','')}",
                     text_color=NAVY, font=("Segoe UI", 16, "bold")).pack(anchor="w")
        dest = ", ".join(item.get("destinos", []))
        ctk.CTkLabel(cont, text=f"{dest}   ·   {usd(item.get('total', 0))}",
                     text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", pady=(0, 8))

        def campo(lbl, val):
            ctk.CTkLabel(cont, text=lbl, text_color=MUTED,
                         font=("Segoe UI", 11)).pack(anchor="w", padx=2)
            v = tk.StringVar(value=val)
            ctk.CTkEntry(cont, textvariable=v, height=32, corner_radius=8,
                         border_color=LINE).pack(fill="x", pady=(0, 8))
            return v
        self.v_cli = campo("Agencia / cliente", item.get("cliente", ""))
        self.v_ase = campo("Asesor", item.get("asesor", ""))
        ctk.CTkLabel(cont, text="Estado del seguimiento", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.v_estado = tk.StringVar(value=item.get("estado", "Pendiente"))
        ctk.CTkOptionMenu(cont, variable=self.v_estado, values=ESTADOS_COT, width=220,
                          height=32, corner_radius=8, fg_color=NAVY, button_color=NAVY2,
                          button_hover_color=BLUE, dropdown_fg_color=CARD,
                          dropdown_text_color=TEXT).pack(anchor="w", pady=(0, 8))
        ctk.CTkLabel(cont, text="Fecha de seguimiento / recordatorio", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.sel_fecha_seg = SelectorFecha(cont, minimo=datetime.date.today())
        self.sel_fecha_seg.pack(anchor="w", pady=(0, 2))
        f0 = parse_fecha(item.get("fecha_seg", ""))
        if f0:
            self.sel_fecha_seg._set(f0)
        ctk.CTkLabel(cont, text="El sistema te avisara al abrir cuando llegue esa fecha.",
                     text_color=MUTED, font=("Segoe UI", 9)).pack(anchor="w", padx=2, pady=(0, 6))
        # correo donde llega el recordatorio de seguimiento
        _app0 = getattr(self.master, "master", None)
        _cfg0 = getattr(_app0, "cfg", None) if _app0 else None
        correo_def = (item.get("correo_seg") or item.get("email")
                      or (_cfg0.get("correo_remitente") if _cfg0 else "") or "")
        self.v_correo_seg = campo("Correo donde llega el recordatorio de seguimiento",
                                  correo_def)
        self.cfg = _cfg0 or cargar_config()

        # --- Correo de seguimiento AL CLIENTE ---
        ctk.CTkLabel(cont, text="CORREO DE SEGUIMIENTO AL CLIENTE", text_color=NAVY,
                     font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(6, 2))
        self.v_correo_cli = campo("Correo del cliente", item.get("email", ""))
        self.v_auto = tk.BooleanVar(value=bool(item.get("auto_correo_seg")))
        ctk.CTkCheckBox(cont, text="Programar: enviarlo automaticamente cuando llegue la fecha "
                        "de seguimiento (al abrir el sistema)", variable=self.v_auto,
                        font=("Segoe UI", 11)).pack(anchor="w", pady=(0, 4))
        fc = ctk.CTkFrame(cont, fg_color="transparent"); fc.pack(fill="x", pady=(0, 8))
        ctk.CTkButton(fc, text="✉ Enviar seguimiento ahora", height=36, fg_color=CYAN,
                      hover_color=BLUE, font=("Segoe UI", 12, "bold"),
                      command=self._enviar_seguimiento).pack(side="left")
        ctk.CTkButton(fc, text="Ver texto", height=36, width=90, fg_color=CARD2, text_color=NAVY,
                      hover_color=LINE, command=self._ver_texto_seg).pack(side="left", padx=8)
        if item.get("correo_seg_enviado"):
            ctk.CTkLabel(fc, text="✓ Enviado " + item.get("correo_seg_enviado", ""),
                         text_color=GREEN_H, font=("Segoe UI", 10, "bold")).pack(side="left", padx=8)

        ctk.CTkLabel(cont, text="Notas", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.txt_notas = ctk.CTkTextbox(cont, height=70, corner_radius=8, border_width=1,
                                        border_color=LINE, fg_color=CARD)
        self.txt_notas.insert("1.0", item.get("notas", "")); self.txt_notas.pack(fill="x", pady=(0, 10))
        # --- tareas de seguimiento ---
        hdr = ctk.CTkFrame(cont, fg_color="transparent"); hdr.pack(fill="x", pady=(4, 2))
        ctk.CTkLabel(hdr, text="Tareas de seguimiento", text_color=NAVY,
                     font=("Segoe UI", 14, "bold")).pack(side="left")
        ctk.CTkButton(hdr, text="+ Agregar tarea", width=140, height=28, corner_radius=8,
                      fg_color=CARD2, text_color=NAVY, hover_color=LINE, border_width=1,
                      border_color=LINE, font=("Segoe UI", 11, "bold"),
                      command=lambda: self._add_tarea()).pack(side="right")
        self.tfr = ctk.CTkFrame(cont, fg_color="transparent"); self.tfr.pack(fill="x", pady=4)
        self.tareas_rows = []
        for t in item.get("tareas", []):
            self._add_tarea(t)
        if not item.get("tareas"):
            self._add_tarea()
        # (Guardar / Cancelar viven en la barra fija del fondo; ver __init__)
        ctk.CTkFrame(cont, fg_color="transparent", height=8).pack()

    def _add_tarea(self, t=None):
        t = t or {}
        row = ctk.CTkFrame(self.tfr, fg_color=CARD2, corner_radius=8)
        row.pack(fill="x", pady=3)
        hecha = tk.BooleanVar(value=bool(t.get("hecha")))
        ctk.CTkCheckBox(row, text="", variable=hecha, width=24, checkbox_width=20,
                        checkbox_height=20, corner_radius=5, fg_color=GREEN,
                        hover_color=GREEN_H).pack(side="left", padx=(8, 2), pady=6)
        v_txt = tk.StringVar(value=t.get("texto", ""))
        ctk.CTkEntry(row, textvariable=v_txt, height=30, corner_radius=6, border_color=LINE,
                     placeholder_text="Que hacer (ej. llamar al cliente)").pack(
            side="left", fill="x", expand=True, padx=3, pady=6)
        v_fec = tk.StringVar(value=t.get("fecha", ""))
        ctk.CTkEntry(row, textvariable=v_fec, width=110, height=30, corner_radius=6,
                     border_color=LINE, placeholder_text="dd/mm/aaaa").pack(side="left", padx=3)
        entry = {"frame": row, "hecha": hecha, "texto": v_txt, "fecha": v_fec}
        ctk.CTkButton(row, text="✕", width=28, height=28, corner_radius=6, fg_color="transparent",
                      text_color=RED, hover_color=LINE,
                      command=lambda: (self.tareas_rows.remove(entry), row.destroy())).pack(
            side="left", padx=(2, 6))
        self.tareas_rows.append(entry)

    def _enviar_seguimiento(self):
        dest = self.v_correo_cli.get().strip()
        if not dest:
            messagebox.showinfo("Correo del cliente",
                                "Escribe el correo del cliente.", parent=self)
            return
        cfgr = cfg_remitente_cotizador(self.cfg, self.item)
        if not (cfgr.get("correo_remitente") and cfgr.get("smtp_password")):
            messagebox.showwarning("Correo no configurado",
                                   "Configura el correo remitente y su contrasena (o el correo "
                                   "de Felipe/Carlos) en 'Datos de mi empresa'.", parent=self)
            return
        if not messagebox.askyesno("Enviar seguimiento",
                                   f"Enviar el correo de seguimiento a {dest}\n"
                                   f"desde {cfgr.get('correo_remitente','')}?", parent=self):
            return
        asunto = (f"Seguimiento de su cotizacion {self.item.get('numero','')} - "
                  f"{cfgr.get('empresa','')}")
        try:
            enviar_correo_texto(cfgr, dest, asunto,
                                cuerpo_seguimiento_cotizacion(self.item, cfgr))
        except Exception as e:
            messagebox.showerror("No se pudo enviar", str(e), parent=self)
            return
        self.item["email"] = dest
        self.item["correo_seg_enviado"] = datetime.date.today().isoformat()
        messagebox.showinfo("Enviado", f"Correo de seguimiento enviado a {dest}.", parent=self)

    def _ver_texto_seg(self):
        top = ctk.CTkToplevel(self); top.title("Texto del correo de seguimiento")
        top.geometry("560x480"); top.configure(fg_color=BG)
        top.transient(self); top.grab_set()
        ctk.CTkLabel(top, text="Asunto: Seguimiento de su cotizacion " +
                     self.item.get("numero", ""), text_color=NAVY,
                     font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=14, pady=(12, 4))
        tb = ctk.CTkTextbox(top, fg_color=CARD, font=("Segoe UI", 12))
        tb.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        tb.insert("1.0", cuerpo_seguimiento_cotizacion(self.item, self.cfg))

    def _guardar(self):
        self.item["cliente"] = self.v_cli.get().strip()
        self.item["asesor"] = self.v_ase.get().strip()
        self.item["estado"] = self.v_estado.get()
        self.item["fecha_seg"] = self.sel_fecha_seg.get_str()
        self.item["correo_seg"] = self.v_correo_seg.get().strip()
        self.item["email"] = self.v_correo_cli.get().strip() or self.item.get("email", "")
        self.item["auto_correo_seg"] = bool(self.v_auto.get())
        self.item["notas"] = self.txt_notas.get("1.0", "end").strip()
        self.item["tareas"] = [
            {"texto": r["texto"].get().strip(), "fecha": r["fecha"].get().strip(),
             "hecha": r["hecha"].get()}
            for r in self.tareas_rows if r["texto"].get().strip()]
        self.on_save()
        # Propagar la gestion (estado, seguimiento, tareas...) a los demas equipos
        try:
            enviar_cotizacion_nube(self.item)
        except Exception:
            pass
        # ofrecer enviar recordatorio de calendario (.ics) para la fecha de seguimiento
        fseg = parse_fecha(self.item.get("fecha_seg", ""))
        if fseg and self.item.get("estado") not in ("Ganada", "Perdida"):
            app = getattr(self.master, "master", None)
            cfg = getattr(app, "cfg", None) if app else None
            correo_dest = (self.item.get("correo_seg", "").strip()
                           or cfg.get("correo_remitente", "") if cfg else "")
            if cfg and cfg.get("correo_remitente") and cfg.get("smtp_password"):
                if messagebox.askyesno(
                        "Recordatorio de seguimiento",
                        f"¿Enviar un recordatorio de calendario para el "
                        f"{self.item.get('fecha_seg')} a {correo_dest}?", parent=self):
                    try:
                        dest = list(dict.fromkeys(
                            [correo_dest, cfg.get("correo_remitente")]))
                        enviar_recordatorio_ics(cfg, dest, self.item, fseg)
                        messagebox.showinfo("Recordatorio",
                                            "Recordatorio de calendario enviado a:\n"
                                            + ", ".join(d for d in dest if d), parent=self)
                    except Exception as e:
                        messagebox.showwarning("Recordatorio",
                                               "No se pudo enviar el recordatorio:\n" + str(e),
                                               parent=self)
        self.destroy()


class VentanaVerCotizacion(ctk.CTkToplevel):
    """Vista de solo lectura de lo que se cotizo (util para las del HTML)."""
    def __init__(self, master, item):
        super().__init__(master)
        self._master = master; self._item = item
        self.title("Detalle cotizacion " + item.get("numero", ""))
        self.geometry("660x660"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        cont = ctk.CTkScrollableFrame(self, fg_color=BG)
        cont.pack(fill="both", expand=True, padx=16, pady=16)
        ctk.CTkLabel(cont, text=f"{item.get('numero','')}   ·   {item.get('estado','')}",
                     text_color=NAVY, font=("Segoe UI", 17, "bold")).pack(anchor="w", pady=(0, 6))
        snap = item.get("snapshot") or {}

        def linea(lbl, val):
            if not val:
                return
            f = ctk.CTkFrame(cont, fg_color="transparent"); f.pack(fill="x", pady=1)
            ctk.CTkLabel(f, text=lbl + ":", text_color=MUTED, font=("Segoe UI", 11, "bold"),
                         width=150, anchor="w").pack(side="left")
            ctk.CTkLabel(f, text=str(val), text_color=TEXT, font=("Segoe UI", 11), anchor="w",
                         justify="left", wraplength=440).pack(side="left", fill="x", expand=True)

        linea("Agencia / cliente", item.get("cliente", ""))
        linea("Asesor", item.get("asesor", ""))
        linea("Cotizado por", item.get("cotizado_por", ""))
        linea("Correo cliente", item.get("email", ""))
        linea("Fecha cotizacion", item.get("fecha", ""))
        linea("Destinos", ", ".join(item.get("destinos", [])))
        fviaje = item.get("fechas_viaje", "") or (
            f"{snap.get('fecha_desde','')} al {snap.get('fecha_hasta','')}"
            if snap.get("fecha_desde") else "")
        linea("Fechas de viaje", fviaje)
        if snap:
            ad = snap.get("adultos", 0); ninos = len(snap.get("ages", []) or [])
            pax = f"{ad} adulto(s)" + (f", {ninos} nino(s)" if ninos else "")
            linea("Pasajeros", pax)
            linea("Alojamiento", snap.get("hab", ""))
        linea("Total para INNOBA (USD)", usd(item.get("total", 0)))
        # Si la cotizacion vino del HTML con margen de la agencia, mostrarlo como referencia
        gan = float(item.get("ganancia_agencia", 0) or 0)
        tcli = float(item.get("total_cliente", 0) or 0)
        if gan > 0 or tcli > 0:
            if gan > 0:
                linea("Margen de la agencia", f"{gan:g}%")
            if tcli > 0:
                linea("Precio final al cliente", usd(tcli) + "  (con margen de la agencia)")
            ctk.CTkLabel(cont, text="El precio que cobramos a la agencia es el 'Total para INNOBA' "
                         "(SIN el margen que la agencia le suma a su cliente).",
                         text_color=MUTED, font=("Segoe UI", 9), wraplength=580,
                         justify="left").pack(anchor="w", pady=(2, 0))

        tramos = snap.get("tramos", [])
        if tramos:
            ctk.CTkLabel(cont, text="SERVICIOS COTIZADOS", text_color=NAVY,
                         font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(10, 2))
            for i, tr in enumerate(tramos, 1):
                card = ctk.CTkFrame(cont, fg_color=CARD, corner_radius=8,
                                    border_width=1, border_color=LINE)
                card.pack(fill="x", pady=3)
                ctk.CTkLabel(card, text=f"Destino {i}: {tr.get('destino','')}  "
                             f"({tr.get('noches','?')} noches)", text_color=NAVY,
                             font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=8, pady=(6, 2))

                def bloque(tit, lst):
                    if lst:
                        ctk.CTkLabel(card, text=f"   {tit}: " + ", ".join(lst), text_color=TEXT,
                                     font=("Segoe UI", 11), anchor="w", justify="left",
                                     wraplength=580).pack(anchor="w", padx=8, pady=(0, 2))
                bloque("Hoteles", tr.get("hoteles", []))
                bloque("Traslados", tr.get("trans", []))
                bloque("Tours / actividades", tr.get("act", []))
                ctk.CTkFrame(card, fg_color="transparent", height=4).pack()

        itin = snap.get("itinerario", "") or item.get("itinerario", "")
        if itin:
            ctk.CTkLabel(cont, text="ITINERARIO", text_color=NAVY,
                         font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(10, 2))
            tb = ctk.CTkTextbox(cont, height=140, fg_color=CARD, font=("Segoe UI", 11))
            tb.pack(fill="x"); tb.insert("1.0", itin); tb.configure(state="disabled")

        if not snap:
            ctk.CTkLabel(cont, text="Esta cotizacion llego del HTML (o de una version anterior) sin "
                         "el detalle completo. Arriba se muestran los datos que si llegaron. Las "
                         "cotizaciones nuevas del HTML ya traen todo el detalle.",
                         text_color=MUTED, font=("Segoe UI", 10), wraplength=580,
                         justify="left").pack(anchor="w", pady=(10, 0))
        botones = ctk.CTkFrame(cont, fg_color="transparent"); botones.pack(pady=14)
        if snap and hasattr(self._master, "_generar_pdf_de"):
            ctk.CTkButton(botones, text="⬇ Descargar PDF", fg_color=GREEN, hover_color=GREEN_H,
                          font=("Segoe UI", 12, "bold"), width=170,
                          command=self._descargar).pack(side="left", padx=6)
        ctk.CTkButton(botones, text="Cerrar", fg_color=NAVY, hover_color=NAVY2,
                      command=self.destroy).pack(side="left", padx=6)

    def _descargar(self):
        # Genera el PDF con la tarifa de INNOBA (precios base, sin el margen de la agencia)
        m = self._master
        self.destroy()
        m._generar_pdf_de(self._item)


class VentanaCotizaciones(ctk.CTkToplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Historial de cotizaciones")
        self.geometry("1240x660"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        # Mostrar YA lo que hay local (rapido); sincronizar con la nube en 2do plano
        # para NO congelar la ventana mientras hay internet lento.
        self._vencidas_borradas = []
        self.data = cargar_cotizaciones()
        self.grid_columnconfigure(0, weight=1); self.grid_rowconfigure(3, weight=1)
        top = ctk.CTkFrame(self, fg_color="transparent")
        top.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 2))
        top.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(top, text="Historial de cotizaciones", text_color=NAVY,
                     font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w")
        botones_top = ctk.CTkFrame(top, fg_color="transparent"); botones_top.grid(row=0, column=1, sticky="e")
        self.seleccion = set()   # numeros seleccionados para borrado en bloque
        self.btn_borrar_sel = ctk.CTkButton(
            botones_top, text="🗑 Eliminar seleccionadas", width=190, height=32, corner_radius=8,
            fg_color=RED, hover_color="#9B2C22", font=("Segoe UI", 11, "bold"),
            command=self._eliminar_seleccionadas)
        self.btn_borrar_sel.pack(side="left", padx=(0, 6))
        ctk.CTkButton(botones_top, text="📊 Reporte de ventas", width=170, height=32, corner_radius=8,
                      fg_color="#7A5AB5", hover_color="#63459A", font=("Segoe UI", 11, "bold"),
                      command=self._reporte_ventas).pack(side="left", padx=(0, 6))
        if WEBHOOK_URL:
            ctk.CTkButton(botones_top, text="☁ Sincronizar", width=150, height=32, corner_radius=8,
                          fg_color=NAVY, hover_color=BLUE, font=("Segoe UI", 11, "bold"),
                          command=self._importar_html).pack(side="left")
        self.var_q = tk.StringVar()
        e = ctk.CTkEntry(self, textvariable=self.var_q, height=36, corner_radius=10,
                         border_color=BLUE, border_width=2, fg_color=CARD, font=("Segoe UI", 12),
                         placeholder_text="Buscar por consecutivo, agencia o asesor...")
        e.grid(row=1, column=0, sticky="ew", padx=16, pady=(2, 6))
        self._q_after = None
        self.var_q.trace_add("write", lambda *a: self._buscar_debounce())
        self.kpis = ctk.CTkFrame(self, fg_color="transparent")
        self.kpis.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 4))
        self.lista = ctk.CTkScrollableFrame(self, fg_color=CARD, corner_radius=12)
        self.lista.grid(row=3, column=0, sticky="nsew", padx=16, pady=(0, 14))
        self.filtro_seg = None   # None / "seg" / "venc" / "ganada" / "perdida"
        self.after(120, e.focus_set)
        self.bind("<Escape>", lambda ev: self.destroy())
        self._pintar()
        if self._vencidas_borradas:
            nums = ", ".join(x.get("numero", "") for x in self._vencidas_borradas[:12])
            extra = "..." if len(self._vencidas_borradas) > 12 else ""
            self.after(250, lambda: messagebox.showinfo(
                "Cotizaciones vencidas eliminadas",
                f"Se eliminaron {len(self._vencidas_borradas)} cotizacion(es) con mas de 1 mes "
                "desde su elaboracion que no se convirtieron en reserva.\n\n"
                f"({nums}{extra})\n\n"
                "No se tocaron las marcadas 'Ganada' ni 'En seguimiento'.", parent=self))
        # Sincronizar con la nube en 2do plano (no congela la ventana)
        self.after(150, self._sync_bg)

    def _buscar_debounce(self):
        # Espera a que el usuario deje de escribir (evita re-dibujar en cada tecla)
        if self._q_after:
            try:
                self.after_cancel(self._q_after)
            except Exception:
                pass
        self._q_after = self.after(280, self._pintar)

    def _sync_bg(self):
        def worker():
            try:
                importar_cotizaciones_html()
            except Exception:
                pass
            try:
                venc = depurar_cotizaciones_vencidas()
            except Exception:
                venc = []
            try:
                subir_cotizaciones_exe()
            except Exception:
                pass
            try:
                self.after(0, lambda: self._fin_sync(venc))
            except Exception:
                pass
        threading.Thread(target=worker, daemon=True).start()

    def _fin_sync(self, venc):
        try:
            if not self.winfo_exists():
                return
        except Exception:
            return
        self.data = cargar_cotizaciones()
        self._pintar()
        if venc:
            nums = ", ".join(x.get("numero", "") for x in venc[:12])
            extra = "..." if len(venc) > 12 else ""
            messagebox.showinfo(
                "Cotizaciones vencidas eliminadas",
                f"Se eliminaron {len(venc)} cotizacion(es) con mas de 1 mes desde su "
                "elaboracion que no se convirtieron en reserva.\n\n"
                f"({nums}{extra})\n\nNo se tocaron 'Ganada' ni 'En seguimiento'.", parent=self)

    def _toggle_filtro(self, f):
        self.filtro_seg = None if self.filtro_seg == f else f
        self._pintar()

    def _pasa_filtro(self, it):
        f = self.filtro_seg
        if not f:
            return True
        estado = it.get("estado", "Pendiente")
        if f == "seg":
            return estado == "En seguimiento"
        if f == "ganada":
            return estado == "Ganada"
        if f == "perdida":
            return estado == "Perdida"
        if f == "venc":
            fseg = parse_fecha(it.get("fecha_seg", ""))
            return bool(fseg and fseg <= datetime.date.today()
                        and estado not in ("Ganada", "Perdida"))
        return True

    def _pintar_kpis(self):
        for w in self.kpis.winfo_children():
            w.destroy()
        ind = indicadores_cotizaciones()
        fila = ctk.CTkFrame(self.kpis, fg_color="transparent"); fila.pack(fill="x")
        _kpi_card(fila, "Cotizaciones", ind["total"], NAVY,
                  on_click=lambda: self._toggle_filtro(None), active=(self.filtro_seg is None))
        _kpi_card(fila, "En seguimiento", ind.get("En seguimiento", 0), "#D9A400",
                  on_click=lambda: self._toggle_filtro("seg"), active=(self.filtro_seg == "seg"))
        _kpi_card(fila, "Ganadas", ind.get("Ganada", 0), GREEN,
                  on_click=lambda: self._toggle_filtro("ganada"), active=(self.filtro_seg == "ganada"))
        _kpi_card(fila, "Perdidas", ind.get("Perdida", 0), RED,
                  on_click=lambda: self._toggle_filtro("perdida"), active=(self.filtro_seg == "perdida"))
        _kpi_card(fila, "Conversion", f"{ind['conversion']}%", "#7A5AB5")
        _kpi_card(fila, "Ganado (USD)", usd(ind["monto_ganado"]), GREEN_H, ancho=170)
        _kpi_card(fila, "Ticket prom.", usd(ind["ticket"]), NAVY, ancho=150)
        _kpi_card(fila, f"Del mes ({ind['mes_n']})", usd(ind["mes_usd"]), BLUE, ancho=160)
        _kpi_card(fila, "Seguim. vencidos", ind["seg_vencidos"],
                  RED if ind["seg_vencidos"] else MUTED,
                  on_click=lambda: self._toggle_filtro("venc"), active=(self.filtro_seg == "venc"))

    def _pintar(self):
        self._pintar_kpis()
        for w in self.lista.winfo_children():
            w.destroy()
        etiquetas = {"seg": "En seguimiento", "venc": "Seguimientos vencidos",
                     "ganada": "Ganadas", "perdida": "Perdidas"}
        if self.filtro_seg in etiquetas:
            bar = ctk.CTkFrame(self.lista, fg_color="#EAF2FD", corner_radius=8)
            bar.pack(fill="x", padx=4, pady=(2, 4))
            ctk.CTkLabel(bar, text="Filtro activo:  " + etiquetas[self.filtro_seg],
                         text_color=NAVY, font=("Segoe UI", 11, "bold")).pack(side="left", padx=10, pady=6)
            ctk.CTkButton(bar, text="Ver todas ✕", width=100, height=28, fg_color=NAVY,
                          hover_color=NAVY2, command=lambda: self._toggle_filtro(None)).pack(
                side="right", padx=8)
        q = _nz(self.var_q.get())
        items = list(reversed(self.data.get("items", [])))   # mas recientes primero
        # Filtrar primero (rapido) y solo DIBUJAR un maximo (dibujar cientos de filas
        # congela la app). El buscador y los filtros de arriba permiten llegar al resto.
        TOPE = 50
        visibles = [it for it in items
                    if (not q or q in _nz(" ".join([it.get("numero", ""), it.get("cliente", ""),
                                                    it.get("asesor", "")])))
                    and self._pasa_filtro(it)]
        total = len(visibles)
        if total > TOPE:
            av = ctk.CTkLabel(self.lista,
                              text=f"Mostrando {TOPE} de {total}. Usa el buscador o los "
                                   "filtros de arriba para encontrar el resto.",
                              text_color="#B7791F", fg_color="#FFF3C4", corner_radius=8,
                              font=("Segoe UI", 11, "bold"))
            av.pack(fill="x", padx=4, pady=(2, 4), ipady=4)
        lista = visibles[:TOPE]
        if not lista:
            msg = ("Aun no hay cotizaciones guardadas.\nGenera un PDF y aparecera aqui."
                   if not self.data.get("items") else "Sin resultados.")
            ctk.CTkLabel(self.lista, text=msg, text_color=MUTED).pack(pady=24)
            self._actualizar_btn_sel()
            return
        # Dibujar por LOTES para que la ventana no se congele (queda responsiva)
        self._render_gen = getattr(self, "_render_gen", 0) + 1
        gen = self._render_gen

        def _chunk(i):
            if gen != self._render_gen:
                return
            try:
                if not self.winfo_exists():
                    return
            except Exception:
                return
            for it in lista[i:i + 6]:
                self._fila_cot(it)
            if i + 6 < len(lista):
                self.after(1, lambda: _chunk(i + 6))
            else:
                self._actualizar_btn_sel()
        _chunk(0)

    def _fila_cot(self, it):
        estado = it.get("estado", "Pendiente")
        fila = ctk.CTkFrame(self.lista, fg_color=ESTADO_FILA.get(estado, CARD2),
                            corner_radius=8, border_width=1, border_color=LINE)
        fila.pack(fill="x", padx=4, pady=2)
        fila.grid_columnconfigure(0, weight=1)
        dest = ", ".join(it.get("destinos", []))
        tareas = it.get("tareas", []) or []
        pend = sum(1 for t in tareas if not t.get("hecha"))
        fseg = parse_fecha(it.get("fecha_seg", ""))
        venc = fseg and fseg <= datetime.date.today() and estado not in ("Ganada", "Perdida")
        cel = ctk.CTkFrame(fila, fg_color="transparent")
        cel.grid(row=0, column=0, sticky="w", padx=10, pady=6)
        top = ctk.CTkFrame(cel, fg_color="transparent"); top.pack(anchor="w")
        chk = ctk.CTkCheckBox(top, text="", width=24, checkbox_width=20, checkbox_height=20,
                              fg_color=RED, hover_color="#9B2C22")
        chk.pack(side="left", padx=(0, 6))
        if it.get("numero") in self.seleccion:
            chk.select()
        chk.configure(command=lambda num=it.get("numero"), c=chk: self._toggle_sel(num, c))
        ctk.CTkLabel(top, text=f"{it.get('numero','')}    {it.get('cliente') or '(sin agencia)'}",
                     text_color=NAVY, anchor="w",
                     font=("Segoe UI", 12, "bold")).pack(side="left")
        ctk.CTkLabel(top, text=" " + estado + " ", text_color="#FFFFFF",
                     fg_color=ESTADO_COLOR.get(estado, MUTED), corner_radius=6,
                     font=("Segoe UI", 9, "bold")).pack(side="left", padx=8)
        if it.get("fecha_seg"):
            ctk.CTkLabel(top, text=("🔔 seguimiento " + it.get("fecha_seg", "")),
                         text_color=(RED if venc else MUTED),
                         font=("Segoe UI", 9, "bold")).pack(side="left", padx=6)
        if pend:
            ctk.CTkLabel(top, text=f"{pend} tarea(s) pend.", text_color=RED,
                         font=("Segoe UI", 9, "bold")).pack(side="left", padx=4)
        linea2 = (f"Asesor: {it.get('asesor') or '-'}     Fecha: {it.get('fecha','')}"
                  f"     {dest}     {usd(it.get('total', 0))}")
        ctk.CTkLabel(cel, text=linea2, text_color=MUTED, anchor="w",
                     font=("Segoe UI", 10)).pack(anchor="w")
        col = 1
        ctk.CTkButton(fila, text="👁 Ver", width=76, height=30, corner_radius=8,
                      fg_color=CYAN, hover_color=BLUE, font=("Segoe UI", 11, "bold"),
                      command=lambda x=it: self._ver(x)).grid(row=0, column=col, padx=4)
        col += 1
        ctk.CTkButton(fila, text="✎ Editar", width=84, height=30,
                      corner_radius=8, fg_color=GREEN, hover_color=GREEN_H,
                      font=("Segoe UI", 11, "bold"),
                      command=lambda x=it: self._editar_cotizacion(x)).grid(
            row=0, column=col, padx=4); col += 1
        ctk.CTkButton(fila, text="Seguimiento", width=100, height=30, corner_radius=8,
                      fg_color=CYAN, hover_color=BLUE, font=("Segoe UI", 11, "bold"),
                      command=lambda x=it: self._detalle(x)).grid(row=0, column=col, padx=4)
        col += 1
        _en_reserva = bool(it.get("reserva_creada"))
        ctk.CTkButton(fila,
                      text=(f"✓ Reserva {it.get('reserva_creada')}" if _en_reserva else "➜ Reserva"),
                      width=(120 if _en_reserva else 98), height=30, corner_radius=8,
                      fg_color=(GREEN if _en_reserva else NAVY),
                      hover_color=(GREEN_H if _en_reserva else NAVY2),
                      font=("Segoe UI", 11, "bold"),
                      command=lambda x=it: self._enviar_a_reserva(x)).grid(row=0, column=col, padx=4)
        col += 1
        if it.get("pdf"):
            ctk.CTkButton(fila, text="Abrir PDF", width=90, height=30, corner_radius=8,
                          fg_color=NAVY, hover_color=BLUE, font=("Segoe UI", 11, "bold"),
                          command=lambda p=it["pdf"]: self._abrir(p)).grid(
                row=0, column=col, padx=4); col += 1
        elif it.get("snapshot"):
            ctk.CTkButton(fila, text="Generar PDF", width=100, height=30, corner_radius=8,
                          fg_color=NAVY, hover_color=BLUE, font=("Segoe UI", 11, "bold"),
                          command=lambda x=it: self._generar_pdf_de(x)).grid(
                row=0, column=col, padx=4); col += 1
        ctk.CTkButton(fila, text="🗑", width=34, height=30, corner_radius=8,
                      fg_color=CARD2, text_color=RED, hover_color=LINE,
                      font=("Segoe UI", 13),
                      command=lambda x=it: self._eliminar(x)).grid(row=0, column=col, padx=(0, 6))

    def _reporte_ventas(self):
        data = cargar_cotizaciones()
        meses = set()
        for it in data.get("items", []):
            if it.get("estado") == "Ganada":
                f = parse_fecha(it.get("fecha", ""))
                if f:
                    meses.add(f.strftime("%Y-%m"))
        DialogoReporteMes(
            self, "Reporte de ventas cerradas",
            "Ventas cerradas (cotizaciones Ganadas) por mes, separadas por Felipe y Carlos.",
            sorted(meses, reverse=True), exportar_reporte_ventas, "Reporte_ventas")

    def _importar_html(self):
        try:
            subidas = subir_cotizaciones_exe()
            n = importar_cotizaciones_html()
        except Exception as e:
            messagebox.showerror("Sincronizar", str(e), parent=self); return
        self.data = cargar_cotizaciones(); self._pintar()
        messagebox.showinfo(
            "Sincronizacion de cotizaciones",
            f"Cotizaciones subidas a la nube: {subidas}\n"
            f"Cotizaciones traidas (HTML + otros equipos): {n}\n\n"
            "Ahora se ven las cotizaciones de Felipe, de Carlos y las del HTML.",
            parent=self)

    def _ver(self, it):
        VentanaVerCotizacion(self, it)

    def _enviar_a_reserva(self, it):
        cfg = getattr(self.master, "cfg", None) or cargar_config()
        if not asesores_reservas(cfg):
            messagebox.showinfo(
                "Configura los asesores",
                "Primero configura los asesores de reservas en el modulo Reservas "
                "(boton 'Asesores') para poder asignar la reserva.", parent=self)
            return
        # Evitar duplicar: revisar si esta cotizacion YA tiene reserva (aca o en otro equipo)
        cot_key = str(it.get("uid") or it.get("web_id") or it.get("numero", ""))
        try:
            importar_reservas_nube()
        except Exception:
            pass
        ya = None
        for r in cargar_reservas().get("items", []):
            if str(r.get("cot_uid", "")) == cot_key and cot_key:
                ya = r
                break
        if ya is None and it.get("reserva_creada"):
            ya = {"numero": it.get("reserva_creada")}
        if ya is not None:
            if not messagebox.askyesno(
                    "Esta cotizacion ya tiene reserva",
                    f"La cotizacion {it.get('numero','')} ya fue enviada a reserva "
                    f"(Reserva N. {ya.get('numero','')}).\n\n"
                    "Si creas otra, quedara DUPLICADA. ¿Seguro que quieres crear otra reserva?",
                    icon="warning", default="no", parent=self):
                return
        elif not messagebox.askyesno(
                "Enviar a reserva",
                f"Crear una reserva a partir de {it.get('numero','')} "
                f"({it.get('cliente','')})?\n\nSe asignara automaticamente a una asesora "
                "de reservas (rotacion). Podras verla y gestionarla en el modulo Reservas.",
                parent=self):
            return
        try:
            rec = reserva_desde_cotizacion(it)
            numero, guardado = registrar_reserva(rec, cfg)
        except Exception as e:
            messagebox.showerror("No se pudo crear la reserva", str(e), parent=self)
            return
        # Marcar la cotizacion como 'ya enviada a reserva' (se guarda y se sincroniza)
        try:
            it["reserva_creada"] = numero
            it["nube_ok"] = False   # forzar re-subida con la marca
            guardar_cotizaciones(self.data)
            if it.get("origen", "") != "HTML (cliente)":
                enviar_cotizacion_nube(it)
            self._pintar()
        except Exception:
            pass
        ase = (guardado.get("asesor", {}) or {})
        messagebox.showinfo(
            "Reserva creada",
            f"Reserva N. {numero} creada desde {it.get('numero','')}.\n"
            f"Asignada a: {ase.get('nombre', '(sin asignar)')}\n\n"
            "Abrela desde el modulo Reservas para gestionarla."
            + _texto_notif_reserva(guardado), parent=self)

    def _editar_cotizacion(self, it):
        app = self.master
        if not hasattr(app, "_cargar_cotizacion"):
            return
        if not it.get("snapshot"):
            messagebox.showinfo(
                "Editar cotizacion",
                f"La cotizacion {it.get('numero','')} no tiene datos guardados para editar "
                "(se creo en una version anterior o llego desde el HTML sin detalle). "
                "Puedes crear una nueva o enviarla a reserva.", parent=self)
            return
        if not messagebox.askyesno(
                "Editar cotizacion",
                f"Se cargara {it.get('numero','')} ({it.get('cliente','')}) en el cotizador "
                "para editarla.\n\nSe reemplazara lo que tengas ahora en pantalla. "
                "Al generar de nuevo se creara una cotizacion nueva.\n\n¿Continuar?",
                parent=self):
            return
        app._cargar_cotizacion(it.get("snapshot"))
        self.destroy()

    def _generar_pdf_de(self, it):
        app = self.master
        if not it.get("snapshot") or not hasattr(app, "_cargar_cotizacion"):
            return
        if not messagebox.askyesno(
                "Generar PDF",
                f"Se cargara {it.get('numero','')} ({it.get('cliente','')}) y se generara "
                "el PDF.\n\n(Puedes revisarla/ajustarla en el cotizador antes de generar "
                "si lo prefieres.)\n\n¿Continuar?", parent=self):
            return
        app._cargar_cotizacion(it.get("snapshot"))
        self.destroy()
        app.after(250, app._generar)

    def _detalle(self, it):
        VentanaCotizacionDetalle(self, it, self._on_guardado)

    def _on_guardado(self):
        guardar_cotizaciones(self.data)
        self._pintar()

    def _abrir(self, p):
        if p and os.path.exists(p):
            try:
                os.startfile(p)
            except Exception as e:
                messagebox.showerror("Error", str(e))
        else:
            messagebox.showinfo("PDF no encontrado",
                                "El archivo PDF ya no esta en:\n" + str(p))

    def _eliminar(self, it):
        if messagebox.askyesno("Quitar del historial",
                               f"¿Quitar {it.get('numero','')} del historial?\n"
                               "Se quitara tambien para los demas equipos (no borra el PDF).",
                               parent=self):
            self.data["items"] = [x for x in self.data["items"] if x is not it]
            guardar_cotizaciones(self.data)
            try:
                enviar_borrado_nube("cotizacion", it.get("uid") or it.get("web_id"))
            except Exception:
                pass
            self.seleccion.discard(it.get("numero"))
            self._pintar()

    def _toggle_sel(self, numero, chk):
        if chk.get():
            self.seleccion.add(numero)
        else:
            self.seleccion.discard(numero)
        self._actualizar_btn_sel()

    def _actualizar_btn_sel(self):
        n = len(self.seleccion)
        try:
            self.btn_borrar_sel.configure(
                text=(f"🗑 Eliminar seleccionadas ({n})" if n else "🗑 Eliminar seleccionadas"),
                state=("normal" if n else "disabled"))
        except Exception:
            pass

    def _eliminar_seleccionadas(self):
        nums = set(self.seleccion)
        if not nums:
            messagebox.showinfo("Eliminar", "No has seleccionado cotizaciones.\n"
                                "Marca las casillas de las que quieres eliminar.", parent=self)
            return
        if not messagebox.askyesno(
                "Eliminar seleccionadas",
                f"¿Eliminar {len(nums)} cotizacion(es) seleccionada(s)?\n\n"
                "Se quitaran del historial y de los demas equipos. Esta accion NO se puede "
                "deshacer.", icon="warning", parent=self):
            return
        borrar = [x for x in self.data["items"] if x.get("numero") in nums]
        claves = [(x.get("uid") or x.get("web_id")) for x in borrar]
        self.data["items"] = [x for x in self.data["items"] if x.get("numero") not in nums]
        guardar_cotizaciones(self.data)
        for k in claves:
            try:
                enviar_borrado_nube("cotizacion", k)
            except Exception:
                pass
        self.seleccion.clear()
        self._pintar()
        messagebox.showinfo("Listo", f"Se eliminaron {len(borrar)} cotizacion(es).", parent=self)


# ============================================================================
# Ventana de Clientes / Empresas (con vendedores)
# ============================================================================
class VentanaClientes(ctk.CTkToplevel):
    def __init__(self, master, on_cambio=None, preseleccion=None):
        super().__init__(master)
        self.on_cambio = on_cambio
        self.title("Clientes / Empresas")
        self.geometry("980x640"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        self.clientes = cargar_clientes()
        self.sel = None            # indice seleccionado (None = nuevo)
        self.vend_rows = []        # filas de vendedores (dicts de StringVars)

        self.geometry("1080x680+60+10")
        self.after(60, lambda: self._maximizar())
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)
        # ---- barra superior (navy) ----
        top = ctk.CTkFrame(self, fg_color=NAVY, corner_radius=0, height=62)
        top.grid(row=0, column=0, columnspan=2, sticky="ew"); top.grid_propagate(False)
        ctk.CTkLabel(top, text="👥   Clientes / Empresas", text_color="#FFFFFF",
                     font=("Segoe UI", 18, "bold")).pack(side="left", padx=22, pady=12)
        ctk.CTkButton(top, text="Importar Excel/CSV", width=160, height=36, corner_radius=10,
                      fg_color=NAVY2, hover_color=BLUE, font=("Segoe UI", 12, "bold"),
                      command=self._importar).pack(side="right", padx=(4, 20))
        ctk.CTkButton(top, text="Exportar Excel", width=140, height=36, corner_radius=10,
                      fg_color=GREEN, hover_color=GREEN_H, font=("Segoe UI", 12, "bold"),
                      command=self._exportar).pack(side="right", padx=4)
        ctk.CTkButton(top, text="+ Nueva empresa", width=150, height=36, corner_radius=10,
                      fg_color="#FFFFFF", text_color=NAVY, hover_color="#E7EEF8",
                      font=("Segoe UI", 12, "bold"),
                      command=self._nuevo).pack(side="right", padx=4)
        # ---- lista izquierda ----
        izq = ctk.CTkFrame(self, fg_color=CARD, corner_radius=16); izq.grid(
            row=1, column=0, sticky="nsew", padx=(14, 7), pady=14)
        izq.grid_rowconfigure(2, weight=1); izq.grid_columnconfigure(0, weight=1)
        cab = ctk.CTkFrame(izq, fg_color="transparent"); cab.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 2))
        ctk.CTkLabel(cab, text="Empresas", text_color=NAVY,
                     font=("Segoe UI", 14, "bold")).pack(side="left")
        self.lbl_count = ctk.CTkLabel(cab, text="", text_color=MUTED, font=("Segoe UI", 11))
        self.lbl_count.pack(side="right")
        self.var_busca = tk.StringVar()
        eb = ctk.CTkEntry(izq, textvariable=self.var_busca, height=38, corner_radius=10,
                          border_color=BLUE, border_width=2, fg_color=CARD2,
                          placeholder_text="🔍  Buscar empresa...", width=280)
        eb.grid(row=1, column=0, sticky="ew", padx=12, pady=(4, 8))
        self.var_busca.trace_add("write", lambda *a: self._rebuild_list())
        self.lista = ctk.CTkScrollableFrame(izq, fg_color=BG, corner_radius=12, width=280)
        self.lista.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
        # ---- form derecha ----
        self.form = ctk.CTkScrollableFrame(self, fg_color=CARD, corner_radius=16)
        self.form.grid(row=1, column=1, sticky="nsew", padx=(7, 14), pady=14)
        self.form.grid_columnconfigure(0, weight=1)
        self._build_form()
        self._rebuild_list()
        self._nuevo()
        if preseleccion:
            idx = next((i for i, c in enumerate(self.clientes)
                        if c.get("empresa", "") == preseleccion), None)
            if idx is not None:
                self.var_busca.set(preseleccion)
                self._cargar(idx)

    def _maximizar(self):
        try:
            self.state("zoomed")
        except Exception:
            pass

    def _build_form(self):
        self.vars = {}

        def campo(parent, clave, etq):
            ctk.CTkLabel(parent, text=etq, text_color=MUTED,
                         font=("Segoe UI", 11)).pack(anchor="w", padx=2, pady=(6, 0))
            v = tk.StringVar(); self.vars[clave] = v
            ctk.CTkEntry(parent, textvariable=v, height=34, corner_radius=8,
                         border_color=LINE).pack(fill="x", padx=2, pady=(0, 2))
            return v

        def pareja(parent, k1, l1, k2, l2):
            f = ctk.CTkFrame(parent, fg_color="transparent"); f.pack(fill="x")
            a = ctk.CTkFrame(f, fg_color="transparent"); a.pack(side="left", fill="x", expand=True, padx=(0, 6))
            b = ctk.CTkFrame(f, fg_color="transparent"); b.pack(side="left", fill="x", expand=True, padx=(6, 0))
            campo(a, k1, l1); campo(b, k2, l2)

        # ---- Tarjeta: datos de la empresa ----
        card = ctk.CTkFrame(self.form, fg_color=CARD2, corner_radius=14,
                            border_width=1, border_color=LINE)
        card.pack(fill="x", padx=12, pady=(10, 8))
        inn = ctk.CTkFrame(card, fg_color="transparent"); inn.pack(fill="x", padx=14, pady=12)
        ctk.CTkLabel(inn, text="🏢  Datos de la empresa", text_color=NAVY,
                     font=("Segoe UI", 15, "bold")).pack(anchor="w", pady=(0, 4))
        campo(inn, "empresa", "Nombre de la empresa *")
        pareja(inn, "nit", "NIT / Documento fiscal", "telefono", "Telefono")
        pareja(inn, "email", "Email", "web", "Sitio web")
        campo(inn, "pais", "Pais")

        # ---- Tarjeta: vendedores / contactos ----
        card2 = ctk.CTkFrame(self.form, fg_color=CARD2, corner_radius=14,
                             border_width=1, border_color=LINE)
        card2.pack(fill="x", padx=12, pady=(0, 8))
        in2 = ctk.CTkFrame(card2, fg_color="transparent"); in2.pack(fill="x", padx=14, pady=12)
        hdr = ctk.CTkFrame(in2, fg_color="transparent"); hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="👤  Vendedores / contactos", text_color=NAVY,
                     font=("Segoe UI", 15, "bold")).pack(side="left")
        ctk.CTkButton(hdr, text="+ Agregar vendedor", width=150, height=30, corner_radius=8,
                      fg_color=BLUE, hover_color=BLUE_H, font=("Segoe UI", 11, "bold"),
                      command=lambda: self._add_vend()).pack(side="right")
        self.vends_frame = ctk.CTkFrame(in2, fg_color="transparent")
        self.vends_frame.pack(fill="x", pady=(8, 2))

        # ---- Botones (barra) ----
        bts = ctk.CTkFrame(self.form, fg_color="transparent"); bts.pack(fill="x", padx=12, pady=14)
        ctk.CTkButton(bts, text="💾  Guardar", height=42, corner_radius=10, fg_color=GREEN,
                      hover_color=GREEN_H, font=("Segoe UI", 13, "bold"),
                      command=self._guardar).pack(side="left")
        ctk.CTkButton(bts, text="Eliminar", height=42, width=110, corner_radius=10,
                      fg_color="#FBE6E6", text_color=RED, hover_color="#F5CFCF",
                      font=("Segoe UI", 12, "bold"), command=self._eliminar).pack(side="left", padx=8)
        self.lbl_estado = ctk.CTkLabel(bts, text="", text_color=GREEN_H,
                                       font=("Segoe UI", 12, "bold")); self.lbl_estado.pack(side="left", padx=10)

    def _add_vend(self, v=None):
        row = ctk.CTkFrame(self.vends_frame, fg_color=CARD, corner_radius=8,
                           border_width=1, border_color=LINE)
        row.pack(fill="x", padx=2, pady=3)
        vv = {k: tk.StringVar(value=(v or {}).get(k, "")) for k in ("nombre", "telefono", "email", "cargo")}
        ctk.CTkEntry(row, textvariable=vv["nombre"], height=30, corner_radius=6,
                     border_color=LINE, placeholder_text="Nombre").pack(
            side="left", fill="x", expand=True, padx=(6, 3), pady=5)
        ctk.CTkEntry(row, textvariable=vv["telefono"], height=30, width=120, corner_radius=6,
                     border_color=LINE, placeholder_text="Telefono").pack(side="left", padx=3, pady=5)
        ctk.CTkEntry(row, textvariable=vv["email"], height=30, width=180, corner_radius=6,
                     border_color=LINE, placeholder_text="Email").pack(side="left", padx=3, pady=5)
        ctk.CTkEntry(row, textvariable=vv["cargo"], height=30, width=110, corner_radius=6,
                     border_color=LINE, placeholder_text="Cargo").pack(side="left", padx=3, pady=5)
        ctk.CTkButton(row, text="✕", width=28, height=28, corner_radius=6, fg_color=RED,
                      text_color="#FFFFFF", hover_color="#9B2C22",
                      command=lambda: (self.vend_rows.remove(entry), row.destroy())).pack(
            side="left", padx=(3, 6))
        entry = {"frame": row, "vars": vv}
        self.vend_rows.append(entry)

    def _clear_vends(self):
        for e in self.vend_rows:
            e["frame"].destroy()
        self.vend_rows = []

    def _rebuild_list(self):
        for w in self.lista.winfo_children():
            w.destroy()
        q = self.var_busca.get().lower()
        mostrados = 0
        for i, c in enumerate(self.clientes):
            if q and q not in c.get("empresa", "").lower():
                continue
            mostrados += 1
            self._item_lista(i, c, i == self.sel)
        try:
            tot = len(self.clientes)
            self.lbl_count.configure(text=(f"{mostrados}/{tot}" if q else f"{tot} empresas"))
        except Exception:
            pass
        if not self.clientes:
            ctk.CTkLabel(self.lista, text="Sin clientes.\nCrea uno o importa un Excel.",
                         text_color=MUTED).pack(pady=20)
        elif mostrados == 0:
            ctk.CTkLabel(self.lista, text="Sin resultados.", text_color=MUTED).pack(pady=16)

    def _item_lista(self, i, c, act):
        item = ctk.CTkFrame(self.lista, fg_color=(NAVY if act else CARD), corner_radius=10,
                            border_width=(0 if act else 1), border_color=LINE)
        item.pack(fill="x", padx=4, pady=3)
        nom = c.get("empresa", "(sin nombre)")
        vends = c.get("vendedores", []) or []
        partes = []
        if c.get("pais"):
            partes.append(c["pais"])
        partes.append(f"{len(vends)} contacto" + ("s" if len(vends) != 1 else ""))
        tc = "#FFFFFF" if act else NAVY
        sc = "#C7D7EE" if act else MUTED
        l1 = ctk.CTkLabel(item, text=nom, anchor="w", text_color=tc, justify="left",
                          font=("Segoe UI", 12, "bold"))
        l1.pack(fill="x", padx=12, pady=(7, 0))
        l2 = ctk.CTkLabel(item, text="   ·   ".join(partes), anchor="w", text_color=sc,
                          font=("Segoe UI", 10))
        l2.pack(fill="x", padx=12, pady=(0, 7))
        for w in (item, l1, l2):
            w.configure(cursor="hand2")
            w.bind("<Button-1>", lambda e, x=i: self._cargar(x))
            if not act:
                w.bind("<Enter>", lambda e, it=item: it.configure(fg_color=CARD2))
                w.bind("<Leave>", lambda e, it=item: it.configure(fg_color=CARD))

    def _cargar(self, idx):
        self.sel = idx
        c = self.clientes[idx]
        for k, v in self.vars.items():
            v.set(c.get(k, ""))
        self._clear_vends()
        for vend in c.get("vendedores", []):
            self._add_vend(vend)
        self.lbl_estado.configure(text="")
        self._rebuild_list()

    def _nuevo(self):
        self.sel = None
        for v in self.vars.values():
            v.set("")
        self._clear_vends(); self._add_vend()
        self.lbl_estado.configure(text="Nueva empresa")
        self._rebuild_list()

    def _recoger(self):
        c = {k: v.get().strip() for k, v in self.vars.items()}
        c["vendedores"] = []
        for e in self.vend_rows:
            d = {k: e["vars"][k].get().strip() for k in ("nombre", "telefono", "email", "cargo")}
            if d["nombre"]:
                c["vendedores"].append(d)
        return c

    def _guardar(self):
        c = self._recoger()
        if not c["empresa"]:
            messagebox.showwarning("Falta el nombre", "Escribe el nombre de la empresa."); return
        if self.sel is None:
            self.clientes.append(c); self.sel = len(self.clientes) - 1
        else:
            self.clientes[self.sel] = c
        guardar_clientes(self.clientes)
        self.lbl_estado.configure(text="Guardado ✓")
        self._rebuild_list()
        if self.on_cambio:
            self.on_cambio()

    def _eliminar(self):
        if self.sel is None:
            return
        if messagebox.askyesno("Eliminar", "¿Eliminar esta empresa?"):
            del self.clientes[self.sel]
            guardar_clientes(self.clientes)
            self._nuevo()
            if self.on_cambio:
                self.on_cambio()

    def _importar(self):
        ruta = filedialog.askopenfilename(title="Importar clientes de Excel/CSV",
                                          filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv *.txt"),
                                                     ("Excel", "*.xlsx *.xlsm"),
                                                     ("CSV", "*.csv *.txt")])
        if not ruta:
            return
        try:
            nuevos = importar_clientes_excel(ruta)
        except Exception as e:
            messagebox.showerror("Error al importar", str(e)); return
        if not nuevos:
            messagebox.showinfo("Importar", "No se encontraron empresas en el archivo."); return
        # combinar por nombre (actualiza / agrega)
        idx = {c["empresa"].lower(): i for i, c in enumerate(self.clientes)}
        add = upd = 0
        for c in nuevos:
            k = c["empresa"].lower()
            if k in idx:
                self.clientes[idx[k]] = c; upd += 1
            else:
                self.clientes.append(c); idx[k] = len(self.clientes) - 1; add += 1
        guardar_clientes(self.clientes)
        self._nuevo()
        if self.on_cambio:
            self.on_cambio()
        messagebox.showinfo("Importado", f"Importados: {add} nuevos, {upd} actualizados.")

    def _exportar(self):
        ruta = filedialog.asksaveasfilename(title="Exportar clientes a Excel",
                                            defaultextension=".xlsx", initialfile="Clientes_Innoba.xlsx",
                                            filetypes=[("Excel", "*.xlsx")])
        if not ruta:
            return
        try:
            exportar_clientes_excel(self.clientes, ruta)
        except Exception as e:
            messagebox.showerror("Error al exportar", str(e)); return
        if messagebox.askyesno("Exportado", "Excel guardado:\n" + ruta + "\n\n¿Abrirlo?"):
            try: os.startfile(ruta)
            except Exception: pass


# ============================================================================
# Calendario (selector de fecha)
# ============================================================================
MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
         "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
DIAS_SEM = ["Lu", "Ma", "Mi", "Ju", "Vi", "Sa", "Do"]


class CalendarioPopup(ctk.CTkToplevel):
    def __init__(self, master, fecha_ini, minimo, on_pick):
        super().__init__(master)
        self.on_pick = on_pick
        self.minimo = minimo
        self.title("Elegir fecha")
        self.configure(fg_color=CARD)
        self.resizable(False, False)
        self.transient(master)
        base = fecha_ini or minimo or datetime.date.today()
        self.y, self.m = base.year, base.month
        self.cont = ctk.CTkFrame(self, fg_color=CARD)
        self.cont.pack(padx=10, pady=10)
        self._build()
        self.after(60, self._centrar)
        self.grab_set()

    def _centrar(self):
        try:
            self.update_idletasks()
            w = self.winfo_width(); h = self.winfo_height()
            sw = self.winfo_screenwidth(); sh = self.winfo_screenheight()
            mx = self.master.winfo_pointerx(); my = self.master.winfo_pointery()
            x = min(max(mx - 120, 10), sw - w - 10)
            y = min(max(my + 12, 10), sh - h - 10)
            self.geometry(f"+{int(x)}+{int(y)}")
        except Exception:
            pass

    def _build(self):
        for w in self.cont.winfo_children():
            w.destroy()
        top = ctk.CTkFrame(self.cont, fg_color=NAVY, corner_radius=8)
        top.pack(fill="x", pady=(0, 8))
        ctk.CTkButton(top, text="◀", width=34, height=30, fg_color=NAVY2,
                      hover_color=BLUE, command=self._prev).pack(side="left", padx=4, pady=4)
        ctk.CTkLabel(top, text=f"{MESES[self.m-1]} {self.y}", text_color="#FFFFFF",
                     font=("Segoe UI", 13, "bold")).pack(side="left", expand=True)
        ctk.CTkButton(top, text="▶", width=34, height=30, fg_color=NAVY2,
                      hover_color=BLUE, command=self._next).pack(side="right", padx=4, pady=4)
        grid = ctk.CTkFrame(self.cont, fg_color=CARD)
        grid.pack()
        for i, d in enumerate(DIAS_SEM):
            ctk.CTkLabel(grid, text=d, text_color=MUTED, width=36,
                         font=("Segoe UI", 10, "bold")).grid(row=0, column=i, padx=1, pady=2)
        cal = calendar.Calendar(firstweekday=0)   # lunes
        for r, semana in enumerate(cal.monthdatescalendar(self.y, self.m), start=1):
            for c, dia in enumerate(semana):
                del_mes = (dia.month == self.m)
                deshab = (self.minimo and dia < self.minimo)
                if del_mes and not deshab:
                    fg, tc = CARD2, NAVY
                else:
                    fg, tc = CARD, MUTED
                b = ctk.CTkButton(grid, text=str(dia.day), width=36, height=30,
                                  corner_radius=6, fg_color=fg, text_color=tc,
                                  hover_color=BLUE, font=("Segoe UI", 11),
                                  command=lambda dd=dia: self._pick(dd))
                if deshab:
                    b.configure(state="disabled")
                b.grid(row=r, column=c, padx=1, pady=1)

    def _prev(self):
        self.m -= 1
        if self.m < 1:
            self.m = 12; self.y -= 1
        self._build()

    def _next(self):
        self.m += 1
        if self.m > 12:
            self.m = 1; self.y += 1
        self._build()

    def _pick(self, dia):
        self.on_pick(dia)
        self.destroy()


class SelectorFecha(ctk.CTkFrame):
    def __init__(self, master, command=None, minimo=None, **kw):
        super().__init__(master, fg_color="transparent", **kw)
        self.command = command
        self.minimo = minimo
        self._fecha = None
        self.btn = ctk.CTkButton(self, text="📅  Elegir...", height=34, corner_radius=8,
                                 fg_color=CARD2, text_color=NAVY, hover_color=LINE,
                                 border_width=1, border_color=LINE, font=("Segoe UI", 12),
                                 anchor="w", command=self._abrir)
        self.btn.pack(fill="x")

    def _abrir(self):
        CalendarioPopup(self.winfo_toplevel(), self._fecha, self.minimo, self._set)

    def _set(self, dia):
        self._fecha = dia
        self.btn.configure(text="📅  " + dia.strftime("%d/%m/%Y"),
                           fg_color="#E7F0FB", text_color=NAVY)
        if self.command:
            self.command()

    def get(self):
        return self._fecha

    def get_str(self):
        return self._fecha.strftime("%d/%m/%Y") if self._fecha else ""

    def clear(self):
        self._fecha = None
        self.btn.configure(text="📅  Elegir...", fg_color=CARD2, text_color=NAVY)


# ============================================================================
# Aplicacion
# ============================================================================
class App(ctk.CTkToplevel):
    def __init__(self, master=None):
        super().__init__(master)
        ctk.set_appearance_mode("light")
        # Escala mas compacta: letra y controles ~15% mas pequenos -> se ven mas
        # opciones de hotel/terrestres y todo cabe mejor en pantalla.
        try:
            ctk.set_widget_scaling(0.85)
        except Exception:
            pass
        self.cfg = cargar_config()
        respaldar_datos(self.cfg)   # respaldo automatico de cotizaciones/clientes
        if not self.cfg.get("logo") or not os.path.exists(self.cfg["logo"]):
            lg = recurso("logo_innoba.png")
            if os.path.exists(lg):
                self.cfg["logo"] = lg
        try:
            self.precios = cargar_precios()
        except Exception as e:
            messagebox.showerror("Error", str(e)); self.destroy(); return
        self.descripciones = cargar_descripciones()
        self.clientes = cargar_clientes()

        self.title(f"Cotizador INNOBA Colombia DMC   v{VERSION}")
        self.geometry("1180x820"); self.minsize(1040, 620)
        self.configure(fg_color=BG)
        try:
            self.iconbitmap(recurso("app.ico"))
        except Exception:
            pass

        self.tramos = []          # itinerario: lista de destinos
        self.activo = None        # indice del destino activo
        self._cargando = False    # evita feedback al cargar widgets
        self.tab = "hotel"        # pestana activa
        self.q = {"hotel": "", "trans": "", "act": ""}   # textos de busqueda
        self._itinerario = ""     # itinerario dia por dia (editable)

        # TRM CRUDA (sin descuento). El descuento se aplica segun la fecha de viaje.
        self._trm = _trm_valida(self.cfg.get("ultima_trm", ""))
        self.var_trm_status = tk.StringVar(value="Consultando tasa...")

        self._construir()
        self._nueva()
        # abrir maximizada para que TODO (incluido el pie con Generar PDF) sea visible
        try:
            self.after(60, lambda: self.state("zoomed"))
        except Exception:
            pass
        self.after(250, lambda: self._actualizar_trm(silencioso=True))
        self.after(1800, self._chequear_actualizacion)   # busca nueva version en el repo
        self.after(1200, self._chequear_seguimientos)    # alerta de cotizaciones a seguir

    # ------------------------------------------------------------------ UI
    def _construir(self):
        self.grid_columnconfigure(0, weight=1)
        # el panel de seleccion tiene una altura minima decente (muestra varias
        # opciones); como la app abre maximizada, el pie tambien cabe.
        self.grid_rowconfigure(5, weight=1, minsize=240)

        # Encabezado
        head = ctk.CTkFrame(self, fg_color=CARD, corner_radius=0, height=60)
        head.grid(row=0, column=0, sticky="ew"); head.grid_propagate(False)
        head.grid_columnconfigure(1, weight=1)
        try:
            img = Image.open(recurso("logo_innoba.png")); w, h = img.size; hh = 40
            self.logo_img = ctk.CTkImage(light_image=img, size=(int(w * hh / h), hh))
            ctk.CTkLabel(head, image=self.logo_img, text="").grid(
                row=0, column=0, padx=(20, 14), pady=8)
        except Exception:
            ctk.CTkLabel(head, text="INNOBA", font=("Segoe UI", 22, "bold"),
                         text_color=NAVY).grid(row=0, column=0, padx=20)
        tit = ctk.CTkFrame(head, fg_color="transparent"); tit.grid(row=0, column=1, sticky="w")
        ctk.CTkLabel(tit, text="Cotizador de Paquetes", text_color=NAVY,
                     font=("Segoe UI", 17, "bold"), height=20).pack(anchor="w")
        ctk.CTkLabel(tit, text=f"INNOBA Colombia DMC  ·  v{VERSION}  ·  Itinerario hasta 5 destinos",
                     text_color=MUTED, font=("Segoe UI", 11), height=15).pack(anchor="w")
        hbtns = ctk.CTkFrame(head, fg_color="transparent"); hbtns.grid(row=0, column=2, padx=20)
        ctk.CTkButton(hbtns, text="⌂ Modulos", width=100, height=36, corner_radius=10,
                      fg_color=NAVY, hover_color=NAVY2, font=("Segoe UI", 12, "bold"),
                      command=self._volver_inicio).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hbtns, text="Cotizaciones", width=120, height=36, corner_radius=10,
                      fg_color=GREEN, hover_color=GREEN_H, font=("Segoe UI", 12, "bold"),
                      command=self._abrir_cotizaciones).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hbtns, text="Itinerario", width=110, height=36, corner_radius=10,
                      fg_color=CYAN, hover_color=BLUE, font=("Segoe UI", 12, "bold"),
                      command=self._abrir_itinerario).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hbtns, text="Clientes", width=110, height=36, corner_radius=10,
                      fg_color=NAVY, hover_color=BLUE, font=("Segoe UI", 12, "bold"),
                      command=self._abrir_clientes).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hbtns, text="Datos de mi empresa", width=170, height=36,
                      corner_radius=10, fg_color=CARD2, text_color=NAVY, hover_color=LINE,
                      border_width=1, border_color=LINE, font=("Segoe UI", 12, "bold"),
                      command=self._abrir_empresa).pack(side="left")

        # Barra de estado de la tasa (SIN mostrar el valor)
        trm = ctk.CTkFrame(self, fg_color=NAVY, corner_radius=0, height=34)
        trm.grid(row=1, column=0, sticky="ew"); trm.grid_propagate(False)
        ins = ctk.CTkFrame(trm, fg_color="transparent"); ins.pack(fill="both", expand=True,
                                                                  padx=20, pady=4)
        ctk.CTkLabel(ins, text="Tasa del dia (USD):", text_color="#BFD4F0",
                     font=("Segoe UI", 12)).pack(side="left")
        ctk.CTkLabel(ins, textvariable=self.var_trm_status, text_color="#FFFFFF",
                     font=("Segoe UI", 12, "bold")).pack(side="left", padx=(6, 0))
        ctk.CTkButton(ins, text="↻ Actualizar", width=110, height=28, corner_radius=8,
                      fg_color=BLUE, hover_color=CYAN, font=("Segoe UI", 11, "bold"),
                      command=lambda: self._actualizar_trm(False)).pack(side="right")

        # Datos globales del viaje
        g = ctk.CTkFrame(self, fg_color=CARD, corner_radius=14)
        g.grid(row=2, column=0, sticky="ew", padx=16, pady=(12, 6))
        for c in range(6):
            g.grid_columnconfigure(c, weight=1)
        ctk.CTkLabel(g, text="Datos del viaje (cliente)", text_color=NAVY,
                     font=("Segoe UI", 14, "bold"), height=18).grid(
            row=0, column=0, columnspan=3, sticky="w", padx=16, pady=(8, 0))
        # Quien realiza la cotizacion
        cotz = ctk.CTkFrame(g, fg_color="transparent")
        cotz.grid(row=0, column=3, columnspan=3, sticky="e", padx=16, pady=(6, 0))
        ctk.CTkLabel(cotz, text="Cotizado por:", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(side="left", padx=(0, 6))
        self._cotz_map = {f"{n}  -  {c}": (n, c) for n, c in COTIZADORES}
        ops = list(self._cotz_map.keys())
        self.var_cotizador = tk.StringVar(value=ops[0])
        ctk.CTkOptionMenu(cotz, variable=self.var_cotizador, values=ops, width=320, height=30,
                          corner_radius=8, fg_color=NAVY, button_color=NAVY2,
                          button_hover_color=BLUE, dropdown_fg_color=CARD,
                          dropdown_text_color=TEXT, font=("Segoe UI", 11)).pack(side="left")
        def lab(t, r, c, span=1):
            ctk.CTkLabel(g, text=t, text_color=MUTED, font=("Segoe UI", 10), height=13).grid(
                row=r, column=c, columnspan=span, sticky="w", padx=16, pady=0)
        lab("EMAIL CLIENTE", 1, 2)
        lab("FECHAS DEL VIAJE (ida - regreso)", 1, 4, 2)
        cli_lab = ctk.CTkFrame(g, fg_color="transparent")
        cli_lab.grid(row=1, column=0, columnspan=2, sticky="ew", padx=16, pady=0)
        ctk.CTkLabel(cli_lab, text="CLIENTE", text_color=MUTED,
                     font=("Segoe UI", 10), height=13).pack(side="left")
        ctk.CTkButton(cli_lab, text="✏ Editar", width=66, height=22, corner_radius=6,
                      fg_color=CARD2, text_color=NAVY, hover_color=LINE, border_width=1,
                      border_color=LINE, font=("Segoe UI", 10, "bold"),
                      command=self._editar_cliente_actual).pack(side="right", padx=(4, 0))
        ctk.CTkButton(cli_lab, text="🔍 Buscar", width=84, height=22, corner_radius=6,
                      fg_color=NAVY, hover_color=BLUE, font=("Segoe UI", 10, "bold"),
                      command=self._buscar_cliente).pack(side="right")
        self.var_cli = tk.StringVar(); self.var_email = tk.StringVar()
        ctk.CTkEntry(g, textvariable=self.var_cli, height=30, corner_radius=8,
                     border_color=LINE, fg_color=CARD2, placeholder_text="Nombre del cliente"
                     ).grid(row=2, column=0, columnspan=2, sticky="ew", padx=16, pady=(0, 6))
        ctk.CTkEntry(g, textvariable=self.var_email, height=30, corner_radius=8,
                     border_color=LINE, fg_color=CARD2,
                     placeholder_text="correo@cliente.com  (se enviara aqui)"
                     ).grid(row=2, column=2, columnspan=2, sticky="ew", padx=16, pady=(0, 6))
        fechas_fr = ctk.CTkFrame(g, fg_color="transparent")
        fechas_fr.grid(row=2, column=4, columnspan=2, sticky="ew", padx=16, pady=(0, 6))
        hoy = datetime.date.today()
        self.sel_desde = SelectorFecha(fechas_fr, command=self._on_fecha_desde, minimo=hoy)
        self.sel_desde.pack(side="left", fill="x", expand=True)
        ctk.CTkLabel(fechas_fr, text="al", text_color=MUTED).pack(side="left", padx=6)
        self.sel_hasta = SelectorFecha(fechas_fr, command=self._recalcular, minimo=hoy)
        self.sel_hasta.pack(side="left", fill="x", expand=True)
        # Asesor / contacto de la empresa
        aso_lab = ctk.CTkFrame(g, fg_color="transparent")
        aso_lab.grid(row=3, column=0, columnspan=2, sticky="ew", padx=16, pady=(2, 0))
        ctk.CTkLabel(aso_lab, text="ASESOR (contacto)", text_color=MUTED,
                     font=("Segoe UI", 10), height=13).pack(side="left")
        self.opt_asesor = ctk.CTkOptionMenu(
            aso_lab, values=["(vendedor)"], width=150, height=22, corner_radius=6,
            fg_color=NAVY, button_color=NAVY2, button_hover_color=BLUE, dropdown_fg_color=CARD,
            dropdown_text_color=TEXT, font=("Segoe UI", 10), command=self._usar_asesor)
        self.opt_asesor.pack(side="right")
        lab("TELEFONO ASESOR", 3, 2, 2)
        self.var_asesor = tk.StringVar(); self.var_aso_tel = tk.StringVar()
        ctk.CTkEntry(g, textvariable=self.var_asesor, height=30, corner_radius=8,
                     border_color=LINE, fg_color=CARD2, placeholder_text="Nombre del asesor"
                     ).grid(row=4, column=0, columnspan=2, sticky="ew", padx=16, pady=(0, 6))
        ctk.CTkEntry(g, textvariable=self.var_aso_tel, height=30, corner_radius=8,
                     border_color=LINE, fg_color=CARD2, placeholder_text="Telefono del asesor"
                     ).grid(row=4, column=2, columnspan=2, sticky="ew", padx=16, pady=(0, 6))
        lab("ADULTOS", 5, 0); lab("NINOS", 5, 1)
        lab("HABITACIONES (Sen/Dob/Tri)", 5, 2, 2)
        lab("EDAD DE CADA NINO", 5, 4, 2)
        self.st_ad = Stepper(g, value=2, minimo=1, maximo=60, command=self._on_pax)
        self.st_ad.grid(row=6, column=0, sticky="w", padx=16, pady=(0, 2))
        self.st_ninos = Stepper(g, value=0, minimo=0, maximo=10, command=self._on_ninos_count)
        self.st_ninos.grid(row=6, column=1, sticky="w", padx=16, pady=(0, 2))
        # Habitaciones al lado de ninos: 3 contadores compactos + Sugerir
        habf = ctk.CTkFrame(g, fg_color="transparent")
        habf.grid(row=6, column=2, columnspan=2, sticky="w", padx=16, pady=(0, 2))
        for attr, val in (("st_hab_s", 0), ("st_hab_d", 1), ("st_hab_t", 0)):
            st = Stepper(habf, value=val, minimo=0, maximo=40, width=84,
                         command=self._on_hab_change)
            st.pack(side="left", padx=(0, 4)); setattr(self, attr, st)
        ctk.CTkButton(habf, text="Sugerir", width=72, height=28, corner_radius=8,
                      fg_color=CARD2, text_color=NAVY, hover_color=LINE, border_width=1,
                      border_color=LINE, font=("Segoe UI", 10, "bold"),
                      command=self._sugerir_hab).pack(side="left", padx=(6, 0))
        self.frame_edades = ctk.CTkFrame(g, fg_color=CARD2, corner_radius=8, height=34)
        self.frame_edades.grid(row=6, column=4, columnspan=2, sticky="ew", padx=16, pady=(0, 2))
        self.frame_edades.grid_propagate(False)
        self.edad_vars = []
        self.lbl_hab = ctk.CTkLabel(g, text="", text_color=MUTED, font=("Segoe UI", 10), height=12)
        self.lbl_hab.grid(row=7, column=0, columnspan=4, sticky="w", padx=16, pady=(0, 4))

        # Barra de destinos (itinerario)
        dst = ctk.CTkFrame(self, fg_color="#E4EDFA", corner_radius=14)
        dst.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 6))
        dst.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(dst, text="Destinos del itinerario (max 5):", text_color=NAVY,
                     font=("Segoe UI", 13, "bold")).grid(row=0, column=0, padx=14, pady=10,
                                                         sticky="w")
        self.chips = ctk.CTkFrame(dst, fg_color="transparent")
        self.chips.grid(row=0, column=1, sticky="w", padx=4, pady=6)
        self.var_add = tk.StringVar(value="+ Agregar destino")
        self.opt_add = ctk.CTkOptionMenu(
            dst, variable=self.var_add, values=list(self.precios.keys()),
            width=190, height=34, corner_radius=8, fg_color=GREEN, button_color=GREEN_H,
            button_hover_color=CYAN, dropdown_fg_color=CARD, dropdown_text_color=TEXT,
            font=("Segoe UI", 12, "bold"), command=self._add_destino)
        self.opt_add.grid(row=0, column=2, padx=14, pady=8)

        # Configuracion del destino activo
        cfgd = ctk.CTkFrame(self, fg_color="transparent")
        cfgd.grid(row=4, column=0, sticky="ew", padx=16)
        self.lbl_activo = ctk.CTkLabel(cfgd, text="", text_color=NAVY,
                                       font=("Segoe UI", 13, "bold"))
        self.lbl_activo.pack(side="left", padx=(4, 16))
        ctk.CTkLabel(cfgd, text="Temporada:", text_color=MUTED).pack(side="left")
        self.var_temp = tk.StringVar()
        self.opt_temp = ctk.CTkOptionMenu(cfgd, variable=self.var_temp, values=["Baja"],
                                          width=140, height=32, corner_radius=8, fg_color=NAVY,
                                          button_color=NAVY2, button_hover_color=BLUE,
                                          dropdown_fg_color=CARD, dropdown_text_color=TEXT,
                                          font=("Segoe UI", 12, "bold"),
                                          command=self._on_temp)
        self.opt_temp.pack(side="left", padx=8)
        ctk.CTkLabel(cfgd, text="Noches:", text_color=MUTED).pack(side="left", padx=(12, 2))
        self.st_noches = Stepper(cfgd, value=3, minimo=1, maximo=60, command=self._on_noches)
        self.st_noches.pack(side="left")

        # Pestanas propias (robustas) + panel del destino activo
        mid = ctk.CTkFrame(self, fg_color="transparent")
        mid.grid(row=5, column=0, sticky="nsew", padx=16, pady=(2, 4))
        mid.grid_columnconfigure(0, weight=1)
        mid.grid_rowconfigure(1, weight=1)
        tabbar = ctk.CTkFrame(mid, fg_color="transparent")
        tabbar.grid(row=0, column=0, sticky="w")
        self._tab_btns = {}
        for key, txt in (("hotel", "  Hotel  "), ("trans", "  Transportes  "),
                         ("act", "  Actividades  ")):
            b = ctk.CTkButton(tabbar, text=txt, height=36, corner_radius=10,
                              font=("Segoe UI", 12, "bold"),
                              command=lambda k=key: self._set_tab(k))
            b.pack(side="left", padx=(0, 4))
            self._tab_btns[key] = b
        self.panel = ctk.CTkFrame(mid, fg_color=CARD, corner_radius=12)
        self.panel.grid(row=1, column=0, sticky="nsew", pady=(4, 0))
        self.panel.grid_columnconfigure(0, weight=1)
        self.panel.grid_rowconfigure(1, weight=1)

        # Barra total
        foot = ctk.CTkFrame(self, fg_color=NAVY, corner_radius=0, height=64)
        foot.grid(row=6, column=0, sticky="ew"); foot.grid_propagate(False)
        fin = ctk.CTkFrame(foot, fg_color="transparent"); fin.pack(fill="both", expand=True,
                                                                   padx=20, pady=6)
        ctk.CTkButton(fin, text="Nueva cotizacion", width=140, height=40, corner_radius=10,
                      fg_color=NAVY2, hover_color=BLUE, font=("Segoe UI", 12, "bold"),
                      command=self._nueva).pack(side="left")
        ctk.CTkButton(fin, text="Generar PDF y enviar  ✉", width=220, height=40,
                      corner_radius=10, fg_color=GREEN, hover_color=GREEN_H,
                      font=("Segoe UI", 14, "bold"),
                      command=self._generar).pack(side="right")
        ct = ctk.CTkFrame(fin, fg_color="transparent"); ct.pack(side="right", padx=24)
        self.lbl_total = ctk.CTkLabel(ct, text="USD 0.00", text_color="#FFFFFF",
                                      font=("Segoe UI", 22, "bold"), height=26); self.lbl_total.pack(anchor="e")
        self.lbl_desglose = ctk.CTkLabel(ct, text="Total del itinerario", text_color="#BFD4F0",
                                         font=("Segoe UI", 10), height=13); self.lbl_desglose.pack(anchor="e")

        self._rebuild_edades()      # muestra "Sin ninos"
        self._refrescar_clientes_picker()
        self._set_tab("hotel")      # estiliza botones y renderiza panel

    OCC = {"sencilla": 1, "doble": 2, "triple": 3}

    def _set_tab(self, key):
        self.tab = key
        for k, b in self._tab_btns.items():
            if k == key:
                b.configure(fg_color=NAVY, text_color="#FFFFFF", hover_color=BLUE)
            else:
                b.configure(fg_color=CARD2, text_color=NAVY, hover_color=LINE)
        self._render_panel()

    def _render_panel(self):
        for w in self.panel.winfo_children():
            w.destroy()
        if self.activo is None:
            ctk.CTkLabel(self.panel, text="Agrega un destino para comenzar.",
                         text_color=MUTED, font=("Segoe UI", 13)).grid(row=0, column=0, pady=30)
            return
        if self.tab == "hotel":
            self._render_hotel()
        else:
            self._render_lista(self.tab)

    # ---- panel HOTEL: lista con precio POR PERSONA ----
    def _render_hotel(self):
        tr = self.tramos[self.activo]
        bar = ctk.CTkFrame(self.panel, fg_color=CARD2, corner_radius=8)
        bar.grid(row=0, column=0, sticky="ew", padx=8, pady=(6, 4))
        ctk.CTkLabel(bar, text="Buscar hotel:", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(side="left", padx=(10, 6), pady=5)
        eb = ctk.CTkEntry(bar, width=220, height=26, corner_radius=8, border_color=LINE,
                          placeholder_text="nombre del hotel...")
        eb.pack(side="left"); eb.insert(0, self.q["hotel"])
        eb.bind("<KeyRelease>", lambda e: (self.q.__setitem__("hotel", eb.get()),
                                           self._fill_hoteles(tr)))
        ctk.CTkLabel(bar, text="Precio POR PERSONA en Sencilla / Doble / Triple",
                     text_color=MUTED, font=("Segoe UI", 9)).pack(side="right", padx=12)
        self._hcont = ctk.CTkScrollableFrame(self.panel, fg_color=CARD2, corner_radius=10)
        self._hcont.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0, 8))
        self._hcont.grid_columnconfigure(0, weight=1)
        self._fill_hoteles(tr)

    def _margenes(self, dd):
        """Margenes efectivos (hotel, terrestre) segun el periodo del viaje."""
        _, mh_over, mt_over = self._periodo()
        mh = mh_over if mh_over else (dd["hoteles"]["margen"] if dd.get("hoteles") else 0.88)
        mt = mt_over if mt_over else (dd["terrestres"]["margen"] if dd.get("terrestres") else 0.75)
        return mh, mt

    def _fill_hoteles(self, tr):
        for w in self._hcont.winfo_children():
            w.destroy()
        tasa = self._tasa()
        dd = self.precios.get(tr["destino"], {})
        mh = self._margenes(dd)[0]
        noches = max(tr["noches"], 1)
        hoteles, nombres = self._hoteles_de(tr["destino"], tr["temporada"])
        # limpiar seleccionados que ya no existen
        tr["hoteles"] = [n for n in tr["hoteles"] if n in nombres]
        info = ctk.CTkLabel(self._hcont,
                            text=f"Marca hasta 5 hoteles como opciones  "
                                 f"({len(tr['hoteles'])}/5 elegidos)",
                            text_color=NAVY, font=("Segoe UI", 10, "bold"))
        info.pack(anchor="w", padx=8, pady=(2, 4))
        q = (self.q["hotel"] or "").lower()
        n_mostrados = 0
        for h, nom in zip(hoteles, nombres):
            if q and q not in nom.lower():
                continue
            sel = (nom in tr["hoteles"])
            cat = h.get("categoria", "")
            row = ctk.CTkFrame(self._hcont, fg_color="#E7F0FB" if sel else CARD,
                               corner_radius=6, border_width=2 if sel else 0, border_color=NAVY)
            row.pack(fill="x", padx=6, pady=1)
            inner = ctk.CTkFrame(row, fg_color="transparent")
            inner.pack(fill="x", padx=8, pady=2)
            ctk.CTkLabel(inner, text=("☑ " if sel else "☐ ") + nom, text_color=NAVY,
                         font=("Segoe UI", 10, "bold"), anchor="w").pack(side="left")
            if cat:
                ctk.CTkLabel(inner, text=cat, text_color="#FFFFFF", fg_color=CYAN, corner_radius=6,
                             font=("Segoe UI", 8, "bold"), width=54, height=15).pack(side="left", padx=6)

            def pp(k, hh=h):
                v = hh.get(k)
                if not v or not tasa:
                    return "N/D"
                return usd(precio_hotel_usd_noche(v, tasa, mh) * noches / self.OCC[k])
            ctk.CTkLabel(inner, text=f"Sen {pp('sencilla')}    Dob {pp('doble')}    Tri {pp('triple')}",
                         text_color=TEXT, font=("Segoe UI", 10), anchor="e").pack(side="right")
            for wdg in (row, inner) + tuple(inner.winfo_children()):
                wdg.bind("<Button-1>", lambda e, n=nom: self._toggle_hotel(n))
            n_mostrados += 1
        if not n_mostrados:
            ctk.CTkLabel(self._hcont, text="No hay hoteles para este filtro.",
                         text_color=MUTED).pack(pady=20)

    def _toggle_hotel(self, nom):
        tr = self.tramos[self.activo]
        if nom in tr["hoteles"]:
            tr["hoteles"].remove(nom)
        else:
            if len(tr["hoteles"]) >= 5:
                messagebox.showinfo("Limite", "Puedes elegir hasta 5 hoteles como opciones.")
                return
            tr["hoteles"].append(nom)
        self._fill_hoteles(tr)
        self._recalcular()

    def _on_hab(self, k, val):
        if self.activo is None:
            return
        self.tramos[self.activo]["hab"][k] = val
        if self.tab == "hotel" and hasattr(self, "_hcont") and self._hcont.winfo_exists():
            self._fill_hoteles(self.tramos[self.activo])
        self._recalcular()

    # ---- panel TRANSPORTES / ACTIVIDADES: lista con precio POR PERSONA ----
    def _render_lista(self, tipo):
        tr = self.tramos[self.activo]
        bar = ctk.CTkFrame(self.panel, fg_color=CARD2, corner_radius=8)
        bar.grid(row=0, column=0, sticky="ew", padx=8, pady=(6, 4))
        ctk.CTkLabel(bar, text="Buscar:", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(side="left", padx=(10, 6), pady=5)
        eb = ctk.CTkEntry(bar, width=240, height=26, corner_radius=8, border_color=LINE,
                          placeholder_text="nombre del servicio...")
        eb.pack(side="left"); eb.insert(0, self.q[tipo])
        eb.bind("<KeyRelease>", lambda e: (self.q.__setitem__(tipo, eb.get()),
                                           self._fill_lista(tr, tipo)))
        ctk.CTkLabel(bar, text="(valor por persona a la derecha)", text_color=MUTED,
                     font=("Segoe UI", 9)).pack(side="right", padx=12)
        self._lcont = ctk.CTkScrollableFrame(self.panel, fg_color=CARD2, corner_radius=10)
        self._lcont.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0, 8))
        self._lcont.grid_columnconfigure(0, weight=1)
        self._fill_lista(tr, tipo)

    def _fill_lista(self, tr, tipo):
        for w in self._lcont.winfo_children():
            w.destroy()
        sel = tr[tipo]
        tasa = self._tasa()
        dd = self.precios.get(tr["destino"], {})
        margen = self._margenes(dd)[1]
        ages = self._ninos_ages()
        N = max(self.st_ad.get() + sum(1 for a in ages if a >= 1), 1)
        q = (self.q[tipo] or "").lower()
        n_mostrados = 0
        for serv in self._servicios_por_tipo(tr["destino"], tipo):
            nombre = serv["nombre"]
            if q and q not in nombre.lower():
                continue
            fila = ctk.CTkFrame(self._lcont, fg_color=CARD, corner_radius=6)
            fila.pack(fill="x", padx=6, pady=1)
            fila.grid_columnconfigure(0, weight=1)
            chk = tk.BooleanVar(value=(nombre in sel))
            txt = nombre + ("   (privado)" if es_privado(nombre) else "")

            def on_toggle(n=nombre, cv=chk, s=sel):
                if cv.get():
                    s.add(n)
                else:
                    s.discard(n)
                self._recalcular()

            ctk.CTkCheckBox(fila, text=txt, variable=chk, onvalue=True, offvalue=False,
                            corner_radius=5, fg_color=NAVY, hover_color=BLUE, text_color=TEXT,
                            font=("Segoe UI", 10), command=on_toggle, height=18,
                            checkbox_width=18, checkbox_height=18).grid(
                row=0, column=0, sticky="w", padx=10, pady=3)
            pp = (precio_terrestre_usd(serv, N, tasa, margen) / N) if tasa else None
            ctk.CTkLabel(fila, text=(usd(pp) + " p/p") if pp else "", text_color=NAVY,
                         font=("Segoe UI", 10, "bold")).grid(row=0, column=1, padx=10)
            n_mostrados += 1
        if not n_mostrados:
            ctk.CTkLabel(self._lcont, text="No hay servicios para este filtro.",
                         text_color=MUTED).pack(pady=20)

    # ------------------------------------------------------------- TRM / periodo
    def _periodo(self):
        """(descuento, margen_hotel, margen_terrestre) segun la fecha de ida."""
        return periodo_por_fecha(self.sel_desde.get() if hasattr(self, "sel_desde") else None)

    def _tasa(self):
        """Dolar aplicado = TRM del dia - descuento del periodo (segun fecha de viaje)."""
        if not self._trm or self._trm < 1000:
            return None
        desc = self._periodo()[0]
        tasa = self._trm - desc
        return tasa if tasa > 0 else None   # nunca tasa negativa -> nunca precios negativos

    def _actualizar_trm(self, silencioso=True):
        self.var_trm_status.set("actualizando...")
        def worker():
            trm = obtener_trm()
            self.after(0, lambda: self._aplicar_trm(trm, silencioso))
        threading.Thread(target=worker, daemon=True).start()

    def _aplicar_trm(self, trm, silencioso):
        hoy = datetime.date.today().strftime("%d/%m/%Y")
        if trm:
            self._trm = trm
            self.var_trm_status.set("actualizada " + hoy + "  ✓")
            self.cfg["ultima_trm"] = str(trm); self.cfg["ultima_trm_fecha"] = hoy
            try: guardar_config(self.cfg)
            except Exception: pass
            if self.activo is not None:
                self._render_panel()
            self._recalcular()
        else:
            if self._trm:
                self.var_trm_status.set("sin conexion (usando ultima del "
                                        + self.cfg.get("ultima_trm_fecha", "?") + ")")
                if self.activo is not None:
                    self._render_panel()
                self._recalcular()
            else:
                self.var_trm_status.set("no disponible - conectate a internet")
            if not silencioso:
                messagebox.showwarning("Tasa no disponible",
                                       "No se pudo consultar la tasa del dia. "
                                       "Verifica tu conexion y pulsa 'Actualizar'.")

    # ------------------------------------------------------- actualizaciones
    def _chequear_actualizacion(self):
        """En segundo plano, compara la version instalada con la del repositorio."""
        def worker():
            info = hay_actualizacion()
            if info:
                self.after(0, lambda: self._ofrecer_actualizacion(info))
        threading.Thread(target=worker, daemon=True).start()

    def _ofrecer_actualizacion(self, info):
        ver = info.get("version", "?")
        notas = (info.get("notas", "") or "").strip()
        url = info.get("installer", "")
        msg = (f"Hay una nueva version disponible: v{ver}\n"
               f"(tienes instalada la v{VERSION})\n")
        if notas:
            msg += f"\nNovedades:\n{notas}\n"
        msg += "\n¿Descargar e instalar ahora?"
        if messagebox.askyesno("Actualizacion disponible", msg):
            self._descargar_e_instalar(url, ver)

    def _descargar_e_instalar(self, url, ver):
        if not url:
            messagebox.showinfo("Actualizacion",
                                "No se encontro el instalador en el repositorio.\n"
                                "Descarga la ultima version desde GitHub.")
            return
        def worker():
            try:
                destino = os.path.join(tempfile.gettempdir(),
                                       f"CotizadorInnoba-Setup-{ver}.exe")
                ctx = ssl.create_default_context()
                req = urllib.request.Request(url,
                                             headers={"User-Agent": "CotizadorInnoba"})
                with urllib.request.urlopen(req, context=ctx, timeout=180) as r, \
                        open(destino, "wb") as f:
                    shutil.copyfileobj(r, f)
                self.after(0, lambda: self._lanzar_instalador(destino))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror(
                    "Error al descargar", f"No se pudo descargar la actualizacion:\n{e}"))
        self.var_trm_status.set("Descargando actualizacion...")
        threading.Thread(target=worker, daemon=True).start()

    def _lanzar_instalador(self, ruta):
        # RESPALDO de seguridad de los datos ANTES de actualizar (cotizaciones, reservas,
        # clientes, tareas). Los datos viven en %APPDATA% y el instalador no los toca, pero
        # dejamos ademas una copia permanente por si acaso.
        try:
            respaldar_datos(self.cfg)
            respaldo_antes_actualizar()
        except Exception:
            pass
        # avisamos ANTES de abrir el instalador
        messagebox.showinfo("Actualizacion",
                            "Tus cotizaciones y reservas estan guardadas y respaldadas; "
                            "la actualizacion no las borra.\n\n"
                            "La aplicacion se cerrara y se abrira el instalador "
                            "para completar la actualizacion.")
        try:
            os.startfile(ruta)   # abre el instalador
        except Exception as e:
            messagebox.showerror("Error al abrir el instalador", str(e))
            return
        # cierre INMEDIATO del proceso para liberar el .exe (evita el error de
        # "acceso denegado" al reemplazarlo). El instalador ademas cierra la app.
        os._exit(0)

    # ------------------------------------------------------------- destinos
    def _temporadas_de(self, destino):
        dd = self.precios.get(destino, {})
        temps = []
        if dd.get("hoteles"):
            for h in dd["hoteles"]["hoteles"]:
                t = (h.get("temporada", "") or "Baja").strip()
                if t not in temps:
                    temps.append(t)
        return temps or ["Baja"]

    def _hoteles_de(self, destino, temporada):
        dd = self.precios.get(destino, {}); hoteles = []; nombres = []
        if dd.get("hoteles"):
            for h in dd["hoteles"]["hoteles"]:
                ht = (h.get("temporada", "") or "Baja").strip()
                if ht == temporada:
                    hoteles.append(h)
                    zona = (" [" + h["zona"] + "]") if h.get("zona") else ""
                    nombres.append(h["nombre"] + zona)
        return hoteles, nombres

    def _add_destino(self, nombre):
        self.var_add.set("+ Agregar destino")
        if nombre not in self.precios:
            return
        for i, tr in enumerate(self.tramos):
            if tr["destino"] == nombre:
                self._set_activo(i); return
        if len(self.tramos) >= 5:
            messagebox.showinfo("Limite", "Puedes cotizar hasta 5 destinos.")
            return
        temp = self._temporadas_de(nombre)[0]
        self.tramos.append({"destino": nombre, "temporada": temp, "noches": 3,
                            "hoteles": [], "hab": {"s": 0, "d": 1, "t": 0},
                            "trans": set(), "act": set()})
        self._rebuild_chips()
        self._set_activo(len(self.tramos) - 1)

    def _remove_destino(self, idx):
        if 0 <= idx < len(self.tramos):
            del self.tramos[idx]
            if not self.tramos:
                self.activo = None
            elif self.activo >= len(self.tramos):
                self.activo = len(self.tramos) - 1
            self._rebuild_chips()
            if self.activo is not None:
                self._cargar_activo()
            else:
                self._limpiar_activo()
            self._recalcular()

    def _rebuild_chips(self):
        for w in self.chips.winfo_children():
            w.destroy()
        for i, tr in enumerate(self.tramos):
            activo = (i == self.activo)
            chip = ctk.CTkFrame(self.chips, fg_color=NAVY if activo else CARD,
                                corner_radius=16)
            chip.pack(side="left", padx=4)
            ctk.CTkButton(chip, text=f"{i+1}. {tr['destino']}",
                          fg_color="transparent", hover_color=BLUE if activo else CARD2,
                          text_color="#FFFFFF" if activo else NAVY, height=30, width=10,
                          font=("Segoe UI", 12, "bold"),
                          command=lambda x=i: self._set_activo(x)).pack(side="left", padx=(8, 0))
            ctk.CTkButton(chip, text="✕", width=24, height=26, corner_radius=13,
                          fg_color="transparent", hover_color=RED,
                          text_color="#FFFFFF" if activo else MUTED,
                          font=("Segoe UI", 12, "bold"),
                          command=lambda x=i: self._remove_destino(x)).pack(side="left", padx=2)
        # limite de agregar
        disponibles = [d for d in self.precios if d not in [t["destino"] for t in self.tramos]]
        self.opt_add.configure(values=disponibles or ["(todos agregados)"])

    def _set_activo(self, idx):
        if idx < 0 or idx >= len(self.tramos):
            return
        self.activo = idx
        self._rebuild_chips()
        self._cargar_activo()
        self._recalcular()

    def _limpiar_activo(self):
        self.lbl_activo.configure(text="Agrega un destino para comenzar")
        self._render_panel()

    def _cargar_activo(self):
        if self.activo is None:
            self._limpiar_activo(); return
        tr = self.tramos[self.activo]
        self._cargando = True
        self.lbl_activo.configure(text="Destino: " + tr["destino"])
        temps = self._temporadas_de(tr["destino"])
        self.opt_temp.configure(values=temps)
        if tr["temporada"] not in temps:
            tr["temporada"] = temps[0]
        self.var_temp.set(tr["temporada"])
        self.st_noches.set(tr["noches"])
        _, nombres = self._hoteles_de(tr["destino"], tr["temporada"])
        tr["hoteles"] = [n for n in tr["hoteles"] if n in nombres]
        self._cargando = False
        self._render_panel()

    # handlers de widgets del destino activo
    def _on_temp(self, *_):
        if self._cargando or self.activo is None:
            return
        tr = self.tramos[self.activo]; tr["temporada"] = self.var_temp.get()
        tr["hoteles"] = []   # los hoteles cambian por temporada
        self._render_panel(); self._recalcular()

    def _on_noches(self, *_):
        if self._cargando or self.activo is None:
            return
        self.tramos[self.activo]["noches"] = self.st_noches.get()
        if self.tab == "hotel":
            self._fill_hoteles(self.tramos[self.activo])   # precios p/p dependen de noches
        self._recalcular()

    def _servicios_por_tipo(self, destino, tipo):
        dd = self.precios.get(destino, {})
        if not dd.get("terrestres"):
            return []
        servs = dd["terrestres"]["servicios"]
        if tipo == "trans":
            return [s for s in servs if es_transporte(s["nombre"])]
        return [s for s in servs if not es_transporte(s["nombre"])]

    # ------------------------------------------------------------- ninos / pax
    def _int(self, s, d=0):
        try:
            return int(float(str(s).replace(",", ".")))
        except Exception:
            return d

    def _on_pax(self):
        """Al cambiar pasajeros: refresca precios por persona de la lista visible."""
        if self.activo is not None and self.tab in ("trans", "act") \
                and hasattr(self, "_lcont") and self._lcont.winfo_exists():
            self._fill_lista(self.tramos[self.activo], self.tab)
        if self.activo is not None and self.tab == "hotel" \
                and hasattr(self, "_hcont") and self._hcont.winfo_exists():
            self._fill_hoteles(self.tramos[self.activo])
        if hasattr(self, "lbl_hab"):
            self._val_hab()
        self._recalcular()

    def _on_ninos_count(self):
        self._rebuild_edades()

    def _rebuild_edades(self):
        prev = [v.get() for v in self.edad_vars]
        for w in self.frame_edades.winfo_children():
            w.destroy()
        self.edad_vars = []
        n = self.st_ninos.get()
        if n == 0:
            ctk.CTkLabel(self.frame_edades, text="  Sin ninos", text_color=MUTED,
                         font=("Segoe UI", 11)).pack(side="left", padx=8, pady=10)
        for i in range(n):
            v = tk.StringVar(value=prev[i] if i < len(prev) else "3 anos")
            self.edad_vars.append(v)
            cont = ctk.CTkFrame(self.frame_edades, fg_color="transparent")
            cont.pack(side="left", padx=4, pady=2)
            ctk.CTkLabel(cont, text=f"Nino {i+1}", text_color=MUTED,
                         font=("Segoe UI", 9)).pack(anchor="w")
            ctk.CTkOptionMenu(cont, variable=v, values=EDAD_OPCIONES, width=100, height=26,
                              corner_radius=8, fg_color=NAVY, button_color=NAVY2,
                              button_hover_color=BLUE, dropdown_fg_color=CARD,
                              dropdown_text_color=TEXT, font=("Segoe UI", 11),
                              command=lambda e: self._on_pax()).pack()
        self._on_pax()

    def _ninos_ages(self):
        ages = []
        for v in self.edad_vars:
            try:
                ages.append(EDAD_OPCIONES.index(v.get()))
            except Exception:
                ages.append(3)
        return ages

    def _pax_total(self):
        return self.st_ad.get() + self.st_ninos.get()

    def _hoteles_sel(self, tr):
        """Lista [(nombre_display, hotel_dict)] de los hoteles elegidos como opciones."""
        hoteles, nombres = self._hoteles_de(tr["destino"], tr["temporada"])
        d = {n: h for h, n in zip(hoteles, nombres)}
        return [(n, d[n]) for n in tr["hoteles"] if n in d]

    def _calcular_tramo(self, tr, tasa):
        """Devuelve un bloque: servicios base + opciones de hotel (cada una con TOTAL)."""
        dd = self.precios.get(tr["destino"], {})
        adultos = max(self.st_ad.get(), 1)
        ages = self._ninos_ages()
        det_pax = f"{adultos} ad" + (f" + {len(ages)} nino(s)" if ages else "")
        mh, mt = self._margenes(dd)
        noches = max(int(tr.get("noches", 1)), 1)
        # --- base: transporte + actividades (no depende del hotel) ---
        base_sec = []; base = 0.0; adult_pp = 0.0
        ninos_serv = [0.0] * len(ages)
        if dd.get("terrestres"):
            servs = {s["nombre"]: s for s in dd["terrestres"]["servicios"]}
            N = max(adultos + sum(1 for a in ages if a >= 1), 1)
            for tipo, titulo in (("trans", "TRANSPORTE TERRESTRE"), ("act", "ACTIVIDADES")):
                filas = []; ss = 0.0
                for nombre in sorted(tr[tipo]):
                    serv = servs.get(nombre)
                    if not serv:
                        continue
                    priv = es_privado(nombre)
                    total_serv = precio_servicio_grupo(serv, adultos, ages, tasa, mt, priv)
                    ss += total_serv
                    pp = precio_terrestre_usd(serv, N, tasa, mt) / N
                    adult_pp += pp
                    for i, a in enumerate(ages):
                        if a == 0:
                            continue
                        elif a <= 2:
                            ninos_serv[i] += (CHILD_PRIVADO_USD if priv else pp)
                        else:
                            ninos_serv[i] += 0.5 * pp
                    desc = ""
                    if tipo == "act":
                        reg = buscar_descripcion(nombre, tr["destino"], self.descripciones)
                        if reg:
                            desc = texto_descripcion(reg)
                    # fila: (concepto, detalle_pax, total, por_pasajero_adulto, descripcion)
                    filas.append((nombre, det_pax, total_serv, pp, desc))
                if filas:
                    base_sec.append((titulo, filas, ss)); base += ss
        # --- ninos: precio fijo (servicios + tarifa hotel 3-9); no varia por hotel ---
        surch = precio_hotel_nino_noche(tasa, mh) * noches
        ninos_tot = [ninos_serv[i] + (surch if a >= 3 else 0.0) for i, a in enumerate(ages)]
        ninos_total = sum(ninos_tot)
        grupos = {}
        for a, pr in zip(ages, ninos_tot):
            g = grupos.setdefault(a, [0, 0.0]); g[0] += 1; g[1] = pr
        ninos_fijo = [(a, c, pr) for a, (c, pr) in sorted(grupos.items())]
        # --- opciones de hotel: precio POR PERSONA en sencilla / doble / triple ---
        opciones = []
        for nom, h in self._hoteles_sel(tr):
            fila = {"nombre": h["nombre"], "categoria": h.get("categoria", "")}
            for k in ("sencilla", "doble", "triple"):
                v = h.get(k)
                fila[k] = (adult_pp + precio_hotel_usd_noche(v, tasa, mh) * noches / self.OCC[k]) \
                    if v else None
            opciones.append(fila)
        return {"destino": tr["destino"],
                "subtitulo": f"{tr['destino']}   ·   Temporada {tr['temporada']}"
                             f"   ·   {noches} noches",
                "base_secciones": base_sec, "base": base,
                "opciones": opciones, "ninos": ninos_fijo, "n_adultos": adultos,
                "solo_servicios": adult_pp}

    def _calcular_todo(self):
        tasa = self._tasa()
        if tasa is None:
            return [], 0.0, False
        bloques = []; ref = 0.0; opc_mode = False
        for tr in self.tramos:
            b = self._calcular_tramo(tr, tasa)
            if len(b["opciones"]) > 1:
                opc_mode = True
            if b["opciones"]:
                o = b["opciones"][0]
                ref += o.get("doble") or o.get("sencilla") or o.get("triple") or 0.0
            else:
                ref += b["base"]
            bloques.append(b)
        return bloques, ref, opc_mode

    # ------------------------------------------------------------- habitaciones
    def _hab(self):
        return {"sencilla": self.st_hab_s.get(),
                "doble": self.st_hab_d.get(),
                "triple": self.st_hab_t.get()}

    def _hab_ocupacion(self, hab=None):
        h = hab or self._hab()
        return h["sencilla"] * 1 + h["doble"] * 2 + h["triple"] * 3

    def _sugerir_hab(self):
        a = max(self.st_ad.get(), 0)
        self.st_hab_t.set(0)
        self.st_hab_d.set(a // 2)
        self.st_hab_s.set(a % 2)
        self._on_hab_change()

    def _val_hab(self):
        try:
            ocup = self._hab_ocupacion(); ad = self.st_ad.get()
            if ocup == ad:
                self.lbl_hab.configure(text=f"OK: {ocup} plazas = {ad} adultos",
                                       text_color=GREEN)
            else:
                self.lbl_hab.configure(
                    text=f"Ojo: {ocup} plazas vs {ad} adultos", text_color=RED)
        except Exception:
            pass

    def _on_hab_change(self):
        self._val_hab()
        self._recalcular()

    def _total_reserva(self, bloques):
        """Precio TOTAL de la reserva usando la 1a opcion de hotel y las
           habitaciones indicadas (sencilla/doble/triple)."""
        hab = self._hab()
        con_op = [b for b in bloques if b["opciones"]]
        if not con_op or self._hab_ocupacion(hab) == 0:
            return None
        total = 0.0
        for b in con_op:
            o = b["opciones"][0]
            for acc, occ in (("sencilla", 1), ("doble", 2), ("triple", 3)):
                pp = o.get(acc)
                if pp and hab[acc]:
                    total += hab[acc] * occ * pp
            total += sum(c * pr for a, c, pr in b["ninos"])
        return total

    # ------------------------------------------------------------- itinerario
    def _itinerario_auto(self):
        """Arma un borrador de itinerario dia por dia a partir de los destinos y
           actividades elegidas (Dia 1 llegada, actividades, ultimo dia salida)."""
        lineas = []
        dia = 1
        if self.tramos:
            d0 = self.tramos[0]["destino"]
            lineas.append(f"DIA {dia:02d}: LLEGADA A {d0.upper()}")
            lineas.append("Recepcion en el aeropuerto por nuestro equipo y traslado al "
                          "hotel seleccionado. Registro y alojamiento.")
            dia += 1
        for tr in self.tramos:
            for act in sorted(tr["act"]):
                reg = buscar_descripcion(act, tr["destino"], self.descripciones)
                d = texto_descripcion(reg) if reg else ""
                lineas.append(f"DIA {dia:02d}: {tr['destino'].upper()} - {act.upper()}")
                lineas.append(d or "Actividad programada. Alojamiento.")
                dia += 1
        lineas.append(f"DIA {dia:02d}: TRASLADO AL AEROPUERTO")
        lineas.append("Desayuno. A la hora indicada realizamos el traslado al aeropuerto. "
                      "Fin de nuestros servicios.")
        return "\n".join(lineas)

    def _abrir_itinerario(self):
        if not self.tramos:
            messagebox.showinfo("Itinerario", "Agrega al menos un destino primero.")
            return
        win = ctk.CTkToplevel(self)
        win.title("Itinerario dia por dia")
        win.geometry("760x620"); win.configure(fg_color=BG)
        win.transient(self); win.grab_set()
        ctk.CTkLabel(win, text="Itinerario dia por dia (editable)", text_color=NAVY,
                     font=("Segoe UI", 16, "bold")).pack(anchor="w", padx=16, pady=(14, 0))
        ctk.CTkLabel(win, text="Se incluira en el PDF. Las lineas que empiezan por 'DIA' "
                     "salen resaltadas.", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=16, pady=(0, 6))
        cont = ctk.CTkFrame(win, fg_color=CARD, corner_radius=10)
        cont.pack(fill="both", expand=True, padx=16, pady=6)
        txt = tk.Text(cont, wrap="word", font=("Segoe UI", 11), bd=0, relief="flat",
                      padx=12, pady=12, background="#FFFFFF", foreground=TEXT)
        txt.pack(fill="both", expand=True, padx=4, pady=4)
        txt.insert("1.0", getattr(self, "_itinerario", "") or self._itinerario_auto())
        bar = ctk.CTkFrame(win, fg_color="transparent"); bar.pack(fill="x", padx=16, pady=(0, 14))
        def regen():
            txt.delete("1.0", "end"); txt.insert("1.0", self._itinerario_auto())
        def guardar():
            self._itinerario = txt.get("1.0", "end").strip(); win.destroy()
        ctk.CTkButton(bar, text="Regenerar automatico", width=180, fg_color=CARD2,
                      text_color=NAVY, hover_color=LINE, border_width=1, border_color=LINE,
                      font=("Segoe UI", 12, "bold"), command=regen).pack(side="left")
        ctk.CTkButton(bar, text="Guardar", width=140, fg_color=GREEN, hover_color=GREEN_H,
                      font=("Segoe UI", 13, "bold"), command=guardar).pack(side="right")

    def _recalcular(self, *a):
        try:
            bloques, total, opc_mode = self._calcular_todo()
        except Exception:
            return
        if self._tasa() is None:
            self.lbl_total.configure(text="USD --")
            self.lbl_desglose.configure(text="Esperando tasa del dia...")
            return
        reserva = self._total_reserva(bloques)
        if reserva is not None:
            self.lbl_total.configure(text=usd(reserva))
            h = self._hab()
            det = ", ".join(f"{h[k]} {k}" for k in ("sencilla", "doble", "triple") if h[k])
            self.lbl_desglose.configure(text=f"Total reserva ({det}) - 1a opcion")
        else:
            self.lbl_total.configure(text=usd(total))
            self.lbl_desglose.configure(text="Indica las habitaciones para el total")

    # ------------------------------------------------------------- acciones
    def _on_fecha_desde(self):
        d = self.sel_desde.get()
        if d:
            self.sel_hasta.minimo = d
            if self.sel_hasta.get() and self.sel_hasta.get() < d:
                self.sel_hasta.clear()
        # el descuento y los margenes dependen de la fecha de viaje -> refrescar
        if self.activo is not None:
            self._render_panel()
        self._recalcular()

    def _abrir_empresa(self):
        VentanaEmpresa(self, self.cfg, self._on_cfg)

    def _on_cfg(self, cfg):
        self.cfg = cfg

    def _abrir_cotizaciones(self):
        VentanaCotizaciones(self)

    def _volver_inicio(self):
        """Cierra el modulo de Cotizacion y vuelve al selector de modulos."""
        lanz = self.master
        try:
            if lanz is not None and hasattr(lanz, "cotizador"):
                lanz.cotizador = None
        except Exception:
            pass
        try:
            self.destroy()
        except Exception:
            pass
        try:
            if lanz is not None:
                lanz.deiconify()
                lanz.lift()
                lanz.focus_force()
        except Exception:
            pass

    def _chequear_seguimientos(self):
        """Al abrir: importa las del HTML (clientes) y alerta de seguimientos vencidos."""
        def worker():
            try:
                importar_cotizaciones_html()
            except Exception:
                pass
            try:
                enviados = procesar_correos_programados(self.cfg)
            except Exception:
                enviados = []
            if enviados:
                self.after(0, lambda: messagebox.showinfo(
                    "Correos de seguimiento programados",
                    f"Se enviaron {len(enviados)} correo(s) de seguimiento a clientes:\n"
                    + ", ".join(enviados)))
            try:
                pend = seguimientos_pendientes()
            except Exception:
                pend = []
            if pend:
                self.after(0, lambda: self._alerta_seguimientos(pend))
        threading.Thread(target=worker, daemon=True).start()

    def _alerta_seguimientos(self, pend):
        lineas = [f"• {it.get('numero','')}  {it.get('cliente','') or '(sin agencia)'}"
                  f"  ->  {motivo}" for it, motivo in pend[:15]]
        extra = f"\n\n(y {len(pend) - 15} mas)" if len(pend) > 15 else ""
        if messagebox.askyesno(
                "Seguimiento de cotizaciones",
                f"Tienes {len(pend)} cotizacion(es) para dar seguimiento hoy:\n\n"
                + "\n".join(lineas) + extra + "\n\n¿Abrir el historial de cotizaciones?"):
            self._abrir_cotizaciones()

    def _abrir_clientes(self):
        VentanaClientes(self, on_cambio=self._on_clientes_cambio)

    def _on_clientes_cambio(self):
        self.clientes = cargar_clientes()

    def _refrescar_clientes_picker(self):
        pass  # el buscador (lupa) lee self.clientes directamente

    def _buscar_cliente(self):
        SelectorContacto(self, self._usar_cliente_contacto)

    def _usar_cliente_contacto(self, c, v):
        """Coloca empresa + contacto (vendedor) elegido en la cotizacion."""
        self.var_cli.set(c.get("empresa", ""))
        self._vendedores = c.get("vendedores", []) or []
        vals = [x.get("nombre", "") for x in self._vendedores if x.get("nombre")]
        self.opt_asesor.configure(values=vals or ["(vendedor)"])
        if v and v.get("nombre"):
            self.opt_asesor.set(v.get("nombre", ""))
            self.var_asesor.set(v.get("nombre", ""))
            self.var_aso_tel.set(v.get("telefono", ""))
            self.var_email.set(v.get("email", "") or c.get("email", ""))
        else:
            self.opt_asesor.set(vals[0] if vals else "(vendedor)")
            if vals:
                self._usar_asesor(vals[0])
            self.var_email.set(c.get("email", "") or
                               (self._vendedores[0].get("email", "") if self._vendedores else ""))

    def _eliminar_cliente_por_nombre(self, nombre):
        # filtra en el mismo objeto lista para que el buscador vea el cambio
        self.clientes[:] = [c for c in self.clientes if c.get("empresa", "") != nombre]
        guardar_clientes(self.clientes)
        if self.var_cli.get().strip() == nombre:
            self.var_cli.set(""); self.var_email.set("")
            self.var_asesor.set(""); self.var_aso_tel.set("")
            self._vendedores = []
            self.opt_asesor.configure(values=["(vendedor)"]); self.opt_asesor.set("(vendedor)")

    def _editar_cliente_actual(self):
        self._editar_cliente_por_nombre(self.var_cli.get().strip())

    def _editar_cliente_por_nombre(self, nombre):
        VentanaClientes(self, on_cambio=self._on_clientes_cambio,
                        preseleccion=nombre or None)

    def _usar_cliente(self, nombre):
        c = next((x for x in self.clientes if x.get("empresa") == nombre), None)
        if not c:
            return
        self.var_cli.set(c.get("empresa", ""))
        email = c.get("email", "")
        if not email and c.get("vendedores"):
            email = c["vendedores"][0].get("email", "")
        if email:
            self.var_email.set(email)
        # llenar el selector de asesores con los vendedores de la empresa
        self._vendedores = c.get("vendedores", []) or []
        vals = [v.get("nombre", "") for v in self._vendedores if v.get("nombre")]
        self.opt_asesor.configure(values=vals or ["(vendedor)"])
        if vals:
            self.opt_asesor.set(vals[0])
            self._usar_asesor(vals[0])
        else:
            self.opt_asesor.set("(vendedor)")
            self.var_asesor.set(""); self.var_aso_tel.set("")

    def _usar_asesor(self, nombre):
        v = next((x for x in getattr(self, "_vendedores", []) if x.get("nombre") == nombre), None)
        self.var_asesor.set(nombre if nombre != "(vendedor)" else "")
        if v:
            self.var_aso_tel.set(v.get("telefono", ""))
            if v.get("email"):
                self.var_email.set(v.get("email", ""))

    def _nueva(self):
        hoy = datetime.date.today()
        self.var_cli.set(""); self.var_email.set("")
        self.var_asesor.set(""); self.var_aso_tel.set("")
        self._vendedores = []
        self.opt_asesor.configure(values=["(vendedor)"]); self.opt_asesor.set("(vendedor)")
        self.sel_desde.clear(); self.sel_hasta.clear()
        self.sel_hasta.minimo = hoy
        self._numero = "COT-" + hoy.strftime("%Y%m%d") + "-001"
        self._fecha = hoy.strftime("%d/%m/%Y")
        self._valida = add_months(hoy, 1).strftime("%d/%m/%Y")
        self.tramos = []; self.activo = None
        self._itinerario = ""
        self._rebuild_chips(); self._limpiar_activo()
        # agregar un primer destino por comodidad
        self._add_destino(list(self.precios.keys())[0])
        if hasattr(self, "st_hab_s"):
            self._sugerir_hab()
        self._recalcular()

    def _snapshot(self):
        """Estado completo de la cotizacion actual (para re-cargarla/editarla)."""
        d = self.sel_desde.get(); h = self.sel_hasta.get()
        return {
            "cliente": self.var_cli.get(), "email": self.var_email.get(),
            "asesor": self.var_asesor.get(), "asesor_tel": self.var_aso_tel.get(),
            "cotizador": self.var_cotizador.get(),
            "fecha_desde": d.isoformat() if d else "",
            "fecha_hasta": h.isoformat() if h else "",
            "adultos": self.st_ad.get(), "ages": self._ninos_ages(),
            "hab": self._hab(), "itinerario": self._itinerario,
            "tramos": [{"destino": t["destino"], "temporada": t["temporada"],
                        "noches": t["noches"], "hoteles": list(t["hoteles"]),
                        "trans": sorted(t["trans"]), "act": sorted(t["act"])}
                       for t in self.tramos],
        }

    def _cargar_cotizacion(self, snap):
        """Carga una cotizacion guardada en el cotizador para editarla."""
        if not snap:
            messagebox.showinfo("Editar cotizacion",
                                "Esta cotizacion no tiene datos para editar "
                                "(se creo en una version anterior o vino del HTML).")
            return
        self.var_cli.set(snap.get("cliente", "")); self.var_email.set(snap.get("email", ""))
        self.var_asesor.set(snap.get("asesor", ""))
        self.var_aso_tel.set(snap.get("asesor_tel", ""))
        if snap.get("cotizador") in self._cotz_map:
            self.var_cotizador.set(snap["cotizador"])
        self.sel_desde.clear(); self.sel_hasta.clear()
        try:
            if snap.get("fecha_desde"):
                self.sel_desde._set(datetime.date.fromisoformat(snap["fecha_desde"]))
            if snap.get("fecha_hasta"):
                self.sel_hasta._set(datetime.date.fromisoformat(snap["fecha_hasta"]))
        except Exception:
            pass
        self.st_ad.set(max(int(snap.get("adultos", 2)), 1))
        ages = snap.get("ages", [])
        self.st_ninos.set(len(ages))
        self._rebuild_edades()
        for i, a in enumerate(ages):
            if i < len(self.edad_vars):
                try:
                    self.edad_vars[i].set(EDAD_OPCIONES[a])
                except Exception:
                    pass
        hab = snap.get("hab", {})
        self.st_hab_s.set(hab.get("sencilla", 0))
        self.st_hab_d.set(hab.get("doble", 0))
        self.st_hab_t.set(hab.get("triple", 0))
        self._itinerario = snap.get("itinerario", "")
        self.tramos = []
        for t in snap.get("tramos", []):
            if t.get("destino") not in self.precios:
                continue
            self.tramos.append({
                "destino": t["destino"], "temporada": t.get("temporada", "Baja"),
                "noches": t.get("noches", 3), "hoteles": list(t.get("hoteles", [])),
                "hab": {"s": 0, "d": 1, "t": 0},
                "trans": set(t.get("trans", [])), "act": set(t.get("act", []))})
        self.activo = 0 if self.tramos else None
        self._rebuild_chips()
        if self.activo is not None:
            self._cargar_activo()
        else:
            self._limpiar_activo()
        self._val_hab()
        self._recalcular()

    def _generar(self):
        if self._tasa() is None:
            messagebox.showwarning("Falta la tasa",
                                   "Aun no hay tasa del dia. Pulsa 'Actualizar'.")
            return
        if not self.tramos:
            messagebox.showwarning("Sin destinos", "Agrega al menos un destino."); return
        bloques, total, opc_mode = self._calcular_todo()
        if total <= 0 or not any(b["base_secciones"] or b["opciones"] for b in bloques):
            messagebox.showwarning("Cotizacion vacia",
                                   "Selecciona al menos un hotel, transporte o actividad."); return
        if not self.sel_desde.get() or not self.sel_hasta.get():
            messagebox.showwarning("Falta la fecha del viaje",
                                   "Elige la FECHA DEL VIAJE (ida y regreso) en el calendario."); return
        if self.sel_hasta.get() < self.sel_desde.get():
            messagebox.showwarning("Fechas invalidas",
                                   "La fecha de regreso no puede ser anterior a la de ida."); return
        # las noches del itinerario deben cuadrar con las fechas del viaje
        total_noches = sum(max(int(t.get("noches", 1)), 1) for t in self.tramos)
        dias = (self.sel_hasta.get() - self.sel_desde.get()).days
        if dias != total_noches:
            if not messagebox.askyesno(
                    "Noches vs fechas del viaje",
                    f"Las noches del itinerario suman {total_noches}, pero entre las fechas "
                    f"elegidas hay {dias} noche(s).\n\nNo coinciden. "
                    "¿Deseas continuar de todos modos?"):
                return
        if not self.var_email.get().strip():
            messagebox.showwarning("Falta el email del cliente",
                                   "Debes ingresar el EMAIL del cliente antes de generar."); return
        ad = self.st_ad.get(); ages = self._ninos_ages()
        pax_txt = f"{ad} adultos"
        if ages:
            det = ", ".join(("bebe" if a == 0 else f"{a} anos") for a in ages)
            pax_txt += f", {len(ages)} ninos ({det})"
        cliente = self.var_cli.get().strip(); email = self.var_email.get().strip()
        firma_nom, firma_cargo = self._cotz_map.get(self.var_cotizador.get(),
                                                    (COTIZADORES[0][0], COTIZADORES[0][1]))
        # consecutivo de la cotizacion (se guarda en el historial al final)
        self._numero = peek_numero_cotizacion()
        reserva = self._total_reserva(bloques)
        total_mostrar = reserva if reserva is not None else total
        datos = {"numero": self._numero, "fecha": self._fecha, "valida_hasta": self._valida,
                 "cliente": cliente, "cli_email": email,
                 "asesor": self.var_asesor.get().strip(),
                 "asesor_tel": self.var_aso_tel.get().strip(),
                 "fechas_viaje": f"{self.sel_desde.get_str()} al {self.sel_hasta.get_str()}",
                 "pax_txt": pax_txt, "opc_mode": opc_mode,
                 "habitaciones": self._hab(),
                 "itinerario": (self._itinerario or self._itinerario_auto()),
                 "firma_nombre": firma_nom, "firma_cargo": firma_cargo,
                 "notas": (f"Vigencia: esta cotizacion tiene una validez de un (1) mes, "
                           f"hasta el {self._valida}. " + self.cfg.get("notas", ""))}
        slug = "".join(c for c in cliente if c.isalnum() or c in " _-").strip().replace(" ", "_")
        destinos_txt = "-".join(t["destino"][:3] for t in self.tramos)
        ruta = filedialog.asksaveasfilename(
            title="Guardar cotizacion PDF", defaultextension=".pdf",
            initialfile=f"Cotizacion_{destinos_txt}_{slug or 'cliente'}.pdf",
            filetypes=[("PDF", "*.pdf")])
        if not ruta:
            return
        try:
            generar_pdf(self.cfg, datos, bloques, total, ruta)
        except Exception as e:
            messagebox.showerror("Error al generar PDF", str(e)); return
        # guardar en el historial de cotizaciones (con consecutivo + estado completo)
        try:
            numero = registrar_cotizacion({
                "cliente": cliente, "asesor": self.var_asesor.get().strip(),
                "asesor_tel": self.var_aso_tel.get().strip(),
                "fecha": self._fecha, "fechas_viaje": datos["fechas_viaje"],
                "cotizado_por": firma_nom, "email": email,
                "destinos": [t["destino"] for t in self.tramos],
                "total": total_mostrar, "pdf": ruta,
                "snapshot": self._snapshot()})
            self.lbl_desglose.configure(text=f"Guardada {numero}")
        except Exception:
            pass
        # enviar por correo
        self._enviar_pdf(ruta, email, cliente, datos.get("asesor", ""))

    def _enviar_pdf(self, ruta, email, cliente, asesor=""):
        smtp_ok = self.cfg.get("correo_remitente") and self.cfg.get("smtp_password")
        if not email:
            if messagebox.askyesno("Sin email",
                                   "El cliente no tiene email, no se puede enviar.\n\n"
                                   "PDF guardado. ¿Abrirlo ahora?"):
                self._abrir(ruta)
            return
        if not smtp_ok:
            messagebox.showinfo("Correo no configurado",
                                "PDF guardado.\n\nPara enviarlo automaticamente, configura el "
                                "'Correo remitente' y su contrasena en 'Datos de mi empresa'.")
            if messagebox.askyesno("PDF", "¿Abrir el PDF ahora?"):
                self._abrir(ruta)
            return
        f_nom, f_cargo = self._cotz_map.get(self.var_cotizador.get(),
                                            (COTIZADORES[0][0], COTIZADORES[0][1]))
        asunto = "Cotizacion de viaje - INNOBA Colombia DMC"
        if cliente:
            asunto += f" - {cliente}"
        saludo = asesor.strip() or cliente or "cliente"
        ref = f"para {cliente}" if (cliente and asesor.strip()) else "para su viaje"
        cuerpo = (f"Estimado(a) {saludo},\n\n"
                  f"Adjunto encontrara la cotizacion solicitada {ref}.\n"
                  "Quedamos atentos a cualquier inquietud.\n\n"
                  "Cordialmente,\n"
                  f"{f_nom}\n{f_cargo}")
        dlg = messagebox.showinfo("Enviando", "Enviando la cotizacion a " + email + "...")
        def worker():
            try:
                enviar_correo(self.cfg, email, asunto, cuerpo, ruta)
                self.after(0, lambda: self._envio_ok(ruta, email))
            except Exception as e:
                self.after(0, lambda: self._envio_error(ruta, str(e)))
        threading.Thread(target=worker, daemon=True).start()

    def _envio_ok(self, ruta, email):
        if messagebox.askyesno("Enviado",
                               "Cotizacion enviada correctamente a:\n" + email +
                               "\n\n¿Abrir el PDF?"):
            self._abrir(ruta)

    def _envio_error(self, ruta, err):
        messagebox.showerror("Error al enviar",
                             "El PDF se guardo, pero no se pudo enviar el correo:\n\n" + err +
                             "\n\nRevisa el correo remitente y la contrasena en "
                             "'Datos de mi empresa'.")
        if messagebox.askyesno("PDF", "¿Abrir el PDF ahora?"):
            self._abrir(ruta)

    def _abrir(self, ruta):
        try: os.startfile(ruta)
        except Exception: pass


class VentanaEmpresa(ctk.CTkToplevel):
    def __init__(self, master, cfg, callback):
        super().__init__(master)
        self.cfg = dict(cfg); self.callback = callback
        self.title("Datos de mi empresa"); self.geometry("580x720"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        cont = ctk.CTkScrollableFrame(self, fg_color=BG); cont.pack(fill="both", expand=True,
                                                                    padx=16, pady=16)
        ctk.CTkLabel(cont, text="Datos de mi empresa", text_color=NAVY,
                     font=("Segoe UI", 18, "bold")).pack(anchor="w", pady=(0, 10))
        self.entradas = {}
        def campo(clave, etq, secreto=False, defecto=""):
            ctk.CTkLabel(cont, text=etq, text_color=MUTED, font=("Segoe UI", 11)).pack(
                anchor="w", padx=2)
            v = tk.StringVar(value=str(self.cfg.get(clave, "") or defecto))
            self.entradas[clave] = v
            e = ctk.CTkEntry(cont, textvariable=v, height=34, corner_radius=8, border_color=LINE,
                             show="*" if secreto else "")
            e.pack(fill="x", pady=(0, 8))
        for clave, etq in [("empresa", "Nombre de la empresa"), ("nit", "NIT / RUC"),
                           ("direccion", "Direccion"), ("telefono", "Telefono"),
                           ("email", "Email"), ("web", "Sitio web"),
                           ("firma_nombre", "Firma - Nombre"), ("firma_cargo", "Firma - Cargo")]:
            campo(clave, etq)
        ctk.CTkLabel(cont, text="—  Envio de correo (Office 365)  —", text_color=NAVY,
                     font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(6, 4))
        campo("correo_remitente", "Correo remitente por defecto (desde donde se envia)")
        campo("smtp_password", "Contrasena del correo (o contrasena de aplicacion)", secreto=True)
        campo("smtp_servidor", "Servidor SMTP")
        campo("smtp_puerto", "Puerto SMTP")
        ctk.CTkLabel(cont, text="—  Correo por cotizador (seguimientos)  —", text_color=NAVY,
                     font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(6, 2))
        ctk.CTkLabel(cont, text="El correo de seguimiento sale del correo del cotizador que hizo "
                     "la cotizacion. Escribe la contrasena (o contrasena de aplicacion) de cada uno.",
                     text_color=MUTED, font=("Segoe UI", 10), wraplength=520,
                     justify="left").pack(anchor="w", padx=2, pady=(0, 4))
        campo("correo_felipe", "Correo de Felipe", defecto="felipe@innobadmc.com")
        campo("pass_felipe", "Contrasena de Felipe", secreto=True)
        campo("correo_carlos", "Correo de Carlos", defecto="directorcomercial@innobadmc.com")
        campo("pass_carlos", "Contrasena de Carlos", secreto=True)
        ctk.CTkLabel(cont, text="Logo (PNG/JPG)", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        lf = ctk.CTkFrame(cont, fg_color="transparent"); lf.pack(fill="x", pady=(0, 8))
        self.logo_var = tk.StringVar(value=self.cfg.get("logo", ""))
        ctk.CTkLabel(lf, textvariable=self.logo_var, text_color=TEXT, font=("Segoe UI", 10),
                     anchor="w").pack(side="left", fill="x", expand=True)
        ctk.CTkButton(lf, text="Elegir", width=70, fg_color=NAVY, hover_color=BLUE,
                      command=self._logo).pack(side="left", padx=4)
        ctk.CTkButton(lf, text="Quitar", width=70, fg_color=CARD2, text_color=NAVY,
                      hover_color=LINE, command=lambda: self.logo_var.set("")).pack(side="left")
        ctk.CTkLabel(cont, text="Notas por defecto", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.txt = ctk.CTkTextbox(cont, height=80, corner_radius=8, border_width=1,
                                  border_color=LINE, fg_color=CARD)
        self.txt.insert("1.0", self.cfg.get("notas", "")); self.txt.pack(fill="x", pady=(0, 12))
        bts = ctk.CTkFrame(cont, fg_color="transparent"); bts.pack(fill="x")
        ctk.CTkButton(bts, text="Guardar", fg_color=GREEN, hover_color=GREEN_H,
                      font=("Segoe UI", 13, "bold"), command=self._guardar).pack(side="right")
        ctk.CTkButton(bts, text="Cancelar", fg_color=CARD2, text_color=NAVY, hover_color=LINE,
                      command=self.destroy).pack(side="right", padx=8)

    def _logo(self):
        r = filedialog.askopenfilename(title="Elegir logo",
                                       filetypes=[("Imagenes", "*.png *.jpg *.jpeg *.gif *.bmp")])
        if r: self.logo_var.set(r)

    def _guardar(self):
        for clave, v in self.entradas.items():
            self.cfg[clave] = v.get().strip()
        self.cfg["logo"] = self.logo_var.get().strip()
        self.cfg["notas"] = self.txt.get("1.0", "end").strip()
        try:
            guardar_config(self.cfg)
        except Exception as e:
            messagebox.showerror("Error", str(e)); return
        self.callback(self.cfg)
        messagebox.showinfo("Guardado", "Datos guardados.")
        self.destroy()


# ============================================================================
# MODULO RESERVAS - Interfaz
# ============================================================================
class DialogoAsesores(ctk.CTkToplevel):
    """Configura los (hasta 3) asesores de reservas para la asignacion rotativa."""
    def __init__(self, master, cfg, on_save=None):
        super().__init__(master)
        self.cfg = cfg; self.on_save = on_save
        self.title("Asesores de reservas")
        self.geometry("540x470"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        ctk.CTkLabel(self, text="Asesores de reservas", font=("Segoe UI", 16, "bold"),
                     text_color=NAVY).pack(pady=(16, 2))
        ctk.CTkLabel(self, text="Cada nueva reserva se asigna automaticamente al siguiente "
                     "asesor (rotacion equitativa).", text_color=MUTED,
                     wraplength=480).pack(pady=(0, 10))
        self.filas = []
        ases = (cfg.get("asesores_reservas") or []) + [{}, {}, {}]
        cont = ctk.CTkFrame(self, fg_color="transparent"); cont.pack(fill="x", padx=20)
        for i in range(3):
            a = ases[i] if i < len(ases) else {}
            f = ctk.CTkFrame(cont, fg_color=CARD, corner_radius=8,
                             border_width=1, border_color=LINE)
            f.pack(fill="x", pady=6)
            ctk.CTkLabel(f, text=f"Asesor {i+1}", text_color=NAVY,
                         font=("Segoe UI", 12, "bold")).grid(
                row=0, column=0, columnspan=2, sticky="w", padx=8, pady=(6, 0))
            vn = tk.StringVar(value=a.get("nombre", ""))
            vm = tk.StringVar(value=a.get("email", ""))
            ctk.CTkEntry(f, textvariable=vn, placeholder_text="Nombre",
                         height=32).grid(row=1, column=0, padx=8, pady=(4, 8), sticky="we")
            ctk.CTkEntry(f, textvariable=vm, placeholder_text="Correo electronico",
                         height=32).grid(row=1, column=1, padx=8, pady=(4, 8), sticky="we")
            f.grid_columnconfigure(0, weight=1); f.grid_columnconfigure(1, weight=1)
            self.filas.append((vn, vm))
        ctk.CTkButton(self, text="Guardar asesores", fg_color=GREEN, hover_color=GREEN_H,
                      height=38, command=self._guardar).pack(pady=14)

    def _guardar(self):
        lst = []
        for vn, vm in self.filas:
            if vn.get().strip():
                lst.append({"nombre": vn.get().strip(), "email": vm.get().strip()})
        self.cfg["asesores_reservas"] = lst
        guardar_config(self.cfg)
        if self.on_save:
            self.on_save()
        messagebox.showinfo("Guardado", f"{len(lst)} asesor(es) guardado(s).")
        self.destroy()


class SelectorCotizacionReserva(ctk.CTkToplevel):
    """Elegir una cotizacion del historial para convertirla en reserva."""
    def __init__(self, master, on_pick):
        super().__init__(master)
        self.on_pick = on_pick
        self.title("Elegir cotizacion")
        self.geometry("760x640"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        ctk.CTkLabel(self, text="Selecciona la cotizacion a convertir en reserva",
                     font=("Segoe UI", 15, "bold"), text_color=NAVY).pack(pady=(14, 6))
        self.q = tk.StringVar()
        bar = ctk.CTkFrame(self, fg_color="transparent"); bar.pack(fill="x", padx=16)
        e = ctk.CTkEntry(bar, textvariable=self.q, height=34,
                         placeholder_text="Buscar por cliente, destino o numero...")
        e.pack(side="left", fill="x", expand=True)
        e.bind("<KeyRelease>", lambda ev: self._pintar())
        self.lista = ctk.CTkScrollableFrame(self, fg_color=BG)
        self.lista.pack(fill="both", expand=True, padx=16, pady=12)
        self._pintar()

    def _pintar(self):
        for w in self.lista.winfo_children():
            w.destroy()
        items = list(reversed(cargar_cotizaciones().get("items", [])))
        q = self.q.get().lower().strip()
        if q:
            items = [it for it in items
                     if q in json.dumps(it, ensure_ascii=False).lower()]
        if not items:
            ctk.CTkLabel(self.lista, text="No hay cotizaciones.",
                         text_color=MUTED).pack(pady=20)
            return
        for it in items:
            row = ctk.CTkFrame(self.lista, fg_color=CARD, corner_radius=8,
                               border_width=1, border_color=LINE)
            row.pack(fill="x", pady=4)
            info = (f"{it.get('numero','')}   ·   {it.get('cliente','')}\n"
                    f"{', '.join(it.get('destinos', []))}   ·   "
                    f"{it.get('fechas_viaje','')}   ·   {usd(it.get('total', 0))}   ·   "
                    f"{it.get('estado','')}")
            ctk.CTkLabel(row, text=info, justify="left", text_color=TEXT,
                         font=("Segoe UI", 11)).pack(side="left", padx=10, pady=8)
            ctk.CTkButton(row, text="Convertir", width=110, fg_color=GREEN,
                          hover_color=GREEN_H,
                          command=lambda x=it: self._elegir(x)).pack(side="right", padx=10)

    def _elegir(self, cot):
        self.on_pick(cot)
        self.destroy()


class VentanaReservaDetalle(ctk.CTkToplevel):
    """Editar una reserva: estado, monto, proveedores y emision de vouchers."""
    def __init__(self, master, res, cfg, on_save=None):
        super().__init__(master)
        self.res = res; self.cfg = cfg; self.on_save = on_save
        self.precios = cargar_precios_seguro()   # para desplegar hoteles por destino
        self.title("Reserva " + res.get("numero", ""))
        self.configure(fg_color=BG)
        # Geometria inicial de respaldo; luego se maximiza para que Windows respete
        # la barra de tareas y el pie (botones) SIEMPRE quede visible.
        self.geometry("980x680+60+10")
        self.transient(master); self.grab_set()
        self.after(60, self._maximizar)
        # Barra de acciones FIJA ARRIBA (siempre visible, no la tapa la barra de tareas)
        self.footer = ctk.CTkFrame(self, fg_color=NAVY, corner_radius=0, height=56)
        self.footer.pack(side="top", fill="x"); self.footer.pack_propagate(False)
        cont = ctk.CTkScrollableFrame(self, fg_color=BG)
        cont.pack(fill="both", expand=True, padx=16, pady=16)

        ase = res.get("asesor", {}) or {}
        ctk.CTkLabel(cont, text=f"Reserva N. {res.get('numero','')}",
                     font=("Segoe UI", 19, "bold"), text_color=NAVY).pack(anchor="w")
        # Asesora asignada: selector de la LISTA OFICIAL (no se escribe a mano)
        self._ases_ofic = asesores_reservas(self.cfg)
        nombres = [a.get("nombre", "") for a in self._ases_ofic if a.get("nombre")]
        opciones = ["(sin asignar)"] + nombres
        actual = ase.get("nombre", "") or "(sin asignar)"
        if actual not in opciones:
            opciones.append(actual)   # conservar el nombre viejo aunque no este en la lista
        fa = ctk.CTkFrame(cont, fg_color="transparent"); fa.pack(anchor="w", fill="x", pady=(2, 2))
        ctk.CTkLabel(fa, text="Asesora asignada:", text_color=BLUE,
                     font=("Segoe UI", 12, "bold")).pack(side="left")
        self.v_asesora = tk.StringVar(value=actual)
        ctk.CTkOptionMenu(fa, variable=self.v_asesora, values=opciones, width=240, height=30,
                          fg_color=NAVY, button_color=NAVY2, button_hover_color=BLUE,
                          dropdown_fg_color=CARD, dropdown_text_color=TEXT,
                          command=self._set_asesora).pack(side="left", padx=8)
        origen = ("cotizacion " + res["cot_origen"]) if res.get("cot_origen") else "manual"
        ctk.CTkLabel(cont, text="Origen: " + origen, text_color=MUTED,
                     font=("Segoe UI", 10)).pack(anchor="w", pady=(0, 6))

        # Datos del cliente / viaje (editables)
        def campo(lbl, valor, ancho=None):
            ctk.CTkLabel(cont, text=lbl, text_color=MUTED,
                         font=("Segoe UI", 11)).pack(anchor="w", padx=2)
            v = tk.StringVar(value=valor)
            ctk.CTkEntry(cont, textvariable=v, height=32, corner_radius=8,
                         border_color=LINE).pack(fill="x", pady=(0, 6))
            return v

        # Cliente / Agencia con buscador de contacto
        ctk.CTkLabel(cont, text="Cliente / Agencia", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        fcli = ctk.CTkFrame(cont, fg_color="transparent"); fcli.pack(fill="x", pady=(0, 6))
        self.v_cli = tk.StringVar(value=res.get("cliente", ""))
        ctk.CTkEntry(fcli, textvariable=self.v_cli, height=32, corner_radius=8,
                     border_color=LINE).pack(side="left", fill="x", expand=True)
        ctk.CTkButton(fcli, text="🔍 Buscar cliente", width=140, height=32, fg_color=NAVY,
                      hover_color=NAVY2, command=self._buscar_cliente_res).pack(side="left", padx=(6, 0))
        f0 = ctk.CTkFrame(cont, fg_color="transparent"); f0.pack(fill="x")
        e1 = ctk.CTkFrame(f0, fg_color="transparent"); e1.pack(side="left", fill="x", expand=True, padx=(0, 6))
        e2 = ctk.CTkFrame(f0, fg_color="transparent"); e2.pack(side="left", fill="x", expand=True, padx=(6, 0))
        ctk.CTkLabel(e1, text="Correo del cliente", text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.v_email = tk.StringVar(value=res.get("email", ""))
        ctk.CTkEntry(e1, textvariable=self.v_email, height=32).pack(fill="x", pady=(0, 6))
        ctk.CTkLabel(e2, text="Contacto (vendedor de la agencia)", text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.v_contacto = tk.StringVar(value=res.get("contacto", ""))
        self.combo_contacto = ctk.CTkComboBox(
            e2, variable=self.v_contacto, height=32,
            values=([res.get("contacto", "")] if res.get("contacto") else []))
        self.combo_contacto.pack(fill="x", pady=(0, 6))
        f2 = ctk.CTkFrame(cont, fg_color="transparent"); f2.pack(fill="x")
        c1 = ctk.CTkFrame(f2, fg_color="transparent"); c1.pack(side="left", fill="x", expand=True, padx=(0, 6))
        c2 = ctk.CTkFrame(f2, fg_color="transparent"); c2.pack(side="left", fill="x", expand=True, padx=(6, 0))
        ctk.CTkLabel(c1, text="Fecha de llegada", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.sel_llegada = SelectorFecha(c1)
        self.sel_llegada.pack(fill="x", pady=(0, 6))
        ctk.CTkLabel(c2, text="Fecha de salida", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.sel_salida = SelectorFecha(c2)
        self.sel_salida.pack(fill="x", pady=(0, 6))
        # inicializar calendarios desde los datos guardados
        _ll = res.get("os_fecha_in", "") or _fechas_in_out(res.get("fechas_viaje", ""))[0]
        _sa = res.get("os_fecha_out", "") or _fechas_in_out(res.get("fechas_viaje", ""))[1]
        dll = _parse_ddmmyyyy(_ll); dsa = _parse_ddmmyyyy(_sa)
        if dll:
            self.sel_llegada._set(dll)
        if dsa:
            self.sel_salida._set(dsa)
        ctk.CTkLabel(cont, text="Pasajeros (resumen, ej. 3 adultos)", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.v_pax = tk.StringVar(value=res.get("pax_txt", ""))
        ctk.CTkEntry(cont, textvariable=self.v_pax, height=32).pack(fill="x", pady=(0, 6))

        # Pasajeros (lista: nombre + pasaporte/documento) -> tabla del voucher
        ph = ctk.CTkFrame(cont, fg_color="transparent"); ph.pack(fill="x", pady=(2, 0))
        ctk.CTkLabel(ph, text="PASAJEROS  (nombre y pasaporte / documento)", text_color=NAVY,
                     font=("Segoe UI", 12, "bold")).pack(side="left")
        ctk.CTkButton(ph, text="+ Agregar pasajero", width=150, height=28, fg_color=BLUE,
                      hover_color=BLUE_H, command=self._agregar_pasajero).pack(side="right")
        self.pax_box = ctk.CTkFrame(cont, fg_color="transparent"); self.pax_box.pack(fill="x", pady=(2, 6))
        self._pintar_pasajeros()

        # Vuelos / tiquetes adjuntos
        vh = ctk.CTkFrame(cont, fg_color="transparent"); vh.pack(fill="x", pady=(2, 0))
        ctk.CTkLabel(vh, text="VUELOS / TIQUETES  (adjuntos)", text_color=NAVY,
                     font=("Segoe UI", 12, "bold")).pack(side="left")
        ctk.CTkButton(vh, text="+ Adjuntar vuelo", width=150, height=28, fg_color=BLUE,
                      hover_color=BLUE_H, command=self._adjuntar_vuelo).pack(side="right")
        self.vuelos_box = ctk.CTkFrame(cont, fg_color="transparent"); self.vuelos_box.pack(fill="x", pady=(2, 6))
        self._pintar_vuelos()

        # Soporte de pago (hasta 3 archivos)
        sh = ctk.CTkFrame(cont, fg_color="transparent"); sh.pack(fill="x", pady=(2, 0))
        ctk.CTkLabel(sh, text="SOPORTE DE PAGO  (adjuntos, hasta 3)", text_color=NAVY,
                     font=("Segoe UI", 12, "bold")).pack(side="left")
        ctk.CTkButton(sh, text="+ Subir soporte", width=150, height=28, fg_color=GREEN,
                      hover_color=GREEN_H, command=self._subir_soporte).pack(side="right")
        self.soportes_box = ctk.CTkFrame(cont, fg_color="transparent"); self.soportes_box.pack(fill="x", pady=(2, 6))
        self._pintar_soportes()

        ctk.CTkLabel(cont, text="Alojamiento / habitaciones (general)", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.v_hab = tk.StringVar(value=res.get("hab", ""))
        ctk.CTkEntry(cont, textvariable=self.v_hab, height=32).pack(fill="x", pady=(0, 8))

        # Estado de la reserva
        fila = ctk.CTkFrame(cont, fg_color="transparent"); fila.pack(fill="x", pady=(10, 4))
        izq = ctk.CTkFrame(fila, fg_color="transparent"); izq.pack(side="left")
        ctk.CTkLabel(izq, text="Estado de la reserva", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w")
        self.v_estado = tk.StringVar(value=res.get("estado", "Confirmada"))
        ctk.CTkOptionMenu(izq, variable=self.v_estado, values=ESTADOS_RES, width=220,
                          height=32, fg_color=NAVY, button_color=NAVY2).pack(anchor="w")

        # Items / servicios cobrados (el total se recalcula solo)
        ih = ctk.CTkFrame(cont, fg_color="transparent"); ih.pack(fill="x", pady=(10, 2))
        ctk.CTkLabel(ih, text="ITEMS DE LA RESERVA  (servicios cobrados)", text_color=NAVY,
                     font=("Segoe UI", 13, "bold")).pack(side="left")
        ctk.CTkButton(ih, text="+ Agregar item", width=130, height=30, fg_color=BLUE,
                      hover_color=BLUE_H, command=self._agregar_item).pack(side="right")
        ctk.CTkButton(ih, text="🏨 Desde tarifario", width=150, height=30, fg_color=NAVY,
                      hover_color=NAVY2, command=self._abrir_tarifario).pack(side="right", padx=(0, 6))
        self.items_box = ctk.CTkFrame(cont, fg_color="transparent"); self.items_box.pack(fill="x")
        tot_bar = ctk.CTkFrame(cont, fg_color=CARD2, corner_radius=8); tot_bar.pack(fill="x", pady=(2, 8))
        self.lbl_total = ctk.CTkLabel(tot_bar, text="Total:  USD 0.00", text_color=NAVY,
                                      font=("Segoe UI", 15, "bold"))
        self.lbl_total.pack(side="right", padx=12, pady=6)
        ctk.CTkLabel(tot_bar, text="El monto negociado = suma de (precio por persona × pasajeros) de cada item.",
                     text_color=MUTED, font=("Segoe UI", 10)).pack(side="left", padx=12)
        self._pintar_items()

        ctk.CTkLabel(cont, text="Notas internas", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", pady=(8, 0))
        self.v_notas = tk.StringVar(value=res.get("notas", ""))
        ctk.CTkEntry(cont, textvariable=self.v_notas, height=32).pack(fill="x", pady=(0, 8))

        # Itinerario de viaje (dia por dia) - EDITOR POR FILAS: Fecha / Actividad / Observaciones
        # Sale tal cual en la tabla "DESCRIPCION DE ACTIVIDADES" de la Orden de Servicio.
        ith = ctk.CTkFrame(cont, fg_color="transparent"); ith.pack(fill="x", pady=(6, 0))
        ctk.CTkLabel(ith, text="ITINERARIO DE VIAJE (dia por dia)", text_color=NAVY,
                     font=("Segoe UI", 13, "bold")).pack(side="left")
        ctk.CTkLabel(ith, text="Sale en la Orden de Servicio del cliente", text_color=MUTED,
                     font=("Segoe UI", 10)).pack(side="left", padx=8)
        ctk.CTkButton(ith, text="+ Agregar dia", width=120, height=28, corner_radius=8,
                      fg_color=BLUE, hover_color=NAVY, font=("Segoe UI", 11, "bold"),
                      command=self._agregar_dia).pack(side="right")
        ctk.CTkButton(ith, text="⟳ Autollenar por fechas", width=180, height=28, corner_radius=8,
                      fg_color=CARD2, text_color=NAVY, hover_color=LINE, border_width=1,
                      border_color=LINE, font=("Segoe UI", 11, "bold"),
                      command=self._autollenar_itinerario).pack(side="right", padx=(0, 6))
        cab = ctk.CTkFrame(cont, fg_color="transparent"); cab.pack(fill="x", pady=(4, 0))
        ctk.CTkLabel(cab, text="Fecha", text_color=MUTED, width=160, anchor="w",
                     font=("Segoe UI", 10, "bold")).pack(side="left", padx=(6, 4))
        ctk.CTkLabel(cab, text="Actividad", text_color=MUTED, anchor="w",
                     font=("Segoe UI", 10, "bold")).pack(side="left", fill="x", expand=True, padx=4)
        ctk.CTkLabel(cab, text="Observaciones", text_color=MUTED, width=210, anchor="w",
                     font=("Segoe UI", 10, "bold")).pack(side="left", padx=4)
        ctk.CTkLabel(cab, text="", width=34).pack(side="left", padx=(4, 6))
        self.itin_widgets = []
        self.itin_box = ctk.CTkFrame(cont, fg_color="transparent"); self.itin_box.pack(fill="x", pady=(0, 8))
        self._pintar_itinerario()

        # ---- Datos de la ORDEN DE SERVICIO (voucher al cliente) ----
        ctk.CTkLabel(cont, text="ORDEN DE SERVICIO (datos del voucher al cliente)",
                     text_color=NAVY, font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(6, 2))
        self.os_vars = {}

        def par(k, lbl):
            v = tk.StringVar(value=res.get(k, ""))
            self.os_vars[k] = v
            return v

        def pareja(k1, l1, k2, l2):
            f = ctk.CTkFrame(cont, fg_color="transparent"); f.pack(fill="x")
            a = ctk.CTkFrame(f, fg_color="transparent"); a.pack(side="left", fill="x", expand=True, padx=(0, 6))
            b = ctk.CTkFrame(f, fg_color="transparent"); b.pack(side="left", fill="x", expand=True, padx=(6, 0))
            ctk.CTkLabel(a, text=l1, text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2)
            ctk.CTkEntry(a, textvariable=par(k1, l1), height=30).pack(fill="x", pady=(0, 5))
            ctk.CTkLabel(b, text=l2, text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2)
            ctk.CTkEntry(b, textvariable=par(k2, l2), height=30).pack(fill="x", pady=(0, 5))

        # Ciudades y hoteles del voucher: MULTIDESTINO (se toman de "Servicios por destino")
        ctk.CTkLabel(cont, text="Ciudades y hoteles del voucher (multidestino)", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.lbl_os_hoteles = ctk.CTkLabel(cont, text="", text_color=NAVY, fg_color=CARD2,
                                           corner_radius=8, anchor="w", justify="left",
                                           font=("Segoe UI", 11))
        self.lbl_os_hoteles.pack(fill="x", pady=(0, 4), ipady=6, ipadx=8)
        ctk.CTkLabel(cont, text="La reserva puede tener varios destinos. Agregalos abajo en "
                     "'SERVICIOS POR DESTINO' con '+ Agregar destino' (hasta 5); el voucher "
                     "listara Ciudad + Hotel de cada uno.", text_color=MUTED,
                     font=("Segoe UI", 10), wraplength=760, justify="left").pack(anchor="w", padx=2, pady=(0, 4))
        ctk.CTkLabel(cont, text="(Las fechas IN/OUT se toman de la llegada y salida de arriba)",
                     text_color=MUTED, font=("Segoe UI", 10)).pack(anchor="w", padx=2)
        pareja("os_habitaciones", "N. Habitaciones", "os_acomodacion", "Acomodacion")
        pareja("os_alimentacion", "Alimentacion", "os_origen", "Origen")
        pareja("os_contacto_principal", "Contacto principal", "os_contacto_secundario", "Segundo contacto")
        pareja("os_vuelo_llegada", "Vuelo de llegada", "os_hora_llegada", "Hora de llegada")
        pareja("os_vuelo_salida", "Vuelo de salida", "os_hora_salida", "Hora de salida")
        pareja("os_vuelo_interno1", "Vuelo interno 1", "os_vuelo_interno2", "Vuelo interno 2")
        ctk.CTkLabel(cont, text="Contacto de emergencia", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        ctk.CTkEntry(cont, textvariable=par("os_contacto_emergencia", ""), height=30).pack(fill="x", pady=(0, 6))

        ctk.CTkLabel(cont, text="Informacion adicional", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.txt_info = ctk.CTkTextbox(cont, height=60, corner_radius=8, border_width=1,
                                       border_color=LINE, fg_color=CARD, font=("Segoe UI", 12))
        self.txt_info.pack(fill="x", pady=(0, 8))
        self.txt_info.insert("1.0", res.get("os_info_adicional", "") or "")

        # Servicios por destino (hasta 5): Hotel, Traslados, Tours
        hdr = ctk.CTkFrame(cont, fg_color="transparent"); hdr.pack(fill="x", pady=(6, 2))
        ctk.CTkLabel(hdr, text="SERVICIOS POR DESTINO (hasta 5)", text_color=NAVY,
                     font=("Segoe UI", 13, "bold")).pack(side="left")
        ctk.CTkButton(hdr, text="+ Agregar destino", width=150, height=30, fg_color=BLUE,
                      hover_color=BLUE_H, command=self._agregar_destino).pack(side="right")
        ctk.CTkLabel(cont, text="En cada destino selecciona el hotel, transporte, guia y actividad. "
                     "Marca el estado de la reserva con cada proveedor y envia su voucher.",
                     text_color=MUTED, font=("Segoe UI", 10), wraplength=740,
                     justify="left").pack(anchor="w", padx=2)
        # Panel de seguimiento (evidencia de gestion de proveedores)
        self.lbl_seg = ctk.CTkLabel(cont, text="", text_color=NAVY, fg_color="#EAF2FD",
                                    corner_radius=8, font=("Segoe UI", 11, "bold"),
                                    anchor="w", justify="left")
        self.lbl_seg.pack(fill="x", pady=(4, 2), ipady=4, ipadx=8)
        ctk.CTkButton(cont, text="📋  Panel de vouchers a proveedores  (enviar todos en un lugar)",
                      height=34, fg_color=NAVY, hover_color=NAVY2, font=("Segoe UI", 12, "bold"),
                      command=self._abrir_panel_vouchers).pack(fill="x", pady=(0, 4))
        self.serv_box = ctk.CTkFrame(cont, fg_color="transparent"); self.serv_box.pack(fill="x")
        self._pintar_servicios()

        # Botones de accion en la barra FIJA superior (self.footer, fondo navy)
        ctk.CTkButton(self.footer, text="💾  Guardar reserva", height=38, corner_radius=10,
                      fg_color=GREEN, hover_color=GREEN_H, font=("Segoe UI", 13, "bold"),
                      command=self._guardar).pack(side="left", padx=(16, 8), pady=9)
        ctk.CTkButton(self.footer, text="Voucher cliente (PDF)", height=38, corner_radius=10,
                      fg_color="#FFFFFF", text_color=NAVY, hover_color="#E7EEF8",
                      font=("Segoe UI", 12, "bold"),
                      command=self._voucher_cliente).pack(side="left", padx=(0, 8), pady=9)
        ctk.CTkButton(self.footer, text="Enviar voucher al cliente", height=38, corner_radius=10,
                      fg_color=CYAN, hover_color=BLUE, font=("Segoe UI", 12, "bold"),
                      command=self._enviar_cliente).pack(side="left", pady=9)
        ctk.CTkButton(self.footer, text="✉ Enviar a la asesora", height=38, corner_radius=10,
                      fg_color="#7A5AB5", hover_color="#63459A", font=("Segoe UI", 12, "bold"),
                      command=self._enviar_asesora).pack(side="left", padx=(8, 0), pady=9)
        ctk.CTkLabel(self.footer, text="Acciones de la reserva", text_color="#BBD0EC",
                     font=("Segoe UI", 11)).pack(side="right", padx=16)

    def _set_asesora(self, nombre):
        """Asigna la asesora (de la lista oficial) al registro de la reserva."""
        if nombre == "(sin asignar)":
            self.res["asesor"] = {}
            return
        info = next((a for a in getattr(self, "_ases_ofic", [])
                     if a.get("nombre") == nombre), None)
        self.res["asesor"] = dict(info) if info else {"nombre": nombre, "email": ""}

    def _enviar_asesora(self):
        # Sincroniza lo editado y notifica por correo a la asesora asignada.
        try:
            self._sync_items()
        except Exception:
            pass
        try:
            self.res["estado"] = self.v_estado.get()
        except Exception:
            pass
        ase = self.res.get("asesor", {}) or {}
        nom = ase.get("nombre", "") if isinstance(ase, dict) else str(ase)
        if not (isinstance(ase, dict) and (ase.get("email") or "").strip()):
            messagebox.showwarning(
                "Sin correo de la asesora",
                f"La asesora asignada ({nom or 'sin asignar'}) no tiene correo configurado.\n"
                "Configuralo en el modulo Reservas > boton 'Asesores'.", parent=self)
            return
        ok, info = notificar_reserva_asesora(self.res, self.cfg)
        if ok:
            messagebox.showinfo("Enviado",
                                f"Se envio la reserva {self.res.get('numero','')} a la asesora "
                                f"{nom} ({info}).", parent=self)
        else:
            messagebox.showerror(
                "No se pudo enviar",
                f"No se pudo enviar el correo a la asesora: {info}\n\n"
                "Verifica el correo remitente y su contrasena en 'Datos de mi empresa'.",
                parent=self)

    # ------------------------------------------------- itinerario (editor por filas)
    def _dias_iniciales(self):
        d = self.res.get("itinerario_dias")
        if isinstance(d, list) and d:
            return [dict(x) for x in d]
        filas = _parse_actividades(self.res.get("os_actividades", ""),
                                   self.res.get("itinerario", ""))
        dias = [{"fecha": f, "actividad": a, "obs": o} for f, a, o in filas]
        return dias or [{"fecha": "", "actividad": "", "obs": ""}]

    def _pintar_itinerario(self):
        for w in self.itin_box.winfo_children():
            w.destroy()
        self.itin_widgets = []
        dias = self.res.get("itinerario_dias")
        if not isinstance(dias, list) or not dias:
            dias = self._dias_iniciales()
            self.res["itinerario_dias"] = dias
        for i, d in enumerate(dias):
            self._fila_dia(i, d)

    def _fila_dia(self, i, d):
        row = ctk.CTkFrame(self.itin_box, fg_color=CARD, corner_radius=8,
                           border_width=1, border_color=LINE)
        row.pack(fill="x", pady=2)
        vf = tk.StringVar(value=d.get("fecha", ""))
        va = tk.StringVar(value=d.get("actividad", ""))
        vo = tk.StringVar(value=d.get("obs", ""))
        ctk.CTkEntry(row, textvariable=vf, width=160, height=30,
                     placeholder_text="30 DE AGOSTO DE 2026").pack(side="left", padx=(6, 4), pady=5)
        ctk.CTkEntry(row, textvariable=va, height=30,
                     placeholder_text="Actividad del dia (ej. City tour + comuna 13)").pack(
            side="left", fill="x", expand=True, padx=4)
        ctk.CTkEntry(row, textvariable=vo, width=210, height=30,
                     placeholder_text="Hora / notas").pack(side="left", padx=4)
        ctk.CTkButton(row, text="✕", width=28, height=28, fg_color=RED, hover_color="#9B2C22",
                      command=lambda idx=i: self._quitar_dia(idx)).pack(side="left", padx=(4, 6))
        self.itin_widgets.append({"fecha": vf, "actividad": va, "obs": vo})

    def _sync_itinerario(self):
        dias = []
        for w in self.itin_widgets:
            f = w["fecha"].get().strip(); a = w["actividad"].get().strip(); o = w["obs"].get().strip()
            if f or a or o:
                dias.append({"fecha": f, "actividad": a, "obs": o})
        self.res["itinerario_dias"] = dias
        # os_actividades alimenta la tabla del voucher; itinerario = texto legado
        self.res["os_actividades"] = "\n".join(
            f"{d['fecha']} | {d['actividad']} | {d['obs']}" for d in dias)
        self.res["itinerario"] = "\n".join(
            (f"{d['fecha']} - {d['actividad']}" + (f" ({d['obs']})" if d['obs'] else ""))
            for d in dias)

    def _agregar_dia(self):
        self._sync_itinerario()
        self.res.setdefault("itinerario_dias", []).append({"fecha": "", "actividad": "", "obs": ""})
        self._pintar_itinerario()

    def _quitar_dia(self, i):
        self._sync_itinerario()
        try:
            del self.res["itinerario_dias"][i]
        except Exception:
            pass
        self._pintar_itinerario()

    def _autollenar_itinerario(self):
        self._sync_itinerario()
        try:
            d1 = self.sel_llegada.get(); d2 = self.sel_salida.get()
        except Exception:
            d1 = d2 = None
        if not d1 or not d2:
            messagebox.showinfo("Fechas del viaje",
                                "Primero elige la fecha de llegada y de salida (arriba).",
                                parent=self)
            return
        if d2 < d1:
            messagebox.showinfo("Fechas del viaje",
                                "La fecha de salida no puede ser anterior a la de llegada.",
                                parent=self)
            return
        prev = self.res.get("itinerario_dias", []) or []
        nuevos = []
        dact = d1; i = 0
        while dact <= d2:
            act = prev[i].get("actividad", "") if i < len(prev) else ""
            obs = prev[i].get("obs", "") if i < len(prev) else ""
            nuevos.append({"fecha": fecha_larga_es(dact), "actividad": act, "obs": obs})
            dact += datetime.timedelta(days=1); i += 1
        self.res["itinerario_dias"] = nuevos
        self._pintar_itinerario()

    def _pintar_pasajeros(self):
        for w in self.pax_box.winfo_children():
            w.destroy()
        self.pax_widgets = []
        lst = self.res.get("pasajeros_list")
        if not isinstance(lst, list) or not lst:
            # migrar desde texto os_pasajeros si existiera
            lst = [{"nombre": n, "documento": d}
                   for n, d in _parse_pasajeros(self.res.get("os_pasajeros", ""))]
            self.res["pasajeros_list"] = lst
        if not lst:
            ctk.CTkLabel(self.pax_box, text="Sin pasajeros. Usa '+ Agregar pasajero' para "
                         "anadir nombre y pasaporte.", text_color=MUTED,
                         font=("Segoe UI", 10)).pack(pady=4)
        for i, p in enumerate(lst):
            self._fila_pasajero(i, p)

    def _fila_pasajero(self, i, p):
        row = ctk.CTkFrame(self.pax_box, fg_color="transparent"); row.pack(fill="x", pady=2)
        v_nom = tk.StringVar(value=p.get("nombre", ""))
        v_doc = tk.StringVar(value=p.get("documento", ""))
        v_tel = tk.StringVar(value=p.get("telefono", ""))
        adj = p.get("adjunto", "")
        ctk.CTkEntry(row, textvariable=v_nom, height=30,
                     placeholder_text="Nombre completo del pasajero").pack(
            side="left", fill="x", expand=True, padx=(0, 6))
        ctk.CTkEntry(row, textvariable=v_doc, width=140, height=30,
                     placeholder_text="Pasaporte / documento").pack(side="left", padx=(0, 6))
        ctk.CTkEntry(row, textvariable=v_tel, width=120, height=30,
                     placeholder_text="Telefono").pack(side="left")
        ctk.CTkButton(row, text="✕", width=30, height=30, fg_color=RED, hover_color="#9B2C22",
                      command=lambda idx=i: self._quitar_pasajero(idx)).pack(side="right", padx=(6, 0))
        if adj:
            nombre_arch = os.path.basename(adj)
            ctk.CTkButton(row, text="🗑", width=30, height=30, fg_color="#9B2C22",
                          hover_color=RED,
                          command=lambda idx=i: self._quitar_adjunto_pax(idx)).pack(side="right", padx=(4, 0))
            ctk.CTkButton(row, text="📄 " + nombre_arch[:12], width=140, height=30, fg_color=GREEN,
                          hover_color=GREEN_H, font=("Segoe UI", 10),
                          command=lambda a=adj: self._abrir_archivo(a)).pack(side="right", padx=(6, 0))
        else:
            ctk.CTkButton(row, text="📎 Pasaporte", width=110, height=30, fg_color=BLUE,
                          hover_color=BLUE_H,
                          command=lambda idx=i: self._adjuntar_pasaporte(idx)).pack(side="right", padx=(6, 0))
        self.pax_widgets.append({"nombre": v_nom, "documento": v_doc, "telefono": v_tel,
                                 "adjunto": adj})

    def _sync_pax(self):
        # Conserva todas las filas (aunque esten vacias) para no perder pasajeros
        # recien agregados; el filtrado de vacios se hace al generar el voucher.
        self.res["pasajeros_list"] = [
            {"nombre": w["nombre"].get().strip(), "documento": w["documento"].get().strip(),
             "telefono": w["telefono"].get().strip(), "adjunto": w.get("adjunto", "")}
            for w in self.pax_widgets]

    def _agregar_pasajero(self):
        self._sync_pax()
        self.res.setdefault("pasajeros_list", []).append(
            {"nombre": "", "documento": "", "telefono": "", "adjunto": ""})
        self._pintar_pasajeros()

    def _quitar_pasajero(self, i):
        self._sync_pax()
        try:
            del self.res["pasajeros_list"][i]
        except Exception:
            pass
        self._pintar_pasajeros()

    def _carpeta_adjuntos(self):
        ruta = os.path.join(datos_dir(), "adjuntos", self.res.get("numero", "tmp"))
        os.makedirs(ruta, exist_ok=True)
        return ruta

    def _copiar_adjunto(self, ruta, prefijo=""):
        base = os.path.basename(ruta)
        destino = os.path.join(self._carpeta_adjuntos(), (prefijo + "_" if prefijo else "") + base)
        try:
            shutil.copyfile(ruta, destino)
        except Exception as e:
            messagebox.showerror("No se pudo adjuntar", str(e)); return ""
        return destino

    def _abrir_archivo(self, ruta):
        try:
            os.startfile(ruta)
        except Exception as e:
            messagebox.showerror("No se pudo abrir", str(e))

    def _adjuntar_pasaporte(self, i):
        self._sync_pax()
        ruta = filedialog.askopenfilename(
            title="Adjuntar pasaporte del pasajero",
            filetypes=[("Documentos", "*.pdf *.jpg *.jpeg *.png"), ("Todos", "*.*")])
        if not ruta:
            return
        destino = self._copiar_adjunto(ruta, prefijo=f"pasaporte_{i+1}")
        if not destino:
            return
        self.res["pasajeros_list"][i]["adjunto"] = destino
        actualizar_reserva(self.res.get("numero", ""),
                           {"pasajeros_list": self.res["pasajeros_list"]})
        self._pintar_pasajeros()

    def _quitar_adjunto_pax(self, i):
        self._sync_pax()
        try:
            self.res["pasajeros_list"][i]["adjunto"] = ""
        except Exception:
            pass
        actualizar_reserva(self.res.get("numero", ""),
                           {"pasajeros_list": self.res["pasajeros_list"]})
        self._pintar_pasajeros()

    def _pintar_vuelos(self):
        for w in self.vuelos_box.winfo_children():
            w.destroy()
        vuelos = self.res.setdefault("vuelos_adjuntos", [])
        if not vuelos:
            ctk.CTkLabel(self.vuelos_box, text="Sin vuelos adjuntos. Usa '+ Adjuntar vuelo' para "
                         "subir tiquetes o itinerarios de vuelo.", text_color=MUTED,
                         font=("Segoe UI", 10)).pack(pady=4)
        for i, v in enumerate(vuelos):
            row = ctk.CTkFrame(self.vuelos_box, fg_color="transparent"); row.pack(fill="x", pady=2)
            ctk.CTkButton(row, text="✈ " + os.path.basename(v), height=30, fg_color=CARD2,
                          text_color=NAVY, hover_color=LINE, anchor="w",
                          command=lambda a=v: self._abrir_archivo(a)).pack(
                side="left", fill="x", expand=True, padx=(0, 6))
            ctk.CTkButton(row, text="🗑", width=32, height=30, fg_color=RED, hover_color="#9B2C22",
                          command=lambda idx=i: self._quitar_vuelo(idx)).pack(side="left")

    def _adjuntar_vuelo(self):
        rutas = filedialog.askopenfilenames(
            title="Adjuntar vuelo(s) / tiquete(s)",
            filetypes=[("Documentos", "*.pdf *.jpg *.jpeg *.png"), ("Todos", "*.*")])
        if not rutas:
            return
        vuelos = self.res.setdefault("vuelos_adjuntos", [])
        for r in rutas:
            destino = self._copiar_adjunto(r, prefijo="vuelo")
            if destino:
                vuelos.append(destino)
        actualizar_reserva(self.res.get("numero", ""), {"vuelos_adjuntos": vuelos})
        self._pintar_vuelos()

    def _quitar_vuelo(self, i):
        try:
            del self.res["vuelos_adjuntos"][i]
        except Exception:
            pass
        actualizar_reserva(self.res.get("numero", ""),
                           {"vuelos_adjuntos": self.res.get("vuelos_adjuntos", [])})
        self._pintar_vuelos()

    def _pintar_soportes(self):
        for w in self.soportes_box.winfo_children():
            w.destroy()
        soportes = self.res.setdefault("soportes_pago", [])
        if not soportes:
            ctk.CTkLabel(self.soportes_box, text="Sin soportes de pago. Usa '+ Subir soporte' "
                         "para adjuntar el comprobante (hasta 3).", text_color=MUTED,
                         font=("Segoe UI", 10)).pack(pady=4)
        for i, v in enumerate(soportes):
            row = ctk.CTkFrame(self.soportes_box, fg_color="transparent"); row.pack(fill="x", pady=2)
            ctk.CTkButton(row, text="💳 " + os.path.basename(v), height=30, fg_color="#E3F5EA",
                          text_color=GREEN_H, hover_color=LINE, anchor="w",
                          command=lambda a=v: self._abrir_archivo(a)).pack(
                side="left", fill="x", expand=True, padx=(0, 6))
            ctk.CTkButton(row, text="🗑", width=32, height=30, fg_color=RED, hover_color="#9B2C22",
                          command=lambda idx=i: self._quitar_soporte(idx)).pack(side="left")

    def _subir_soporte(self):
        soportes = self.res.setdefault("soportes_pago", [])
        if len(soportes) >= 3:
            messagebox.showinfo("Soporte de pago", "Ya hay 3 soportes. Quita alguno para subir otro.")
            return
        rutas = filedialog.askopenfilenames(
            title="Subir soporte(s) de pago",
            filetypes=[("Documentos", "*.pdf *.jpg *.jpeg *.png"), ("Todos", "*.*")])
        if not rutas:
            return
        for r in rutas:
            if len(soportes) >= 3:
                messagebox.showinfo("Soporte de pago", "Solo se permiten 3 soportes; se omitieron los demas.")
                break
            destino = self._copiar_adjunto(r, prefijo="pago")
            if destino:
                soportes.append(destino)
        actualizar_reserva(self.res.get("numero", ""), {"soportes_pago": soportes})
        self._pintar_soportes()

    def _quitar_soporte(self, i):
        try:
            del self.res["soportes_pago"][i]
        except Exception:
            pass
        actualizar_reserva(self.res.get("numero", ""),
                           {"soportes_pago": self.res.get("soportes_pago", [])})
        self._pintar_soportes()

    def _pintar_servicios(self):
        for w in self.serv_box.winfo_children():
            w.destroy()
        self.serv_widgets = []      # (di, cat, si, vars)
        self.dest_nom_vars = {}     # di -> StringVar del nombre del destino
        dd = destinos_detalle_de(self.res)
        if not dd:
            ctk.CTkLabel(self.serv_box, text="Sin destinos. Usa '+ Agregar destino' para "
                         "anadir un destino con su hotel, transporte, guia y actividad.",
                         text_color=MUTED).pack(pady=8)
        for di, dest in enumerate(dd):
            self._card_destino(di, dest)
        self._refrescar_resumen()
        self._refrescar_os_hoteles()

    def _pintar_items(self):
        for w in self.items_box.winfo_children():
            w.destroy()
        self.item_widgets = []
        items = self.res.setdefault("items_cobro", [])
        # Migracion: si no hay items pero hay monto, sembrar un item con ese monto
        if not items and float(self.res.get("monto", 0) or 0) > 0:
            items.append({"desc": "Servicios de la reserva",
                          "valor": float(self.res.get("monto", 0) or 0), "cant": 1})
        if not items:
            ctk.CTkLabel(self.items_box, text="Sin items. Usa '+ Agregar item' para ingresar "
                         "los servicios cobrados (descripcion y valor).", text_color=MUTED,
                         font=("Segoe UI", 10)).pack(pady=4)
        for i, it in enumerate(items):
            self._fila_item(i, it)
        self._recalcular_total()

    def _fila_item(self, i, it):
        row = ctk.CTkFrame(self.items_box, fg_color=CARD, corner_radius=8,
                           border_width=1, border_color=LINE)
        row.pack(fill="x", pady=2)
        v_desc = tk.StringVar(value=it.get("desc", ""))
        v_val = tk.StringVar(value=f"{float(it.get('valor', 0) or 0):.2f}")
        # cantidad de pasajeros: si el item no la trae, se asume el # de pasajeros de la
        # reserva cuando el precio es "por persona" (p.p.), o 1 si es un valor total.
        cant0 = it.get("cant")
        if cant0 in (None, ""):
            cant0 = self._grupo_reserva() if "p.p." in (it.get("desc", "") or "").lower() else 1
        try:
            cant0 = max(1, int(float(cant0)))
        except Exception:
            cant0 = 1
        v_cant = tk.StringVar(value=str(cant0))
        ctk.CTkEntry(row, textvariable=v_desc, height=30,
                     placeholder_text="Descripcion del servicio").pack(
            side="left", fill="x", expand=True, padx=(6, 4), pady=5)
        ctk.CTkLabel(row, text="USD p.p.", text_color=MUTED, font=("Segoe UI", 10)).pack(side="left")
        e = ctk.CTkEntry(row, textvariable=v_val, width=84, height=30, placeholder_text="0.00")
        e.pack(side="left", padx=2)
        e.bind("<KeyRelease>", lambda ev: self._recalcular_total())
        ctk.CTkLabel(row, text="× pax", text_color=MUTED, font=("Segoe UI", 10)).pack(side="left")
        ec = ctk.CTkEntry(row, textvariable=v_cant, width=46, height=30, placeholder_text="2")
        ec.pack(side="left", padx=2)
        ec.bind("<KeyRelease>", lambda ev: self._recalcular_total())
        lbl_sub = ctk.CTkLabel(row, text="= USD 0.00", text_color=GREEN_H, width=118,
                               anchor="e", font=("Segoe UI", 11, "bold"))
        lbl_sub.pack(side="left", padx=4)
        ctk.CTkButton(row, text="✕", width=28, height=28, fg_color=RED, hover_color="#9B2C22",
                      command=lambda idx=i: self._quitar_item(idx)).pack(side="left", padx=(4, 6))
        self.item_widgets.append({"desc": v_desc, "valor": v_val, "cant": v_cant, "sub": lbl_sub})

    @staticmethod
    def _num_pos(sv, defecto):
        try:
            v = float(str(sv.get()).replace(",", "").strip() or defecto)
            return v if v > 0 else defecto
        except Exception:
            return defecto

    def _recalcular_total(self):
        tot = 0.0
        for w in self.item_widgets:
            try:
                val = float(str(w["valor"].get()).replace(",", "").strip() or 0)
            except Exception:
                val = 0.0
            cant = self._num_pos(w["cant"], 1)
            sub = val * cant
            tot += sub
            try:
                w["sub"].configure(text=f"= USD {sub:,.2f}")
            except Exception:
                pass
        self.res["monto"] = round(tot, 2)
        try:
            self.lbl_total.configure(text=f"Total:  USD {tot:,.2f}")
        except Exception:
            pass

    def _sync_items(self):
        items = []
        for w in self.item_widgets:
            desc = w["desc"].get().strip()
            try:
                val = float(str(w["valor"].get()).replace(",", "").strip() or 0)
            except Exception:
                val = 0.0
            cant = int(self._num_pos(w["cant"], 1))
            if desc or val:
                items.append({"desc": desc, "valor": val, "cant": cant})
        self.res["items_cobro"] = items
        self.res["monto"] = round(sum(it["valor"] * it.get("cant", 1) for it in items), 2)

    def _agregar_item(self):
        self._sync_items()
        self.res.setdefault("items_cobro", []).append({"desc": "", "valor": 0.0, "cant": 1})
        self._pintar_items()

    def _noches_reserva(self):
        try:
            d1 = self.sel_llegada.get(); d2 = self.sel_salida.get()
            if d1 and d2:
                return max(1, (d2 - d1).days)
        except Exception:
            pass
        return 1

    def _grupo_reserva(self):
        n = len([p for p in self.res.get("pasajeros_list", [])
                 if (p.get("nombre") or p.get("documento"))])
        if n >= 1:
            return n
        txt = self.v_pax.get() if hasattr(self, "v_pax") else ""
        num = ""
        for ch in txt:
            if ch.isdigit():
                num += ch
            elif num:
                break
        return int(num) if num else 2

    def _abrir_tarifario(self):
        self._sync_items()
        dd = destinos_detalle_de(self.res)
        destino0 = dd[0].get("nombre", "") if dd else ""
        SelectorTarifario(self, self.precios, self.cfg, destino0,
                          self._noches_reserva(), self._grupo_reserva(),
                          on_add=self._add_item_tarifa)

    def _add_item_tarifa(self, desc, valor):
        self._sync_items()
        # Los precios del tarifario son POR PERSONA: se multiplican por el # de pasajeros.
        self.res.setdefault("items_cobro", []).append(
            {"desc": desc, "valor": round(float(valor or 0), 2), "cant": self._grupo_reserva()})
        self._pintar_items()

    def _quitar_item(self, i):
        self._sync_items()
        try:
            del self.res["items_cobro"][i]
        except Exception:
            pass
        self._pintar_items()

    def _maximizar(self):
        ok = False
        try:
            self.state("zoomed")
            ok = self.state() == "zoomed"
        except Exception:
            ok = False
        if not ok:
            # Respaldo: ajustar al area util (sin barra de tareas)
            try:
                wah = _alto_util_pantalla(fallback=(self.winfo_screenheight() - 70))
                alto = max(480, wah - 70)
                sw = self.winfo_screenwidth()
                self.geometry(f"{min(1200, sw - 60)}x{alto}+30+8")
            except Exception:
                pass

    def _refrescar_os_hoteles(self):
        if not hasattr(self, "lbl_os_hoteles"):
            return
        try:
            dd = destinos_detalle_de(self.res)
            lineas = []
            for d in dd:
                nom = d.get("nombre", "") or "(destino sin nombre)"
                hoteles = [h.get("servicio", "") for h in d.get("hotel", []) if h.get("servicio")]
                lineas.append(f"•  {nom}:  " + (" / ".join(hoteles) if hoteles else "(sin hotel)"))
            txt = ("\n".join(lineas) if lineas else
                   "Aun no hay destinos. Agrega uno abajo en 'Servicios por destino'.")
            self.lbl_os_hoteles.configure(text=txt)
        except Exception:
            pass

    def _buscar_cliente_res(self):
        SelectorContacto(self, self._usar_cliente_res)

    def _usar_cliente_res(self, c, v):
        self.v_cli.set(c.get("empresa", ""))
        vends = c.get("vendedores", []) or []
        nombres = [x.get("nombre", "") for x in vends if x.get("nombre")]
        try:
            self.combo_contacto.configure(values=nombres or [""])
        except Exception:
            pass
        if v and v.get("nombre"):
            self.v_contacto.set(v.get("nombre", ""))
            if v.get("email"):
                self.v_email.set(v.get("email", ""))
            elif c.get("email"):
                self.v_email.set(c.get("email", ""))
        else:
            if nombres:
                self.v_contacto.set(nombres[0])
            if c.get("email"):
                self.v_email.set(c.get("email", ""))

    def _abrir_panel_vouchers(self):
        self._sync()
        VentanaVouchersProveedores(self, self.res, self.cfg, on_change=self._pintar_servicios)

    def _refrescar_resumen(self):
        try:
            total, enviados, por_estado = resumen_seguimiento(self.res)
        except Exception:
            return
        if total == 0:
            txt = "Seguimiento de proveedores: aun no hay servicios."
        else:
            txt = (f"Seguimiento de proveedores:  {enviados}/{total} con voucher enviado   ·   "
                   f"✓ {por_estado.get('Reservado con pago', 0)} reservado con pago   ·   "
                   f"◐ {por_estado.get('Reservado sin pago', 0)} sin pago   ·   "
                   f"○ {por_estado.get('Pendiente', 0)} pendiente")
        try:
            self.lbl_seg.configure(text=txt)
        except Exception:
            pass

    def _on_estado_prov(self, val, menu, di, cat, si):
        try:
            self.res["destinos_detalle"][di][cat][si]["estado_prov"] = val
        except Exception:
            pass
        try:
            menu.configure(fg_color=ESTADO_PROV_COLOR.get(val, MUTED))
        except Exception:
            pass
        self._refrescar_resumen()

    def _card_destino(self, di, dest):
        card = ctk.CTkFrame(self.serv_box, fg_color=CARD2, corner_radius=12,
                            border_width=1, border_color=LINE)
        card.pack(fill="x", pady=6)
        head = ctk.CTkFrame(card, fg_color="transparent"); head.pack(fill="x", padx=8, pady=(8, 2))
        ctk.CTkLabel(head, text=f"Destino {di+1}", text_color=NAVY,
                     font=("Segoe UI", 13, "bold")).pack(side="left")
        v_nom = tk.StringVar(value=dest.get("nombre", ""))
        self.dest_nom_vars[di] = v_nom
        ctk.CTkEntry(head, textvariable=v_nom, width=230, height=30,
                     placeholder_text="Ciudad / destino").pack(side="left", padx=8)
        ctk.CTkButton(head, text="Quitar destino", width=120, height=30, fg_color=RED,
                      hover_color="#9B2C22",
                      command=lambda i=di: self._quitar_destino(i)).pack(side="right")
        for cat, etiqueta, _tipo in CATEGORIAS_SERV:
            sec = ctk.CTkFrame(card, fg_color="transparent"); sec.pack(fill="x", padx=10, pady=(4, 0))
            ctk.CTkLabel(sec, text=etiqueta, text_color=BLUE,
                         font=("Segoe UI", 11, "bold")).pack(side="left")
            ctk.CTkButton(sec, text="+ " + etiqueta.split(" ")[0], width=90, height=26,
                          fg_color=BLUE, hover_color=BLUE_H,
                          command=lambda i=di, c=cat: self._agregar_servicio(i, c)).pack(side="right")
            # Transporte y Guia: UN solo voucher por proveedor con TODOS sus servicios
            if cat in ("transporte", "guia"):
                ctk.CTkButton(sec, text="✉ Enviar (todos)", width=120, height=26, fg_color=GREEN,
                              hover_color=GREEN_H, font=("Segoe UI", 10, "bold"),
                              command=lambda i=di, c=cat: self._enviar_grupo(i, c)).pack(side="right", padx=4)
                ctk.CTkButton(sec, text="📄 Voucher (todos)", width=130, height=26, fg_color=NAVY,
                              hover_color=NAVY2, font=("Segoe UI", 10, "bold"),
                              command=lambda i=di, c=cat: self._voucher_grupo(i, c)).pack(side="right", padx=4)
            for si, s in enumerate(dest.get(cat, [])):
                self._fila_servicio(card, di, cat, si, s)
        ctk.CTkFrame(card, fg_color="transparent", height=4).pack()

    def _fila_servicio(self, parent, di, cat, si, s):
        row = ctk.CTkFrame(parent, fg_color=CARD, corner_radius=8,
                           border_width=1, border_color=LINE)
        row.pack(fill="x", padx=12, pady=3)
        v = {"servicio": tk.StringVar(value=s.get("servicio", "")),
             "proveedor": tk.StringVar(value=s.get("proveedor", "")),
             "correo": tk.StringVar(value=s.get("correo", "")),
             "estado_prov": tk.StringVar(value=s.get("estado_prov", "Pendiente")),
             "hora": tk.StringVar(value=s.get("hora", "")),
             "origen": tk.StringVar(value=s.get("origen", "")),
             "vehiculo": tk.StringVar(value=s.get("vehiculo", "")),
             "observacion": tk.StringVar(value=s.get("observacion", ""))}
        # Linea 1: servicio + estado de la reserva con el proveedor (seguimiento)
        l1 = ctk.CTkFrame(row, fg_color="transparent"); l1.pack(fill="x", padx=6, pady=(5, 2))
        if cat == "hotel":
            nombre_dest = ""
            try:
                nombre_dest = self.res["destinos_detalle"][di].get("nombre", "")
            except Exception:
                pass
            opciones = hoteles_por_destino(self.precios, nombre_dest)
            if opciones:
                ctk.CTkComboBox(l1, variable=v["servicio"], values=opciones, height=28,
                                dropdown_font=("Segoe UI", 11)).pack(side="left", fill="x", expand=True)
            else:
                ctk.CTkEntry(l1, textvariable=v["servicio"], height=28,
                             placeholder_text="Hotel").pack(side="left", fill="x", expand=True)
        else:
            ctk.CTkEntry(l1, textvariable=v["servicio"], height=28,
                         placeholder_text="Servicio").pack(side="left", fill="x", expand=True)
        om = ctk.CTkOptionMenu(l1, variable=v["estado_prov"], values=ESTADOS_PROV, width=190, height=28,
                               fg_color=ESTADO_PROV_COLOR.get(s.get("estado_prov", "Pendiente"), MUTED),
                               button_color=NAVY2)
        om.configure(command=lambda val, o=om, dd=di, cc=cat, ss=si:
                     self._on_estado_prov(val, o, dd, cc, ss))
        om.pack(side="left", padx=(6, 0))
        # Linea 2: proveedor + correo
        l2 = ctk.CTkFrame(row, fg_color="transparent"); l2.pack(fill="x", padx=6, pady=(0, 2))
        ctk.CTkLabel(l2, text="Proveedor", text_color=MUTED, width=62).pack(side="left")
        ctk.CTkEntry(l2, textvariable=v["proveedor"], placeholder_text="Nombre del proveedor",
                     width=190, height=28).pack(side="left", padx=4)
        ctk.CTkLabel(l2, text="Correo", text_color=MUTED, width=44).pack(side="left")
        ctk.CTkEntry(l2, textvariable=v["correo"], placeholder_text="correo@proveedor.com",
                     height=28).pack(side="left", padx=4, fill="x", expand=True)
        # Linea 3: datos del voucher (fecha / hora / origen / vehiculo / observacion)
        l3 = ctk.CTkFrame(row, fg_color="transparent"); l3.pack(fill="x", padx=6, pady=(0, 2))
        ctk.CTkLabel(l3, text="Fecha", text_color=MUTED, width=38).pack(side="left")
        sel_f = SelectorFecha(l3)
        sel_f.btn.configure(height=28, font=("Segoe UI", 11))
        sel_f.pack(side="left", padx=(0, 6))
        f0 = parse_fecha(s.get("fecha", ""))
        if f0:
            sel_f._set(f0)
        v["fecha_sel"] = sel_f
        ctk.CTkLabel(l3, text="Hora", text_color=MUTED, width=34).pack(side="left")
        ctk.CTkEntry(l3, textvariable=v["hora"], width=80, height=28,
                     placeholder_text="8:00 am").pack(side="left", padx=(0, 6))
        if cat == "transporte":
            ctk.CTkLabel(l3, text="Origen", text_color=MUTED, width=46).pack(side="left")
            ctk.CTkEntry(l3, textvariable=v["origen"], width=130, height=28,
                         placeholder_text="Aeropuerto...").pack(side="left", padx=(0, 6))
            ctk.CTkLabel(l3, text="Vehiculo", text_color=MUTED, width=54).pack(side="left")
            ctk.CTkEntry(l3, textvariable=v["vehiculo"], width=110, height=28,
                         placeholder_text="Van / Bus").pack(side="left", padx=(0, 6))
        ctk.CTkLabel(l3, text="Detalle", text_color=MUTED, width=48).pack(side="left")
        ctk.CTkEntry(l3, textvariable=v["observacion"], height=28,
                     placeholder_text="Incluye / observaciones").pack(side="left", fill="x", expand=True)
        # Linea 4: evidencia + acciones
        l4 = ctk.CTkFrame(row, fg_color="transparent"); l4.pack(fill="x", padx=6, pady=(0, 5))
        est = ctk.CTkLabel(l4, text=("✓ Voucher enviado " + s.get("fecha_envio", "")
                                     if s.get("enviado") else "Voucher sin enviar"),
                           text_color=(GREEN if s.get("enviado") else MUTED),
                           font=("Segoe UI", 10, "bold" if s.get("enviado") else "normal"))
        est.pack(side="left")
        ctk.CTkButton(l4, text="Quitar", width=58, height=26, fg_color=RED, hover_color="#9B2C22",
                      command=lambda: self._quitar_servicio(di, cat, si)).pack(side="right", padx=3)
        ctk.CTkButton(l4, text="Enviar al proveedor", width=140, height=26, fg_color=GREEN,
                      hover_color=GREEN_H,
                      command=lambda l=est: self._enviar_serv(di, cat, si, l)).pack(side="right", padx=3)
        ctk.CTkButton(l4, text="Generar voucher", width=124, height=26, fg_color=BLUE,
                      hover_color=BLUE_H,
                      command=lambda: self._voucher_serv(di, cat, si)).pack(side="right", padx=3)
        self.serv_widgets.append((di, cat, si, v))

    def _sync_serv(self):
        dd = self.res.setdefault("destinos_detalle", [])
        for di, cat, si, w in self.serv_widgets:
            try:
                s = dd[di][cat][si]
                for k in ("servicio", "proveedor", "correo", "estado_prov",
                          "hora", "origen", "vehiculo", "observacion"):
                    s[k] = w[k].get().strip()
                if w.get("fecha_sel") is not None:
                    s["fecha"] = w["fecha_sel"].get_str()
            except Exception:
                pass
        for di, val in self.dest_nom_vars.items():
            try:
                dd[di]["nombre"] = val.get().strip()
            except Exception:
                pass
        self.res["destinos"] = [d.get("nombre", "") for d in dd if d.get("nombre", "").strip()]

    def _sync(self):
        self._sync_serv()
        self._sync_pax()
        self.res["cliente"] = self.v_cli.get().strip()
        self.res["email"] = self.v_email.get().strip()
        self.res["contacto"] = self.v_contacto.get().strip()
        ll = self.sel_llegada.get_str(); sa = self.sel_salida.get_str()
        self.res["os_fecha_in"] = ll
        self.res["os_fecha_out"] = sa
        self.res["fechas_viaje"] = (ll + " al " + sa) if (ll and sa) else (ll or sa)
        self.res["pax_txt"] = self.v_pax.get().strip()
        self.res["hab"] = self.v_hab.get().strip()
        self.res["estado"] = self.v_estado.get()
        self._sync_items()
        self.res["notas"] = self.v_notas.get().strip()
        try:
            self._sync_itinerario()
            for k, v in self.os_vars.items():
                self.res[k] = v.get().strip()
            self.res["os_info_adicional"] = self.txt_info.get("1.0", "end").strip()
        except Exception:
            pass

    def _agregar_destino(self):
        self._sync()
        dd = self.res.setdefault("destinos_detalle", [])
        if len(dd) >= MAX_DESTINOS_RES:
            messagebox.showinfo("Limite", f"Maximo {MAX_DESTINOS_RES} destinos por reserva.")
            return
        dd.append(_destino_vacio())
        self._pintar_servicios()

    def _quitar_destino(self, di):
        dd = self.res.get("destinos_detalle", [])
        if di >= len(dd):
            return
        nom = dd[di].get("nombre", "") or f"Destino {di+1}"
        if messagebox.askyesno("Quitar destino",
                               f"Quitar '{nom}' y todos sus servicios de la reserva?"):
            self._sync()
            try:
                del self.res["destinos_detalle"][di]
            except Exception:
                pass
            self._pintar_servicios()

    def _agregar_servicio(self, di, cat):
        self._sync()
        try:
            self.res["destinos_detalle"][di][cat].append(_servicio_vacio())
        except Exception:
            pass
        self._pintar_servicios()

    def _quitar_servicio(self, di, cat, si):
        self._sync()
        try:
            del self.res["destinos_detalle"][di][cat][si]
        except Exception:
            pass
        self._pintar_servicios()

    def _carpeta_vouchers(self):
        ruta = os.path.join(datos_dir(), "vouchers")
        os.makedirs(ruta, exist_ok=True)
        return ruta

    def _reng_de(self, di, cat, si):
        tipo = next((t for k, _l, t in CATEGORIAS_SERV if k == cat), cat)
        dest = self.res["destinos_detalle"][di]
        s = dest[cat][si]
        reng = {"tipo": tipo, "destino": dest.get("nombre", ""),
                "servicio": s.get("servicio", ""), "proveedor": s.get("proveedor", ""),
                "correo": s.get("correo", ""), "hora": s.get("hora", ""),
                "origen": s.get("origen", ""), "vehiculo": s.get("vehiculo", ""),
                "observacion": s.get("observacion", "")}
        return reng, s

    def _voucher_serv(self, di, cat, si, abrir=True):
        self._sync()
        try:
            reng, s = self._reng_de(di, cat, si)
        except Exception:
            return None
        if not (reng["proveedor"] or reng["servicio"]):
            messagebox.showinfo("Datos incompletos",
                                "Escribe al menos el proveedor o el servicio.")
            return None
        fn = os.path.join(self._carpeta_vouchers(),
                          f"Voucher_prov_{self.res.get('numero','')}_{di+1}_{cat}_{si+1}.pdf")
        try:
            generar_voucher_proveedor(self.cfg, self.res, reng, fn)
        except Exception as e:
            messagebox.showerror("Error al generar el voucher", str(e))
            return None
        if abrir:
            try:
                os.startfile(fn)
            except Exception:
                pass
        return fn

    def _enviar_serv(self, di, cat, si, lbl):
        fn = self._voucher_serv(di, cat, si, abrir=False)
        if not fn:
            return
        reng, s = self._reng_de(di, cat, si)
        if not reng["correo"]:
            messagebox.showinfo("Correo del proveedor",
                                "Escribe el correo del proveedor para enviarle el voucher.")
            return
        try:
            asunto = (f"Reserva {self.res.get('numero','')} - {reng['servicio']} - "
                      f"{self.cfg.get('empresa','')}")
            cuerpo = (f"Estimado {reng['proveedor']}:\n\n"
                      f"Adjuntamos el voucher de la reserva {self.res.get('numero','')} "
                      f"para {self.res.get('cliente','')}.\n"
                      f"Destino: {reng['destino']}\n"
                      f"Fechas de viaje: {self.res.get('fechas_viaje','')}\n"
                      f"Pasajeros: {self.res.get('pax_txt','')}\n\n"
                      f"Favor confirmar disponibilidad y remitir la facturacion a nombre de "
                      f"{self.cfg.get('empresa','')}.\n\nCordialmente,\n{self.cfg.get('empresa','')}")
            enviar_correo(self.cfg, reng["correo"], asunto, cuerpo, fn)
        except Exception as e:
            messagebox.showerror("No se pudo enviar", str(e))
            return
        s["enviado"] = True
        s["fecha_envio"] = datetime.date.today().strftime("%d/%m/%Y")
        lbl.configure(text="✓ Voucher enviado " + s["fecha_envio"], text_color=GREEN)
        actualizar_reserva(self.res.get("numero", ""),
                           {"destinos_detalle": self.res["destinos_detalle"]})
        self._refrescar_resumen()
        messagebox.showinfo("Enviado", f"Voucher enviado a {reng['correo']}.")

    # ---- Voucher CONSOLIDADO por proveedor (transporte / guia) ----
    def _voucher_grupo(self, di, cat):
        self._sync()
        grupos = [g for g in servicios_por_proveedor(self.res, di, cat) if g[2]]
        if not grupos:
            messagebox.showinfo("Voucher", "No hay servicios de esta categoria en el destino.",
                                parent=self)
            return
        generados = []
        for prov, correo, items in grupos:
            try:
                fn = generar_voucher_consolidado_archivo(self.cfg, self.res, di, cat, prov, items)
                generados.append(fn)
            except Exception as e:
                messagebox.showerror("Error al generar el voucher", str(e), parent=self)
                return
        if generados:
            try:
                os.startfile(generados[0])
            except Exception:
                pass
        messagebox.showinfo(
            "Voucher del proveedor",
            f"Se genero {len(generados)} voucher(s): uno por cada proveedor, con TODOS "
            f"sus servicios de este destino.", parent=self)

    def _enviar_grupo(self, di, cat):
        self._sync()
        grupos = [g for g in servicios_por_proveedor(self.res, di, cat) if g[2]]
        if not grupos:
            messagebox.showinfo("Enviar", "No hay servicios de esta categoria en el destino.",
                                parent=self)
            return
        enviados = 0; errores = []
        for prov, correo, items in grupos:
            try:
                enviar_voucher_consolidado(self.cfg, self.res, di, cat, prov, correo, items)
                enviados += 1
            except Exception as e:
                errores.append(f"• {prov or '(sin proveedor)'}: {e}")
        self._pintar_servicios(); self._refrescar_resumen()
        msg = f"Vouchers enviados: {enviados} (uno por proveedor)."
        if errores:
            msg += "\n\nNo se pudieron enviar:\n" + "\n".join(errores)
        messagebox.showinfo("Enviar al proveedor", msg, parent=self)

    def _voucher_cliente(self, abrir=True):
        self._sync()
        fn = os.path.join(self._carpeta_vouchers(),
                          f"Voucher_cliente_{self.res.get('numero','')}.pdf")
        try:
            generar_voucher_cliente(self.cfg, self.res, fn)
        except Exception as e:
            messagebox.showerror("Error al generar el voucher", str(e))
            return None
        self.res["voucher_cliente"] = fn
        actualizar_reserva(self.res.get("numero", ""), {"voucher_cliente": fn})
        if abrir:
            try:
                os.startfile(fn)
            except Exception:
                pass
        return fn

    def _enviar_cliente(self):
        fn = self._voucher_cliente(abrir=False)
        if not fn:
            return
        dest = self.res.get("email", "")
        if not dest:
            messagebox.showinfo("Correo del cliente",
                                "La reserva no tiene correo del cliente.")
            return
        try:
            asunto = (f"Confirmacion de reserva {self.res.get('numero','')} - "
                      f"{self.cfg.get('empresa','')}")
            cuerpo = (f"Estimado(a) {self.res.get('cliente','')}:\n\n"
                      f"Adjuntamos el voucher de confirmacion de su reserva "
                      f"{self.res.get('numero','')}.\n\n"
                      f"Feliz viaje.\n{self.cfg.get('empresa','')}")
            enviar_correo(self.cfg, dest, asunto, cuerpo, fn)
        except Exception as e:
            messagebox.showerror("No se pudo enviar", str(e))
            return
        messagebox.showinfo("Enviado", f"Voucher de cliente enviado a {dest}.")

    def _guardar(self):
        self._sync()
        cambios = {
            "cliente": self.res.get("cliente", ""), "email": self.res.get("email", ""),
            "contacto": self.res.get("contacto", ""),
            "destinos": self.res.get("destinos", []),
            "fechas_viaje": self.res.get("fechas_viaje", ""),
            "pax_txt": self.res.get("pax_txt", ""), "hab": self.res.get("hab", ""),
            "estado": self.res["estado"], "monto": self.res.get("monto", 0),
            "asesor": self.res.get("asesor", {}),
            "items_cobro": self.res.get("items_cobro", []),
            "notas": self.res.get("notas", ""), "itinerario": self.res.get("itinerario", ""),
            "itinerario_dias": self.res.get("itinerario_dias", []),
            "destinos_detalle": self.res.get("destinos_detalle", []),
            "pasajeros_list": self.res.get("pasajeros_list", []),
            "vuelos_adjuntos": self.res.get("vuelos_adjuntos", []),
            "soportes_pago": self.res.get("soportes_pago", [])}
        for k in self.res:
            if k.startswith("os_"):
                cambios[k] = self.res[k]
        actualizar_reserva(self.res.get("numero", ""), cambios)
        if self.on_save:
            self.on_save()
        messagebox.showinfo("Guardado", "Reserva actualizada.")
        self.destroy()


class DialogoReporteMes(ctk.CTkToplevel):
    """Elegir mes (o todos) y descargar un reporte en Excel."""
    def __init__(self, master, titulo, subtitulo, meses, export_fn, base_nombre):
        super().__init__(master)
        self.export_fn = export_fn; self.base = base_nombre
        self.title(titulo)
        self.geometry("500x260"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        ctk.CTkLabel(self, text=titulo, font=("Segoe UI", 16, "bold"),
                     text_color=NAVY).pack(pady=(18, 2))
        ctk.CTkLabel(self, text=subtitulo, text_color=MUTED,
                     font=("Segoe UI", 11), wraplength=440).pack(pady=(0, 12))
        opciones = ["Todos los meses"] + list(meses)
        self.v_mes = tk.StringVar(value=opciones[0])
        ctk.CTkLabel(self, text="Mes", text_color=MUTED, font=("Segoe UI", 11)).pack()
        ctk.CTkOptionMenu(self, variable=self.v_mes, values=opciones, width=240, height=34,
                          fg_color=NAVY, button_color=NAVY2).pack(pady=(2, 16))
        ctk.CTkButton(self, text="⬇  Descargar Excel", height=42, corner_radius=10, fg_color=GREEN,
                      hover_color=GREEN_H, font=("Segoe UI", 13, "bold"),
                      command=self._exportar).pack()

    def _exportar(self):
        mes = None if self.v_mes.get().startswith("Todos") else self.v_mes.get()
        ruta = filedialog.asksaveasfilename(
            title="Guardar reporte", defaultextension=".xlsx",
            initialfile=f"{self.base}_{mes or 'todos'}.xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not ruta:
            return
        try:
            n = self.export_fn(ruta, mes)
        except Exception as e:
            messagebox.showerror("Error al generar el reporte", str(e), parent=self)
            return
        try:
            os.startfile(ruta)
        except Exception:
            pass
        messagebox.showinfo("Reporte generado",
                            f"Reporte generado con {n} registro(s).\n\n{ruta}", parent=self)
        self.destroy()


class SelectorTarifario(ctk.CTkToplevel):
    """Elegir hoteles y tours del tarifario (precios_2026) por destino y agregarlos
       como items de la reserva, con su precio en USD por persona."""
    def __init__(self, master, precios, cfg, destino, noches, grupo, on_add):
        super().__init__(master)
        self.precios = precios; self.cfg = cfg; self.on_add = on_add
        self.title("Agregar items desde el tarifario")
        self.geometry("860x660"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        ctk.CTkLabel(self, text="Agregar items desde el tarifario",
                     text_color=NAVY, font=("Segoe UI", 16, "bold")).pack(anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(self, text="Elige el destino; se listan hoteles y tours con su precio por "
                     "persona (USD). Ajusta acomodacion, noches y grupo si hace falta.",
                     text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=16)

        destinos = sorted([d for d in (precios or {}).keys()])
        dest0 = next((d for d in destinos if d.strip().lower() == (destino or "").strip().lower()),
                     (destinos[0] if destinos else ""))
        bar = ctk.CTkFrame(self, fg_color=CARD, corner_radius=10); bar.pack(fill="x", padx=16, pady=10)
        row = ctk.CTkFrame(bar, fg_color="transparent"); row.pack(fill="x", padx=10, pady=8)
        ctk.CTkLabel(row, text="Destino", text_color=MUTED).pack(side="left")
        self.v_dest = tk.StringVar(value=dest0)
        ctk.CTkOptionMenu(row, variable=self.v_dest, values=destinos, width=180, height=30,
                          fg_color=NAVY, button_color=NAVY2,
                          command=lambda _v=None: self._rebuild()).pack(side="left", padx=(4, 14))
        ctk.CTkLabel(row, text="Acomodacion", text_color=MUTED).pack(side="left")
        self.v_acom = tk.StringVar(value="doble")
        ctk.CTkOptionMenu(row, variable=self.v_acom, values=["sencilla", "doble", "triple"],
                          width=110, height=30, fg_color=NAVY, button_color=NAVY2,
                          command=lambda _v=None: self._rebuild()).pack(side="left", padx=(4, 14))
        ctk.CTkLabel(row, text="Noches", text_color=MUTED).pack(side="left")
        self.v_noches = tk.StringVar(value=str(max(1, int(noches or 1))))
        e1 = ctk.CTkEntry(row, textvariable=self.v_noches, width=54, height=30); e1.pack(side="left", padx=(4, 14))
        ctk.CTkLabel(row, text="Grupo (pax)", text_color=MUTED).pack(side="left")
        self.v_grupo = tk.StringVar(value=str(max(1, int(grupo or 2))))
        e2 = ctk.CTkEntry(row, textvariable=self.v_grupo, width=54, height=30); e2.pack(side="left", padx=4)
        ctk.CTkButton(row, text="Actualizar precios", width=150, height=30, fg_color=BLUE,
                      hover_color=BLUE_H, command=self._rebuild).pack(side="right")
        for e in (e1, e2):
            e.bind("<Return>", lambda ev: self._rebuild())

        self.lbl_msg = ctk.CTkLabel(self, text="", text_color=GREEN_H, font=("Segoe UI", 11, "bold"))
        self.lbl_msg.pack(anchor="w", padx=16)
        self.box = ctk.CTkScrollableFrame(self, fg_color=BG)
        self.box.pack(fill="both", expand=True, padx=16, pady=(4, 14))
        self._rebuild()

    def _num(self, var, defecto):
        try:
            return max(1, int(float(str(var.get()).strip())))
        except Exception:
            return defecto

    def _rebuild(self):
        for w in self.box.winfo_children():
            w.destroy()
        dest = self.v_dest.get()
        acom = self.v_acom.get()
        noches = self._num(self.v_noches, 1)
        grupo = self._num(self.v_grupo, 2)

        self._banda(f"HOTELES  ·  precio por persona en {acom}, {noches} noche(s)")
        hoteles = hoteles_detalle(self.precios, dest)
        if not hoteles:
            ctk.CTkLabel(self.box, text="Sin hoteles para este destino.", text_color=MUTED).pack(pady=4)
        for h in hoteles:
            if not float(h.get(acom, 0) or 0):
                continue
            usd = precio_hotel_usd_pp(self.precios, dest, h, acom, noches, self.cfg)
            temp = h.get("temporada", "")
            cat = h.get("categoria", "")
            sub = "  ·  ".join(x for x in [temp, cat] if x)
            desc = f"{dest} - {h.get('nombre','')} ({acom}, {noches}N, p.p.)"
            self._fila(h.get("nombre", ""), sub, usd, desc)

        self._banda(f"TOURS / SERVICIOS  ·  precio por persona, grupo de {grupo}")
        servicios = servicios_terrestres(self.precios, dest)
        if not servicios:
            ctk.CTkLabel(self.box, text="Sin servicios para este destino.", text_color=MUTED).pack(pady=4)
        for s in servicios:
            usd = precio_servicio_usd_pp(self.precios, dest, s, grupo, self.cfg)
            desc = f"{dest} - {s.get('nombre','')} (p.p., grupo {grupo})"
            self._fila(s.get("nombre", ""), "", usd, desc)

    def _banda(self, texto):
        b = ctk.CTkLabel(self.box, text=texto, text_color="#FFFFFF", fg_color=NAVY,
                         corner_radius=6, anchor="w", font=("Segoe UI", 11, "bold"))
        b.pack(fill="x", pady=(8, 2), ipady=4, ipadx=8)

    def _fila(self, nombre, sub, usd, desc):
        row = ctk.CTkFrame(self.box, fg_color=CARD, corner_radius=8,
                           border_width=1, border_color=LINE)
        row.pack(fill="x", pady=2)
        izq = ctk.CTkFrame(row, fg_color="transparent"); izq.pack(side="left", fill="x", expand=True, padx=10, pady=6)
        ctk.CTkLabel(izq, text=nombre, text_color=NAVY, anchor="w",
                     font=("Segoe UI", 12, "bold")).pack(anchor="w")
        if sub:
            ctk.CTkLabel(izq, text=sub, text_color=MUTED, anchor="w",
                         font=("Segoe UI", 10)).pack(anchor="w")
        ctk.CTkLabel(row, text=f"USD {usd:,.2f}", text_color=GREEN_H,
                     font=("Segoe UI", 12, "bold")).pack(side="left", padx=10)
        ctk.CTkButton(row, text="+ Agregar", width=100, height=30, fg_color=GREEN,
                      hover_color=GREEN_H, font=("Segoe UI", 11, "bold"),
                      command=lambda: self._agregar(desc, usd)).pack(side="right", padx=10)

    def _agregar(self, desc, usd):
        try:
            self.on_add(desc, usd)
            self.lbl_msg.configure(text=f"Agregado:  {desc}   (USD {usd:,.2f})")
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)


class VentanaVouchersProveedores(ctk.CTkToplevel):
    """Tablero para enviar los vouchers a todos los proveedores de una reserva
       (hotel, transporte, guia, actividad) con su estado, en un solo lugar."""
    def __init__(self, master, res, cfg, on_change=None):
        super().__init__(master)
        self.res = res; self.cfg = cfg; self.on_change = on_change
        self.title("Vouchers a proveedores - Reserva " + res.get("numero", ""))
        self.geometry("1040x600"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        ctk.CTkLabel(self, text=f"Vouchers a proveedores  ·  Reserva N. {res.get('numero','')}"
                     f"  ·  {res.get('cliente','')}", font=("Segoe UI", 16, "bold"),
                     text_color=NAVY).pack(anchor="w", padx=16, pady=(14, 2))
        self.lbl_seg = ctk.CTkLabel(self, text="", text_color=NAVY, fg_color="#EAF2FD",
                                    corner_radius=8, font=("Segoe UI", 11, "bold"),
                                    anchor="w")
        self.lbl_seg.pack(fill="x", padx=16, pady=(0, 6), ipady=4, ipadx=8)
        top = ctk.CTkFrame(self, fg_color="transparent"); top.pack(fill="x", padx=16)
        ctk.CTkButton(top, text="✉ Enviar TODOS los pendientes", height=34, fg_color=GREEN,
                      hover_color=GREEN_H, font=("Segoe UI", 12, "bold"),
                      command=self._enviar_todos).pack(side="left")
        ctk.CTkLabel(top, text="(cada servicio necesita proveedor y correo)",
                     text_color=MUTED, font=("Segoe UI", 10)).pack(side="left", padx=10)
        self.box = ctk.CTkScrollableFrame(self, fg_color=BG)
        self.box.pack(fill="both", expand=True, padx=16, pady=(8, 14))
        self._pintar()

    def _servicios(self):
        out = []
        for di, d in enumerate(destinos_detalle_de(self.res)):
            for cat, etiqueta, tipo in CATEGORIAS_SERV:
                for si, s in enumerate(d.get(cat, [])):
                    if s.get("servicio") or s.get("proveedor"):
                        out.append((di, cat, si, tipo, d.get("nombre", ""), s))
        return out

    def _pintar(self):
        for w in self.box.winfo_children():
            w.destroy()
        servicios = self._servicios()
        if not servicios:
            ctk.CTkLabel(self.box, text="Esta reserva aun no tiene servicios con proveedor. "
                         "Agrega destinos y servicios en la reserva.", text_color=MUTED).pack(pady=24)
        for di, cat, si, tipo, dest, s in servicios:
            self._fila(di, cat, si, tipo, dest, s)
        self._resumen()

    def _fila(self, di, cat, si, tipo, dest, s):
        card = ctk.CTkFrame(self.box, fg_color=CARD, corner_radius=8,
                            border_width=1, border_color=LINE)
        card.pack(fill="x", pady=3)
        # encabezado: tipo + destino + servicio
        cab = ctk.CTkFrame(card, fg_color="transparent"); cab.pack(fill="x", padx=8, pady=(6, 2))
        ctk.CTkLabel(cab, text=f"{tipo}", text_color="#FFFFFF", fg_color=NAVY, corner_radius=6,
                     font=("Segoe UI", 10, "bold")).pack(side="left", ipadx=6, ipady=1)
        ctk.CTkLabel(cab, text=f"  {dest}  ·  {s.get('servicio','(sin nombre)')}", text_color=NAVY,
                     font=("Segoe UI", 12, "bold")).pack(side="left")
        env = s.get("enviado")
        ctk.CTkLabel(cab, text=("✓ enviado " + s.get("fecha_envio", "") if env else "sin enviar"),
                     text_color=(GREEN if env else MUTED),
                     font=("Segoe UI", 10, "bold" if env else "normal")).pack(side="right")
        # proveedor + correo + estado
        v_prov = tk.StringVar(value=s.get("proveedor", ""))
        v_mail = tk.StringVar(value=s.get("correo", ""))
        v_est = tk.StringVar(value=s.get("estado_prov", "Pendiente"))
        fila = ctk.CTkFrame(card, fg_color="transparent"); fila.pack(fill="x", padx=8, pady=(0, 4))
        ctk.CTkLabel(fila, text="Proveedor", text_color=MUTED, width=60).pack(side="left")
        ctk.CTkEntry(fila, textvariable=v_prov, width=180, height=28).pack(side="left", padx=(0, 6))
        ctk.CTkLabel(fila, text="Correo", text_color=MUTED, width=44).pack(side="left")
        ctk.CTkEntry(fila, textvariable=v_mail, height=28).pack(side="left", padx=(0, 6),
                                                                fill="x", expand=True)
        om = ctk.CTkOptionMenu(fila, variable=v_est, values=ESTADOS_PROV, width=170, height=28,
                               fg_color=ESTADO_PROV_COLOR.get(s.get("estado_prov", "Pendiente"), MUTED),
                               button_color=NAVY2)
        om.configure(command=lambda val, o=om, ss=s: self._set_estado(ss, val, o))
        om.pack(side="left")
        # acciones
        acc = ctk.CTkFrame(card, fg_color="transparent"); acc.pack(fill="x", padx=8, pady=(0, 6))
        ctk.CTkButton(acc, text="Generar voucher", width=130, height=28, fg_color=BLUE,
                      hover_color=BLUE_H,
                      command=lambda: self._generar(di, cat, si, v_prov, v_mail)).pack(side="right", padx=3)
        ctk.CTkButton(acc, text="Enviar al proveedor", width=150, height=28, fg_color=GREEN,
                      hover_color=GREEN_H,
                      command=lambda: self._enviar(di, cat, si, v_prov, v_mail)).pack(side="right", padx=3)

    def _aplicar(self, di, cat, si, v_prov, v_mail):
        s = self.res["destinos_detalle"][di][cat][si]
        s["proveedor"] = v_prov.get().strip()
        s["correo"] = v_mail.get().strip()
        return s

    def _set_estado(self, s, val, menu):
        s["estado_prov"] = val
        try:
            menu.configure(fg_color=ESTADO_PROV_COLOR.get(val, MUTED))
        except Exception:
            pass
        actualizar_reserva(self.res.get("numero", ""),
                           {"destinos_detalle": self.res["destinos_detalle"]})
        self._resumen()
        if self.on_change:
            self.on_change()

    def _generar(self, di, cat, si, v_prov, v_mail):
        self._aplicar(di, cat, si, v_prov, v_mail)
        try:
            fn, _r, _s = generar_voucher_prov_archivo(self.cfg, self.res, di, cat, si)
            os.startfile(fn)
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)

    def _enviar(self, di, cat, si, v_prov, v_mail):
        self._aplicar(di, cat, si, v_prov, v_mail)
        try:
            enviar_voucher_prov(self.cfg, self.res, di, cat, si)
        except Exception as e:
            messagebox.showerror("No se pudo enviar", str(e), parent=self)
            return
        self._pintar()
        if self.on_change:
            self.on_change()
        messagebox.showinfo("Enviado", "Voucher enviado al proveedor.", parent=self)

    def _enviar_todos(self):
        pendientes = [(di, cat, si) for di, cat, si, tipo, dest, s in self._servicios()
                      if not s.get("enviado") and s.get("correo")]
        sin_correo = [s for di, cat, si, tipo, dest, s in self._servicios()
                      if not s.get("enviado") and not s.get("correo")]
        if not pendientes:
            messagebox.showinfo("Enviar todos", "No hay servicios pendientes con correo para enviar.",
                                parent=self)
            return
        if not messagebox.askyesno("Enviar todos",
                                   f"Enviar {len(pendientes)} voucher(s) a los proveedores ahora?",
                                   parent=self):
            return
        ok = 0; errores = []
        for di, cat, si in pendientes:
            try:
                enviar_voucher_prov(self.cfg, self.res, di, cat, si); ok += 1
            except Exception as e:
                errores.append(str(e))
        self._pintar()
        if self.on_change:
            self.on_change()
        msg = f"Enviados: {ok}."
        if sin_correo:
            msg += f"\nSin correo (no enviados): {len(sin_correo)}."
        if errores:
            msg += "\nErrores: " + "; ".join(errores[:3])
        messagebox.showinfo("Envio de vouchers", msg, parent=self)

    def _resumen(self):
        try:
            total, enviados, por_estado = resumen_seguimiento(self.res)
        except Exception:
            return
        self.lbl_seg.configure(
            text=(f"{enviados}/{total} con voucher enviado   ·   "
                  f"✓ {por_estado.get('Reservado con pago', 0)} con pago   ·   "
                  f"◐ {por_estado.get('Reservado sin pago', 0)} sin pago   ·   "
                  f"○ {por_estado.get('Pendiente', 0)} pendiente"))


class ModuloReservas(ctk.CTkToplevel):
    """Ventana principal del modulo de Reservas (historial + acciones)."""
    def __init__(self, master=None):
        super().__init__(master)
        ctk.set_appearance_mode("light")
        try:
            ctk.set_widget_scaling(0.85)
        except Exception:
            pass
        self.cfg = cargar_config()
        try:
            self.iconbitmap(recurso("app.ico"))
        except Exception:
            pass
        self.title(f"Reservas - INNOBA Colombia DMC   v{VERSION}")
        self.configure(fg_color=BG)
        self.geometry("1180x720")
        # Reparar local (rapido) y mostrar YA; traer de la nube en 2do plano para
        # no congelar la ventana con internet lento.
        try:
            reparar_reservas()
        except Exception:
            pass
        self._build()
        self.after(60, lambda: self._max())
        self.after(150, self._sync_bg)

    def _sync_bg(self):
        def worker():
            try:
                n = importar_reservas_nube()
            except Exception:
                n = 0
            if n:
                try:
                    self.after(0, self._pintar)
                except Exception:
                    pass
        threading.Thread(target=worker, daemon=True).start()

    def _max(self):
        try:
            self.state("zoomed")
        except Exception:
            pass

    def _volver_inicio(self):
        lanz = self.master
        try:
            if lanz is not None and hasattr(lanz, "reservas"):
                lanz.reservas = None
        except Exception:
            pass
        try:
            self.destroy()
        except Exception:
            pass
        try:
            if lanz is not None:
                lanz.deiconify(); lanz.lift(); lanz.focus_force()
        except Exception:
            pass

    def _build(self):
        head = ctk.CTkFrame(self, fg_color=CARD, corner_radius=0, height=60)
        head.pack(fill="x"); head.pack_propagate(False)
        head.grid_columnconfigure(1, weight=1)
        try:
            img = Image.open(recurso("logo_innoba.png")); w, h = img.size; hh = 38
            self.logo_img = ctk.CTkImage(light_image=img, size=(int(w * hh / h), hh))
            ctk.CTkLabel(head, image=self.logo_img, text="").grid(row=0, column=0, padx=(18, 12), pady=8)
        except Exception:
            ctk.CTkLabel(head, text="INNOBA", font=("Segoe UI", 20, "bold"),
                         text_color=NAVY).grid(row=0, column=0, padx=18)
        tit = ctk.CTkFrame(head, fg_color="transparent"); tit.grid(row=0, column=1, sticky="w")
        ctk.CTkLabel(tit, text="Modulo de Reservas", text_color=NAVY,
                     font=("Segoe UI", 17, "bold"), height=20).pack(anchor="w")
        ctk.CTkLabel(tit, text=f"INNOBA Colombia DMC  ·  v{VERSION}", text_color=MUTED,
                     font=("Segoe UI", 11), height=15).pack(anchor="w")
        hb = ctk.CTkFrame(head, fg_color="transparent"); hb.grid(row=0, column=2, padx=18)
        ctk.CTkButton(hb, text="⌂ Modulos", width=100, height=36, corner_radius=10,
                      fg_color=NAVY, hover_color=NAVY2, font=("Segoe UI", 12, "bold"),
                      command=self._volver_inicio).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hb, text="+ Nueva reserva", width=160, height=36, corner_radius=10,
                      fg_color=GREEN, hover_color=GREEN_H, font=("Segoe UI", 12, "bold"),
                      command=self._nueva_desde_cot).pack(side="left", padx=(0, 8))
        if WEBHOOK_URL:
            ctk.CTkButton(hb, text="☁ Sincronizar", width=130, height=36, corner_radius=10,
                          fg_color=BLUE, hover_color=NAVY, font=("Segoe UI", 12, "bold"),
                          command=self._sincronizar).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hb, text="🔢 Reordenar N°", width=140, height=36, corner_radius=10,
                      fg_color="#0E7C6B", hover_color="#0A5C50", font=("Segoe UI", 12, "bold"),
                      command=self._reordenar).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hb, text="💰 Liquidacion", width=140, height=36, corner_radius=10,
                      fg_color="#0E7C6B", hover_color="#0A5C50", font=("Segoe UI", 12, "bold"),
                      command=self._liquidacion).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hb, text="📊 Reporte", width=110, height=36, corner_radius=10,
                      fg_color="#7A5AB5", hover_color="#63459A", font=("Segoe UI", 12, "bold"),
                      command=self._reporte).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hb, text="Asesores", width=100, height=36, corner_radius=10,
                      fg_color=CYAN, hover_color=BLUE, font=("Segoe UI", 12, "bold"),
                      command=self._config_asesores).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hb, text="Datos de mi empresa", width=170, height=36, corner_radius=10,
                      fg_color=NAVY2, hover_color=NAVY, font=("Segoe UI", 12, "bold"),
                      command=self._config_empresa).pack(side="left")

        # Indicadores
        self.kpis = ctk.CTkFrame(self, fg_color="transparent"); self.kpis.pack(fill="x", padx=18, pady=(12, 2))

        bar = ctk.CTkFrame(self, fg_color="transparent"); bar.pack(fill="x", padx=18, pady=(4, 4))
        self.q = tk.StringVar()
        e = ctk.CTkEntry(bar, textvariable=self.q, height=36, corner_radius=10,
                         placeholder_text="Buscar reserva por numero, cliente, asesor, estado...")
        e.pack(side="left", fill="x", expand=True)
        e.bind("<KeyRelease>", lambda ev: self._pintar())
        # Filtro por asesora (para que cada una vea solo sus reservas)
        nombres_ase = [a.get("nombre", "") for a in asesores_reservas(self.cfg) if a.get("nombre")]
        self.v_filtro_ase = tk.StringVar(value="Todas")
        ctk.CTkLabel(bar, text="Ver:", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(side="left", padx=(10, 2))
        ctk.CTkOptionMenu(bar, variable=self.v_filtro_ase, values=["Todas"] + nombres_ase,
                          width=200, height=36, fg_color=NAVY, button_color=NAVY2,
                          button_hover_color=BLUE, dropdown_fg_color=CARD,
                          dropdown_text_color=TEXT,
                          command=lambda *_: self._pintar()).pack(side="left", padx=(0, 6))
        self.lbl_tot = ctk.CTkLabel(bar, text="", text_color=MUTED, font=("Segoe UI", 11))
        self.lbl_tot.pack(side="right", padx=10)

        self.lista = ctk.CTkScrollableFrame(self, fg_color=BG)
        self.lista.pack(fill="both", expand=True, padx=18, pady=(4, 14))
        self._pintar()

    def _pintar_kpis(self):
        for w in self.kpis.winfo_children():
            w.destroy()
        ind = indicadores_reservas()
        fila = ctk.CTkFrame(self.kpis, fg_color="transparent"); fila.pack(fill="x")
        _kpi_card(fila, "Reservas totales", ind["total"], NAVY)
        _kpi_card(fila, "Aprobadas (con pago)", ind.get("Confirmada con pago", 0), GREEN)
        _kpi_card(fila, "En seguimiento", ind.get("Confirmada", 0) + ind.get("Aplazada", 0), "#D9A400")
        _kpi_card(fila, "Anuladas", ind.get("Anulada", 0), RED)
        _kpi_card(fila, "Negociado (activo)", usd(ind["monto_total"]), NAVY, ancho=170)
        _kpi_card(fila, "Cobrado (con pago)", usd(ind["con_pago_usd"]), GREEN_H, ancho=170)
        _kpi_card(fila, f"Del mes ({ind['mes_n']})", usd(ind["mes_usd"]), BLUE, ancho=160)
        _kpi_card(fila, "Vouchers enviados",
                  f"{ind['serv_enviados']}/{ind['serv_total']}", "#7A5AB5", ancho=150)

    def _pintar(self):
        self._pintar_kpis()
        for w in self.lista.winfo_children():
            w.destroy()
        # Ordenar por consecutivo DESCENDENTE (la ultima reserva de primero)
        def _num_res(it):
            try:
                return int(re.sub(r"\D", "", str(it.get("numero", "0"))) or 0)
            except Exception:
                return 0
        items = sorted(cargar_reservas().get("items", []), key=_num_res, reverse=True)
        fase = self.v_filtro_ase.get() if hasattr(self, "v_filtro_ase") else "Todas"
        if fase and fase != "Todas":
            items = [it for it in items
                     if (it.get("asesor", {}) or {}).get("nombre", "") == fase]
        q = self.q.get().lower().strip()
        if q:
            def _match(it):
                blob = " ".join(str(it.get(k, "")) for k in
                                ("numero", "cliente", "contacto", "fechas_viaje", "estado")).lower()
                ase = it.get("asesor", {})
                if isinstance(ase, dict):
                    blob += " " + str(ase.get("nombre", "")).lower()
                blob += " " + " ".join(str(d.get("nombre", "")) for d in it.get("destinos_detalle", [])).lower()
                return q in blob
            items = [it for it in items if _match(it)]
        # totales negociados por moneda (excluye anuladas)
        tot = sum(float(it.get("monto", 0) or 0) for it in items
                  if it.get("estado") != "Anulada")
        self.lbl_tot.configure(text=f"{len(items)} reserva(s)  ·  Negociado: {usd(tot)}")
        if not items:
            ctk.CTkLabel(self.lista, text="Aun no hay reservas. Crea una con "
                         "'+ Nueva desde cotizacion'.", text_color=MUTED).pack(pady=24)
            return
        TOPE = 50
        if len(items) > TOPE:
            ctk.CTkLabel(self.lista,
                         text=f"Mostrando {TOPE} de {len(items)}. Usa el buscador o el filtro "
                              "por asesora para ver el resto.",
                         text_color="#B7791F", fg_color="#FFF3C4", corner_radius=8,
                         font=("Segoe UI", 11, "bold")).pack(fill="x", padx=2, pady=(2, 4), ipady=4)
        lista = items[:TOPE]
        self._render_gen = getattr(self, "_render_gen", 0) + 1
        gen = self._render_gen

        def _chunk(i):
            if gen != self._render_gen:
                return
            try:
                if not self.winfo_exists():
                    return
            except Exception:
                return
            for it in lista[i:i + 6]:
                self._fila(it)
            if i + 6 < len(lista):
                self.after(1, lambda: _chunk(i + 6))
        _chunk(0)

    def _fila(self, it):
        estado = it.get("estado", "Confirmada")
        ase = it.get("asesor", {}) or {}
        fila = ctk.CTkFrame(self.lista, fg_color=ESTADO_RES_FILA.get(estado, CARD2),
                            corner_radius=10)
        fila.pack(fill="x", pady=4, padx=2)
        fila.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(fila, text="N. " + it.get("numero", ""), text_color=NAVY,
                     font=("Segoe UI", 16, "bold"), width=90).grid(
            row=0, column=0, rowspan=2, padx=(12, 6), pady=8)
        info = ctk.CTkFrame(fila, fg_color="transparent"); info.grid(row=0, column=1, rowspan=2, sticky="w")
        ctk.CTkLabel(info, text=it.get("cliente", ""), text_color=TEXT,
                     font=("Segoe UI", 13, "bold")).pack(anchor="w")
        ctk.CTkLabel(info, text=f"{', '.join(it.get('destinos', []))}  ·  "
                     f"{it.get('fechas_viaje','')}", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w")
        ctk.CTkLabel(info, text="Asesor: " + (ase.get("nombre", "") or "(sin asignar)"),
                     text_color=BLUE, font=("Segoe UI", 11)).pack(anchor="w")
        ctk.CTkLabel(fila, text=usd(it.get("monto", 0)), text_color=NAVY,
                     font=("Segoe UI", 13, "bold")).grid(row=0, column=2, rowspan=2, padx=10)
        badge = ctk.CTkLabel(fila, text=estado, fg_color=ESTADO_RES_COLOR.get(estado, MUTED),
                             text_color="#FFFFFF", corner_radius=6,
                             font=("Segoe UI", 11, "bold"))
        badge.grid(row=0, column=3, rowspan=2, padx=8, ipadx=8, ipady=3)
        btns = ctk.CTkFrame(fila, fg_color="transparent"); btns.grid(row=0, column=4, rowspan=2, padx=10)
        f1 = ctk.CTkFrame(btns, fg_color="transparent"); f1.pack()
        ctk.CTkButton(f1, text="Abrir", width=90, height=32, fg_color=NAVY,
                      hover_color=NAVY2, command=lambda x=it: self._abrir(x)).pack(side="left", pady=2, padx=2)
        ctk.CTkButton(f1, text="🗑", width=36, height=32, fg_color=RED,
                      hover_color="#9B2C22", command=lambda x=it: self._eliminar(x)).pack(side="left", pady=2, padx=2)
        ctk.CTkButton(btns, text="Voucher cliente", width=130, height=30, fg_color=CYAN,
                      hover_color=BLUE, command=lambda x=it: self._voucher_cli(x)).pack(pady=2)

    def _abrir(self, it):
        VentanaReservaDetalle(self, it, self.cfg, on_save=self._pintar)

    def _eliminar(self, it):
        if messagebox.askyesno("Eliminar reserva",
                               f"Eliminar la reserva N. {it.get('numero','')} de "
                               f"{it.get('cliente','')}?\n\nEsta accion no se puede deshacer."):
            eliminar_reserva(it.get("numero", ""), it.get("uid"))
            self._pintar()

    def _voucher_cli(self, it):
        fn = os.path.join(datos_dir(), "vouchers")
        os.makedirs(fn, exist_ok=True)
        ruta = os.path.join(fn, f"Voucher_cliente_{it.get('numero','')}.pdf")
        try:
            generar_voucher_cliente(self.cfg, it, ruta)
            os.startfile(ruta)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _reporte(self):
        data = cargar_reservas()
        meses = sorted({_mes_de_iso(it.get("fecha_creacion", "")) for it in data.get("items", [])
                        if it.get("fecha_creacion")}, reverse=True)
        DialogoReporteMes(self, "Reporte de reservas por mes",
                          "Descarga un Excel con el detalle de las reservas y un resumen por mes.",
                          meses, exportar_reporte_reservas, "Reporte_reservas")

    def _sincronizar(self):
        try:
            subidas, importadas = sincronizar_reservas_nube()
        except Exception as e:
            messagebox.showerror("Sincronizar", str(e), parent=self); return
        self._pintar()
        messagebox.showinfo(
            "Sincronizacion de reservas",
            f"Reservas subidas a la nube: {subidas}\n"
            f"Reservas traidas/actualizadas: {importadas}\n\n"
            "Las reservas ahora se comparten entre los equipos que usan el sistema.",
            parent=self)

    def _liquidacion(self):
        VentanaLiquidacion(self, self.cfg)

    def _reordenar(self):
        if not messagebox.askyesno(
                "Reordenar consecutivo",
                "Se renumeraran TODAS las reservas de forma continua (2989, 2990, 2991, ...) "
                "en orden de creacion, para quitar saltos en el consecutivo.\n\n"
                "Hazlo en UN solo equipo; los demas adoptaran los numeros al Sincronizar.\n\n"
                "¿Continuar?", icon="warning", parent=self):
            return
        try:
            ultimo = compactar_consecutivo_reservas()
        except Exception as e:
            messagebox.showerror("Reordenar", str(e), parent=self); return
        self._pintar()
        messagebox.showinfo(
            "Consecutivo reordenado",
            f"Listo. Las reservas quedaron numeradas de forma continua hasta {ultimo}.\n"
            "En los demas equipos, entra a Reservas y dale '☁ Sincronizar' para que "
            "adopten los mismos numeros.", parent=self)

    def _config_asesores(self):
        DialogoAsesores(self, self.cfg, on_save=self._recargar_cfg)

    def _config_empresa(self):
        VentanaEmpresa(self, self.cfg, self._on_cfg_empresa)

    def _on_cfg_empresa(self, cfg):
        self.cfg = cfg

    def _recargar_cfg(self):
        self.cfg = cargar_config()

    def _nueva_desde_cot(self):
        if not asesores_reservas(self.cfg):
            messagebox.showinfo("Configura los asesores",
                                "Primero configura los asesores de reservas para poder "
                                "asignarlas automaticamente.")
            DialogoAsesores(self, self.cfg, on_save=self._recargar_cfg)
            return
        # Elegir origen: desde una cotizacion o en blanco (manual)
        dlg = ctk.CTkToplevel(self); dlg.title("Nueva reserva")
        dlg.geometry("460x250"); dlg.configure(fg_color=BG)
        dlg.transient(self); dlg.grab_set()
        try:
            dlg.after(80, lambda: (dlg.lift(), dlg.focus_force()))
        except Exception:
            pass
        ctk.CTkLabel(dlg, text="Nueva reserva", font=("Segoe UI", 16, "bold"),
                     text_color=NAVY).pack(pady=(22, 2))
        ctk.CTkLabel(dlg, text="Como quieres crear la reserva?", text_color=MUTED,
                     font=("Segoe UI", 12)).pack(pady=(0, 16))

        def desde():
            dlg.destroy()
            SelectorCotizacionReserva(self, self._crear_desde)

        def blanco():
            dlg.destroy()
            self._crear_blanco()

        ctk.CTkButton(dlg, text="Desde una cotizacion", height=46, corner_radius=10,
                      fg_color=GREEN, hover_color=GREEN_H, font=("Segoe UI", 13, "bold"),
                      command=desde).pack(fill="x", padx=34, pady=6)
        ctk.CTkButton(dlg, text="En blanco (manual)", height=46, corner_radius=10,
                      fg_color=NAVY, hover_color=NAVY2, font=("Segoe UI", 13, "bold"),
                      command=blanco).pack(fill="x", padx=34, pady=6)

    def _crear_desde(self, cot):
        rec = reserva_desde_cotizacion(cot)
        numero, guardado = registrar_reserva(rec, self.cfg)
        self._pintar()
        ase = guardado.get("asesor", {}) or {}
        messagebox.showinfo("Reserva creada",
                            f"Reserva N. {numero} creada.\n"
                            f"Asesor asignado: {ase.get('nombre','(sin asignar)')}"
                            + _texto_notif_reserva(guardado))
        VentanaReservaDetalle(self, guardado, self.cfg, on_save=self._pintar)

    def _crear_blanco(self):
        rec = {"cot_origen": "", "cliente": "", "email": "", "destinos": [],
               "fechas_viaje": "", "pax_txt": "", "hab": "", "estado": "Confirmada",
               "monto": 0.0, "moneda": "USD", "destinos_detalle": [],
               "itinerario": "", "notas": "", "voucher_cliente": "",
               "fecha_creacion": datetime.date.today().isoformat()}
        rec.update(_voucher_defaults())
        numero, guardado = registrar_reserva(rec, self.cfg)
        self._pintar()
        ase = guardado.get("asesor", {}) or {}
        messagebox.showinfo("Reserva creada",
                            f"Reserva N. {numero} creada (manual).\n"
                            f"Asesor asignado: {ase.get('nombre','(sin asignar)')}\n\n"
                            "Completa los datos del cliente y agrega los servicios."
                            + _texto_notif_reserva(guardado))
        VentanaReservaDetalle(self, guardado, self.cfg, on_save=self._pintar)


# ============================================================================
# MODULO COMERCIAL - Interfaz
# ============================================================================
class VentanaTareaDetalle(ctk.CTkToplevel):
    """Crear/editar una tarea comercial con su gestion (checklist, estado, etc.)."""
    def __init__(self, master, tarea, cfg, on_save=None):
        super().__init__(master)
        self.tarea = tarea; self.cfg = cfg; self.on_save = on_save
        self.title("Tarea " + tarea.get("numero", "nueva"))
        self.configure(fg_color=BG)
        self.geometry("860x720+80+10")
        self.transient(master); self.grab_set()
        self.after(60, self._max)
        # Barra de acciones FIJA ARRIBA (siempre visible, no depende del alto ni del DPI)
        self.footer = ctk.CTkFrame(self, fg_color=CARD, corner_radius=0, height=58)
        self.footer.pack(side="top", fill="x"); self.footer.pack_propagate(False)
        cont = ctk.CTkScrollableFrame(self, fg_color=BG)
        cont.pack(fill="both", expand=True, padx=16, pady=16)

        ctk.CTkLabel(cont, text="Tarea " + (tarea.get("numero", "") or "(nueva)"),
                     font=("Segoe UI", 18, "bold"), text_color=NAVY).pack(anchor="w")

        def campo(lbl, val):
            ctk.CTkLabel(cont, text=lbl, text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2, pady=(8, 0))
            v = tk.StringVar(value=val)
            ctk.CTkEntry(cont, textvariable=v, height=34, corner_radius=8, border_color=LINE).pack(fill="x", padx=2)
            return v

        self.v_titulo = campo("Titulo de la tarea *", tarea.get("titulo", ""))

        # cliente (empresas) + responsable
        f2 = ctk.CTkFrame(cont, fg_color="transparent"); f2.pack(fill="x")
        a = ctk.CTkFrame(f2, fg_color="transparent"); a.pack(side="left", fill="x", expand=True, padx=(0, 6))
        b = ctk.CTkFrame(f2, fg_color="transparent"); b.pack(side="left", fill="x", expand=True, padx=(6, 0))
        ctk.CTkLabel(a, text="Cliente / empresa", text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        try:
            empresas = sorted([c.get("empresa", "") for c in cargar_clientes() if c.get("empresa")])
        except Exception:
            empresas = []
        acli = ctk.CTkFrame(a, fg_color="transparent"); acli.pack(fill="x", padx=2)
        self.v_cliente = tk.StringVar(value=tarea.get("cliente", ""))
        self.combo_cliente = ctk.CTkComboBox(acli, variable=self.v_cliente, values=empresas, height=32)
        self.combo_cliente.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(acli, text="🔍", width=38, height=32, fg_color=NAVY, hover_color=NAVY2,
                      command=self._buscar_cliente_tarea).pack(side="left", padx=(4, 0))
        ctk.CTkButton(acli, text="+ Nuevo", width=70, height=32, fg_color=GREEN, hover_color=GREEN_H,
                      font=("Segoe UI", 11, "bold"),
                      command=self._nuevo_cliente_tarea).pack(side="left", padx=(4, 0))
        ctk.CTkLabel(b, text="Contacto (vendedor)", text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.v_contacto = tk.StringVar(value=tarea.get("contacto", ""))
        self.combo_contacto = ctk.CTkComboBox(
            b, variable=self.v_contacto, height=32,
            values=([tarea.get("contacto", "")] if tarea.get("contacto") else []))
        self.combo_contacto.pack(fill="x", padx=2)
        f2b = ctk.CTkFrame(cont, fg_color="transparent"); f2b.pack(fill="x", pady=(6, 0))
        ra = ctk.CTkFrame(f2b, fg_color="transparent"); ra.pack(side="left", fill="x", expand=True, padx=(0, 6))
        rb = ctk.CTkFrame(f2b, fg_color="transparent"); rb.pack(side="left", fill="x", expand=True, padx=(6, 0))
        ctk.CTkLabel(ra, text="Responsable", text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        resp = [x.get("nombre", "") for x in asesores_reservas(cfg)] + [c[0] for c in COTIZADORES]
        resp = sorted(set([r for r in resp if r]))
        self.v_resp = tk.StringVar(value=tarea.get("responsable", ""))
        ctk.CTkComboBox(ra, variable=self.v_resp, values=resp, height=32).pack(fill="x", padx=2)
        ctk.CTkLabel(rb, text="Estado del cliente", text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.v_cli_estado = tk.StringVar(value=tarea.get("cliente_estado", "Sin clasificar"))
        ctk.CTkOptionMenu(rb, variable=self.v_cli_estado, values=ESTADOS_CLIENTE, height=32,
                          fg_color=NAVY, button_color=NAVY2).pack(fill="x", padx=2)
        ctk.CTkLabel(cont, text="Motivo / observacion del estado del cliente", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=2, pady=(6, 0))
        self.v_cli_motivo = tk.StringVar(value=tarea.get("cliente_motivo", ""))
        ctk.CTkEntry(cont, textvariable=self.v_cli_motivo, height=32,
                     placeholder_text="Ej. compra activa mensual / pendiente de decision / precio alto...").pack(
            fill="x", padx=2, pady=(0, 4))

        # estado + prioridad + fecha limite
        f3 = ctk.CTkFrame(cont, fg_color="transparent"); f3.pack(fill="x", pady=(8, 0))
        c1 = ctk.CTkFrame(f3, fg_color="transparent"); c1.pack(side="left", padx=(0, 12))
        ctk.CTkLabel(c1, text="Estado", text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w")
        self.v_estado = tk.StringVar(value=tarea.get("estado", "Pendiente"))
        ctk.CTkOptionMenu(c1, variable=self.v_estado, values=ESTADOS_TAREA, width=160, height=32,
                          fg_color=NAVY, button_color=NAVY2).pack()
        c2 = ctk.CTkFrame(f3, fg_color="transparent"); c2.pack(side="left", padx=(0, 12))
        ctk.CTkLabel(c2, text="Prioridad", text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w")
        self.v_prio = tk.StringVar(value=tarea.get("prioridad", "Media"))
        ctk.CTkOptionMenu(c2, variable=self.v_prio, values=PRIORIDADES_TAREA, width=130, height=32,
                          fg_color=NAVY, button_color=NAVY2).pack()
        c3 = ctk.CTkFrame(f3, fg_color="transparent"); c3.pack(side="left")
        ctk.CTkLabel(c3, text="Fecha limite", text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w")
        self.sel_fecha = SelectorFecha(c3)
        self.sel_fecha.pack()
        fl = _parse_fecha_iso(tarea.get("fecha_limite", ""))
        if fl:
            self.sel_fecha._set(fl)

        ctk.CTkLabel(cont, text="Descripcion", text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2, pady=(8, 0))
        self.txt_desc = ctk.CTkTextbox(cont, height=80, corner_radius=8, border_width=1,
                                       border_color=LINE, fg_color=CARD, font=("Segoe UI", 12))
        self.txt_desc.pack(fill="x", padx=2, pady=(0, 6))
        self.txt_desc.insert("1.0", tarea.get("descripcion", "") or "")

        # checklist (gestion)
        ch = ctk.CTkFrame(cont, fg_color="transparent"); ch.pack(fill="x", pady=(6, 0))
        ctk.CTkLabel(ch, text="LISTA DE VERIFICACION (gestion)", text_color=NAVY,
                     font=("Segoe UI", 13, "bold")).pack(side="left")
        ctk.CTkButton(ch, text="+ Agregar elemento", width=160, height=28, fg_color=BLUE,
                      hover_color=BLUE_H, command=self._agregar_check).pack(side="right")
        self.check_box = ctk.CTkFrame(cont, fg_color="transparent"); self.check_box.pack(fill="x", pady=(4, 6))
        self._pintar_checklist()

        ctk.CTkLabel(cont, text="Notas de gestion", text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2, pady=(6, 0))
        self.txt_notas = ctk.CTkTextbox(cont, height=80, corner_radius=8, border_width=1,
                                        border_color=LINE, fg_color=CARD, font=("Segoe UI", 12))
        self.txt_notas.pack(fill="x", padx=2, pady=(0, 8))
        self.txt_notas.insert("1.0", tarea.get("notas", "") or "")

        ctk.CTkButton(self.footer, text="💾  Guardar tarea", height=40, width=190, fg_color=GREEN,
                      hover_color=GREEN_H, font=("Segoe UI", 13, "bold"),
                      command=self._guardar).pack(side="right", padx=(8, 16), pady=9)
        ctk.CTkButton(self.footer, text="✓ Marcar completada", height=40, fg_color=NAVY,
                      hover_color=NAVY2, font=("Segoe UI", 12, "bold"),
                      command=self._completar).pack(side="right", padx=8, pady=9)
        ctk.CTkButton(self.footer, text="Cancelar", height=40, width=100, fg_color=CARD2,
                      text_color=NAVY, hover_color=LINE, command=self.destroy).pack(
            side="left", padx=(16, 8), pady=9)

    def _max(self):
        # El gestor de ventanas maximiza respetando la barra de tareas; como los
        # botones estan en la barra fija de ARRIBA, siempre quedan visibles.
        try:
            self.state("zoomed")
        except Exception:
            pass

    def _pintar_checklist(self):
        for w in self.check_box.winfo_children():
            w.destroy()
        self.check_widgets = []
        ch = self.tarea.setdefault("checklist", [])
        if not ch:
            ctk.CTkLabel(self.check_box, text="Sin elementos. Agrega pasos de gestion (ej. "
                         "'Solicitar cita', 'Enviar propuesta', 'Hacer seguimiento').",
                         text_color=MUTED, font=("Segoe UI", 10)).pack(pady=4)
        for i, it in enumerate(ch):
            row = ctk.CTkFrame(self.check_box, fg_color=CARD, corner_radius=8,
                               border_width=1, border_color=LINE)
            row.pack(fill="x", pady=2)
            v_ok = tk.BooleanVar(value=bool(it.get("hecha")))
            v_txt = tk.StringVar(value=it.get("texto", ""))
            ctk.CTkCheckBox(row, text="", variable=v_ok, width=28,
                            command=self._sync_checklist).pack(side="left", padx=(8, 2), pady=5)
            ctk.CTkEntry(row, textvariable=v_txt, height=30, placeholder_text="Paso de gestion").pack(
                side="left", fill="x", expand=True, padx=4)
            ctk.CTkButton(row, text="✕", width=28, height=28, fg_color=RED, hover_color="#9B2C22",
                          command=lambda idx=i: self._quitar_check(idx)).pack(side="left", padx=(2, 6))
            self.check_widgets.append({"hecha": v_ok, "texto": v_txt})

    def _sync_checklist(self):
        self.tarea["checklist"] = [{"texto": w["texto"].get().strip(), "hecha": bool(w["hecha"].get())}
                                   for w in self.check_widgets]

    def _agregar_check(self):
        self._sync_checklist()
        self.tarea.setdefault("checklist", []).append({"texto": "", "hecha": False})
        self._pintar_checklist()

    def _quitar_check(self, i):
        self._sync_checklist()
        try:
            del self.tarea["checklist"][i]
        except Exception:
            pass
        self._pintar_checklist()

    def _buscar_cliente_tarea(self):
        SelectorContacto(self, self._usar_cliente_tarea)

    def _usar_cliente_tarea(self, c, v):
        self.v_cliente.set(c.get("empresa", ""))
        vends = c.get("vendedores", []) or []
        nombres = [x.get("nombre", "") for x in vends if x.get("nombre")]
        try:
            self.combo_contacto.configure(values=nombres or [""])
        except Exception:
            pass
        if v and v.get("nombre"):
            self.v_contacto.set(v.get("nombre", ""))
        elif nombres:
            self.v_contacto.set(nombres[0])

    def _nuevo_cliente_tarea(self):
        VentanaClientes(self, on_cambio=self._refrescar_empresas,
                        preseleccion=self.v_cliente.get().strip() or None)

    def _refrescar_empresas(self):
        try:
            empresas = sorted([c.get("empresa", "") for c in cargar_clientes() if c.get("empresa")])
            self.combo_cliente.configure(values=empresas)
        except Exception:
            pass

    def _recoger(self):
        self._sync_checklist()
        return {
            "titulo": self.v_titulo.get().strip(),
            "cliente": self.v_cliente.get().strip(),
            "contacto": self.v_contacto.get().strip(),
            "cliente_estado": self.v_cli_estado.get(),
            "cliente_motivo": self.v_cli_motivo.get().strip(),
            "responsable": self.v_resp.get().strip(),
            "estado": self.v_estado.get(),
            "prioridad": self.v_prio.get(),
            "fecha_limite": (self.sel_fecha.get().isoformat() if self.sel_fecha.get() else ""),
            "descripcion": self.txt_desc.get("1.0", "end").strip(),
            "notas": self.txt_notas.get("1.0", "end").strip(),
            "checklist": self.tarea.get("checklist", []),
        }

    def _guardar(self, estado=None):
        datos = self._recoger()
        if estado:
            datos["estado"] = estado
        if not datos["titulo"]:
            messagebox.showwarning("Falta el titulo", "Escribe el titulo de la tarea.", parent=self)
            return
        if self.tarea.get("numero"):
            actualizar_tarea(self.tarea["numero"], datos)
        else:
            self.tarea = registrar_tarea(datos)
        if self.on_save:
            self.on_save()
        messagebox.showinfo("Guardado", "Tarea guardada.", parent=self)
        self.destroy()

    def _completar(self):
        self._guardar(estado="Completada")


class ModuloComercial(ctk.CTkToplevel):
    """Modulo Comercial: tareas de gestion + indicadores."""
    def __init__(self, master=None):
        super().__init__(master)
        ctk.set_appearance_mode("light")
        try:
            ctk.set_widget_scaling(0.85)
        except Exception:
            pass
        self.cfg = cargar_config()
        try:
            self.iconbitmap(recurso("app.ico"))
        except Exception:
            pass
        self.title(f"Comercial - INNOBA Colombia DMC   v{VERSION}")
        self.configure(fg_color=BG)
        self.geometry("1180x720")
        self.filtro = tk.StringVar(value="Todas")
        self._build()
        self.after(60, self._max)

    def _max(self):
        try:
            self.state("zoomed")
        except Exception:
            pass

    def _volver_inicio(self):
        lanz = self.master
        try:
            if lanz is not None and hasattr(lanz, "comercial"):
                lanz.comercial = None
        except Exception:
            pass
        try:
            self.destroy()
        except Exception:
            pass
        try:
            if lanz is not None:
                lanz.deiconify(); lanz.lift(); lanz.focus_force()
        except Exception:
            pass

    def _build(self):
        head = ctk.CTkFrame(self, fg_color=CARD, corner_radius=0, height=60)
        head.pack(fill="x"); head.pack_propagate(False); head.grid_columnconfigure(1, weight=1)
        try:
            img = Image.open(recurso("logo_innoba.png")); w, h = img.size; hh = 38
            self.logo_img = ctk.CTkImage(light_image=img, size=(int(w * hh / h), hh))
            ctk.CTkLabel(head, image=self.logo_img, text="").grid(row=0, column=0, padx=(18, 12), pady=8)
        except Exception:
            ctk.CTkLabel(head, text="INNOBA", font=("Segoe UI", 20, "bold"), text_color=NAVY).grid(row=0, column=0, padx=18)
        tit = ctk.CTkFrame(head, fg_color="transparent"); tit.grid(row=0, column=1, sticky="w")
        ctk.CTkLabel(tit, text="Modulo Comercial", text_color=NAVY,
                     font=("Segoe UI", 17, "bold"), height=20).pack(anchor="w")
        ctk.CTkLabel(tit, text=f"INNOBA Colombia DMC  ·  v{VERSION}", text_color=MUTED,
                     font=("Segoe UI", 11), height=15).pack(anchor="w")
        hb = ctk.CTkFrame(head, fg_color="transparent"); hb.grid(row=0, column=2, padx=18)
        ctk.CTkButton(hb, text="⌂ Modulos", width=100, height=36, corner_radius=10, fg_color=NAVY,
                      hover_color=NAVY2, font=("Segoe UI", 12, "bold"),
                      command=self._volver_inicio).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hb, text="+ Nueva tarea", width=140, height=36, corner_radius=10, fg_color=GREEN,
                      hover_color=GREEN_H, font=("Segoe UI", 12, "bold"),
                      command=self._nueva_tarea).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hb, text="🏆 Contactos", width=130, height=36, corner_radius=10,
                      fg_color="#D9A400", hover_color="#B7791F", font=("Segoe UI", 12, "bold"),
                      command=self._ranking_contactos).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hb, text="📊 Reporte", width=110, height=36, corner_radius=10, fg_color="#7A5AB5",
                      hover_color="#63459A", font=("Segoe UI", 12, "bold"),
                      command=self._reporte).pack(side="left")

        # Indicadores
        self.kpis = ctk.CTkFrame(self, fg_color="transparent"); self.kpis.pack(fill="x", padx=18, pady=(12, 4))
        # Barra de busqueda / filtro
        bar = ctk.CTkFrame(self, fg_color="transparent"); bar.pack(fill="x", padx=18, pady=(4, 4))
        self.q = tk.StringVar()
        e = ctk.CTkEntry(bar, textvariable=self.q, height=36, corner_radius=10,
                         placeholder_text="Buscar tarea por titulo, cliente o responsable...")
        e.pack(side="left", fill="x", expand=True)
        e.bind("<KeyRelease>", lambda ev: self._pintar())
        ctk.CTkLabel(bar, text="Ver:", text_color=MUTED).pack(side="left", padx=(10, 4))
        ctk.CTkOptionMenu(bar, variable=self.filtro,
                          values=["Todas", "Pendiente", "En progreso", "Vencida", "Completada"],
                          width=150, height=36, fg_color=NAVY, button_color=NAVY2,
                          command=lambda _v=None: self._pintar()).pack(side="left")

        self.lista = ctk.CTkScrollableFrame(self, fg_color=BG); self.lista.pack(fill="both", expand=True, padx=18, pady=(4, 14))
        self._refrescar()

    def _refrescar(self):
        self._pintar_kpis()
        self._pintar()

    def _kpi(self, parent, titulo, valor, color, ancho=150):
        return _kpi_card(parent, titulo, valor, color, ancho=ancho)

    def _pintar_kpis(self):
        for w in self.kpis.winfo_children():
            w.destroy()
        ind = indicadores_comerciales()
        fila1 = ctk.CTkFrame(self.kpis, fg_color="transparent"); fila1.pack(fill="x")
        self._kpi(fila1, "Tareas totales", ind["tareas_total"], NAVY)
        self._kpi(fila1, "Pendientes", ind["pendientes"], MUTED)
        self._kpi(fila1, "En progreso", ind["en_progreso"], BLUE)
        self._kpi(fila1, "Vencidas", ind["vencidas"], RED)
        self._kpi(fila1, "Completadas", ind["completadas"], GREEN)
        self._kpi(fila1, f"Ventas del mes ({ind['ventas_mes_n']})",
                  usd(ind["ventas_mes_usd"]), GREEN_H, ancho=180)
        self._kpi(fila1, "Conversion cotiz.", f"{ind['conversion']}%", "#7A5AB5")
        self._kpi(fila1, f"Reservas del mes ({ind['reservas_mes_n']})",
                  usd(ind["reservas_mes_usd"]), NAVY, ancho=180)

    def _pintar(self):
        for w in self.lista.winfo_children():
            w.destroy()
        items = list(reversed(cargar_tareas().get("items", [])))
        q = self.q.get().lower().strip()
        filt = self.filtro.get()
        vistos = 0
        for t in items:
            est = estado_tarea_efectivo(t)
            if filt != "Todas" and est != filt:
                continue
            if q and q not in json.dumps(t, ensure_ascii=False).lower():
                continue
            vistos += 1
            self._fila(t, est)
        if vistos == 0:
            ctk.CTkLabel(self.lista, text="Sin tareas. Crea una con '+ Nueva tarea'.",
                         text_color=MUTED).pack(pady=24)

    def _fila(self, t, est):
        fila = ctk.CTkFrame(self.lista, fg_color=ESTADO_TAREA_FILA.get(est, CARD2), corner_radius=10)
        fila.pack(fill="x", pady=4, padx=2); fila.grid_columnconfigure(1, weight=1)
        prio = t.get("prioridad", "Media")
        ctk.CTkLabel(fila, text=prio.upper(), text_color="#FFFFFF",
                     fg_color=PRIORIDAD_COLOR.get(prio, MUTED), corner_radius=6,
                     font=("Segoe UI", 9, "bold"), width=60).grid(row=0, column=0, rowspan=2, padx=(12, 6), pady=8, ipady=4)
        info = ctk.CTkFrame(fila, fg_color="transparent"); info.grid(row=0, column=1, rowspan=2, sticky="w")
        ctk.CTkLabel(info, text=t.get("titulo", "(sin titulo)"), text_color=TEXT,
                     font=("Segoe UI", 13, "bold")).pack(anchor="w")
        h, n = _progreso_checklist(t)
        con = t.get("contacto", "")
        sub = f"{t.get('cliente','') or '(sin cliente)'}"
        if con:
            sub += f" ({con})"
        sub += f"   ·   Resp: {t.get('responsable','') or '-'}"
        if n:
            sub += f"   ·   Checklist {h}/{n}"
        ctk.CTkLabel(info, text=sub, text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w")
        ce = t.get("cliente_estado", "")
        if ce and ce != "Sin clasificar":
            col = {"Cliente actual (vigente en compra)": GREEN, "Descartado": RED}.get(ce, "#D9A400")
            ctk.CTkLabel(info, text="● " + ce, text_color=col, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        fl = _parse_fecha_iso(t.get("fecha_limite", ""))
        ftxt = ("Vence: " + fl.strftime("%d/%m/%Y")) if fl else "Sin fecha"
        ctk.CTkLabel(fila, text=ftxt, text_color=(RED if est == "Vencida" else MUTED),
                     font=("Segoe UI", 11, "bold" if est == "Vencida" else "normal")).grid(row=0, column=2, rowspan=2, padx=10)
        ctk.CTkLabel(fila, text=est, fg_color=ESTADO_TAREA_COLOR.get(est, MUTED), text_color="#FFFFFF",
                     corner_radius=6, font=("Segoe UI", 11, "bold")).grid(row=0, column=3, rowspan=2, padx=8, ipadx=8, ipady=3)
        btns = ctk.CTkFrame(fila, fg_color="transparent"); btns.grid(row=0, column=4, rowspan=2, padx=10)
        ctk.CTkButton(btns, text="Abrir", width=80, height=30, fg_color=NAVY, hover_color=NAVY2,
                      command=lambda x=t: self._abrir(x)).pack(pady=2)
        b2 = ctk.CTkFrame(btns, fg_color="transparent"); b2.pack()
        if est != "Completada":
            ctk.CTkButton(b2, text="✓", width=36, height=28, fg_color=GREEN, hover_color=GREEN_H,
                          command=lambda x=t: self._completar(x)).pack(side="left", padx=2)
        ctk.CTkButton(b2, text="🗑", width=36, height=28, fg_color=RED, hover_color="#9B2C22",
                      command=lambda x=t: self._eliminar(x)).pack(side="left", padx=2)

    def _abrir(self, t):
        VentanaTareaDetalle(self, dict(t), self.cfg, on_save=self._refrescar)

    def _completar(self, t):
        actualizar_tarea(t.get("numero", ""), {"estado": "Completada"})
        self._refrescar()

    def _eliminar(self, t):
        if messagebox.askyesno("Eliminar tarea", f"Eliminar la tarea '{t.get('titulo','')}'?"):
            eliminar_tarea(t.get("numero", ""))
            self._refrescar()

    def _nueva_tarea(self):
        VentanaTareaDetalle(self, {"estado": "Pendiente", "prioridad": "Media", "checklist": []},
                            self.cfg, on_save=self._refrescar)

    def _ranking_contactos(self):
        VentanaRankingContactos(self)

    def _reporte(self):
        data = cargar_tareas()
        meses = sorted({_mes_de_iso(it.get("fecha_creacion", "")) for it in data.get("items", [])
                        if it.get("fecha_creacion")}, reverse=True)
        DialogoReporteMes(self, "Reporte de tareas comerciales",
                          "Descarga un Excel con las tareas de gestion.",
                          meses, exportar_reporte_tareas, "Reporte_tareas")


class VentanaRankingContactos(ctk.CTkToplevel):
    """Ranking de contactos por reservas (mes / ano) para premiar a los mejores."""
    def __init__(self, master):
        super().__init__(master)
        self.title("Ranking de contactos")
        self.geometry("900x640"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        top = ctk.CTkFrame(self, fg_color="transparent"); top.pack(fill="x", padx=16, pady=(14, 4))
        ctk.CTkLabel(top, text="🏆  Ranking de contactos por reservas",
                     text_color=NAVY, font=("Segoe UI", 16, "bold")).pack(side="left")
        ctk.CTkButton(top, text="⬇ Descargar Excel", height=34, fg_color=GREEN, hover_color=GREEN_H,
                      command=self._exportar).pack(side="right")
        ctk.CTkLabel(self, text="Cuenta las reservas que cada contacto (vendedor de agencia) ha "
                     "hecho con nosotros. Ordenado por reservas del ano.", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w", padx=16)
        # encabezado
        hdr = ctk.CTkFrame(self, fg_color=NAVY, corner_radius=6); hdr.pack(fill="x", padx=16, pady=(8, 0))
        cols = [("#", 40), ("Contacto", 220), ("Agencia(s)", 260), ("Mes", 70),
                ("Ano", 70), ("Monto ano", 120)]
        for txt, w in cols:
            ctk.CTkLabel(hdr, text=txt, text_color="#FFFFFF", font=("Segoe UI", 11, "bold"),
                         width=w, anchor="w").pack(side="left", padx=6, pady=6)
        self.box = ctk.CTkScrollableFrame(self, fg_color=CARD)
        self.box.pack(fill="both", expand=True, padx=16, pady=(2, 14))
        self._pintar()

    def _pintar(self):
        for w in self.box.winfo_children():
            w.destroy()
        filas = ranking_contactos()
        if not filas:
            ctk.CTkLabel(self.box, text="Aun no hay reservas con contacto asignado.",
                         text_color=MUTED).pack(pady=24)
            return
        medallas = {1: "🥇", 2: "🥈", 3: "🥉"}
        for i, f in enumerate(filas, 1):
            fg = "#FFF7DB" if i <= 3 else CARD2
            row = ctk.CTkFrame(self.box, fg_color=fg, corner_radius=8); row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=medallas.get(i, str(i)), text_color=NAVY,
                         font=("Segoe UI", 13, "bold"), width=40, anchor="w").pack(side="left", padx=6, pady=6)
            ctk.CTkLabel(row, text=f["contacto"], text_color=NAVY, font=("Segoe UI", 12, "bold"),
                         width=220, anchor="w").pack(side="left", padx=6)
            ctk.CTkLabel(row, text=f["empresas"] or "-", text_color=MUTED, font=("Segoe UI", 10),
                         width=260, anchor="w", wraplength=250, justify="left").pack(side="left", padx=6)
            ctk.CTkLabel(row, text=str(f["mes_n"]), text_color=BLUE, font=("Segoe UI", 12, "bold"),
                         width=70, anchor="w").pack(side="left", padx=6)
            ctk.CTkLabel(row, text=str(f["ano_n"]), text_color=GREEN_H, font=("Segoe UI", 12, "bold"),
                         width=70, anchor="w").pack(side="left", padx=6)
            ctk.CTkLabel(row, text=usd(f["ano_usd"]), text_color=NAVY, font=("Segoe UI", 11),
                         width=120, anchor="w").pack(side="left", padx=6)

    def _exportar(self):
        ruta = filedialog.asksaveasfilename(
            title="Guardar ranking", defaultextension=".xlsx",
            initialfile="Ranking_contactos.xlsx", filetypes=[("Excel", "*.xlsx")])
        if not ruta:
            return
        try:
            exportar_reporte_contactos(ruta)
            os.startfile(ruta)
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)


# ============================================================================
# LIQUIDACION DE RENTABILIDAD (dentro de Reservas)
# ============================================================================
def exportar_liquidacion_excel(ruta, mes, cfg):
    from openpyxl import Workbook
    trm_mes = trm_del_mes(cfg, mes)
    reservas = [r for r in cargar_reservas().get("items", [])
                if _mes_de_iso(r.get("fecha_creacion", "")) == mes
                and r.get("estado") != "Anulada"]
    wb = Workbook(); ws = wb.active; ws.title = "Liquidacion"
    heads = (["N. Reserva", "Cliente", "Responsable", "Monto USD", "TRM", "Ventas (pesos)"]
             + [lbl for _, lbl in GASTOS_OPER]
             + ["Comision", "IVA Comision", "Retefuente", "ReteICA", "4x1000",
                "Total Gasto", "Utilidad", "Rent %", "Bonif 3%"])
    ws.append(heads)
    tot = {"ventas": 0.0, "total_gasto": 0.0, "utilidad": 0.0, "bonif": 0.0}
    porase = {}
    for r in reservas:
        L = liquidar_reserva(r, cfg, trm_default=trm_mes); liq = r.get("liq", {}) or {}
        ase = (r.get("asesor", {}) or {}).get("nombre", "") if isinstance(r.get("asesor"), dict) else str(r.get("asesor", ""))
        fila = [r.get("numero", ""), r.get("cliente", ""), ase,
                float(r.get("monto", 0) or 0), round(L["trm"], 2), round(L["ventas"])]
        fila += [float(liq.get(k, 0) or 0) for k, _ in GASTOS_OPER]
        fila += [round(L["comision"]), round(L["iva_comision"]), round(L["retefuente"]),
                 round(L["reteica"]), round(L["cuatromil"]), round(L["total_gasto"]),
                 round(L["utilidad"]), round(L["rent"] * 100, 2), round(L["bonif"])]
        ws.append(fila)
        for k in tot:
            tot[k] += L[k]
        porase[ase] = porase.get(ase, 0.0) + L["bonif"]
    ws.append([])
    fila_tot = ["TOTALES", "", "", "", "", round(tot["ventas"])] + [""] * len(GASTOS_OPER)
    fila_tot += ["", "", "", "", "", round(tot["total_gasto"]), round(tot["utilidad"]), "", round(tot["bonif"])]
    ws.append(fila_tot)
    ws.append([])
    ws.append(["BONIFICACION POR ASESORA (3% de la utilidad)"])
    for ase, b in sorted(porase.items(), key=lambda x: -x[1]):
        ws.append([ase or "(sin asignar)", round(b)])
    wb.save(ruta)
    return ruta


class DialogoGastosReserva(ctk.CTkToplevel):
    """Editar el TRM y los gastos operativos de una reserva y ver su liquidacion en vivo."""
    def __init__(self, master, res, cfg, on_save=None, trm_default=None):
        super().__init__(master)
        self.res = res; self.cfg = cfg; self.on_save = on_save
        self.trm_default = trm_default
        self.title("Liquidacion reserva " + res.get("numero", ""))
        self.geometry("560x680"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        cont = ctk.CTkScrollableFrame(self, fg_color=BG); cont.pack(fill="both", expand=True, padx=16, pady=14)
        ctk.CTkLabel(cont, text=f"Reserva N. {res.get('numero','')}  ·  {res.get('cliente','')}",
                     text_color=NAVY, font=("Segoe UI", 15, "bold")).pack(anchor="w")
        ctk.CTkLabel(cont, text=f"Monto de venta: {usd(res.get('monto',0))}",
                     text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", pady=(0, 8))
        liq = res.get("liq", {}) or {}
        self.vars = {}
        self.evid = dict(liq.get("evid", {}) or {})   # soporte por gasto
        self.evid_lbls = {}

        def campo(clave, etq, val, ph=None, adjunto=None):
            f = ctk.CTkFrame(cont, fg_color="transparent"); f.pack(fill="x", pady=2)
            ctk.CTkLabel(f, text=etq, text_color=TEXT, width=130, anchor="w",
                         font=("Segoe UI", 11)).pack(side="left")
            v = tk.StringVar(value=str(val if val not in (None, "") else "")); self.vars[clave] = v
            e = ctk.CTkEntry(f, textvariable=v, height=30, width=130,
                             placeholder_text=(ph or ""))
            e.pack(side="left"); e.bind("<KeyRelease>", lambda ev: self._recalc())
            if adjunto:
                ctk.CTkButton(f, text="📎 Soportes", width=90, height=30, fg_color=CARD2,
                              text_color=NAVY, hover_color=LINE, border_width=1, border_color=LINE,
                              font=("Segoe UI", 10, "bold"),
                              command=lambda kk=adjunto: self._gestionar_evid(kk)).pack(side="left", padx=4)
                lb = ctk.CTkLabel(f, text="", text_color=GREEN_H, font=("Segoe UI", 10, "bold"),
                                  cursor="hand2")
                lb.pack(side="left")
                lb.bind("<Button-1>", lambda ev, kk=adjunto: self._gestionar_evid(kk))
                self.evid_lbls[adjunto] = lb
            return v

        trm_prom = self.trm_default or cfg.get("trm_liq", 4000)
        campo("trm", "TRM (vacio = prom. mes)", liq.get("trm", ""),
              ph=f"Prom. mes: {trm_prom:g}")
        ctk.CTkLabel(cont, text="GASTOS OPERATIVOS (en pesos)  ·  📎 adjunta el soporte de cada uno",
                     text_color=NAVY, font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(8, 2))
        for k, lbl in GASTOS_OPER:
            campo(k, lbl, liq.get(k, ""), adjunto=k)
        self._refrescar_evid()
        ctk.CTkLabel(cont, text="RESULTADO", text_color=NAVY,
                     font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(10, 2))
        self.box_res = ctk.CTkFrame(cont, fg_color=CARD2, corner_radius=10)
        self.box_res.pack(fill="x", pady=(0, 6))
        self.lbls = {}
        for k, lbl in [("ventas", "Ventas"), ("operativos", "Gastos operativos"),
                       ("financieros", "Gastos financieros"), ("total_gasto", "Total gasto"),
                       ("utilidad", "UTILIDAD"), ("rent", "Rentabilidad"), ("bonif", "Bonificacion 3%")]:
            f = ctk.CTkFrame(self.box_res, fg_color="transparent"); f.pack(fill="x", padx=10, pady=1)
            ctk.CTkLabel(f, text=lbl, text_color=TEXT, width=180, anchor="w",
                         font=("Segoe UI", 11, "bold" if k in ("utilidad", "bonif") else "normal")).pack(side="left")
            self.lbls[k] = ctk.CTkLabel(f, text="", text_color=(GREEN_H if k in ("utilidad", "bonif") else NAVY),
                                        anchor="e", font=("Segoe UI", 12, "bold"))
            self.lbls[k].pack(side="right", padx=8)
        bts = ctk.CTkFrame(cont, fg_color="transparent"); bts.pack(fill="x", pady=10)
        ctk.CTkButton(bts, text="💾 Guardar", fg_color=GREEN, hover_color=GREEN_H,
                      font=("Segoe UI", 13, "bold"), command=self._guardar).pack(side="right")
        ctk.CTkButton(bts, text="Cancelar", fg_color=CARD2, text_color=NAVY, hover_color=LINE,
                      command=self.destroy).pack(side="right", padx=8)
        self._recalc()

    def _carpeta_evid(self):
        ruta = os.path.join(datos_dir(), "adjuntos",
                            self.res.get("numero", "tmp"), "gastos")
        os.makedirs(ruta, exist_ok=True)
        return ruta

    def _evid_list(self, k):
        v = self.evid.get(k)
        if not v:
            return []
        return [v] if isinstance(v, str) else list(v)

    def _adjuntar_gasto(self, k):
        etq = dict(GASTOS_OPER).get(k, k)
        rutas = filedialog.askopenfilenames(
            title=f"Evidencia(s) del gasto: {etq}  (puedes elegir varias)",
            filetypes=[("Documentos", "*.pdf *.jpg *.jpeg *.png"), ("Todos", "*.*")])
        if not rutas:
            return
        lst = self._evid_list(k)
        for ruta in rutas:
            destino = os.path.join(self._carpeta_evid(),
                                   f"{k}_{len(lst)+1}_{os.path.basename(ruta)}")
            try:
                shutil.copyfile(ruta, destino)
                lst.append(destino)
            except Exception as e:
                messagebox.showerror("No se pudo adjuntar", str(e), parent=self)
        self.evid[k] = lst
        self._refrescar_evid()

    def _gestionar_evid(self, k):
        etq = dict(GASTOS_OPER).get(k, k)
        win = ctk.CTkToplevel(self); win.title("Soportes: " + etq)
        win.geometry("560x440"); win.configure(fg_color=BG)
        win.transient(self); win.grab_set()
        ctk.CTkLabel(win, text="Soportes del gasto: " + etq, text_color=NAVY,
                     font=("Segoe UI", 15, "bold")).pack(anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(win, text="Adjunta las facturas / recibos que respaldan este gasto.",
                     text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=16, pady=(0, 8))
        box = ctk.CTkScrollableFrame(win, fg_color=CARD, corner_radius=10)
        box.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        def pintar():
            for w in box.winfo_children():
                w.destroy()
            lst = self._evid_list(k)
            if not lst:
                ctk.CTkLabel(box, text="Sin soportes. Usa '+ Agregar soporte'.",
                             text_color=MUTED).pack(pady=16)
            for i, p in enumerate(lst):
                row = ctk.CTkFrame(box, fg_color=CARD2, corner_radius=8); row.pack(fill="x", pady=2, padx=2)
                ctk.CTkLabel(row, text=f"{i+1}. {os.path.basename(p)}", text_color=NAVY, anchor="w",
                             font=("Segoe UI", 11)).pack(side="left", fill="x", expand=True, padx=8, pady=6)
                ctk.CTkButton(row, text="Abrir", width=64, height=28, fg_color=NAVY, hover_color=NAVY2,
                              command=lambda pp=p: self._abrir_evid(pp)).pack(side="left", padx=3)
                ctk.CTkButton(row, text="Quitar", width=64, height=28, fg_color=RED, hover_color="#9B2C22",
                              command=lambda idx=i: quitar(idx)).pack(side="left", padx=(3, 8))

        def agregar():
            self._adjuntar_gasto(k); pintar()

        def quitar(idx):
            lst = self._evid_list(k)
            try:
                del lst[idx]
            except Exception:
                pass
            self.evid[k] = lst; self._refrescar_evid(); pintar()

        bar = ctk.CTkFrame(win, fg_color="transparent"); bar.pack(fill="x", padx=16, pady=(0, 14))
        ctk.CTkButton(bar, text="+ Agregar soporte", height=36, fg_color=GREEN, hover_color=GREEN_H,
                      font=("Segoe UI", 12, "bold"), command=agregar).pack(side="left")
        ctk.CTkButton(bar, text="Cerrar", height=36, fg_color=NAVY, hover_color=NAVY2,
                      command=win.destroy).pack(side="right")
        pintar()

    def _abrir_evid(self, p):
        if p and os.path.exists(p):
            try:
                os.startfile(p)
            except Exception as e:
                messagebox.showerror("No se pudo abrir", str(e), parent=self)
        else:
            messagebox.showinfo("Soporte", "El archivo ya no existe en la ruta guardada.", parent=self)

    def _refrescar_evid(self):
        for k, lb in self.evid_lbls.items():
            n = len(self._evid_list(k))
            lb.configure(text=(f"✓ {n} soporte(s)" if n else ""))

    def _leer(self):
        liq = {}
        for k, v in self.vars.items():
            liq[k] = _num(v.get())
        liq["evid"] = self.evid
        return liq

    def _recalc(self):
        tmp = dict(self.res); tmp["liq"] = self._leer()
        L = liquidar_reserva(tmp, self.cfg, trm_default=self.trm_default)
        self.lbls["ventas"].configure(text=_cop(L["ventas"]))
        self.lbls["operativos"].configure(text=_cop(L["operativos"]))
        self.lbls["financieros"].configure(text=_cop(L["financieros"]))
        self.lbls["total_gasto"].configure(text=_cop(L["total_gasto"]))
        self.lbls["utilidad"].configure(text=_cop(L["utilidad"]),
                                         text_color=(GREEN_H if L["utilidad"] >= 0 else RED))
        self.lbls["rent"].configure(text=f"{L['rent'] * 100:,.2f}%")
        self.lbls["bonif"].configure(text=_cop(L["bonif"]))

    def _guardar(self):
        self.res["liq"] = self._leer()
        actualizar_reserva(self.res.get("numero", ""), {"liq": self.res["liq"]})
        if self.on_save:
            self.on_save()
        self.destroy()


class VentanaLiquidacion(ctk.CTkToplevel):
    """Liquidacion de rentabilidad por mes: ventas - gastos - financieros = utilidad,
       rentabilidad % y bonificacion (3%). Alimentada por las reservas."""
    def __init__(self, master, cfg):
        super().__init__(master)
        self.cfg = cfg
        self.title("Liquidacion de rentabilidad"); self.configure(fg_color=BG)
        self.geometry("1200x720+30+10"); self.transient(master); self.grab_set()
        self.after(60, self._max)
        head = ctk.CTkFrame(self, fg_color=CARD, corner_radius=0, height=56)
        head.pack(fill="x"); head.pack_propagate(False)
        ctk.CTkLabel(head, text="Liquidacion de rentabilidad", text_color=NAVY,
                     font=("Segoe UI", 16, "bold")).pack(side="left", padx=18)
        ctk.CTkLabel(head, text="Mes:", text_color=MUTED, font=("Segoe UI", 12)).pack(side="left", padx=(20, 4))
        meses = sorted({_mes_de_iso(r.get("fecha_creacion", "")) for r in cargar_reservas().get("items", [])
                        if r.get("fecha_creacion")}, reverse=True)
        if not meses:
            meses = [datetime.date.today().strftime("%Y-%m")]
        self.v_mes = tk.StringVar(value=meses[0])
        ctk.CTkOptionMenu(head, variable=self.v_mes, values=meses, width=140, height=34,
                          fg_color=NAVY, button_color=NAVY2, command=lambda *_: self._cambiar_mes()).pack(side="left")
        ctk.CTkLabel(head, text="TRM del mes:", text_color=MUTED,
                     font=("Segoe UI", 12)).pack(side="left", padx=(18, 4))
        self.v_trm = tk.StringVar(value=str(trm_del_mes(cfg, self.v_mes.get())))
        et = ctk.CTkEntry(head, textvariable=self.v_trm, width=90, height=34)
        et.pack(side="left"); et.bind("<KeyRelease>", lambda e: self._guardar_trm())
        ctk.CTkButton(head, text="↻ TRM hoy", width=90, height=34, fg_color=CYAN, hover_color=BLUE,
                      font=("Segoe UI", 11, "bold"), command=self._trm_hoy).pack(side="left", padx=(4, 0))
        ctk.CTkButton(head, text="⬇ Exportar Excel", height=34, fg_color=GREEN, hover_color=GREEN_H,
                      font=("Segoe UI", 12, "bold"), command=self._exportar).pack(side="right", padx=18)
        self.kpis = ctk.CTkFrame(self, fg_color="transparent"); self.kpis.pack(fill="x", padx=16, pady=(10, 2))
        self.lista = ctk.CTkScrollableFrame(self, fg_color=BG); self.lista.pack(fill="both", expand=True, padx=16, pady=(4, 6))
        self.bonif_box = ctk.CTkFrame(self, fg_color=CARD, corner_radius=10); self.bonif_box.pack(fill="x", padx=16, pady=(0, 12))
        self._pintar()

    def _max(self):
        try:
            self.state("zoomed")
        except Exception:
            pass

    def _cambiar_mes(self):
        self.v_trm.set(str(trm_del_mes(self.cfg, self.v_mes.get())))
        self._pintar()

    def _guardar_trm(self):
        tm = self.cfg.setdefault("trm_mensual", {})
        tm[self.v_mes.get()] = _num(self.v_trm.get())
        try:
            guardar_config(self.cfg)
        except Exception:
            pass
        self._pintar()

    def _trm_hoy(self):
        try:
            v = obtener_trm()
        except Exception:
            v = None
        if v:
            self.v_trm.set(str(int(round(v))))
            self._guardar_trm()
        else:
            messagebox.showinfo("TRM", "No se pudo consultar la TRM de hoy. Escribela a mano.", parent=self)

    def _trm_mes(self):
        try:
            return float(_num(self.v_trm.get())) or trm_del_mes(self.cfg, self.v_mes.get())
        except Exception:
            return trm_del_mes(self.cfg, self.v_mes.get())

    def _reservas_mes(self):
        mes = self.v_mes.get()
        return [r for r in cargar_reservas().get("items", [])
                if _mes_de_iso(r.get("fecha_creacion", "")) == mes and r.get("estado") != "Anulada"]

    def _pintar(self):
        for w in self.lista.winfo_children():
            w.destroy()
        for w in self.kpis.winfo_children():
            w.destroy()
        for w in self.bonif_box.winfo_children():
            w.destroy()
        reservas = self._reservas_mes()
        trm_mes = self._trm_mes()
        tot_v = tot_g = tot_u = tot_b = 0.0
        porase = {}
        # encabezado de tabla
        cab = ctk.CTkFrame(self.lista, fg_color=NAVY, corner_radius=6); cab.pack(fill="x", pady=(0, 3))
        for txt, w in (("N.", 70), ("Cliente", 150), ("Responsable", 120), ("USD", 70),
                       ("Ventas", 110), ("Total gasto", 110), ("Utilidad", 110), ("Rent%", 60),
                       ("Bonif 3%", 100), ("", 90)):
            ctk.CTkLabel(cab, text=txt, text_color="#FFFFFF", font=("Segoe UI", 10, "bold"),
                         width=w, anchor="w").pack(side="left", padx=3, pady=5)
        for r in reservas:
            L = liquidar_reserva(r, self.cfg, trm_default=trm_mes)
            tot_v += L["ventas"]; tot_g += L["total_gasto"]; tot_u += L["utilidad"]; tot_b += L["bonif"]
            ase = (r.get("asesor", {}) or {}).get("nombre", "") if isinstance(r.get("asesor"), dict) else str(r.get("asesor", ""))
            porase[ase] = porase.get(ase, 0.0) + L["bonif"]
            row = ctk.CTkFrame(self.lista, fg_color=CARD, corner_radius=6); row.pack(fill="x", pady=1)
            def cel(txt, w, color=TEXT, bold=False):
                ctk.CTkLabel(row, text=txt, text_color=color, width=w, anchor="w",
                             font=("Segoe UI", 10, "bold" if bold else "normal")).pack(side="left", padx=3, pady=4)
            cel(r.get("numero", ""), 70, NAVY, True)
            cel((r.get("cliente", "") or "")[:22], 150)
            cel((ase or "-")[:18], 120, BLUE)
            cel(f"{float(r.get('monto',0) or 0):,.0f}", 70)
            cel(_cop(L["ventas"]), 110)
            cel(_cop(L["total_gasto"]), 110, "#B7791F")
            cel(_cop(L["utilidad"]), 110, GREEN_H if L["utilidad"] >= 0 else RED, True)
            cel(f"{L['rent']*100:,.1f}%", 60)
            cel(_cop(L["bonif"]), 100, "#7A5AB5", True)
            ctk.CTkButton(row, text="Editar gastos", width=90, height=26, fg_color=CYAN,
                          hover_color=BLUE, font=("Segoe UI", 10),
                          command=lambda x=r: self._editar(x)).pack(side="left", padx=3)
        if not reservas:
            ctk.CTkLabel(self.lista, text="No hay reservas en este mes.", text_color=MUTED).pack(pady=20)
        # KPIs
        fila = ctk.CTkFrame(self.kpis, fg_color="transparent"); fila.pack(fill="x")
        _kpi_card(fila, "Reservas del mes", len(reservas), NAVY)
        _kpi_card(fila, "Ventas (pesos)", _cop(tot_v), NAVY, ancho=170)
        _kpi_card(fila, "Utilidad (pesos)", _cop(tot_u), GREEN_H, ancho=170)
        _kpi_card(fila, "Rentabilidad prom.", f"{(tot_u/tot_v*100) if tot_v else 0:,.1f}%", BLUE, ancho=150)
        _kpi_card(fila, "Bonificacion total", _cop(tot_b), "#7A5AB5", ancho=170)
        # bonificacion por asesora
        ctk.CTkLabel(self.bonif_box, text="Bonificacion por asesora (3% de la utilidad)",
                     text_color=NAVY, font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=(8, 2))
        fb = ctk.CTkFrame(self.bonif_box, fg_color="transparent"); fb.pack(fill="x", padx=12, pady=(0, 10))
        for ase, b in sorted(porase.items(), key=lambda x: -x[1]):
            chip = ctk.CTkFrame(fb, fg_color=CARD2, corner_radius=8); chip.pack(side="left", padx=4)
            ctk.CTkLabel(chip, text=(ase or "(sin asignar)"), text_color=NAVY,
                         font=("Segoe UI", 11, "bold")).pack(padx=10, pady=(5, 0))
            ctk.CTkLabel(chip, text=_cop(b), text_color="#7A5AB5",
                         font=("Segoe UI", 13, "bold")).pack(padx=10, pady=(0, 5))

    def _editar(self, r):
        DialogoGastosReserva(self, r, self.cfg, on_save=self._pintar, trm_default=self._trm_mes())

    def _exportar(self):
        ruta = filedialog.asksaveasfilename(
            title="Guardar liquidacion", defaultextension=".xlsx",
            initialfile=f"Liquidacion_{self.v_mes.get()}.xlsx", filetypes=[("Excel", "*.xlsx")])
        if not ruta:
            return
        try:
            exportar_liquidacion_excel(ruta, self.v_mes.get(), self.cfg)
            os.startfile(ruta)
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)


# ============================================================================
# MODULO MICE / EVENTOS - cotizacion de eventos corporativos (minuto a minuto)
# ============================================================================
MICE_PATH = os.path.join(datos_dir(), "mice.json")
MICE_BIB_PATH = os.path.join(datos_dir(), "mice_biblioteca.json")
CIUDADES_MICE = ["Cartagena", "Bogota", "Medellin", "Santa Marta", "Cali",
                 "Eje Cafetero", "San Andres", "Barranquilla", "Otra"]
ALIMENTACION_MICE = ["Desayuno", "Media pension", "Pension completa", "Todo incluido"]
CATEGORIAS_MICE_ITEM = ["Traslados", "Cenas", "Salones", "Audiovisuales",
                        "Personal", "Tours", "Decoracion", "Impuestos", "Otros"]


def cargar_mice():
    if os.path.exists(MICE_PATH):
        try:
            with open(MICE_PATH, "r", encoding="utf-8") as f:
                d = json.load(f)
            if isinstance(d, dict) and "items" in d:
                return d
        except Exception:
            pass
    return {"seq": 0, "items": []}


def guardar_mice(data):
    with open(MICE_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def registrar_mice(rec):
    data = cargar_mice()
    data["seq"] = int(data.get("seq", 0)) + 1
    rec = dict(rec); rec["numero"] = f"MICE-{data['seq']:05d}"
    if not rec.get("uid"):
        rec["uid"] = "MICE-" + uuid.uuid4().hex[:12]
    data["items"].append(rec); guardar_mice(data)
    return rec["numero"]


def actualizar_mice(numero, cambios):
    data = cargar_mice()
    for it in data["items"]:
        if it.get("numero") == numero:
            it.update(cambios); break
    guardar_mice(data)


def eliminar_mice(numero):
    data = cargar_mice()
    data["items"] = [x for x in data["items"] if x.get("numero") != numero]
    guardar_mice(data)


def total_opcion_mice(op):
    t = 0.0
    for ln in op.get("lineas", []):
        try:
            t += float(ln.get("cantidad", 0) or 0) * float(ln.get("unitario", 0) or 0)
        except Exception:
            pass
    return round(t, 2)


def total_mice(rec):
    ops = rec.get("opciones", [])
    return total_opcion_mice(ops[0]) if ops else 0.0


# ---- Biblioteca de items y hoteles frecuentes ----
def _mice_bib_semilla():
    return {"items": [
        {"categoria": "Traslados", "item": "Traslado in/out aeropuerto",
         "descripcion": "Traslado de llegada y de salida con asistencia en los aeropuertos", "unitario": 12},
        {"categoria": "Cenas", "item": "Cena inaugural",
         "descripcion": "Cena de bienvenida en restaurante - entrada, plato fuerte y postre", "unitario": 0},
        {"categoria": "Cenas", "item": "Cena de cierre",
         "descripcion": "Cena de cierre de evento con open bar de 2 horas", "unitario": 0},
        {"categoria": "Cenas", "item": "Cena en restaurante (3 tiempos)",
         "descripcion": "Cena en restaurante de la ciudad: entrada, plato fuerte y postre", "unitario": 0},
        {"categoria": "Salones", "item": "Alquiler de salon para evento",
         "descripcion": "Montaje del salon para el evento privado", "unitario": 0},
        {"categoria": "Audiovisuales", "item": "Video beam", "descripcion": "Alquiler de video beam", "unitario": 0},
        {"categoria": "Audiovisuales", "item": "Papelografo", "descripcion": "Alquiler de papelografo", "unitario": 0},
        {"categoria": "Audiovisuales", "item": "Podium", "descripcion": "Alquiler de podium", "unitario": 0},
        {"categoria": "Audiovisuales", "item": "Sonido y microfonia",
         "descripcion": "Equipo de sonido y microfonos para el evento", "unitario": 0},
        {"categoria": "Personal", "item": "Meseros", "descripcion": "Servicio de meseros para el evento", "unitario": 0},
        {"categoria": "Tours", "item": "City tour + Castillo de San Felipe",
         "descripcion": "Recorrido en transporte climatizado con guia profesional bilingue", "unitario": 0},
        {"categoria": "Tours", "item": "Chiva Rumbera",
         "descripcion": "Recorrido nocturno festivo por la ciudad", "unitario": 0},
        {"categoria": "Impuestos", "item": "Propina voluntaria 10%",
         "descripcion": "Propina voluntaria 10% sobre alimentos y bebidas", "unitario": 0},
    ], "hoteles": []}


def cargar_mice_bib():
    if os.path.exists(MICE_BIB_PATH):
        try:
            with open(MICE_BIB_PATH, "r", encoding="utf-8") as f:
                d = json.load(f)
            if isinstance(d, dict):
                d.setdefault("items", []); d.setdefault("hoteles", [])
                return d
        except Exception:
            pass
    d = _mice_bib_semilla()
    try:
        guardar_mice_bib(d)
    except Exception:
        pass
    return d


def guardar_mice_bib(d):
    with open(MICE_BIB_PATH, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)


def bib_agregar_item(item):
    d = cargar_mice_bib()
    for x in d["items"]:
        if x.get("item", "").strip().lower() == item.get("item", "").strip().lower():
            x.update(item); guardar_mice_bib(d); return
    d["items"].append(item); guardar_mice_bib(d)


def bib_agregar_hotel(h):
    d = cargar_mice_bib()
    for x in d["hoteles"]:
        if x.get("hotel", "").strip().lower() == h.get("hotel", "").strip().lower():
            x.update(h); guardar_mice_bib(d); return
    d["hoteles"].append(h); guardar_mice_bib(d)


# ---- PDF de la cotizacion MICE (una seccion por opcion) ----
class MICEPDF(CotizacionPDF):
    pass


def generar_pdf_mice(cfg, rec, ruta):
    pdf = MICEPDF(cfg); T = pdf._t
    ops = rec.get("opciones", []) or [{}]
    for i, op in enumerate(ops):
        pdf.add_page()
        pdf.ln(2)
        pdf.set_fill_color(*PDF_PRIM); pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 12)
        tit = op.get("nombre") or op.get("ciudad") or f"OPCION {i + 1}"
        pdf.cell(0, 9, T("COTIZACION MICE - " + tit), ln=1, fill=True, align="C")
        pdf.set_text_color(*PDF_TXT); pdf.ln(2)
        info = [("Empresa / Cliente", rec.get("empresa", "")), ("Contacto", rec.get("contacto", "")),
                ("Evento", rec.get("evento", "")), ("Ciudad", op.get("ciudad", "")),
                ("No. de pasajeros", str(rec.get("pax", "") or "")),
                ("Fechas del evento", rec.get("fechas_evento", "")),
                ("No. de cotizacion", rec.get("numero", "")), ("Fecha", rec.get("fecha", "")),
                ("Asesor", rec.get("cotizado_por", ""))]
        for et, val in info:
            if val:
                pdf.set_font("Helvetica", "B", 9); pdf.cell(46, 5.6, T(et + ":"), border=0)
                pdf.set_font("Helvetica", "", 9); pdf.multi_cell(0, 5.6, T(str(val)))
        # Hoteles
        hoteles = op.get("hoteles", [])
        if hoteles:
            pdf.ln(2); pdf.set_fill_color(*PDF_CLARO); pdf.set_text_color(*PDF_PRIM)
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(64, 7, T(" HOTEL"), border=1, fill=True)
            pdf.cell(30, 7, T("ALIMENTACION"), border=1, fill=True, align="C")
            pdf.cell(26, 7, T("CATEGORIA"), border=1, fill=True, align="C")
            pdf.cell(20, 7, T("SENCILLA"), border=1, fill=True, align="R")
            pdf.cell(20, 7, T("DOBLE"), border=1, fill=True, align="R")
            pdf.cell(0, 7, T("TRIPLE"), border=1, fill=True, align="R", ln=1)
            pdf.set_text_color(*PDF_TXT); pdf.set_font("Helvetica", "", 9)
            for h in hoteles:
                pdf.cell(64, 6.5, T(" " + str(h.get("hotel", ""))[:44]), border=1)
                pdf.cell(30, 6.5, T(str(h.get("alimentacion", ""))), border=1, align="C")
                pdf.cell(26, 6.5, T(str(h.get("categoria", ""))), border=1, align="C")
                pdf.cell(20, 6.5, T(usd(h.get("sencilla", 0))), border=1, align="R")
                pdf.cell(20, 6.5, T(usd(h.get("doble", 0))), border=1, align="R")
                pdf.cell(0, 6.5, T(usd(h.get("triple", 0))), border=1, align="R", ln=1)
            nn = hoteles[0].get("noches", "")
            if nn:
                pdf.set_font("Helvetica", "I", 8)
                pdf.cell(0, 5, T(f"Tarifa por {nn} noche(s), por persona."), ln=1)
        # Minuto a minuto
        pdf.ln(2); pdf.set_fill_color(*PDF_PRIM); pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(0, 7, T(" MINUTO A MINUTO"), ln=1, fill=True)
        pdf.set_fill_color(*PDF_CLARO); pdf.set_text_color(*PDF_PRIM); pdf.set_font("Helvetica", "B", 8.5)
        W = [22, 34, 66, 12, 22, 24]
        pdf.cell(W[0], 6.5, T(" FECHA"), border=1, fill=True)
        pdf.cell(W[1], 6.5, T("ITEM"), border=1, fill=True)
        pdf.cell(W[2], 6.5, T("DESCRIPCION"), border=1, fill=True)
        pdf.cell(W[3], 6.5, T("CANT"), border=1, fill=True, align="C")
        pdf.cell(W[4], 6.5, T("UNIT."), border=1, fill=True, align="R")
        pdf.cell(W[5], 6.5, T("TOTAL"), border=1, fill=True, align="R", ln=1)
        pdf.set_text_color(*PDF_TXT); pdf.set_font("Helvetica", "", 8.5)
        x0 = pdf.l_margin
        for ln in op.get("lineas", []):
            cant = float(ln.get("cantidad", 0) or 0); unit = float(ln.get("unitario", 0) or 0)
            tot = cant * unit
            y0 = pdf.get_y()
            pdf.set_xy(x0, y0); pdf.multi_cell(W[0], 5, T(" " + str(ln.get("dia", ""))), border=1)
            h1 = pdf.get_y() - y0
            pdf.set_xy(x0 + W[0], y0); pdf.multi_cell(W[1], 5, T(" " + str(ln.get("item", ""))), border=1)
            h2 = pdf.get_y() - y0
            pdf.set_xy(x0 + W[0] + W[1], y0); pdf.multi_cell(W[2], 5, T(" " + str(ln.get("descripcion", ""))), border=1)
            h3 = pdf.get_y() - y0
            hh = max(h1, h2, h3, 5)
            pdf.set_xy(x0 + W[0] + W[1] + W[2], y0); pdf.cell(W[3], hh, T(f"{cant:g}"), border=1, align="C")
            pdf.cell(W[4], hh, T(usd(unit)), border=1, align="R")
            pdf.cell(W[5], hh, T(usd(tot)), border=1, align="R", ln=1)
            pdf.set_y(y0 + hh)
        pdf.ln(1); pdf.set_font("Helvetica", "B", 11); pdf.set_text_color(*PDF_PRIM)
        pdf.cell(0, 8, T(f"TOTAL {tit}:   {usd(total_opcion_mice(op))}"), ln=1, align="R")
        notas = rec.get("notas", "")
        if notas:
            pdf.ln(1); pdf.set_font("Helvetica", "", 8.5); pdf.set_text_color(*PDF_TXT)
            pdf.multi_cell(0, 5, T("Notas: " + notas))
    pdf.output(ruta)
    return ruta


class SelectorMICEBib(ctk.CTkToplevel):
    """Elegir un item (o hotel) de la biblioteca para insertarlo en la cotizacion."""
    def __init__(self, master, tipo, on_pick):
        super().__init__(master)
        self.tipo = tipo; self.on_pick = on_pick
        self.title("Biblioteca de " + ("hoteles" if tipo == "hoteles" else "items"))
        self.geometry("640x560"); self.configure(fg_color=BG)
        self.transient(master); self.grab_set()
        ctk.CTkLabel(self, text="Biblioteca de " + ("hoteles" if tipo == "hoteles" else "items"),
                     text_color=NAVY, font=("Segoe UI", 16, "bold")).pack(anchor="w", padx=16, pady=(14, 2))
        self.q = tk.StringVar()
        e = ctk.CTkEntry(self, textvariable=self.q, height=36, corner_radius=10,
                         border_color=BLUE, border_width=2,
                         placeholder_text="Buscar...")
        e.pack(fill="x", padx=16, pady=(2, 8)); self.q.trace_add("write", lambda *a: self._pintar())
        self.box = ctk.CTkScrollableFrame(self, fg_color=CARD, corner_radius=12)
        self.box.pack(fill="both", expand=True, padx=16, pady=(0, 14))
        self._pintar(); self.after(120, e.focus_set)
        self.bind("<Escape>", lambda ev: self.destroy())

    def _pintar(self):
        for w in self.box.winfo_children():
            w.destroy()
        d = cargar_mice_bib()
        lst = d["hoteles"] if self.tipo == "hoteles" else d["items"]
        q = _nz(self.q.get())
        for x in lst:
            txt = x.get("hotel", "") if self.tipo == "hoteles" else x.get("item", "")
            sub = (f"{x.get('ciudad','')} · {x.get('categoria','')}" if self.tipo == "hoteles"
                   else f"{x.get('categoria','')} · {x.get('descripcion','')[:60]}")
            if q and q not in _nz(txt + " " + sub):
                continue
            row = ctk.CTkFrame(self.box, fg_color=CARD2, corner_radius=8)
            row.pack(fill="x", pady=2, padx=2)
            izq = ctk.CTkFrame(row, fg_color="transparent"); izq.pack(side="left", fill="x", expand=True, padx=8, pady=5)
            ctk.CTkLabel(izq, text=txt or "(sin nombre)", text_color=NAVY, anchor="w",
                         font=("Segoe UI", 12, "bold")).pack(anchor="w")
            ctk.CTkLabel(izq, text=sub, text_color=MUTED, anchor="w",
                         font=("Segoe UI", 10), wraplength=440, justify="left").pack(anchor="w")
            ctk.CTkButton(row, text="+ Insertar", width=100, height=30, fg_color=GREEN,
                          hover_color=GREEN_H, font=("Segoe UI", 11, "bold"),
                          command=lambda xx=x: self._pick(xx)).pack(side="right", padx=8)
        if not self.box.winfo_children():
            ctk.CTkLabel(self.box, text="Sin elementos. Los que guardes iran apareciendo aqui.",
                         text_color=MUTED).pack(pady=20)

    def _pick(self, x):
        try:
            self.on_pick(dict(x))
        finally:
            self.destroy()


class VentanaMICEDetalle(ctk.CTkToplevel):
    """Editor de una cotizacion MICE: datos del evento + varias opciones (minuto a minuto)."""
    def __init__(self, master, rec, cfg, on_save=None):
        super().__init__(master)
        self.res = rec; self.cfg = cfg; self.on_save = on_save
        self.title("Cotizacion MICE " + rec.get("numero", "nueva"))
        self.configure(fg_color=BG); self.geometry("1120x740+40+10")
        self.transient(master); self.grab_set()
        self.after(60, lambda: self._safe_zoom())
        bar = ctk.CTkFrame(self, fg_color=NAVY, corner_radius=0, height=56)
        bar.pack(side="top", fill="x"); bar.pack_propagate(False)
        ctk.CTkButton(bar, text="💾  Guardar", height=38, width=150, fg_color=GREEN,
                      hover_color=GREEN_H, font=("Segoe UI", 13, "bold"),
                      command=self._guardar).pack(side="left", padx=(16, 8), pady=9)
        ctk.CTkButton(bar, text="📄 Generar PDF", height=38, width=160, fg_color="#FFFFFF",
                      text_color=NAVY, hover_color="#E7EEF8", font=("Segoe UI", 12, "bold"),
                      command=self._pdf).pack(side="left", padx=(0, 8), pady=9)
        ctk.CTkButton(bar, text="Cerrar", height=38, width=100, fg_color=NAVY2,
                      hover_color=NAVY, font=("Segoe UI", 12, "bold"),
                      command=self.destroy).pack(side="right", padx=16, pady=9)
        self.lbl_tot = ctk.CTkLabel(bar, text="", text_color="#BBD0EC", font=("Segoe UI", 12, "bold"))
        self.lbl_tot.pack(side="right", padx=8)
        cont = ctk.CTkScrollableFrame(self, fg_color=BG)
        cont.pack(fill="both", expand=True, padx=16, pady=12)

        ctk.CTkLabel(cont, text="Datos del evento", text_color=NAVY,
                     font=("Segoe UI", 15, "bold")).pack(anchor="w")
        self.v = {}

        def campo(parent, clave, etq, valor=""):
            ctk.CTkLabel(parent, text=etq, text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2)
            var = tk.StringVar(value=str(valor)); self.v[clave] = var
            ctk.CTkEntry(parent, textvariable=var, height=32, corner_radius=8, border_color=LINE).pack(fill="x", pady=(0, 6))
            return var

        f1 = ctk.CTkFrame(cont, fg_color="transparent"); f1.pack(fill="x")
        a = ctk.CTkFrame(f1, fg_color="transparent"); a.pack(side="left", fill="x", expand=True, padx=(0, 6))
        b = ctk.CTkFrame(f1, fg_color="transparent"); b.pack(side="left", fill="x", expand=True, padx=(6, 0))
        campo(a, "empresa", "Empresa / Cliente", rec.get("empresa", ""))
        campo(b, "contacto", "Contacto", rec.get("contacto", ""))
        f2 = ctk.CTkFrame(cont, fg_color="transparent"); f2.pack(fill="x")
        a2 = ctk.CTkFrame(f2, fg_color="transparent"); a2.pack(side="left", fill="x", expand=True, padx=(0, 6))
        b2 = ctk.CTkFrame(f2, fg_color="transparent"); b2.pack(side="left", fill="x", expand=True, padx=(6, 0))
        campo(a2, "email", "Correo", rec.get("email", ""))
        campo(b2, "evento", "Nombre del evento", rec.get("evento", ""))
        f3 = ctk.CTkFrame(cont, fg_color="transparent"); f3.pack(fill="x")
        a3 = ctk.CTkFrame(f3, fg_color="transparent"); a3.pack(side="left", fill="x", expand=True, padx=(0, 6))
        b3 = ctk.CTkFrame(f3, fg_color="transparent"); b3.pack(side="left", fill="x", expand=True, padx=(6, 0))
        campo(a3, "pax", "No. de pasajeros", rec.get("pax", ""))
        campo(b3, "fechas_evento", "Fechas del evento", rec.get("fechas_evento", ""))
        f4 = ctk.CTkFrame(cont, fg_color="transparent"); f4.pack(fill="x")
        a4 = ctk.CTkFrame(f4, fg_color="transparent"); a4.pack(side="left", fill="x", expand=True, padx=(0, 6))
        b4 = ctk.CTkFrame(f4, fg_color="transparent"); b4.pack(side="left", fill="x", expand=True, padx=(6, 0))
        ctk.CTkLabel(a4, text="Cotizado por", text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.v_cotpor = tk.StringVar(value=rec.get("cotizado_por", "") or COTIZADORES[0][0])
        ctk.CTkOptionMenu(a4, variable=self.v_cotpor, values=[c[0] for c in COTIZADORES],
                          height=32, fg_color=NAVY, button_color=NAVY2).pack(fill="x", pady=(0, 6))
        ctk.CTkLabel(b4, text="Estado", text_color=MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=2)
        self.v_estado = tk.StringVar(value=rec.get("estado", "Pendiente"))
        ctk.CTkOptionMenu(b4, variable=self.v_estado, values=ESTADOS_COT,
                          height=32, fg_color=NAVY, button_color=NAVY2).pack(fill="x", pady=(0, 6))
        campo(cont, "notas", "Notas (salen en el PDF)", rec.get("notas", ""))

        oh = ctk.CTkFrame(cont, fg_color="transparent"); oh.pack(fill="x", pady=(8, 2))
        ctk.CTkLabel(oh, text="OPCIONES DE LA COTIZACION", text_color=NAVY,
                     font=("Segoe UI", 15, "bold")).pack(side="left")
        ctk.CTkButton(oh, text="+ Agregar opcion", width=150, height=30, fg_color=BLUE,
                      hover_color=NAVY, font=("Segoe UI", 11, "bold"),
                      command=self._agregar_opcion).pack(side="right")
        self.ops_box = ctk.CTkFrame(cont, fg_color="transparent"); self.ops_box.pack(fill="x")
        self.op_widgets = []
        if not self.res.get("opciones"):
            self.res["opciones"] = [self._opcion_vacia()]
        self._pintar_opciones()

    def _safe_zoom(self):
        try:
            self.state("zoomed")
        except Exception:
            pass

    def _opcion_vacia(self):
        return {"nombre": "", "ciudad": "Cartagena", "hoteles": [], "lineas": []}

    # ---------- render de opciones ----------
    def _pintar_opciones(self):
        for w in self.ops_box.winfo_children():
            w.destroy()
        self.op_widgets = []
        for i, op in enumerate(self.res.get("opciones", [])):
            self._render_opcion(i, op)
        self._recalc()

    def _render_opcion(self, i, op):
        card = ctk.CTkFrame(self.ops_box, fg_color=CARD, corner_radius=12, border_width=1, border_color=LINE)
        card.pack(fill="x", pady=6)
        top = ctk.CTkFrame(card, fg_color="transparent"); top.pack(fill="x", padx=10, pady=(8, 2))
        ctk.CTkLabel(top, text=f"Opcion {i + 1}", text_color=NAVY, font=("Segoe UI", 13, "bold")).pack(side="left")
        ctk.CTkButton(top, text="🗑 Quitar opcion", width=130, height=28, fg_color=CARD2, text_color=RED,
                      hover_color=LINE, border_width=1, border_color=LINE, font=("Segoe UI", 11),
                      command=lambda idx=i: self._quitar_opcion(idx)).pack(side="right")
        fr = ctk.CTkFrame(card, fg_color="transparent"); fr.pack(fill="x", padx=10)
        na = ctk.CTkFrame(fr, fg_color="transparent"); na.pack(side="left", fill="x", expand=True, padx=(0, 6))
        nc = ctk.CTkFrame(fr, fg_color="transparent"); nc.pack(side="left", padx=(6, 0))
        ctk.CTkLabel(na, text="Nombre de la opcion", text_color=MUTED, font=("Segoe UI", 10)).pack(anchor="w")
        v_nom = tk.StringVar(value=op.get("nombre", ""))
        ctk.CTkEntry(na, textvariable=v_nom, height=30, placeholder_text="Ej. OPCION CARTAGENA").pack(fill="x")
        ctk.CTkLabel(nc, text="Ciudad", text_color=MUTED, font=("Segoe UI", 10)).pack(anchor="w")
        v_ciu = tk.StringVar(value=op.get("ciudad", "Cartagena"))
        ctk.CTkOptionMenu(nc, variable=v_ciu, values=CIUDADES_MICE, width=160, height=30,
                          fg_color=NAVY, button_color=NAVY2).pack()
        ow = {"nombre": v_nom, "ciudad": v_ciu, "hoteles": [], "lineas": [],
              "hbox": None, "lbox": None, "tot": None}
        self.op_widgets.append(ow)
        # Hoteles
        hh = ctk.CTkFrame(card, fg_color="transparent"); hh.pack(fill="x", padx=10, pady=(8, 0))
        ctk.CTkLabel(hh, text="Hoteles (tarifa por persona)", text_color=NAVY,
                     font=("Segoe UI", 12, "bold")).pack(side="left")
        ctk.CTkButton(hh, text="+ Hotel", width=80, height=26, fg_color=BLUE, hover_color=NAVY,
                      font=("Segoe UI", 10, "bold"),
                      command=lambda idx=i: self._add_hotel(idx)).pack(side="right", padx=(4, 0))
        ctk.CTkButton(hh, text="📚 De biblioteca", width=120, height=26, fg_color=CARD2, text_color=NAVY,
                      hover_color=LINE, border_width=1, border_color=LINE, font=("Segoe UI", 10, "bold"),
                      command=lambda idx=i: self._hotel_bib(idx)).pack(side="right")
        ow["hbox"] = ctk.CTkFrame(card, fg_color="transparent"); ow["hbox"].pack(fill="x", padx=10)
        for h in op.get("hoteles", []):
            self._fila_hotel(ow, h)
        # Minuto a minuto
        lh = ctk.CTkFrame(card, fg_color="transparent"); lh.pack(fill="x", padx=10, pady=(8, 0))
        ctk.CTkLabel(lh, text="Minuto a minuto", text_color=NAVY,
                     font=("Segoe UI", 12, "bold")).pack(side="left")
        ctk.CTkButton(lh, text="+ Linea", width=80, height=26, fg_color=BLUE, hover_color=NAVY,
                      font=("Segoe UI", 10, "bold"),
                      command=lambda idx=i: self._add_linea(idx)).pack(side="right", padx=(4, 0))
        ctk.CTkButton(lh, text="📚 De biblioteca", width=120, height=26, fg_color=CARD2, text_color=NAVY,
                      hover_color=LINE, border_width=1, border_color=LINE, font=("Segoe UI", 10, "bold"),
                      command=lambda idx=i: self._linea_bib(idx)).pack(side="right")
        cab = ctk.CTkFrame(card, fg_color="transparent"); cab.pack(fill="x", padx=10)
        for txt, w in (("Dia", 70), ("Item", 150), ("Descripcion", 0), ("Cant", 52),
                       ("Unit.", 70), ("Total", 80), ("", 30)):
            ctk.CTkLabel(cab, text=txt, text_color=MUTED, font=("Segoe UI", 9, "bold"),
                         width=w, anchor="w").pack(side="left", fill=("x" if w == 0 else None),
                                                   expand=(w == 0), padx=1)
        ow["lbox"] = ctk.CTkFrame(card, fg_color="transparent"); ow["lbox"].pack(fill="x", padx=10)
        for ln in op.get("lineas", []):
            self._fila_linea(ow, ln)
        tb = ctk.CTkFrame(card, fg_color=CARD2, corner_radius=8); tb.pack(fill="x", padx=10, pady=(6, 10))
        ow["tot"] = ctk.CTkLabel(tb, text="Total opcion: USD 0.00", text_color=NAVY,
                                 font=("Segoe UI", 13, "bold"))
        ow["tot"].pack(side="right", padx=12, pady=6)

    def _fila_hotel(self, ow, h):
        row = ctk.CTkFrame(ow["hbox"], fg_color=CARD2, corner_radius=8); row.pack(fill="x", pady=2)
        v = {}
        v["hotel"] = tk.StringVar(value=h.get("hotel", ""))
        v["alimentacion"] = tk.StringVar(value=h.get("alimentacion", "Desayuno"))
        v["categoria"] = tk.StringVar(value=h.get("categoria", ""))
        v["noches"] = tk.StringVar(value=str(h.get("noches", "") or ""))
        v["sencilla"] = tk.StringVar(value=str(h.get("sencilla", "") or ""))
        v["doble"] = tk.StringVar(value=str(h.get("doble", "") or ""))
        v["triple"] = tk.StringVar(value=str(h.get("triple", "") or ""))
        ctk.CTkEntry(row, textvariable=v["hotel"], height=28, placeholder_text="Hotel").pack(side="left", fill="x", expand=True, padx=(6, 2), pady=4)
        ctk.CTkOptionMenu(row, variable=v["alimentacion"], values=ALIMENTACION_MICE, width=110, height=28,
                          fg_color=NAVY, button_color=NAVY2).pack(side="left", padx=2)
        ctk.CTkEntry(row, textvariable=v["categoria"], height=28, width=80, placeholder_text="Cat.").pack(side="left", padx=2)
        ctk.CTkEntry(row, textvariable=v["noches"], height=28, width=44, placeholder_text="Noch").pack(side="left", padx=2)
        ctk.CTkEntry(row, textvariable=v["sencilla"], height=28, width=60, placeholder_text="Sencilla").pack(side="left", padx=2)
        ctk.CTkEntry(row, textvariable=v["doble"], height=28, width=60, placeholder_text="Doble").pack(side="left", padx=2)
        ctk.CTkEntry(row, textvariable=v["triple"], height=28, width=60, placeholder_text="Triple").pack(side="left", padx=2)
        ctk.CTkButton(row, text="✕", width=26, height=26, fg_color=RED, hover_color="#9B2C22",
                      command=lambda: self._del_hotel(ow, v)).pack(side="left", padx=(2, 6))
        ow["hoteles"].append(v)

    def _fila_linea(self, ow, ln):
        row = ctk.CTkFrame(ow["lbox"], fg_color=CARD2, corner_radius=8); row.pack(fill="x", pady=2)
        v = {}
        v["dia"] = tk.StringVar(value=ln.get("dia", ""))
        v["item"] = tk.StringVar(value=ln.get("item", ""))
        v["descripcion"] = tk.StringVar(value=ln.get("descripcion", ""))
        v["cantidad"] = tk.StringVar(value=str(ln.get("cantidad", "") or ""))
        v["unitario"] = tk.StringVar(value=str(ln.get("unitario", "") or ""))
        ctk.CTkEntry(row, textvariable=v["dia"], height=28, width=70, placeholder_text="DIA 1").pack(side="left", padx=(6, 1), pady=4)
        ctk.CTkEntry(row, textvariable=v["item"], height=28, width=150, placeholder_text="Item").pack(side="left", padx=1)
        ctk.CTkEntry(row, textvariable=v["descripcion"], height=28, placeholder_text="Descripcion").pack(side="left", fill="x", expand=True, padx=1)
        ec = ctk.CTkEntry(row, textvariable=v["cantidad"], height=28, width=52, placeholder_text="Cant")
        ec.pack(side="left", padx=1); ec.bind("<KeyRelease>", lambda e: self._recalc())
        eu = ctk.CTkEntry(row, textvariable=v["unitario"], height=28, width=70, placeholder_text="Unit")
        eu.pack(side="left", padx=1); eu.bind("<KeyRelease>", lambda e: self._recalc())
        v["tot_lbl"] = ctk.CTkLabel(row, text="0.00", text_color=GREEN_H, width=80, anchor="e",
                                    font=("Segoe UI", 11, "bold"))
        v["tot_lbl"].pack(side="left", padx=2)
        b = ctk.CTkFrame(row, fg_color="transparent"); b.pack(side="left", padx=(2, 6))
        ctk.CTkButton(b, text="★", width=26, height=26, fg_color="#D9A400", hover_color="#B7791F",
                      command=lambda: self._guardar_linea_bib(v)).pack(side="left", padx=1)
        ctk.CTkButton(b, text="✕", width=26, height=26, fg_color=RED, hover_color="#9B2C22",
                      command=lambda: self._del_linea(ow, v)).pack(side="left", padx=1)
        ow["lineas"].append(v)

    # ---------- acciones ----------
    def _sync(self):
        ops = []
        for ow in self.op_widgets:
            hoteles = []
            for h in ow["hoteles"]:
                hoteles.append({"hotel": h["hotel"].get().strip(),
                                "alimentacion": h["alimentacion"].get(),
                                "categoria": h["categoria"].get().strip(),
                                "noches": _num(h["noches"].get()),
                                "sencilla": _num(h["sencilla"].get()),
                                "doble": _num(h["doble"].get()),
                                "triple": _num(h["triple"].get())})
            lineas = []
            for l in ow["lineas"]:
                lineas.append({"dia": l["dia"].get().strip(), "item": l["item"].get().strip(),
                               "descripcion": l["descripcion"].get().strip(),
                               "cantidad": _num(l["cantidad"].get()),
                               "unitario": _num(l["unitario"].get())})
            ops.append({"nombre": ow["nombre"].get().strip(), "ciudad": ow["ciudad"].get(),
                        "hoteles": hoteles, "lineas": lineas})
        self.res["opciones"] = ops
        self.res["empresa"] = self.v["empresa"].get().strip()
        self.res["contacto"] = self.v["contacto"].get().strip()
        self.res["email"] = self.v["email"].get().strip()
        self.res["evento"] = self.v["evento"].get().strip()
        self.res["pax"] = _num(self.v["pax"].get())
        self.res["fechas_evento"] = self.v["fechas_evento"].get().strip()
        self.res["notas"] = self.v["notas"].get().strip()
        self.res["cotizado_por"] = self.v_cotpor.get()
        self.res["estado"] = self.v_estado.get()

    def _recalc(self):
        gran = 0.0
        for ow in self.op_widgets:
            t = 0.0
            for l in ow["lineas"]:
                sub = _num(l["cantidad"].get()) * _num(l["unitario"].get())
                t += sub
                try:
                    l["tot_lbl"].configure(text=f"{sub:,.2f}")
                except Exception:
                    pass
            try:
                ow["tot"].configure(text=f"Total opcion: {usd(t)}")
            except Exception:
                pass
            if gran == 0.0:
                gran = t
        try:
            self.lbl_tot.configure(text="Total 1a opcion: " + usd(gran))
        except Exception:
            pass

    def _agregar_opcion(self):
        self._sync()
        self.res["opciones"].append(self._opcion_vacia())
        self._pintar_opciones()

    def _quitar_opcion(self, i):
        self._sync()
        try:
            del self.res["opciones"][i]
        except Exception:
            pass
        if not self.res["opciones"]:
            self.res["opciones"] = [self._opcion_vacia()]
        self._pintar_opciones()

    def _add_hotel(self, i):
        self._sync()
        self.res["opciones"][i].setdefault("hoteles", []).append(
            {"hotel": "", "alimentacion": "Desayuno", "categoria": "", "noches": "",
             "sencilla": "", "doble": "", "triple": ""})
        self._pintar_opciones()

    def _del_hotel(self, ow, v):
        ow["hoteles"] = [x for x in ow["hoteles"] if x is not v]
        self._sync(); self._pintar_opciones()

    def _hotel_bib(self, i):
        def pick(h):
            self._sync()
            self.res["opciones"][i].setdefault("hoteles", []).append({
                "hotel": h.get("hotel", ""), "alimentacion": h.get("alimentacion", "Desayuno"),
                "categoria": h.get("categoria", ""), "noches": h.get("noches", ""),
                "sencilla": h.get("sencilla", ""), "doble": h.get("doble", ""),
                "triple": h.get("triple", "")})
            self._pintar_opciones()
        SelectorMICEBib(self, "hoteles", pick)

    def _add_linea(self, i):
        self._sync()
        self.res["opciones"][i].setdefault("lineas", []).append(
            {"dia": "", "item": "", "descripcion": "", "cantidad": self.res.get("pax", "") or "", "unitario": ""})
        self._pintar_opciones()

    def _del_linea(self, ow, v):
        ow["lineas"] = [x for x in ow["lineas"] if x is not v]
        self._sync(); self._pintar_opciones()

    def _linea_bib(self, i):
        def pick(it):
            self._sync()
            self.res["opciones"][i].setdefault("lineas", []).append({
                "dia": "", "item": it.get("item", ""), "descripcion": it.get("descripcion", ""),
                "cantidad": self.res.get("pax", "") or "", "unitario": it.get("unitario", "")})
            self._pintar_opciones()
        SelectorMICEBib(self, "items", pick)

    def _guardar_linea_bib(self, v):
        it = {"categoria": "Otros", "item": v["item"].get().strip(),
              "descripcion": v["descripcion"].get().strip(),
              "unitario": _num(v["unitario"].get())}
        if not it["item"]:
            messagebox.showinfo("Biblioteca", "Escribe el nombre del item antes de guardarlo.", parent=self)
            return
        bib_agregar_item(it)
        messagebox.showinfo("Biblioteca", f"'{it['item']}' guardado en la biblioteca.", parent=self)

    def _guardar(self):
        self._sync()
        if self.res.get("numero"):
            actualizar_mice(self.res["numero"], self.res)
        else:
            self.res["fecha"] = datetime.date.today().strftime("%d/%m/%Y")
            registrar_mice(self.res)
        if self.on_save:
            self.on_save()
        messagebox.showinfo("Guardado", "Cotizacion MICE guardada.", parent=self)
        self.destroy()

    def _pdf(self):
        self._sync()
        if not self.res.get("numero"):
            self.res["fecha"] = datetime.date.today().strftime("%d/%m/%Y")
            registrar_mice(self.res)
            if self.on_save:
                self.on_save()
        ruta = filedialog.asksaveasfilename(
            title="Guardar cotizacion MICE (PDF)", defaultextension=".pdf",
            initialfile=f"MICE_{self.res.get('numero','')}_{_nz(self.res.get('empresa',''))[:20]}.pdf",
            filetypes=[("PDF", "*.pdf")])
        if not ruta:
            return
        try:
            generar_pdf_mice(self.cfg, self.res, ruta)
            os.startfile(ruta)
        except Exception as e:
            messagebox.showerror("Error al generar PDF", str(e), parent=self)


class ModuloMICE(ctk.CTkToplevel):
    """Ventana principal del modulo MICE / Eventos."""
    def __init__(self, master=None):
        super().__init__(master)
        ctk.set_appearance_mode("light")
        try:
            ctk.set_widget_scaling(0.85)
        except Exception:
            pass
        self.cfg = cargar_config()
        try:
            self.iconbitmap(recurso("app.ico"))
        except Exception:
            pass
        self.title(f"MICE / Eventos - INNOBA Colombia DMC   v{VERSION}")
        self.configure(fg_color=BG); self.geometry("1180x720")
        self._build()
        self.after(60, self._max)

    def _max(self):
        try:
            self.state("zoomed")
        except Exception:
            pass

    def _volver(self):
        lanz = self.master
        try:
            if lanz is not None and hasattr(lanz, "mice"):
                lanz.mice = None
        except Exception:
            pass
        try:
            self.destroy()
        except Exception:
            pass
        try:
            if lanz is not None:
                lanz.deiconify(); lanz.lift(); lanz.focus_force()
        except Exception:
            pass

    def _build(self):
        head = ctk.CTkFrame(self, fg_color=CARD, corner_radius=0, height=60)
        head.pack(fill="x"); head.pack_propagate(False)
        try:
            img = Image.open(recurso("logo_innoba.png")); w, h = img.size; hh = 38
            self.logo_img = ctk.CTkImage(light_image=img, size=(int(w * hh / h), hh))
            ctk.CTkLabel(head, image=self.logo_img, text="").pack(side="left", padx=(18, 12), pady=8)
        except Exception:
            ctk.CTkLabel(head, text="INNOBA", font=("Segoe UI", 20, "bold"), text_color=NAVY).pack(side="left", padx=18)
        tit = ctk.CTkFrame(head, fg_color="transparent"); tit.pack(side="left")
        ctk.CTkLabel(tit, text="Modulo MICE / Eventos", text_color=NAVY,
                     font=("Segoe UI", 17, "bold"), height=20).pack(anchor="w")
        ctk.CTkLabel(tit, text=f"Eventos corporativos  ·  v{VERSION}", text_color=MUTED,
                     font=("Segoe UI", 11), height=15).pack(anchor="w")
        hb = ctk.CTkFrame(head, fg_color="transparent"); hb.pack(side="right", padx=18)
        ctk.CTkButton(hb, text="⌂ Modulos", width=100, height=36, corner_radius=10, fg_color=NAVY,
                      hover_color=NAVY2, font=("Segoe UI", 12, "bold"), command=self._volver).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hb, text="+ Nueva cotizacion", width=170, height=36, corner_radius=10, fg_color=GREEN,
                      hover_color=GREEN_H, font=("Segoe UI", 12, "bold"), command=self._nueva).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hb, text="📚 Biblioteca", width=120, height=36, corner_radius=10, fg_color=CYAN,
                      hover_color=BLUE, font=("Segoe UI", 12, "bold"), command=self._biblioteca).pack(side="left", padx=(0, 8))
        ctk.CTkButton(hb, text="Datos de mi empresa", width=170, height=36, corner_radius=10, fg_color=NAVY2,
                      hover_color=NAVY, font=("Segoe UI", 12, "bold"), command=self._config_empresa).pack(side="left")

        bar = ctk.CTkFrame(self, fg_color="transparent"); bar.pack(fill="x", padx=18, pady=(10, 4))
        self.q = tk.StringVar()
        e = ctk.CTkEntry(bar, textvariable=self.q, height=36, corner_radius=10,
                         placeholder_text="Buscar por numero, empresa o evento...")
        e.pack(side="left", fill="x", expand=True); e.bind("<KeyRelease>", lambda ev: self._pintar())
        self.lbl_tot = ctk.CTkLabel(bar, text="", text_color=MUTED, font=("Segoe UI", 11)); self.lbl_tot.pack(side="right", padx=10)
        self.lista = ctk.CTkScrollableFrame(self, fg_color=BG); self.lista.pack(fill="both", expand=True, padx=18, pady=(4, 14))
        self._pintar()

    def _pintar(self):
        for w in self.lista.winfo_children():
            w.destroy()
        items = list(reversed(cargar_mice().get("items", [])))
        q = self.q.get().lower().strip()
        if q:
            items = [it for it in items if q in json.dumps(it, ensure_ascii=False).lower()]
        self.lbl_tot.configure(text=f"{len(items)} cotizacion(es) MICE")
        if not items:
            ctk.CTkLabel(self.lista, text="Aun no hay cotizaciones MICE. Crea una con '+ Nueva cotizacion'.",
                         text_color=MUTED).pack(pady=24)
            return
        for it in items:
            self._fila(it)

    def _fila(self, it):
        fila = ctk.CTkFrame(self.lista, fg_color=CARD, corner_radius=10, border_width=1, border_color=LINE)
        fila.pack(fill="x", pady=4, padx=2); fila.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(fila, text=it.get("numero", ""), text_color=NAVY, font=("Segoe UI", 13, "bold"),
                     width=110).grid(row=0, column=0, rowspan=2, padx=(12, 6), pady=8)
        info = ctk.CTkFrame(fila, fg_color="transparent"); info.grid(row=0, column=1, rowspan=2, sticky="w")
        ctk.CTkLabel(info, text=(it.get("empresa", "") or "(sin empresa)"), text_color=TEXT,
                     font=("Segoe UI", 13, "bold")).pack(anchor="w")
        ctk.CTkLabel(info, text=f"{it.get('evento','')}  ·  {it.get('pax','')} pax  ·  "
                     f"{len(it.get('opciones', []))} opcion(es)", text_color=MUTED,
                     font=("Segoe UI", 11)).pack(anchor="w")
        ctk.CTkLabel(fila, text=usd(total_mice(it)), text_color=NAVY,
                     font=("Segoe UI", 13, "bold")).grid(row=0, column=2, rowspan=2, padx=10)
        ctk.CTkLabel(fila, text=it.get("estado", "Pendiente"),
                     fg_color=ESTADO_COLOR.get(it.get("estado", "Pendiente"), MUTED), text_color="#FFFFFF",
                     corner_radius=6, font=("Segoe UI", 10, "bold")).grid(row=0, column=3, rowspan=2, padx=8, ipadx=8, ipady=3)
        btns = ctk.CTkFrame(fila, fg_color="transparent"); btns.grid(row=0, column=4, rowspan=2, padx=10)
        ctk.CTkButton(btns, text="Abrir", width=80, height=32, fg_color=NAVY, hover_color=NAVY2,
                      command=lambda x=it: self._abrir(x)).pack(side="left", padx=2)
        ctk.CTkButton(btns, text="PDF", width=64, height=32, fg_color=BLUE, hover_color=NAVY,
                      command=lambda x=it: self._pdf(x)).pack(side="left", padx=2)
        ctk.CTkButton(btns, text="🗑", width=36, height=32, fg_color=RED, hover_color="#9B2C22",
                      command=lambda x=it: self._eliminar(x)).pack(side="left", padx=2)

    def _nueva(self):
        rec = {"empresa": "", "contacto": "", "email": "", "evento": "", "pax": "",
               "fechas_evento": "", "cotizado_por": COTIZADORES[0][0], "moneda": "USD",
               "estado": "Pendiente", "notas": "", "opciones": []}
        VentanaMICEDetalle(self, rec, self.cfg, on_save=self._pintar)

    def _abrir(self, it):
        VentanaMICEDetalle(self, dict(it), self.cfg, on_save=self._pintar)

    def _pdf(self, it):
        ruta = filedialog.asksaveasfilename(
            title="Guardar cotizacion MICE (PDF)", defaultextension=".pdf",
            initialfile=f"MICE_{it.get('numero','')}.pdf", filetypes=[("PDF", "*.pdf")])
        if not ruta:
            return
        try:
            generar_pdf_mice(self.cfg, it, ruta); os.startfile(ruta)
        except Exception as e:
            messagebox.showerror("Error al generar PDF", str(e), parent=self)

    def _eliminar(self, it):
        if messagebox.askyesno("Eliminar", f"¿Eliminar la cotizacion MICE {it.get('numero','')}?", parent=self):
            eliminar_mice(it.get("numero", "")); self._pintar()

    def _biblioteca(self):
        SelectorMICEBib(self, "items", lambda x: None)

    def _config_empresa(self):
        VentanaEmpresa(self, self.cfg, lambda cfg: setattr(self, "cfg", cfg))


class Launcher(ctk.CTk):
    """Pantalla de inicio del .exe: permite elegir uno de los tres modulos
    (Cotizacion, Reservas, Comercial). Solo Cotizacion esta desarrollado; los
    otros dos quedan como modulos futuros."""

    MODULOS = [
        ("Cotizacion", "📄",
         "Crear, guardar y dar seguimiento a las cotizaciones. Genera el PDF y las\n"
         "importa desde la version HTML de los clientes.", GREEN, GREEN_H, True),
        ("Reservas", "🧳",
         "Convertir una cotizacion en reserva confirmada: proveedores, vouchers a\n"
         "cliente y proveedor, asignacion por asesor y control de estados.", NAVY, NAVY2, True),
        ("Comercial", "📊",
         "Tareas de gestion comercial (con checklist, cliente y responsable) e\n"
         "indicadores: ventas del mes, conversion y reservas.", CYAN, BLUE_H, True),
        ("MICE / Eventos", "🎤",
         "Cotizar eventos corporativos (MICE): hoteles, cenas, alquiler de salones\n"
         "con audiovisuales y tours, minuto a minuto y con varias opciones.", "#7A5AB5", "#63459A", True),
    ]

    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("light")
        try:
            ctk.set_widget_scaling(0.9)
        except Exception:
            pass
        self.title("INNOBA Colombia DMC  ·  Sistema de Gestion")
        self.configure(fg_color=BG)
        self.geometry("1220x640")
        self.minsize(1000, 560)
        try:
            self.iconbitmap(recurso("app.ico"))
        except Exception:
            pass
        self.cotizador = None
        self.reservas = None
        self.comercial = None
        self.mice = None
        self._construir()
        self._centrar()

    def _centrar(self):
        try:
            self.update_idletasks()
            w, h = 1220, 640
            x = (self.winfo_screenwidth() - w) // 2
            y = (self.winfo_screenheight() - h) // 2
            self.geometry(f"{w}x{h}+{max(0,x)}+{max(0,y)}")
        except Exception:
            pass

    def _construir(self):
        # Encabezado con banda de marca (navy) + logo grande
        head = ctk.CTkFrame(self, fg_color=NAVY, corner_radius=0, height=104)
        head.pack(fill="x"); head.pack_propagate(False)
        izq = ctk.CTkFrame(head, fg_color="transparent"); izq.pack(side="left", padx=30)
        try:
            img = Image.open(recurso("logo_innoba.png")); w, h = img.size; hh = 68
            self.logo_img = ctk.CTkImage(light_image=img, size=(int(w * hh / h), hh))
            chip = ctk.CTkFrame(izq, fg_color="#FFFFFF", corner_radius=12)
            chip.pack(side="left", pady=18)
            ctk.CTkLabel(chip, image=self.logo_img, text="").pack(padx=14, pady=8)
        except Exception:
            ctk.CTkLabel(izq, text="INNOBA", font=("Segoe UI", 30, "bold"),
                         text_color="#FFFFFF").pack(side="left", pady=18)
        der = ctk.CTkFrame(head, fg_color="transparent"); der.pack(side="right", padx=30)
        ctk.CTkLabel(der, text="Sistema de Gestion", text_color="#FFFFFF",
                     font=("Segoe UI", 16, "bold")).pack(anchor="e", pady=(30, 0))
        ctk.CTkLabel(der, text=f"INNOBA Colombia DMC  ·  v{VERSION}", text_color="#BBD0EC",
                     font=("Segoe UI", 12)).pack(anchor="e")

        # Titulo central
        ctk.CTkLabel(self, text="Bienvenido", text_color=NAVY,
                     font=("Segoe UI", 30, "bold")).pack(pady=(30, 2))
        ctk.CTkLabel(self, text="Elige el modulo con el que quieres trabajar",
                     text_color=MUTED, font=("Segoe UI", 15)).pack(pady=(0, 26))

        # Tarjetas de modulo
        cont = ctk.CTkFrame(self, fg_color="transparent")
        cont.pack(fill="both", expand=True, padx=30, pady=(0, 18))
        for i in range(len(self.MODULOS)):
            cont.grid_columnconfigure(i, weight=1, uniform="mod")
        cont.grid_rowconfigure(0, weight=1)
        for i, (nombre, icono, desc, col, colh, activo) in enumerate(self.MODULOS):
            self._tarjeta(cont, i, nombre, icono, desc, col, colh, activo)

        # Pie
        ctk.CTkLabel(self, text="INNOBA Colombia DMC   ·   Sistema interno",
                     text_color=MUTED, font=("Segoe UI", 11)).pack(pady=(0, 14))

    def _tarjeta(self, parent, col, nombre, icono, desc, color, colorh, activo):
        card = ctk.CTkFrame(parent, fg_color=CARD, corner_radius=20,
                            border_width=1, border_color=LINE)
        card.grid(row=0, column=col, padx=16, sticky="nsew")
        card.grid_rowconfigure(5, weight=1)
        card.grid_columnconfigure(0, weight=1)

        # Franja de color superior (barra de acento)
        acento = ctk.CTkFrame(card, fg_color=color, height=6, corner_radius=20)
        acento.grid(row=0, column=0, sticky="new", padx=26, pady=(0, 0))

        # Badge circular con el icono sobre tinte del color del modulo
        badge = ctk.CTkFrame(card, width=104, height=104, corner_radius=52,
                             fg_color=aclarar(color, 0.85))
        badge.grid(row=1, column=0, pady=(28, 8)); badge.grid_propagate(False)
        ctk.CTkLabel(badge, text=icono, font=("Segoe UI Emoji", 50)).place(
            relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(card, text=nombre, text_color=NAVY,
                     font=("Segoe UI", 22, "bold")).grid(row=2, column=0, pady=(0, 4))
        if activo:
            ctk.CTkLabel(card, text="  DISPONIBLE  ", text_color=GREEN_H,
                         fg_color="#E3F5EA", corner_radius=10,
                         font=("Segoe UI", 10, "bold")).grid(row=3, column=0, pady=(0, 8),
                                                             ipady=3)
        else:
            ctk.CTkLabel(card, text="  EN DESARROLLO  ", text_color="#B7791F",
                         fg_color="#FFF3C4", corner_radius=10,
                         font=("Segoe UI", 10, "bold")).grid(row=3, column=0, pady=(0, 8),
                                                             ipady=3)
        ctk.CTkLabel(card, text=desc, text_color=MUTED, justify="center",
                     wraplength=280, font=("Segoe UI", 12)).grid(row=4, column=0, padx=22, sticky="n")

        txt = "Abrir modulo  →" if activo else "Proximamente"
        btn = ctk.CTkButton(card, text=txt, height=46, corner_radius=12,
                            fg_color=color, hover_color=colorh,
                            font=("Segoe UI", 14, "bold"),
                            command=lambda n=nombre: self._abrir(n))
        btn.grid(row=5, column=0, padx=22, pady=(10, 24), sticky="ews")
        if not activo:
            btn.configure(fg_color="#CBD5E1", hover_color="#B8C4D6", text_color="#4B5563")

        # Efecto hover: resalta el borde de la tarjeta
        def _enter(_e, c=card, col=color):
            try: c.configure(border_color=col, border_width=2)
            except Exception: pass
        def _leave(_e, c=card):
            try: c.configure(border_color=LINE, border_width=1)
            except Exception: pass
        for wdg in (card, badge):
            wdg.bind("<Enter>", _enter); wdg.bind("<Leave>", _leave)

    def _abrir(self, nombre):
        if nombre == "Cotizacion":
            self._abrir_cotizacion()
        elif nombre == "Reservas":
            self._abrir_reservas()
        elif nombre == "Comercial":
            self._abrir_comercial()
        elif nombre == "MICE / Eventos":
            self._abrir_mice()
        else:
            messagebox.showinfo(
                nombre,
                f"El modulo '{nombre}' esta en desarrollo.\n\n"
                "Pronto podras usarlo desde aqui.")

    def _abrir_mice(self):
        try:
            if self.mice is not None and self.mice.winfo_exists():
                self.mice.deiconify(); self.mice.lift(); return
        except Exception:
            self.mice = None
        try:
            self.mice = ModuloMICE(self)
            self.mice.protocol("WM_DELETE_WINDOW", self._cerrar_mice)
        except Exception as e:
            self.mice = None
            messagebox.showerror("Error", f"No se pudo abrir MICE:\n{e}")

    def _cerrar_mice(self):
        try:
            if self.mice is not None:
                self.mice.destroy()
        except Exception:
            pass
        self.mice = None

    def _abrir_cotizacion(self):
        try:
            if self.cotizador is not None and self.cotizador.winfo_exists():
                self.cotizador.deiconify(); self.cotizador.lift(); return
        except Exception:
            self.cotizador = None
        self.withdraw()
        try:
            self.cotizador = App(self)
            self.cotizador.protocol("WM_DELETE_WINDOW", self._cerrar_cotizador)
        except Exception as e:
            self.cotizador = None
            self.deiconify()
            messagebox.showerror("Error", f"No se pudo abrir Cotizacion:\n{e}")

    def _cerrar_cotizador(self):
        try:
            if self.cotizador is not None:
                self.cotizador.destroy()
        except Exception:
            pass
        self.cotizador = None
        try:
            self.deiconify(); self.lift(); self.focus_force()
        except Exception:
            pass

    def _abrir_reservas(self):
        try:
            if self.reservas is not None and self.reservas.winfo_exists():
                self.reservas.deiconify(); self.reservas.lift(); return
        except Exception:
            self.reservas = None
        self.withdraw()
        try:
            self.reservas = ModuloReservas(self)
            self.reservas.protocol("WM_DELETE_WINDOW", self._cerrar_reservas)
        except Exception as e:
            self.reservas = None
            self.deiconify()
            messagebox.showerror("Error", f"No se pudo abrir Reservas:\n{e}")

    def _cerrar_reservas(self):
        try:
            if self.reservas is not None:
                self.reservas.destroy()
        except Exception:
            pass
        self.reservas = None
        try:
            self.deiconify(); self.lift(); self.focus_force()
        except Exception:
            pass

    def _abrir_comercial(self):
        try:
            if self.comercial is not None and self.comercial.winfo_exists():
                self.comercial.deiconify(); self.comercial.lift(); return
        except Exception:
            self.comercial = None
        self.withdraw()
        try:
            self.comercial = ModuloComercial(self)
            self.comercial.protocol("WM_DELETE_WINDOW", self._cerrar_comercial)
        except Exception as e:
            self.comercial = None
            self.deiconify()
            messagebox.showerror("Error", f"No se pudo abrir Comercial:\n{e}")

    def _cerrar_comercial(self):
        try:
            if self.comercial is not None:
                self.comercial.destroy()
        except Exception:
            pass
        self.comercial = None
        try:
            self.deiconify(); self.lift(); self.focus_force()
        except Exception:
            pass


if __name__ == "__main__":
    app = Launcher()
    app.mainloop()
